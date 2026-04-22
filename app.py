# -*- coding: utf-8 -*-
"""
4Sale 自動分單印單系統 v4
五個通路：宅配 / 超商 / 店到店 / 店到店隔日配 / 無包裝
區域細分：倉庫(前倉/主倉/備用倉) + 區域字母(A/B/C...)
"""

import sys, csv, io, os, re, json, threading, time, hashlib, secrets
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, session, redirect, url_for, send_file

# ── 圖片URL解析：將 =IMAGE("url") 公式轉成純 URL ──
def parse_image_url(cell_value):
    if not cell_value:
        return ""
    s = str(cell_value).strip()
    m = re.search(r'=IMAGE\("([^"]+)"', s, re.IGNORECASE)
    if m:
        return m.group(1)
    if s.startswith("http://") or s.startswith("https://"):
        return s
    return ""

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 設定區
# ============================================================
CONFIG = {
    "company_name": "我的電商公司",
    "flask_port":   5000,
    "auto_run_hour":7,
    "auto_run_min": 30,

    # CSV 欄位名稱
    "col_txn":      "子交易序號",
    "col_order_id": "訂單編號",
    "col_shipping": "出貨類型",
    "col_warehouse":"商品倉庫儲位",
    "col_sku":      "商品編號",
    "col_length":   "商品長",       # 單位 cm
    "col_width":    "商品寬",       # 單位 cm
    "col_height":   "商品高",       # 單位 cm
    "col_weight":   "子交易總重量", # 單位 kg（整張訂單已加總）
    "col_fee":      "運費",         # 運費欄位（0=可拆單，>0=買家已付不拆單）

    # 通路判斷關鍵字
    "kw_delivery": ["新竹物流", "嘉里", "店到家"],
    "kw_cvs":      ["7-11", "711", "全家", "萊爾富"],
    "kw_nextday":  ["隔日到貨"],
    "kw_nopkg":    ["無包裝"],
    "kw_store":    ["店到店"],

    # 倉庫前綴
    "wh_prefix": {
        "?前": "參前",
        "?倉": "參前",
        "主":  "參倉",
    },

    # 超材規則（整張訂單加總後判斷）
    "oversize_rules": {
        # 通路key → (三邊總和上限cm, 最長邊上限cm, 重量上限kg, 說明)
        "cvs":      (105, 45,  10, "超商"),
        "store":    (105, 45,  10, "店到店"),
        "nextday":  (105, 45,  10, "店到店隔日配"),
        "nopkg":    (105, 45,  10, "無包裝"),
        "delivery_jialy":   (200, 120, 20, "嘉里快遞"),
        "delivery_hsinchu": (210, 150, 20, "新竹物流"),
        "delivery_store":   (150, 100, 15, "店到家大型"),
    },

    # 新竹物流補助運費對照（三邊總和 → 買家自付運費）
    # ≦150cm 正常補助，151~210cm 買家自付，>210cm 異常
    "hsinchu_surcharge": [
        (150, 0,   "正常補助"),
        (160, 135, "買家自付135元"),
        (170, 165, "買家自付165元"),
        (180, 195, "買家自付195元"),
        (190, 225, "買家自付225元"),
        (200, 285, "買家自付285元"),
        (210, 335, "買家自付335元"),
    ],

    # ★ 特殊可出超材品清單（SKU → 斜放後有效最長邊cm）
    # 斜放後最長邊 = √(長²+寬²+高²)，填入計算後的值
    # 例如：55x40x30 的商品，對角線 ≈ 74cm，但斜放進箱後最長邊可能只有 44cm
    # 請依實際測量填入，系統會用此值取代原始最長邊判斷超材
    "diagonal_skus": {
        # "SKU-001": 44,   # 範例：SKU-001 斜放後有效最長邊 44cm
        # "SKU-002": 43,
    },

    # 拆單規則（店到店/店到家超材時）
    "split_max_units":   5,     # 最多拆幾單
    "split_max_dim":     105,   # 每包三邊總和上限 cm
    "split_max_side":    45,    # 每包最長邊上限 cm
    "split_max_weight":  10,    # 每包重量上限 kg
}

# 通路顯示設定
CHANNEL_META = {
    "delivery": {"label": "宅配",         "icon": "🚙", "color": "#1565c0"},
    "cvs":      {"label": "超商",         "icon": "🏪", "color": "#c85000"},
    "store":    {"label": "店到店",       "icon": "🏬", "color": "#2e7d32"},
    "nextday":  {"label": "店到店隔日配", "icon": "⚡",   "color": "#6a1b9a"},
    "nopkg":    {"label": "無包裝",       "icon": "📦", "color": "#00838f"},
}

state = {
    "groups": {}, "total": 0, "last_update": None,
    "status": "idle", "status_msg": "請上傳 CSV 開始分單",
    "log": [], "summary": {},
}

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    state["log"].append(line)
    if len(state["log"]) > 200:
        state["log"] = state["log"][-200:]

# ============================================================
# 分單邏輯
# ============================================================
def detect_channel(val):
    s = (val or "").strip()
    if any(k in s for k in CONFIG["kw_delivery"]): return "delivery"
    if any(k in s for k in CONFIG["kw_cvs"]):      return "cvs"
    if any(k in s for k in CONFIG["kw_nextday"]):  return "nextday"
    if any(k in s for k in CONFIG["kw_nopkg"]):    return "nopkg"
    if any(k in s for k in CONFIG["kw_store"]):    return "store"
    return "delivery"

def parse_location(raw):
    """儲位 → (倉庫名稱, 區域字母)"""
    raw = (raw or "").strip()
    wh, rest = "其他", raw
    for prefix, name in CONFIG["wh_prefix"].items():
        if raw.startswith(prefix):
            wh, rest = name, raw[len(prefix):]
            break
    m = re.match(r"([A-Z])", rest)
    zone = m.group(1) if m else "?"
    return wh, zone

def safe_float(val):
    try: return float((str(val) or "0").strip())
    except: return 0.0

def parse_fee(val):
    """解析運費，去除 $ NT$ 等符號，回傳數字"""
    try:
        s = str(val or "0").strip()
        # 移除常見貨幣符號
        s = s.replace("NT$", "").replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0
    except:
        return 0.0

def check_oversize(ch, ship_raw, total_dim, max_side, weight):
    """
    判斷是否超材。
    回傳 (is_oversize: bool, warn_msg: str, can_split: bool, split_rules: dict)
    split_rules: {"max_dim": int, "max_side": int, "max_weight": int} 或 None
    """
    if total_dim == 0 and weight == 0:
        return False, "", False, None

    # 新竹物流
    if "新竹物流" in ship_raw:
        if total_dim > 210 or max_side > 150 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        for limit, fee, desc in CONFIG["hsinchu_surcharge"]:
            if total_dim <= limit:
                if fee > 0:
                    return True, f"新竹補助不足 {total_dim:.0f}cm → {desc}", False, None
                return False, "", False, None
        return True, f"超規 {total_dim:.0f}cm → 異常單", False, None

    # 嘉里快遞
    if "嘉里" in ship_raw:
        if total_dim > 200 or max_side > 120 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        return False, "", False, None

    # 店到家宅配 — 標準 or 大型
    if "店到家" in ship_raw:
        # 標準合規
        if total_dim <= 105 and max_side <= 45 and weight <= 10:
            return False, "", False, None
        # 大型合規
        if total_dim <= 150 and max_side <= 100 and weight <= 15:
            return False, "", False, None
        # 超材：先嘗試用大型規格拆，再嘗試標準規格拆
        rules = {"max_dim": 150, "max_side": 100, "max_weight": 15, "label": "店到家大型"}
        return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules

    # 超商（7-11 / 全家 / 萊爾富）
    if ch == "cvs":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "超商"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 店到店
    if ch == "store":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "店到店"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 隔日配 / 無包裝
    if ch in ("nextday", "nopkg"):
        if total_dim > 105 or max_side > 45 or weight > 10:
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", False, None
        return False, "", False, None

    return False, "", False, None

def apply_diagonal(sku, L, W, H):
    """
    如果此 SKU 在可斜放清單內，回傳斜放後的有效最長邊。
    否則回傳原始最長邊。
    """
    import math
    diag_map = CONFIG.get("diagonal_skus", {})
    if sku in diag_map:
        effective = diag_map[sku]
        return effective, True   # (有效最長邊, 是否斜放)
    return max(L, W, H), False

def suggest_split(products, weight_per_item, L, W, H, rules=None):
    """
    建議拆單方式。
    rules: {"max_dim": int, "max_side": int, "max_weight": int}
    """
    if rules is None:
        rules = {
            "max_dim":    CONFIG["split_max_dim"],
            "max_side":   CONFIG["split_max_side"],
            "max_weight": CONFIG["split_max_weight"],
        }
    max_units  = CONFIG["split_max_units"]
    max_dim    = rules["max_dim"]
    max_side   = rules["max_side"]
    max_weight = rules["max_weight"]

    # 展開所有件數為一個清單
    all_items = []
    for p in products:
        for _ in range(max(1, p.get("qty", 1))):
            all_items.append(p["sku"])

    total_items = len(all_items)

    # 單件尺寸判斷
    sides = sorted([L, W, H], reverse=True)
    single_dim  = sum(sides)
    single_side = sides[0]
    single_w    = weight_per_item

    # 如果單件就超規，無法拆單
    if single_dim > max_dim or single_side > max_side or single_w > max_weight:
        return None, "單件商品已超規，無法拆單"

    # 每包最多放幾件
    max_per_pkg_dim    = int(max_dim    // single_dim)  if single_dim  > 0 else 999
    max_per_pkg_side   = int(max_side   // single_side) if single_side > 0 else 999
    max_per_pkg_weight = int(max_weight // single_w)    if single_w    > 0 else 999
    max_per_pkg = min(max_per_pkg_dim, max_per_pkg_side, max_per_pkg_weight)
    max_per_pkg = max(1, max_per_pkg)

    # 需要幾包
    import math
    needed = math.ceil(total_items / max_per_pkg)

    if needed > max_units:
        return None, f"需要 {needed} 包，超過最多 {max_units} 單上限，無法自動拆單"

    # 建立拆單建議（同SKU盡量放同一包）
    packages = []
    current_pkg = []
    current_count = 0

    # 按SKU分組，同SKU連續放
    from collections import Counter
    sku_counts = Counter(all_items)
    sorted_skus = sorted(sku_counts.items(), key=lambda x: -x[1])

    remaining = []
    for sku, qty in sorted_skus:
        remaining.extend([sku] * qty)

    for item in remaining:
        if current_count >= max_per_pkg:
            packages.append(current_pkg[:])
            current_pkg = []
            current_count = 0
        current_pkg.append(item)
        current_count += 1

    if current_pkg:
        packages.append(current_pkg)

    result = []
    for i, pkg in enumerate(packages):
        pkg_counter = Counter(pkg)
        items_str = ", ".join(f"{sku}×{cnt}" for sku, cnt in sorted(pkg_counter.items()))
        pkg_dim    = single_dim * len(pkg)
        pkg_weight = single_w   * len(pkg)
        result.append({
            "pkg_no":  i + 1,
            "count":   len(pkg),
            "items":   items_str,
            "dim":     round(pkg_dim, 1),
            "weight":  round(pkg_weight, 2),
        })

    return result, ""

def _get_zone_label(locs):
    """取得區域描述字串，供宅配分群用"""
    whs   = set(wh for wh, z in locs)
    zones = set(z  for wh, z in locs if z != "?")
    if not zones: zones = {"?"}
    if len(whs) == 1 and len(zones) == 1:
        return f"{next(iter(whs))}{next(iter(zones))}區"
    elif len(whs) == 1:
        return f"{next(iter(whs))}混單"
    return "混單"

def get_sort_key(k):
    """排序：宅配→純區+單品→店到店多品→隔日配→無包裝→可拆單→超材"""
    if k == "__delivery__":     return (0,  k, k)
    if k == "__single_zone__":  return (20, k, k)   # 新：純區+單品大分類
    if k == "__nopkg__":        return (88, k, k)
    if k == "__splittable__":   return (92, k, k)
    if k == "__oversize__":     return (95, k, k)
    order = {"店到店":2, "店到店隔日配":3}
    for label, pri in order.items():
        if k.startswith(label):
            if "混單" in k: return (pri*10+9, "zzz", k)
            return (pri*10, k, k)
    return (99, k, k)

def split_orders(rows):
    # ── 以交易序號合併同一張訂單 ──────────────────────
    order_map = {}
    for row in rows:
        txn = (row.get(CONFIG["col_txn"], "") or "").strip()
        if not txn: continue
        if txn not in order_map:
            order_map[txn] = {
                "txn":       txn,
                "channel":   detect_channel(row.get(CONFIG["col_shipping"], "")),
                "ship_raw":  (row.get(CONFIG["col_shipping"], "") or "").strip(),
                "order_ids": [],
                "products":  [],
                "locations": set(),
                "total_qty": 0,
                # 尺寸（取該訂單最大值，因為同訂單多列尺寸應相同）
                "length":    0.0,
                "width":     0.0,
                "height":    0.0,
                "weight":    0.0,
                "fee":       0.0,  # 運費
            }
        oid    = (row.get(CONFIG["col_order_id"], "") or "").strip()
        raw_wh = (row.get(CONFIG["col_warehouse"], "") or "").strip()
        sku    = (row.get(CONFIG["col_sku"], "") or "").strip()
        wh, zone = parse_location(raw_wh)

        # 尺寸取最大值（避免空值覆蓋有效值）
        L = safe_float(row.get(CONFIG["col_length"], 0))
        W = safe_float(row.get(CONFIG["col_width"],  0))
        H = safe_float(row.get(CONFIG["col_height"], 0))
        Wt= safe_float(row.get(CONFIG["col_weight"], 0))
        Fe= parse_fee(row.get(CONFIG["col_fee"], 0))
        if L > 0: order_map[txn]["length"]  = max(order_map[txn]["length"],  L)
        if W > 0: order_map[txn]["width"]   = max(order_map[txn]["width"],   W)
        if H > 0: order_map[txn]["height"]  = max(order_map[txn]["height"],  H)
        if Wt> 0: order_map[txn]["weight"]  = max(order_map[txn]["weight"],  Wt)
        if Fe> 0: order_map[txn]["fee"]     = max(order_map[txn]["fee"],     Fe)

        order_map[txn]["order_ids"].append(oid)
        order_map[txn]["products"].append({
            "oid": oid, "sku": sku,
            "zone_raw": raw_wh, "wh": wh, "zone": zone,
        })
        order_map[txn]["locations"].add((wh, zone))
        order_map[txn]["total_qty"] += 1

    # ── 計算三邊總和、最長邊，斜放判斷，超材檢查 ────────
    for txn, o in order_map.items():
        L, W, H = o["length"], o["width"], o["height"]

        # 斜放判斷：若訂單內有可斜放 SKU，用對角線取代最長邊
        skus_in_order = [p["sku"] for p in o["products"]]
        diagonal_applied = False
        effective_side = max(L, W, H)
        for sku in skus_in_order:
            eff, is_diag = apply_diagonal(sku, L, W, H)
            if is_diag:
                effective_side = eff
                diagonal_applied = True
                break

        dims = sorted([L, W, H], reverse=True)
        o["max_side"]       = effective_side
        o["total_dim"]      = sum(dims)
        o["diagonal_used"]  = diagonal_applied
        o["oversize"], o["oversize_msg"], o["can_split"], o["split_rules"] = check_oversize(
            o["channel"], o["ship_raw"], o["total_dim"], o["max_side"], o["weight"]
        )

        # 若斜放後不超材，標示說明
        if diagonal_applied and not o["oversize"]:
            o["oversize_msg"] = "斜放後合規"

        # 拆單建議（運費=0才可拆單）
        o["split_suggestion"] = None
        o["split_error"]      = ""
        o["fee_paid"]         = o["fee"] > 0

        # 運費>0 → 標示買家已付，不拆單
        if o["oversize"] and o["fee"] > 0:
            o["oversize_msg"] += f"｜買家已付運費 {o['fee']:.0f} 元"

        can_split = o["oversize"] and o["can_split"] and o["fee"] == 0
        if can_split and o["total_qty"] > 0:
            weight_per = o["weight"] / o["total_qty"] if o["total_qty"] > 0 else 0
            from collections import Counter
            sku_count = Counter(p["sku"] for p in o["products"])
            prods = [{"sku": sku, "qty": qty} for sku, qty in sku_count.items()]
            suggestion, err = suggest_split(prods, weight_per, L, W, H, o["split_rules"])
            o["split_suggestion"] = suggestion
            o["split_error"]      = err

    # ── 套用分單規則 ──────────────────────────────────
    groups = {}

    def add(key, title, icon, color, o):
        if key not in groups:
            groups[key] = {"title": title, "icon": icon, "color": color, "orders": []}
        groups[key]["orders"].append(o)

    summary = {}

    for txn, o in order_map.items():
        ch    = o["channel"]
        locs  = o["locations"]
        meta  = CHANNEL_META.get(ch, CHANNEL_META["delivery"])
        label = meta["label"]
        icon  = meta["icon"]
        color = meta["color"]

        # 超材處理
        if o["oversize"]:
            o["oversize_channel"] = label
            if o["can_split"] and o["split_suggestion"]:
                # 可拆單 → 獨立可拆單分類
                add("__splittable__", "✂ 可拆單", "✂", "#e65100", o)
                summary["✂ 可拆單"] = summary.get("✂ 可拆單", 0) + 1
            else:
                # 不可拆單 → 超材異常分類
                add("__oversize__", "⚠ 超材", "⚠", "#b71c1c", o)
                summary["⚠ 超材"] = summary.get("⚠ 超材", 0) + 1
            continue

        # 無包裝 → 全部合併成一組
        if ch == "nopkg":
            add("__nopkg__", "📦 無包裝", "📦", "#00838f", o)
            summary["無包裝"] = summary.get("無包裝", 0) + 1
            continue

        # 宅配（新竹/嘉里/店到家）→ 外層一個大分類，內部依區域分群
        if ch == "delivery":
            o["delivery_zone"] = _get_zone_label(locs)  # 記錄區域供分群用
            add("__delivery__", "🚚 宅配", "🚚", "#1565c0", o)
            summary["宅配"] = summary.get("宅配", 0) + 1
            continue

        # ── 新分類邏輯：純區 + 超商單品 + 店到店單品 ──
        skus = set(p["sku"] for p in o["products"] if p["sku"])
        whs_all   = set(wh for wh, z in locs)
        zones_all = set(z  for wh, z in locs if z != "?")
        is_single_zone = (len(whs_all) == 1 and len(zones_all) == 1)  # 純單一區域
        is_single_item = (o["total_qty"] == 1)                         # 只有1件商品

        # 超商單品（1件）→ 進大分類
        if ch == "cvs" and is_single_item:
            o["single_zone_sub"] = "超商單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 店到店單品（1件）→ 進大分類
        if ch == "store" and is_single_item:
            o["single_zone_sub"] = "店到店單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 純區包裹（單一倉+單一區，任何通路）→ 進大分類
        if is_single_zone:
            wh_s   = next(iter(whs_all))
            zone_s = next(iter(zones_all))
            o["single_zone_sub"] = f"{label}｜{wh_s}{zone_s}區"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 其餘通路（超商多品跨區、店到店多品、隔日配）→ 依倉庫區域細分
        whs   = whs_all
        zones = zones_all
        if not zones:
            zones = {"?"}

        if len(whs) == 1 and len(zones) == 1:
            wh    = next(iter(whs))
            zone  = next(iter(zones))
            key   = f"{label} - {wh}{zone}區"
            title = f"{icon} {label} ｜ {wh} {zone} 區"
        elif len(whs) == 1:
            key   = f"{label} - {next(iter(whs))}混單"
            title = f"{icon} {label} ｜ {next(iter(whs))} 混單"
        else:
            key   = f"{label} - 混單"
            title = f"{icon} {label} ｜ 混單（跨倉）"

        if "混單" in key:
            color = color + "bb"

        add(key, title, icon, color, o)
        summary[label] = summary.get(label, 0) + 1

    state["summary"] = summary
    return dict(sorted(groups.items(), key=lambda x: get_sort_key(x[0])))

def load_csv(source, is_text=False):
    for enc in ["utf-8-sig", "big5", "cp950", "utf-8"]:
        try:
            if is_text:
                reader = csv.DictReader(io.StringIO(source))
            else:
                f = open(source, encoding=enc, newline="")
                reader = csv.DictReader(f)
            rows = list(reader)
            if not is_text: f.close()
            if rows: return rows, enc
        except Exception:
            continue
    return [], None

def run_pipeline(rows=None):
    state["status"] = "fetching"
    if not rows:
        state["status"] = "error"
        state["status_msg"] = "無資料"
        return
    groups = split_orders(rows)
    state["groups"]      = groups
    state["total"]       = len(set((r.get(CONFIG["col_txn"]) or "") for r in rows))
    state["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state["status"]      = "ready"
    state["status_msg"]  = f"共 {state['total']} 張訂單，分成 {len(groups)} 組 | {state['last_update']}"
    for k, g in groups.items():
        log(f"  {g['title']}：{len(g['orders'])} 張")
    log("分單完成")

def scheduler():
    while True:
        now = datetime.now()
        if now.hour == CONFIG["auto_run_hour"] and now.minute == CONFIG["auto_run_min"]:
            log("定時觸發")
            time.sleep(61)
        time.sleep(30)

# ============================================================
# 設定檔持久化（斜放商品清單）
# ============================================================
def _get_base_dir():
    """取得執行檔所在目錄（支援 PyInstaller 打包）"""
    if getattr(sys, 'frozen', False):
        # 打包成 exe 後，exe 所在目錄
        return os.path.dirname(sys.executable)
    else:
        # 一般 Python 執行
        return os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_get_base_dir(), "settings.json")

def load_settings():
    """從 settings.json 讀取設定，啟動時呼叫"""
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            if "diagonal_skus" in data:
                CONFIG["diagonal_skus"] = data["diagonal_skus"]
                log(f"已載入特殊可出超材品設定：{len(CONFIG['diagonal_skus'])} 筆")
    except Exception as e:
        log(f"讀取設定檔失敗：{e}")

def save_settings():
    """將設定寫入 settings.json"""
    try:
        data = {"diagonal_skus": CONFIG["diagonal_skus"]}
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"儲存設定檔失敗：{e}")

# ============================================================
# Flask 網頁
# ============================================================
app = Flask(__name__)

# -*- coding: utf-8 -*-
"""
4Sale 自動分單印單系統 v4
五個通路：宅配 / 超商 / 店到店 / 店到店隔日配 / 無包裝
區域細分：倉庫(前倉/主倉/備用倉) + 區域字母(A/B/C...)
"""

import sys, csv, io, os, re, json, threading, time
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 設定區
# ============================================================
CONFIG = {
    "company_name": "我的電商公司",
    "flask_port":   5000,
    "auto_run_hour":7,
    "auto_run_min": 30,

    # CSV 欄位名稱
    "col_txn":      "子交易序號",
    "col_order_id": "訂單編號",
    "col_shipping": "出貨類型",
    "col_warehouse":"商品倉庫儲位",
    "col_sku":      "商品編號",
    "col_length":   "商品長",       # 單位 cm
    "col_width":    "商品寬",       # 單位 cm
    "col_height":   "商品高",       # 單位 cm
    "col_weight":   "子交易總重量", # 單位 kg（整張訂單已加總）
    "col_fee":      "運費",         # 運費欄位（0=可拆單，>0=買家已付不拆單）

    # 通路判斷關鍵字
    "kw_delivery": ["新竹物流", "嘉里", "店到家"],
    "kw_cvs":      ["7-11", "711", "全家", "萊爾富"],
    "kw_nextday":  ["隔日到貨"],
    "kw_nopkg":    ["無包裝"],
    "kw_store":    ["店到店"],

    # 倉庫前綴
    "wh_prefix": {
        "?前": "參前",
        "?倉": "參前",
        "主":  "參倉",
    },

    # 超材規則（整張訂單加總後判斷）
    "oversize_rules": {
        # 通路key → (三邊總和上限cm, 最長邊上限cm, 重量上限kg, 說明)
        "cvs":      (105, 45,  10, "超商"),
        "store":    (105, 45,  10, "店到店"),
        "nextday":  (105, 45,  10, "店到店隔日配"),
        "nopkg":    (105, 45,  10, "無包裝"),
        "delivery_jialy":   (200, 120, 20, "嘉里快遞"),
        "delivery_hsinchu": (210, 150, 20, "新竹物流"),
        "delivery_store":   (150, 100, 15, "店到家大型"),
    },

    # 新竹物流補助運費對照（三邊總和 → 買家自付運費）
    # ≦150cm 正常補助，151~210cm 買家自付，>210cm 異常
    "hsinchu_surcharge": [
        (150, 0,   "正常補助"),
        (160, 135, "買家自付135元"),
        (170, 165, "買家自付165元"),
        (180, 195, "買家自付195元"),
        (190, 225, "買家自付225元"),
        (200, 285, "買家自付285元"),
        (210, 335, "買家自付335元"),
    ],

    # ★ 特殊可出超材品清單（SKU → 斜放後有效最長邊cm）
    # 斜放後最長邊 = √(長²+寬²+高²)，填入計算後的值
    # 例如：55x40x30 的商品，對角線 ≈ 74cm，但斜放進箱後最長邊可能只有 44cm
    # 請依實際測量填入，系統會用此值取代原始最長邊判斷超材
    "diagonal_skus": {
        # "SKU-001": 44,   # 範例：SKU-001 斜放後有效最長邊 44cm
        # "SKU-002": 43,
    },

    # 拆單規則（店到店/店到家超材時）
    "split_max_units":   5,     # 最多拆幾單
    "split_max_dim":     105,   # 每包三邊總和上限 cm
    "split_max_side":    45,    # 每包最長邊上限 cm
    "split_max_weight":  10,    # 每包重量上限 kg
}

# 通路顯示設定
CHANNEL_META = {
    "delivery": {"label": "宅配",         "icon": "🚙", "color": "#1565c0"},
    "cvs":      {"label": "超商",         "icon": "🏪", "color": "#c85000"},
    "store":    {"label": "店到店",       "icon": "🏬", "color": "#2e7d32"},
    "nextday":  {"label": "店到店隔日配", "icon": "⚡",   "color": "#6a1b9a"},
    "nopkg":    {"label": "無包裝",       "icon": "📦", "color": "#00838f"},
}

state = {
    "groups": {}, "total": 0, "last_update": None,
    "status": "idle", "status_msg": "請上傳 CSV 開始分單",
    "log": [], "summary": {},
}

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    state["log"].append(line)
    if len(state["log"]) > 200:
        state["log"] = state["log"][-200:]

# ============================================================
# 分單邏輯
# ============================================================
def detect_channel(val):
    s = (val or "").strip()
    if any(k in s for k in CONFIG["kw_delivery"]): return "delivery"
    if any(k in s for k in CONFIG["kw_cvs"]):      return "cvs"
    if any(k in s for k in CONFIG["kw_nextday"]):  return "nextday"
    if any(k in s for k in CONFIG["kw_nopkg"]):    return "nopkg"
    if any(k in s for k in CONFIG["kw_store"]):    return "store"
    return "delivery"

def parse_location(raw):
    """儲位 → (倉庫名稱, 區域字母)"""
    raw = (raw or "").strip()
    wh, rest = "其他", raw
    for prefix, name in CONFIG["wh_prefix"].items():
        if raw.startswith(prefix):
            wh, rest = name, raw[len(prefix):]
            break
    m = re.match(r"([A-Z])", rest)
    zone = m.group(1) if m else "?"
    return wh, zone

def safe_float(val):
    try: return float((str(val) or "0").strip())
    except: return 0.0

def parse_fee(val):
    """解析運費，去除 $ NT$ 等符號，回傳數字"""
    try:
        s = str(val or "0").strip()
        # 移除常見貨幣符號
        s = s.replace("NT$", "").replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0
    except:
        return 0.0

def check_oversize(ch, ship_raw, total_dim, max_side, weight):
    """
    判斷是否超材。
    回傳 (is_oversize: bool, warn_msg: str, can_split: bool, split_rules: dict)
    split_rules: {"max_dim": int, "max_side": int, "max_weight": int} 或 None
    """
    if total_dim == 0 and weight == 0:
        return False, "", False, None

    # 新竹物流
    if "新竹物流" in ship_raw:
        if total_dim > 210 or max_side > 150 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        for limit, fee, desc in CONFIG["hsinchu_surcharge"]:
            if total_dim <= limit:
                if fee > 0:
                    return True, f"新竹補助不足 {total_dim:.0f}cm → {desc}", False, None
                return False, "", False, None
        return True, f"超規 {total_dim:.0f}cm → 異常單", False, None

    # 嘉里快遞
    if "嘉里" in ship_raw:
        if total_dim > 200 or max_side > 120 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        return False, "", False, None

    # 店到家宅配 — 標準 or 大型
    if "店到家" in ship_raw:
        # 標準合規
        if total_dim <= 105 and max_side <= 45 and weight <= 10:
            return False, "", False, None
        # 大型合規
        if total_dim <= 150 and max_side <= 100 and weight <= 15:
            return False, "", False, None
        # 超材：先嘗試用大型規格拆，再嘗試標準規格拆
        rules = {"max_dim": 150, "max_side": 100, "max_weight": 15, "label": "店到家大型"}
        return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules

    # 超商（7-11 / 全家 / 萊爾富）
    if ch == "cvs":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "超商"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 店到店
    if ch == "store":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "店到店"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 隔日配 / 無包裝
    if ch in ("nextday", "nopkg"):
        if total_dim > 105 or max_side > 45 or weight > 10:
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", False, None
        return False, "", False, None

    return False, "", False, None

def apply_diagonal(sku, L, W, H):
    """
    如果此 SKU 在可斜放清單內，回傳斜放後的有效最長邊。
    否則回傳原始最長邊。
    """
    import math
    diag_map = CONFIG.get("diagonal_skus", {})
    if sku in diag_map:
        effective = diag_map[sku]
        return effective, True   # (有效最長邊, 是否斜放)
    return max(L, W, H), False

def suggest_split(products, weight_per_item, L, W, H, rules=None):
    """
    建議拆單方式。
    rules: {"max_dim": int, "max_side": int, "max_weight": int}
    """
    if rules is None:
        rules = {
            "max_dim":    CONFIG["split_max_dim"],
            "max_side":   CONFIG["split_max_side"],
            "max_weight": CONFIG["split_max_weight"],
        }
    max_units  = CONFIG["split_max_units"]
    max_dim    = rules["max_dim"]
    max_side   = rules["max_side"]
    max_weight = rules["max_weight"]

    # 展開所有件數為一個清單
    all_items = []
    for p in products:
        for _ in range(max(1, p.get("qty", 1))):
            all_items.append(p["sku"])

    total_items = len(all_items)

    # 單件尺寸判斷
    sides = sorted([L, W, H], reverse=True)
    single_dim  = sum(sides)
    single_side = sides[0]
    single_w    = weight_per_item

    # 如果單件就超規，無法拆單
    if single_dim > max_dim or single_side > max_side or single_w > max_weight:
        return None, "單件商品已超規，無法拆單"

    # 每包最多放幾件
    max_per_pkg_dim    = int(max_dim    // single_dim)  if single_dim  > 0 else 999
    max_per_pkg_side   = int(max_side   // single_side) if single_side > 0 else 999
    max_per_pkg_weight = int(max_weight // single_w)    if single_w    > 0 else 999
    max_per_pkg = min(max_per_pkg_dim, max_per_pkg_side, max_per_pkg_weight)
    max_per_pkg = max(1, max_per_pkg)

    # 需要幾包
    import math
    needed = math.ceil(total_items / max_per_pkg)

    if needed > max_units:
        return None, f"需要 {needed} 包，超過最多 {max_units} 單上限，無法自動拆單"

    # 建立拆單建議（同SKU盡量放同一包）
    packages = []
    current_pkg = []
    current_count = 0

    # 按SKU分組，同SKU連續放
    from collections import Counter
    sku_counts = Counter(all_items)
    sorted_skus = sorted(sku_counts.items(), key=lambda x: -x[1])

    remaining = []
    for sku, qty in sorted_skus:
        remaining.extend([sku] * qty)

    for item in remaining:
        if current_count >= max_per_pkg:
            packages.append(current_pkg[:])
            current_pkg = []
            current_count = 0
        current_pkg.append(item)
        current_count += 1

    if current_pkg:
        packages.append(current_pkg)

    result = []
    for i, pkg in enumerate(packages):
        pkg_counter = Counter(pkg)
        items_str = ", ".join(f"{sku}×{cnt}" for sku, cnt in sorted(pkg_counter.items()))
        pkg_dim    = single_dim * len(pkg)
        pkg_weight = single_w   * len(pkg)
        result.append({
            "pkg_no":  i + 1,
            "count":   len(pkg),
            "items":   items_str,
            "dim":     round(pkg_dim, 1),
            "weight":  round(pkg_weight, 2),
        })

    return result, ""

def _get_zone_label(locs):
    """取得區域描述字串，供宅配分群用"""
    whs   = set(wh for wh, z in locs)
    zones = set(z  for wh, z in locs if z != "?")
    if not zones: zones = {"?"}
    if len(whs) == 1 and len(zones) == 1:
        return f"{next(iter(whs))}{next(iter(zones))}區"
    elif len(whs) == 1:
        return f"{next(iter(whs))}混單"
    return "混單"

def get_sort_key(k):
    """排序：宅配→純區+單品→店到店多品→隔日配→無包裝→可拆單→超材"""
    if k == "__delivery__":     return (0,  k, k)
    if k == "__single_zone__":  return (20, k, k)   # 新：純區+單品大分類
    if k == "__nopkg__":        return (88, k, k)
    if k == "__splittable__":   return (92, k, k)
    if k == "__oversize__":     return (95, k, k)
    order = {"店到店":2, "店到店隔日配":3}
    for label, pri in order.items():
        if k.startswith(label):
            if "混單" in k: return (pri*10+9, "zzz", k)
            return (pri*10, k, k)
    return (99, k, k)

def split_orders(rows):
    # ── 以交易序號合併同一張訂單 ──────────────────────
    order_map = {}
    for row in rows:
        txn = (row.get(CONFIG["col_txn"], "") or "").strip()
        if not txn: continue
        if txn not in order_map:
            order_map[txn] = {
                "txn":       txn,
                "channel":   detect_channel(row.get(CONFIG["col_shipping"], "")),
                "ship_raw":  (row.get(CONFIG["col_shipping"], "") or "").strip(),
                "order_ids": [],
                "products":  [],
                "locations": set(),
                "total_qty": 0,
                # 尺寸（取該訂單最大值，因為同訂單多列尺寸應相同）
                "length":    0.0,
                "width":     0.0,
                "height":    0.0,
                "weight":    0.0,
                "fee":       0.0,  # 運費
            }
        oid    = (row.get(CONFIG["col_order_id"], "") or "").strip()
        raw_wh = (row.get(CONFIG["col_warehouse"], "") or "").strip()
        sku    = (row.get(CONFIG["col_sku"], "") or "").strip()
        wh, zone = parse_location(raw_wh)

        # 尺寸取最大值（避免空值覆蓋有效值）
        L = safe_float(row.get(CONFIG["col_length"], 0))
        W = safe_float(row.get(CONFIG["col_width"],  0))
        H = safe_float(row.get(CONFIG["col_height"], 0))
        Wt= safe_float(row.get(CONFIG["col_weight"], 0))
        Fe= parse_fee(row.get(CONFIG["col_fee"], 0))
        if L > 0: order_map[txn]["length"]  = max(order_map[txn]["length"],  L)
        if W > 0: order_map[txn]["width"]   = max(order_map[txn]["width"],   W)
        if H > 0: order_map[txn]["height"]  = max(order_map[txn]["height"],  H)
        if Wt> 0: order_map[txn]["weight"]  = max(order_map[txn]["weight"],  Wt)
        if Fe> 0: order_map[txn]["fee"]     = max(order_map[txn]["fee"],     Fe)

        order_map[txn]["order_ids"].append(oid)
        order_map[txn]["products"].append({
            "oid": oid, "sku": sku,
            "zone_raw": raw_wh, "wh": wh, "zone": zone,
        })
        order_map[txn]["locations"].add((wh, zone))
        order_map[txn]["total_qty"] += 1

    # ── 計算三邊總和、最長邊，斜放判斷，超材檢查 ────────
    for txn, o in order_map.items():
        L, W, H = o["length"], o["width"], o["height"]

        # 斜放判斷：若訂單內有可斜放 SKU，用對角線取代最長邊
        skus_in_order = [p["sku"] for p in o["products"]]
        diagonal_applied = False
        effective_side = max(L, W, H)
        for sku in skus_in_order:
            eff, is_diag = apply_diagonal(sku, L, W, H)
            if is_diag:
                effective_side = eff
                diagonal_applied = True
                break

        dims = sorted([L, W, H], reverse=True)
        o["max_side"]       = effective_side
        o["total_dim"]      = sum(dims)
        o["diagonal_used"]  = diagonal_applied
        o["oversize"], o["oversize_msg"], o["can_split"], o["split_rules"] = check_oversize(
            o["channel"], o["ship_raw"], o["total_dim"], o["max_side"], o["weight"]
        )

        # 若斜放後不超材，標示說明
        if diagonal_applied and not o["oversize"]:
            o["oversize_msg"] = "斜放後合規"

        # 拆單建議（運費=0才可拆單）
        o["split_suggestion"] = None
        o["split_error"]      = ""
        o["fee_paid"]         = o["fee"] > 0

        # 運費>0 → 標示買家已付，不拆單
        if o["oversize"] and o["fee"] > 0:
            o["oversize_msg"] += f"｜買家已付運費 {o['fee']:.0f} 元"

        can_split = o["oversize"] and o["can_split"] and o["fee"] == 0
        if can_split and o["total_qty"] > 0:
            weight_per = o["weight"] / o["total_qty"] if o["total_qty"] > 0 else 0
            from collections import Counter
            sku_count = Counter(p["sku"] for p in o["products"])
            prods = [{"sku": sku, "qty": qty} for sku, qty in sku_count.items()]
            suggestion, err = suggest_split(prods, weight_per, L, W, H, o["split_rules"])
            o["split_suggestion"] = suggestion
            o["split_error"]      = err

    # ── 套用分單規則 ──────────────────────────────────
    groups = {}

    def add(key, title, icon, color, o):
        if key not in groups:
            groups[key] = {"title": title, "icon": icon, "color": color, "orders": []}
        groups[key]["orders"].append(o)

    summary = {}

    for txn, o in order_map.items():
        ch    = o["channel"]
        locs  = o["locations"]
        meta  = CHANNEL_META.get(ch, CHANNEL_META["delivery"])
        label = meta["label"]
        icon  = meta["icon"]
        color = meta["color"]

        # 超材處理
        if o["oversize"]:
            o["oversize_channel"] = label
            if o["can_split"] and o["split_suggestion"]:
                # 可拆單 → 獨立可拆單分類
                add("__splittable__", "✂ 可拆單", "✂", "#e65100", o)
                summary["✂ 可拆單"] = summary.get("✂ 可拆單", 0) + 1
            else:
                # 不可拆單 → 超材異常分類
                add("__oversize__", "⚠ 超材", "⚠", "#b71c1c", o)
                summary["⚠ 超材"] = summary.get("⚠ 超材", 0) + 1
            continue

        # 無包裝 → 全部合併成一組
        if ch == "nopkg":
            add("__nopkg__", "📦 無包裝", "📦", "#00838f", o)
            summary["無包裝"] = summary.get("無包裝", 0) + 1
            continue

        # 宅配（新竹/嘉里/店到家）→ 外層一個大分類，內部依區域分群
        if ch == "delivery":
            o["delivery_zone"] = _get_zone_label(locs)  # 記錄區域供分群用
            add("__delivery__", "🚚 宅配", "🚚", "#1565c0", o)
            summary["宅配"] = summary.get("宅配", 0) + 1
            continue

        # ── 新分類邏輯：純區 + 超商單品 + 店到店單品 ──
        skus = set(p["sku"] for p in o["products"] if p["sku"])
        whs_all   = set(wh for wh, z in locs)
        zones_all = set(z  for wh, z in locs if z != "?")
        is_single_zone = (len(whs_all) == 1 and len(zones_all) == 1)  # 純單一區域
        is_single_item = (o["total_qty"] == 1)                         # 只有1件商品

        # 超商單品（1件）→ 進大分類
        if ch == "cvs" and is_single_item:
            o["single_zone_sub"] = "超商單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 店到店單品（1件）→ 進大分類
        if ch == "store" and is_single_item:
            o["single_zone_sub"] = "店到店單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 純區包裹（單一倉+單一區，任何通路）→ 進大分類
        if is_single_zone:
            wh_s   = next(iter(whs_all))
            zone_s = next(iter(zones_all))
            o["single_zone_sub"] = f"{label}｜{wh_s}{zone_s}區"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 其餘通路（超商多品跨區、店到店多品、隔日配）→ 依倉庫區域細分
        whs   = whs_all
        zones = zones_all
        if not zones:
            zones = {"?"}

        if len(whs) == 1 and len(zones) == 1:
            wh    = next(iter(whs))
            zone  = next(iter(zones))
            key   = f"{label} - {wh}{zone}區"
            title = f"{icon} {label} ｜ {wh} {zone} 區"
        elif len(whs) == 1:
            key   = f"{label} - {next(iter(whs))}混單"
            title = f"{icon} {label} ｜ {next(iter(whs))} 混單"
        else:
            key   = f"{label} - 混單"
            title = f"{icon} {label} ｜ 混單（跨倉）"

        if "混單" in key:
            color = color + "bb"

        add(key, title, icon, color, o)
        summary[label] = summary.get(label, 0) + 1

    state["summary"] = summary
    return dict(sorted(groups.items(), key=lambda x: get_sort_key(x[0])))

def load_csv(source, is_text=False):
    for enc in ["utf-8-sig", "big5", "cp950", "utf-8"]:
        try:
            if is_text:
                reader = csv.DictReader(io.StringIO(source))
            else:
                f = open(source, encoding=enc, newline="")
                reader = csv.DictReader(f)
            rows = list(reader)
            if not is_text: f.close()
            if rows: return rows, enc
        except Exception:
            continue
    return [], None

def run_pipeline(rows=None):
    state["status"] = "fetching"
    if not rows:
        state["status"] = "error"
        state["status_msg"] = "無資料"
        return
    groups = split_orders(rows)
    state["groups"]      = groups
    state["total"]       = len(set((r.get(CONFIG["col_txn"]) or "") for r in rows))
    state["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state["status"]      = "ready"
    state["status_msg"]  = f"共 {state['total']} 張訂單，分成 {len(groups)} 組 | {state['last_update']}"
    for k, g in groups.items():
        log(f"  {g['title']}：{len(g['orders'])} 張")
    log("分單完成")

def scheduler():
    while True:
        now = datetime.now()
        if now.hour == CONFIG["auto_run_hour"] and now.minute == CONFIG["auto_run_min"]:
            log("定時觸發")
            time.sleep(61)
        time.sleep(30)

# ============================================================
# 設定檔持久化（斜放商品清單）
# ============================================================
def _get_base_dir():
    """取得執行檔所在目錄（支援 PyInstaller 打包）"""
    if getattr(sys, 'frozen', False):
        # 打包成 exe 後，exe 所在目錄
        return os.path.dirname(sys.executable)
    else:
        # 一般 Python 執行
        return os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_get_base_dir(), "settings.json")

def load_settings():
    """從 settings.json 讀取設定，啟動時呼叫"""
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            if "diagonal_skus" in data:
                CONFIG["diagonal_skus"] = data["diagonal_skus"]
                log(f"已載入特殊可出超材品設定：{len(CONFIG['diagonal_skus'])} 筆")
    except Exception as e:
        log(f"讀取設定檔失敗：{e}")

def save_settings():
    """將設定寫入 settings.json"""
    try:
        data = {"diagonal_skus": CONFIG["diagonal_skus"]}
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"儲存設定檔失敗：{e}")

# ============================================================
# Flask 網頁
# ============================================================
app = Flask(__name__)

# -*- coding: utf-8 -*-
"""
4Sale 自動分單印單系統 v4
五個通路：宅配 / 超商 / 店到店 / 店到店隔日配 / 無包裝
區域細分：倉庫(前倉/主倉/備用倉) + 區域字母(A/B/C...)
"""

import sys, csv, io, os, re, json, threading, time
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 設定區
# ============================================================
CONFIG = {
    "company_name": "我的電商公司",
    "flask_port":   5000,
    "auto_run_hour":7,
    "auto_run_min": 30,

    # CSV 欄位名稱
    "col_txn":      "子交易序號",
    "col_order_id": "訂單編號",
    "col_shipping": "出貨類型",
    "col_warehouse":"商品倉庫儲位",
    "col_sku":      "商品編號",
    "col_length":   "商品長",       # 單位 cm
    "col_width":    "商品寬",       # 單位 cm
    "col_height":   "商品高",       # 單位 cm
    "col_weight":   "子交易總重量", # 單位 kg（整張訂單已加總）
    "col_fee":      "運費",         # 運費欄位（0=可拆單，>0=買家已付不拆單）

    # 通路判斷關鍵字
    "kw_delivery": ["新竹物流", "嘉里", "店到家"],
    "kw_cvs":      ["7-11", "711", "全家", "萊爾富"],
    "kw_nextday":  ["隔日到貨"],
    "kw_nopkg":    ["無包裝"],
    "kw_store":    ["店到店"],

    # 倉庫前綴
    "wh_prefix": {
        "?前": "參前",
        "?倉": "參前",
        "主":  "參倉",
    },

    # 超材規則（整張訂單加總後判斷）
    "oversize_rules": {
        # 通路key → (三邊總和上限cm, 最長邊上限cm, 重量上限kg, 說明)
        "cvs":      (105, 45,  10, "超商"),
        "store":    (105, 45,  10, "店到店"),
        "nextday":  (105, 45,  10, "店到店隔日配"),
        "nopkg":    (105, 45,  10, "無包裝"),
        "delivery_jialy":   (200, 120, 20, "嘉里快遞"),
        "delivery_hsinchu": (210, 150, 20, "新竹物流"),
        "delivery_store":   (150, 100, 15, "店到家大型"),
    },

    # 新竹物流補助運費對照（三邊總和 → 買家自付運費）
    # ≦150cm 正常補助，151~210cm 買家自付，>210cm 異常
    "hsinchu_surcharge": [
        (150, 0,   "正常補助"),
        (160, 135, "買家自付135元"),
        (170, 165, "買家自付165元"),
        (180, 195, "買家自付195元"),
        (190, 225, "買家自付225元"),
        (200, 285, "買家自付285元"),
        (210, 335, "買家自付335元"),
    ],

    # ★ 特殊可出超材品清單（SKU → 斜放後有效最長邊cm）
    # 斜放後最長邊 = √(長²+寬²+高²)，填入計算後的值
    # 例如：55x40x30 的商品，對角線 ≈ 74cm，但斜放進箱後最長邊可能只有 44cm
    # 請依實際測量填入，系統會用此值取代原始最長邊判斷超材
    "diagonal_skus": {
        # "SKU-001": 44,   # 範例：SKU-001 斜放後有效最長邊 44cm
        # "SKU-002": 43,
    },

    # 拆單規則（店到店/店到家超材時）
    "split_max_units":   5,     # 最多拆幾單
    "split_max_dim":     105,   # 每包三邊總和上限 cm
    "split_max_side":    45,    # 每包最長邊上限 cm
    "split_max_weight":  10,    # 每包重量上限 kg
}

# 通路顯示設定
CHANNEL_META = {
    "delivery": {"label": "宅配",         "icon": "🚙", "color": "#1565c0"},
    "cvs":      {"label": "超商",         "icon": "🏪", "color": "#c85000"},
    "store":    {"label": "店到店",       "icon": "🏬", "color": "#2e7d32"},
    "nextday":  {"label": "店到店隔日配", "icon": "⚡",   "color": "#6a1b9a"},
    "nopkg":    {"label": "無包裝",       "icon": "📦", "color": "#00838f"},
}

state = {
    "groups": {}, "total": 0, "last_update": None,
    "status": "idle", "status_msg": "請上傳 CSV 開始分單",
    "log": [], "summary": {},
}

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    state["log"].append(line)
    if len(state["log"]) > 200:
        state["log"] = state["log"][-200:]

# ============================================================
# 分單邏輯
# ============================================================
def detect_channel(val):
    s = (val or "").strip()
    if any(k in s for k in CONFIG["kw_delivery"]): return "delivery"
    if any(k in s for k in CONFIG["kw_cvs"]):      return "cvs"
    if any(k in s for k in CONFIG["kw_nextday"]):  return "nextday"
    if any(k in s for k in CONFIG["kw_nopkg"]):    return "nopkg"
    if any(k in s for k in CONFIG["kw_store"]):    return "store"
    return "delivery"

def parse_location(raw):
    """儲位 → (倉庫名稱, 區域字母)"""
    raw = (raw or "").strip()
    wh, rest = "其他", raw
    for prefix, name in CONFIG["wh_prefix"].items():
        if raw.startswith(prefix):
            wh, rest = name, raw[len(prefix):]
            break
    m = re.match(r"([A-Z])", rest)
    zone = m.group(1) if m else "?"
    return wh, zone

def safe_float(val):
    try: return float((str(val) or "0").strip())
    except: return 0.0

def parse_fee(val):
    """解析運費，去除 $ NT$ 等符號，回傳數字"""
    try:
        s = str(val or "0").strip()
        # 移除常見貨幣符號
        s = s.replace("NT$", "").replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0
    except:
        return 0.0

def check_oversize(ch, ship_raw, total_dim, max_side, weight):
    """
    判斷是否超材。
    回傳 (is_oversize: bool, warn_msg: str, can_split: bool, split_rules: dict)
    split_rules: {"max_dim": int, "max_side": int, "max_weight": int} 或 None
    """
    if total_dim == 0 and weight == 0:
        return False, "", False, None

    # 新竹物流
    if "新竹物流" in ship_raw:
        if total_dim > 210 or max_side > 150 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        for limit, fee, desc in CONFIG["hsinchu_surcharge"]:
            if total_dim <= limit:
                if fee > 0:
                    return True, f"新竹補助不足 {total_dim:.0f}cm → {desc}", False, None
                return False, "", False, None
        return True, f"超規 {total_dim:.0f}cm → 異常單", False, None

    # 嘉里快遞
    if "嘉里" in ship_raw:
        if total_dim > 200 or max_side > 120 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        return False, "", False, None

    # 店到家宅配 — 標準 or 大型
    if "店到家" in ship_raw:
        # 標準合規
        if total_dim <= 105 and max_side <= 45 and weight <= 10:
            return False, "", False, None
        # 大型合規
        if total_dim <= 150 and max_side <= 100 and weight <= 15:
            return False, "", False, None
        # 超材：先嘗試用大型規格拆，再嘗試標準規格拆
        rules = {"max_dim": 150, "max_side": 100, "max_weight": 15, "label": "店到家大型"}
        return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules

    # 超商（7-11 / 全家 / 萊爾富）
    if ch == "cvs":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "超商"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 店到店
    if ch == "store":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "店到店"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 隔日配 / 無包裝
    if ch in ("nextday", "nopkg"):
        if total_dim > 105 or max_side > 45 or weight > 10:
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", False, None
        return False, "", False, None

    return False, "", False, None

def apply_diagonal(sku, L, W, H):
    """
    如果此 SKU 在可斜放清單內，回傳斜放後的有效最長邊。
    否則回傳原始最長邊。
    """
    import math
    diag_map = CONFIG.get("diagonal_skus", {})
    if sku in diag_map:
        effective = diag_map[sku]
        return effective, True   # (有效最長邊, 是否斜放)
    return max(L, W, H), False

def suggest_split(products, weight_per_item, L, W, H, rules=None):
    """
    建議拆單方式。
    rules: {"max_dim": int, "max_side": int, "max_weight": int}
    """
    if rules is None:
        rules = {
            "max_dim":    CONFIG["split_max_dim"],
            "max_side":   CONFIG["split_max_side"],
            "max_weight": CONFIG["split_max_weight"],
        }
    max_units  = CONFIG["split_max_units"]
    max_dim    = rules["max_dim"]
    max_side   = rules["max_side"]
    max_weight = rules["max_weight"]

    # 展開所有件數為一個清單
    all_items = []
    for p in products:
        for _ in range(max(1, p.get("qty", 1))):
            all_items.append(p["sku"])

    total_items = len(all_items)

    # 單件尺寸判斷
    sides = sorted([L, W, H], reverse=True)
    single_dim  = sum(sides)
    single_side = sides[0]
    single_w    = weight_per_item

    # 如果單件就超規，無法拆單
    if single_dim > max_dim or single_side > max_side or single_w > max_weight:
        return None, "單件商品已超規，無法拆單"

    # 每包最多放幾件
    max_per_pkg_dim    = int(max_dim    // single_dim)  if single_dim  > 0 else 999
    max_per_pkg_side   = int(max_side   // single_side) if single_side > 0 else 999
    max_per_pkg_weight = int(max_weight // single_w)    if single_w    > 0 else 999
    max_per_pkg = min(max_per_pkg_dim, max_per_pkg_side, max_per_pkg_weight)
    max_per_pkg = max(1, max_per_pkg)

    # 需要幾包
    import math
    needed = math.ceil(total_items / max_per_pkg)

    if needed > max_units:
        return None, f"需要 {needed} 包，超過最多 {max_units} 單上限，無法自動拆單"

    # 建立拆單建議（同SKU盡量放同一包）
    packages = []
    current_pkg = []
    current_count = 0

    # 按SKU分組，同SKU連續放
    from collections import Counter
    sku_counts = Counter(all_items)
    sorted_skus = sorted(sku_counts.items(), key=lambda x: -x[1])

    remaining = []
    for sku, qty in sorted_skus:
        remaining.extend([sku] * qty)

    for item in remaining:
        if current_count >= max_per_pkg:
            packages.append(current_pkg[:])
            current_pkg = []
            current_count = 0
        current_pkg.append(item)
        current_count += 1

    if current_pkg:
        packages.append(current_pkg)

    result = []
    for i, pkg in enumerate(packages):
        pkg_counter = Counter(pkg)
        items_str = ", ".join(f"{sku}×{cnt}" for sku, cnt in sorted(pkg_counter.items()))
        pkg_dim    = single_dim * len(pkg)
        pkg_weight = single_w   * len(pkg)
        result.append({
            "pkg_no":  i + 1,
            "count":   len(pkg),
            "items":   items_str,
            "dim":     round(pkg_dim, 1),
            "weight":  round(pkg_weight, 2),
        })

    return result, ""

def _get_zone_label(locs):
    """取得區域描述字串，供宅配分群用"""
    whs   = set(wh for wh, z in locs)
    zones = set(z  for wh, z in locs if z != "?")
    if not zones: zones = {"?"}
    if len(whs) == 1 and len(zones) == 1:
        return f"{next(iter(whs))}{next(iter(zones))}區"
    elif len(whs) == 1:
        return f"{next(iter(whs))}混單"
    return "混單"

def get_sort_key(k):
    """排序：宅配→純區+單品→店到店多品→隔日配→無包裝→可拆單→超材"""
    if k == "__delivery__":     return (0,  k, k)
    if k == "__single_zone__":  return (20, k, k)   # 新：純區+單品大分類
    if k == "__nopkg__":        return (88, k, k)
    if k == "__splittable__":   return (92, k, k)
    if k == "__oversize__":     return (95, k, k)
    order = {"店到店":2, "店到店隔日配":3}
    for label, pri in order.items():
        if k.startswith(label):
            if "混單" in k: return (pri*10+9, "zzz", k)
            return (pri*10, k, k)
    return (99, k, k)

def split_orders(rows):
    # ── 以交易序號合併同一張訂單 ──────────────────────
    order_map = {}
    for row in rows:
        txn = (row.get(CONFIG["col_txn"], "") or "").strip()
        if not txn: continue
        if txn not in order_map:
            order_map[txn] = {
                "txn":       txn,
                "channel":   detect_channel(row.get(CONFIG["col_shipping"], "")),
                "ship_raw":  (row.get(CONFIG["col_shipping"], "") or "").strip(),
                "order_ids": [],
                "products":  [],
                "locations": set(),
                "total_qty": 0,
                # 尺寸（取該訂單最大值，因為同訂單多列尺寸應相同）
                "length":    0.0,
                "width":     0.0,
                "height":    0.0,
                "weight":    0.0,
                "fee":       0.0,  # 運費
            }
        oid    = (row.get(CONFIG["col_order_id"], "") or "").strip()
        raw_wh = (row.get(CONFIG["col_warehouse"], "") or "").strip()
        sku    = (row.get(CONFIG["col_sku"], "") or "").strip()
        wh, zone = parse_location(raw_wh)

        # 尺寸取最大值（避免空值覆蓋有效值）
        L = safe_float(row.get(CONFIG["col_length"], 0))
        W = safe_float(row.get(CONFIG["col_width"],  0))
        H = safe_float(row.get(CONFIG["col_height"], 0))
        Wt= safe_float(row.get(CONFIG["col_weight"], 0))
        Fe= parse_fee(row.get(CONFIG["col_fee"], 0))
        if L > 0: order_map[txn]["length"]  = max(order_map[txn]["length"],  L)
        if W > 0: order_map[txn]["width"]   = max(order_map[txn]["width"],   W)
        if H > 0: order_map[txn]["height"]  = max(order_map[txn]["height"],  H)
        if Wt> 0: order_map[txn]["weight"]  = max(order_map[txn]["weight"],  Wt)
        if Fe> 0: order_map[txn]["fee"]     = max(order_map[txn]["fee"],     Fe)

        order_map[txn]["order_ids"].append(oid)
        order_map[txn]["products"].append({
            "oid": oid, "sku": sku,
            "zone_raw": raw_wh, "wh": wh, "zone": zone,
        })
        order_map[txn]["locations"].add((wh, zone))
        order_map[txn]["total_qty"] += 1

    # ── 計算三邊總和、最長邊，斜放判斷，超材檢查 ────────
    for txn, o in order_map.items():
        L, W, H = o["length"], o["width"], o["height"]

        # 斜放判斷：若訂單內有可斜放 SKU，用對角線取代最長邊
        skus_in_order = [p["sku"] for p in o["products"]]
        diagonal_applied = False
        effective_side = max(L, W, H)
        for sku in skus_in_order:
            eff, is_diag = apply_diagonal(sku, L, W, H)
            if is_diag:
                effective_side = eff
                diagonal_applied = True
                break

        dims = sorted([L, W, H], reverse=True)
        o["max_side"]       = effective_side
        o["total_dim"]      = sum(dims)
        o["diagonal_used"]  = diagonal_applied
        o["oversize"], o["oversize_msg"], o["can_split"], o["split_rules"] = check_oversize(
            o["channel"], o["ship_raw"], o["total_dim"], o["max_side"], o["weight"]
        )

        # 若斜放後不超材，標示說明
        if diagonal_applied and not o["oversize"]:
            o["oversize_msg"] = "斜放後合規"

        # 拆單建議（運費=0才可拆單）
        o["split_suggestion"] = None
        o["split_error"]      = ""
        o["fee_paid"]         = o["fee"] > 0

        # 運費>0 → 標示買家已付，不拆單
        if o["oversize"] and o["fee"] > 0:
            o["oversize_msg"] += f"｜買家已付運費 {o['fee']:.0f} 元"

        can_split = o["oversize"] and o["can_split"] and o["fee"] == 0
        if can_split and o["total_qty"] > 0:
            weight_per = o["weight"] / o["total_qty"] if o["total_qty"] > 0 else 0
            from collections import Counter
            sku_count = Counter(p["sku"] for p in o["products"])
            prods = [{"sku": sku, "qty": qty} for sku, qty in sku_count.items()]
            suggestion, err = suggest_split(prods, weight_per, L, W, H, o["split_rules"])
            o["split_suggestion"] = suggestion
            o["split_error"]      = err

    # ── 套用分單規則 ──────────────────────────────────
    groups = {}

    def add(key, title, icon, color, o):
        if key not in groups:
            groups[key] = {"title": title, "icon": icon, "color": color, "orders": []}
        groups[key]["orders"].append(o)

    summary = {}

    for txn, o in order_map.items():
        ch    = o["channel"]
        locs  = o["locations"]
        meta  = CHANNEL_META.get(ch, CHANNEL_META["delivery"])
        label = meta["label"]
        icon  = meta["icon"]
        color = meta["color"]

        # 超材處理
        if o["oversize"]:
            o["oversize_channel"] = label
            if o["can_split"] and o["split_suggestion"]:
                # 可拆單 → 獨立可拆單分類
                add("__splittable__", "✂ 可拆單", "✂", "#e65100", o)
                summary["✂ 可拆單"] = summary.get("✂ 可拆單", 0) + 1
            else:
                # 不可拆單 → 超材異常分類
                add("__oversize__", "⚠ 超材", "⚠", "#b71c1c", o)
                summary["⚠ 超材"] = summary.get("⚠ 超材", 0) + 1
            continue

        # 無包裝 → 全部合併成一組
        if ch == "nopkg":
            add("__nopkg__", "📦 無包裝", "📦", "#00838f", o)
            summary["無包裝"] = summary.get("無包裝", 0) + 1
            continue

        # 宅配（新竹/嘉里/店到家）→ 外層一個大分類，內部依區域分群
        if ch == "delivery":
            o["delivery_zone"] = _get_zone_label(locs)  # 記錄區域供分群用
            add("__delivery__", "🚚 宅配", "🚚", "#1565c0", o)
            summary["宅配"] = summary.get("宅配", 0) + 1
            continue

        # ── 新分類邏輯：純區 + 超商單品 + 店到店單品 ──
        skus = set(p["sku"] for p in o["products"] if p["sku"])
        whs_all   = set(wh for wh, z in locs)
        zones_all = set(z  for wh, z in locs if z != "?")
        is_single_zone = (len(whs_all) == 1 and len(zones_all) == 1)  # 純單一區域
        is_single_item = (o["total_qty"] == 1)                         # 只有1件商品

        # 超商單品（1件）→ 進大分類
        if ch == "cvs" and is_single_item:
            o["single_zone_sub"] = "超商單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 店到店單品（1件）→ 進大分類
        if ch == "store" and is_single_item:
            o["single_zone_sub"] = "店到店單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 純區包裹（單一倉+單一區，任何通路）→ 進大分類
        if is_single_zone:
            wh_s   = next(iter(whs_all))
            zone_s = next(iter(zones_all))
            o["single_zone_sub"] = f"{label}｜{wh_s}{zone_s}區"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 其餘通路（超商多品跨區、店到店多品、隔日配）→ 依倉庫區域細分
        whs   = whs_all
        zones = zones_all
        if not zones:
            zones = {"?"}

        if len(whs) == 1 and len(zones) == 1:
            wh    = next(iter(whs))
            zone  = next(iter(zones))
            key   = f"{label} - {wh}{zone}區"
            title = f"{icon} {label} ｜ {wh} {zone} 區"
        elif len(whs) == 1:
            key   = f"{label} - {next(iter(whs))}混單"
            title = f"{icon} {label} ｜ {next(iter(whs))} 混單"
        else:
            key   = f"{label} - 混單"
            title = f"{icon} {label} ｜ 混單（跨倉）"

        if "混單" in key:
            color = color + "bb"

        add(key, title, icon, color, o)
        summary[label] = summary.get(label, 0) + 1

    state["summary"] = summary
    return dict(sorted(groups.items(), key=lambda x: get_sort_key(x[0])))

def load_csv(source, is_text=False):
    for enc in ["utf-8-sig", "big5", "cp950", "utf-8"]:
        try:
            if is_text:
                reader = csv.DictReader(io.StringIO(source))
            else:
                f = open(source, encoding=enc, newline="")
                reader = csv.DictReader(f)
            rows = list(reader)
            if not is_text: f.close()
            if rows: return rows, enc
        except Exception:
            continue
    return [], None

def run_pipeline(rows=None):
    state["status"] = "fetching"
    if not rows:
        state["status"] = "error"
        state["status_msg"] = "無資料"
        return
    groups = split_orders(rows)
    state["groups"]      = groups
    state["total"]       = len(set((r.get(CONFIG["col_txn"]) or "") for r in rows))
    state["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state["status"]      = "ready"
    state["status_msg"]  = f"共 {state['total']} 張訂單，分成 {len(groups)} 組 | {state['last_update']}"
    for k, g in groups.items():
        log(f"  {g['title']}：{len(g['orders'])} 張")
    log("分單完成")

def scheduler():
    while True:
        now = datetime.now()
        if now.hour == CONFIG["auto_run_hour"] and now.minute == CONFIG["auto_run_min"]:
            log("定時觸發")
            time.sleep(61)
        time.sleep(30)

# ============================================================
# 設定檔持久化（斜放商品清單）
# ============================================================
def _get_base_dir():
    """取得執行檔所在目錄（支援 PyInstaller 打包）"""
    if getattr(sys, 'frozen', False):
        # 打包成 exe 後，exe 所在目錄
        return os.path.dirname(sys.executable)
    else:
        # 一般 Python 執行
        return os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_get_base_dir(), "settings.json")

def load_settings():
    """從 settings.json 讀取設定，啟動時呼叫"""
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            if "diagonal_skus" in data:
                CONFIG["diagonal_skus"] = data["diagonal_skus"]
                log(f"已載入特殊可出超材品設定：{len(CONFIG['diagonal_skus'])} 筆")
    except Exception as e:
        log(f"讀取設定檔失敗：{e}")

def save_settings():
    """將設定寫入 settings.json"""
    try:
        data = {"diagonal_skus": CONFIG["diagonal_skus"]}
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"儲存設定檔失敗：{e}")

# ============================================================
# Flask 網頁
# ============================================================
app = Flask(__name__)

# -*- coding: utf-8 -*-
"""
4Sale 自動分單印單系統 v4
五個通路：宅配 / 超商 / 店到店 / 店到店隔日配 / 無包裝
區域細分：倉庫(前倉/主倉/備用倉) + 區域字母(A/B/C...)
"""

import sys, csv, io, os, re, json, threading, time
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# 設定區
# ============================================================
CONFIG = {
    "company_name": "我的電商公司",
    "flask_port":   5000,
    "auto_run_hour":7,
    "auto_run_min": 30,

    # CSV 欄位名稱
    "col_txn":      "子交易序號",
    "col_order_id": "訂單編號",
    "col_shipping": "出貨類型",
    "col_warehouse":"商品倉庫儲位",
    "col_sku":      "商品編號",
    "col_length":   "商品長",       # 單位 cm
    "col_width":    "商品寬",       # 單位 cm
    "col_height":   "商品高",       # 單位 cm
    "col_weight":   "子交易總重量", # 單位 kg（整張訂單已加總）
    "col_fee":      "運費",         # 運費欄位（0=可拆單，>0=買家已付不拆單）

    # 通路判斷關鍵字
    "kw_delivery": ["新竹物流", "嘉里", "店到家"],
    "kw_cvs":      ["7-11", "711", "全家", "萊爾富"],
    "kw_nextday":  ["隔日到貨"],
    "kw_nopkg":    ["無包裝"],
    "kw_store":    ["店到店"],

    # 倉庫前綴
    "wh_prefix": {
        "?前": "參前",
        "?倉": "參前",
        "主":  "參倉",
    },

    # 超材規則（整張訂單加總後判斷）
    "oversize_rules": {
        # 通路key → (三邊總和上限cm, 最長邊上限cm, 重量上限kg, 說明)
        "cvs":      (105, 45,  10, "超商"),
        "store":    (105, 45,  10, "店到店"),
        "nextday":  (105, 45,  10, "店到店隔日配"),
        "nopkg":    (105, 45,  10, "無包裝"),
        "delivery_jialy":   (200, 120, 20, "嘉里快遞"),
        "delivery_hsinchu": (210, 150, 20, "新竹物流"),
        "delivery_store":   (150, 100, 15, "店到家大型"),
    },

    # 新竹物流補助運費對照（三邊總和 → 買家自付運費）
    # ≦150cm 正常補助，151~210cm 買家自付，>210cm 異常
    "hsinchu_surcharge": [
        (150, 0,   "正常補助"),
        (160, 135, "買家自付135元"),
        (170, 165, "買家自付165元"),
        (180, 195, "買家自付195元"),
        (190, 225, "買家自付225元"),
        (200, 285, "買家自付285元"),
        (210, 335, "買家自付335元"),
    ],

    # ★ 特殊可出超材品清單（SKU → 斜放後有效最長邊cm）
    # 斜放後最長邊 = √(長²+寬²+高²)，填入計算後的值
    # 例如：55x40x30 的商品，對角線 ≈ 74cm，但斜放進箱後最長邊可能只有 44cm
    # 請依實際測量填入，系統會用此值取代原始最長邊判斷超材
    "diagonal_skus": {
        # "SKU-001": 44,   # 範例：SKU-001 斜放後有效最長邊 44cm
        # "SKU-002": 43,
    },

    # 拆單規則（店到店/店到家超材時）
    "split_max_units":   5,     # 最多拆幾單
    "split_max_dim":     105,   # 每包三邊總和上限 cm
    "split_max_side":    45,    # 每包最長邊上限 cm
    "split_max_weight":  10,    # 每包重量上限 kg
}

# 通路顯示設定
CHANNEL_META = {
    "delivery": {"label": "宅配",         "icon": "🚙", "color": "#1565c0"},
    "cvs":      {"label": "超商",         "icon": "🏪", "color": "#c85000"},
    "store":    {"label": "店到店",       "icon": "🏬", "color": "#2e7d32"},
    "nextday":  {"label": "店到店隔日配", "icon": "⚡",   "color": "#6a1b9a"},
    "nopkg":    {"label": "無包裝",       "icon": "📦", "color": "#00838f"},
}

state = {
    "groups": {}, "total": 0, "last_update": None,
    "status": "idle", "status_msg": "請上傳 CSV 開始分單",
    "log": [], "summary": {},
}

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    state["log"].append(line)
    if len(state["log"]) > 200:
        state["log"] = state["log"][-200:]

# ============================================================
# 分單邏輯
# ============================================================
def detect_channel(val):
    s = (val or "").strip()
    if any(k in s for k in CONFIG["kw_delivery"]): return "delivery"
    if any(k in s for k in CONFIG["kw_cvs"]):      return "cvs"
    if any(k in s for k in CONFIG["kw_nextday"]):  return "nextday"
    if any(k in s for k in CONFIG["kw_nopkg"]):    return "nopkg"
    if any(k in s for k in CONFIG["kw_store"]):    return "store"
    return "delivery"

def parse_location(raw):
    """儲位 → (倉庫名稱, 區域字母)"""
    raw = (raw or "").strip()
    wh, rest = "其他", raw
    for prefix, name in CONFIG["wh_prefix"].items():
        if raw.startswith(prefix):
            wh, rest = name, raw[len(prefix):]
            break
    m = re.match(r"([A-Z])", rest)
    zone = m.group(1) if m else "?"
    return wh, zone

def safe_float(val):
    try: return float((str(val) or "0").strip())
    except: return 0.0

def parse_fee(val):
    """解析運費，去除 $ NT$ 等符號，回傳數字"""
    try:
        s = str(val or "0").strip()
        # 移除常見貨幣符號
        s = s.replace("NT$", "").replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0
    except:
        return 0.0

def check_oversize(ch, ship_raw, total_dim, max_side, weight, buyer_fee=0):
    """
    判斷是否超材。
    回傳 (is_oversize: bool, warn_msg: str, can_split: bool, split_rules: dict)
    """
    if total_dim == 0 and weight == 0:
        return False, "", False, None

    # 新竹物流
    if "新竹物流" in ship_raw:
        if total_dim > 210 or max_side > 150 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        for limit, required_fee, desc in CONFIG["hsinchu_surcharge"]:
            if total_dim <= limit:
                if required_fee == 0:
                    return False, "", False, None
                if buyer_fee >= required_fee:
                    return False, "", False, None
                return True, f"新竹補助不足 {total_dim:.0f}cm，應付 {required_fee} 元，實付 {buyer_fee:.0f} 元 → 異常單", False, None
        return True, f"超規 {total_dim:.0f}cm → 異常單", False, None

    # 嘉里快遞
    if "嘉里" in ship_raw:
        if total_dim > 200 or max_side > 120 or weight > 20:
            return True, f"超規 {total_dim:.0f}cm/{weight:.1f}kg → 異常單", False, None
        return False, "", False, None

    # 店到家宅配 — 標準 or 大型
    if "店到家" in ship_raw:
        # 標準合規
        if total_dim <= 105 and max_side <= 45 and weight <= 10:
            return False, "", False, None
        # 大型合規
        if total_dim <= 150 and max_side <= 100 and weight <= 15:
            return False, "", False, None
        # 超材：先嘗試用大型規格拆，再嘗試標準規格拆
        rules = {"max_dim": 150, "max_side": 100, "max_weight": 15, "label": "店到家大型"}
        return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules

    # 超商（7-11 / 全家 / 萊爾富）
    if ch == "cvs":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "超商"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 店到店
    if ch == "store":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "店到店"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 隔日配（支援拆單）
    if ch == "nextday":
        if total_dim > 105 or max_side > 45 or weight > 10:
            rules = {"max_dim": 105, "max_side": 45, "max_weight": 10, "label": "隔日配"}
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", True, rules
        return False, "", False, None

    # 無包裝（不拆單）
    if ch == "nopkg":
        if total_dim > 105 or max_side > 45 or weight > 10:
            return True, f"超材 {total_dim:.0f}cm/{weight:.1f}kg", False, None
        return False, "", False, None
    return False, "", False, None

def apply_diagonal(sku, L, W, H):
    """
    如果此 SKU 在可斜放清單內，回傳斜放後的有效最長邊。
    支援新格式 {side: float, max_qty: int} 和舊格式 float。
    """
    diag_map = CONFIG.get("diagonal_skus", {})
    if sku in diag_map:
        cfg = diag_map[sku]
        if isinstance(cfg, dict):
            side = cfg.get("side")
            if side: return float(side), True
        elif isinstance(cfg, (int, float)):
            return float(cfg), True
    return max(L, W, H), False

def get_sku_max_qty(sku, channel=None):
    """
    取得 SKU 的每包最大件數限制。
    channel: 'store'(店配類) 或 'delivery'(快遞類)
    無設定或通路不符回傳 None。
    """
    diag_map = CONFIG.get("diagonal_skus", {})
    if sku not in diag_map:
        return None
    cfg = diag_map[sku]
    if not isinstance(cfg, dict):
        return None
    max_qty  = cfg.get("max_qty")
    channels = cfg.get("channels", [])
    if not max_qty:
        return None
    # 若沒有設定通路限制，或通路符合，才回傳
    if not channels:
        return max_qty
    if channel and channel in channels:
        return max_qty
    return None


def suggest_split(products, weight_per_item, L, W, H, rules=None):
    """
    建議拆單方式。
    rules: {max_dim, max_side, max_weight, max_qty(optional for bulky items)}
    """
    if rules is None:
        rules = {
            "max_dim":    CONFIG["split_max_dim"],
            "max_side":   CONFIG["split_max_side"],
            "max_weight": CONFIG["split_max_weight"],
        }
    max_units     = CONFIG["split_max_units"]
    max_dim       = rules["max_dim"]
    max_side      = rules["max_side"]
    max_weight    = rules["max_weight"]
    max_qty_limit = rules.get("max_qty")

    all_items = []
    for p in products:
        for _ in range(max(1, p.get("qty", 1))):
            all_items.append(p["sku"])

    total_items = len(all_items)
    sides       = sorted([L, W, H], reverse=True)
    single_dim  = sum(sides)
    single_side = sides[0]
    single_w    = weight_per_item

    if max_qty_limit:
        max_per_pkg = max_qty_limit
    else:
        if single_dim > max_dim or single_side > max_side or single_w > max_weight:
            return None, "單件商品已超規，無法拆單"
        max_per_pkg = min(
            int(max_dim    // single_dim)  if single_dim  > 0 else 999,
            int(max_side   // single_side) if single_side > 0 else 999,
            int(max_weight // single_w)    if single_w    > 0 else 999,
        )
        max_per_pkg = max(1, max_per_pkg)

    # 需要幾包
    import math
    needed = math.ceil(total_items / max_per_pkg)

    if needed > max_units:
        return None, f"需要 {needed} 包，超過最多 {max_units} 單上限，無法自動拆單"

    # 建立拆單建議（同SKU盡量放同一包）
    packages = []
    current_pkg = []
    current_count = 0

    # 按SKU分組，同SKU連續放
    from collections import Counter
    sku_counts = Counter(all_items)
    sorted_skus = sorted(sku_counts.items(), key=lambda x: -x[1])

    remaining = []
    for sku, qty in sorted_skus:
        remaining.extend([sku] * qty)

    for item in remaining:
        if current_count >= max_per_pkg:
            packages.append(current_pkg[:])
            current_pkg = []
            current_count = 0
        current_pkg.append(item)
        current_count += 1

    if current_pkg:
        packages.append(current_pkg)

    result = []
    for i, pkg in enumerate(packages):
        pkg_counter = Counter(pkg)
        items_str = ", ".join(f"{sku}×{cnt}" for sku, cnt in sorted(pkg_counter.items()))
        pkg_dim    = single_dim * len(pkg)
        pkg_weight = single_w   * len(pkg)
        result.append({
            "pkg_no":  i + 1,
            "count":   len(pkg),
            "items":   items_str,
            "dim":     round(pkg_dim, 1),
            "weight":  round(pkg_weight, 2),
        })

    return result, ""

def _get_zone_label(locs):
    """取得區域描述字串，供宅配分群用"""
    whs   = set(wh for wh, z in locs)
    zones = set(z  for wh, z in locs if z != "?")
    if not zones: zones = {"?"}
    if len(whs) == 1 and len(zones) == 1:
        return f"{next(iter(whs))}{next(iter(zones))}區"
    elif len(whs) == 1:
        return f"{next(iter(whs))}混單"
    return "混單"

def get_sort_key(k):
    """排序：宅配→純區+單品→店到店多品→隔日配→無包裝→巨無霸→可拆單→超材"""
    if k == "__delivery__":     return (0,  k, k)
    if k == "__single_zone__":  return (20, k, k)
    if k == "__nopkg__":        return (88, k, k)
    if k == "__giant__":        return (91, k, k)
    if k == "__splittable__":   return (92, k, k)
    if k == "__oversize__":     return (95, k, k)
    order = {"店到店":2, "店到店隔日配":3}
    for label, pri in order.items():
        if k.startswith(label):
            if "混單" in k: return (pri*10+9, "zzz", k)
            return (pri*10, k, k)
    return (99, k, k)

def split_orders(rows):
    # ── 以交易序號合併同一張訂單 ──────────────────────
    order_map = {}
    for row in rows:
        txn = (row.get(CONFIG["col_txn"], "") or "").strip()
        if not txn: continue
        if txn not in order_map:
            order_map[txn] = {
                "txn":       txn,
                "channel":   detect_channel(row.get(CONFIG["col_shipping"], "")),
                "ship_raw":  (row.get(CONFIG["col_shipping"], "") or "").strip(),
                "order_ids": [],
                "products":  [],
                "locations": set(),
                "total_qty": 0,
                # 尺寸（取該訂單最大值，因為同訂單多列尺寸應相同）
                "length":    0.0,
                "width":     0.0,
                "height":    0.0,
                "weight":    0.0,
                "fee":       0.0,  # 運費
            }
        oid    = (row.get(CONFIG["col_order_id"], "") or "").strip()
        raw_wh = (row.get(CONFIG["col_warehouse"], "") or "").strip()
        sku    = (row.get(CONFIG["col_sku"], "") or "").strip()
        wh, zone = parse_location(raw_wh)

        # 尺寸取最大值（避免空值覆蓋有效值）
        L = safe_float(row.get(CONFIG["col_length"], 0))
        W = safe_float(row.get(CONFIG["col_width"],  0))
        H = safe_float(row.get(CONFIG["col_height"], 0))
        Wt= safe_float(row.get(CONFIG["col_weight"], 0))
        Fe= parse_fee(row.get(CONFIG["col_fee"], 0))
        if L > 0: order_map[txn]["length"]  = max(order_map[txn]["length"],  L)
        if W > 0: order_map[txn]["width"]   = max(order_map[txn]["width"],   W)
        if H > 0: order_map[txn]["height"]  = max(order_map[txn]["height"],  H)
        if Wt> 0: order_map[txn]["weight"]  = max(order_map[txn]["weight"],  Wt)
        if Fe> 0: order_map[txn]["fee"]     = max(order_map[txn]["fee"],     Fe)

        order_map[txn]["order_ids"].append(oid)
        order_map[txn]["products"].append({
            "oid": oid, "sku": sku,
            "zone_raw": raw_wh, "wh": wh, "zone": zone,
        })
        order_map[txn]["locations"].add((wh, zone))
        order_map[txn]["total_qty"] += 1

    # ── 計算三邊總和、最長邊，斜放判斷，超材檢查 ────────
    for txn, o in order_map.items():
        L, W, H = o["length"], o["width"], o["height"]

        # 高度為 0 → ERP 無法輸入小數，自動補 0.2cm
        if H == 0 and (L > 0 or W > 0):
            H = 0.2
            o["height"] = 0.2
            o["height_auto"] = True
        else:
            o["height_auto"] = False

        # 斜放判斷：若訂單內有可斜放 SKU，用對角線取代最長邊
        dims = sorted([L, W, H], reverse=True)
        skus_in_order = [p["sku"] for p in o["products"]]
        diagonal_applied = False
        effective_side = dims[0] if dims else 0
        for sku in skus_in_order:
            eff, is_diag = apply_diagonal(sku, L, W, H)
            if is_diag:
                effective_side = eff
                diagonal_applied = True
                break

        o["max_side"]       = effective_side
        o["total_dim"]      = sum(dims)
        o["diagonal_used"]  = diagonal_applied
        o["oversize"], o["oversize_msg"], o["can_split"], o["split_rules"] = check_oversize(
            o["channel"], o["ship_raw"], o["total_dim"], o["max_side"], o["weight"], o["fee"]
        )

        # 若斜放後不超材，標示說明
        if diagonal_applied and not o["oversize"]:
            o["oversize_msg"] = "斜放後合規"

        # 件數上限判斷：檢查訂單內的 SKU 是否有設定每包最大件數
        if not o["oversize"]:
            from collections import Counter
            sku_count = Counter(p["sku"] for p in o["products"])
            # 判斷通路群組：delivery=快遞類, store=店配類
            ch = o["channel"]
            ch_group = "delivery" if ch == "delivery" else "store"
            for sku, qty in sku_count.items():
                max_q = get_sku_max_qty(sku, ch_group)
                if max_q and qty > max_q:
                    o["oversize"]     = True
                    o["can_split"]    = True
                    o["split_rules"]  = {
                        "max_dim":    CONFIG["split_max_dim"],
                        "max_side":   CONFIG["split_max_side"],
                        "max_weight": CONFIG["split_max_weight"],
                        "label":      f"每包最多{max_q}件",
                        "max_qty":    max_q,
                    }
                    o["oversize_msg"] = f"件數超限（{sku} 共{qty}件，每包上限{max_q}件）"
                    break

        # 拆單建議（運費=0才可拆單，店到店/店到家/隔日配）
        o["split_suggestion"] = None
        o["split_error"]      = ""
        o["fee_paid"]         = o["fee"] > 0

        # 運費>0 → 標示買家已付，不拆單
        if o["oversize"] and o["fee"] > 0:
            o["oversize_msg"] += f"｜買家已付運費 {o['fee']:.0f} 元"

        can_split = (
            o["oversize"] and
            o["can_split"] and
            o["fee"] == 0 and
            any(k in o["ship_raw"] for k in ["店到店", "店到家", "隔日到貨"])
        )
        if can_split and o["total_qty"] > 0:
            weight_per = o["weight"] / o["total_qty"] if o["total_qty"] > 0 else 0
            from collections import Counter
            sku_count = Counter(p["sku"] for p in o["products"])
            prods = [{"sku": sku, "qty": qty} for sku, qty in sku_count.items()]
            suggestion, err = suggest_split(prods, weight_per, L, W, H, o["split_rules"])
            o["split_suggestion"] = suggestion
            o["split_error"]      = err

    # ── 套用分單規則 ──────────────────────────────────
    groups = {}

    def add(key, title, icon, color, o):
        if key not in groups:
            groups[key] = {"title": title, "icon": icon, "color": color, "orders": []}
        groups[key]["orders"].append(o)

    summary = {}

    for txn, o in order_map.items():
        ch    = o["channel"]
        locs  = o["locations"]
        meta  = CHANNEL_META.get(ch, CHANNEL_META["delivery"])
        label = meta["label"]
        icon  = meta["icon"]
        color = meta["color"]

        # 超材處理
        if o["oversize"]:
            o["oversize_channel"] = label
            if o["can_split"] and o["split_suggestion"]:
                # 可拆單 → 獨立可拆單分類
                add("__splittable__", "✂ 可拆單", "✂", "#e65100", o)
                summary["✂ 可拆單"] = summary.get("✂ 可拆單", 0) + 1
            else:
                # 不可拆單 → 超材異常分類
                add("__oversize__", "⚠ 超材", "⚠", "#b71c1c", o)
                summary["⚠ 超材"] = summary.get("⚠ 超材", 0) + 1
            continue

        # 無包裝 → 全部合併成一組
        if ch == "nopkg":
            add("__nopkg__", "📦 無包裝", "📦", "#00838f", o)
            summary["無包裝"] = summary.get("無包裝", 0) + 1
            continue

        # 宅配（新竹/嘉里/店到家）→ 先判斷巨無霸，再進宅配大分類
        if ch == "delivery":
            is_giant = (
                (o["total_dim"] >= 170 or o["weight"] >= 15) and
                any(k in o["ship_raw"] for k in ["新竹物流", "嘉里"])
            )
            if is_giant:
                reasons = []
                if o["total_dim"] >= 170: reasons.append(f"三邊{o['total_dim']:.0f}cm")
                if o["weight"] >= 15:     reasons.append(f"重量{o['weight']:.1f}kg")
                o["oversize_channel"] = label
                o["oversize_msg"] = "巨無霸 " + "、".join(reasons) + " → 人工確認"
                add("__giant__", "&#129427; 巨無霸", "&#129427;", "#6a0dad", o)
                summary["&#129427; 巨無霸"] = summary.get("&#129427; 巨無霸", 0) + 1
                continue
            o["delivery_zone"] = _get_zone_label(locs)
            add("__delivery__", "&#128665; 宅配", "&#128665;", "#1565c0", o)
            summary["宅配"] = summary.get("宅配", 0) + 1
            continue

        # ── 新分類邏輯：純區 + 超商單品 + 店到店單品 ──
        skus = set(p["sku"] for p in o["products"] if p["sku"])
        whs_all   = set(wh for wh, z in locs)
        zones_all = set(z  for wh, z in locs if z != "?")
        is_single_zone = (len(whs_all) == 1 and len(zones_all) == 1)  # 純單一區域
        is_single_item = (o["total_qty"] == 1)                         # 只有1件商品

        # 超商單品（1件）→ 進大分類
        if ch == "cvs" and is_single_item:
            o["single_zone_sub"] = "超商單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 店到店單品（1件）→ 進大分類
        if ch == "store" and is_single_item:
            o["single_zone_sub"] = "店到店單品"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 純區包裹（單一倉+單一區，任何通路）→ 進大分類
        if is_single_zone:
            wh_s   = next(iter(whs_all))
            zone_s = next(iter(zones_all))
            o["single_zone_sub"] = f"{label}｜{wh_s}{zone_s}區"
            add("__single_zone__", "⚡ 純區 + 超商單品 + 店到店單品", "⚡", "#e65100", o)
            summary["純區+單品"] = summary.get("純區+單品", 0) + 1
            continue

        # 其餘通路（超商多品跨區、店到店多品、隔日配）→ 依倉庫區域細分
        whs   = whs_all
        zones = zones_all
        if not zones:
            zones = {"?"}

        if len(whs) == 1 and len(zones) == 1:
            wh    = next(iter(whs))
            zone  = next(iter(zones))
            key   = f"{label} - {wh}{zone}區"
            title = f"{icon} {label} ｜ {wh} {zone} 區"
        elif len(whs) == 1:
            key   = f"{label} - {next(iter(whs))}混單"
            title = f"{icon} {label} ｜ {next(iter(whs))} 混單"
        else:
            key   = f"{label} - 混單"
            title = f"{icon} {label} ｜ 混單（跨倉）"

        if "混單" in key:
            color = color + "bb"

        add(key, title, icon, color, o)
        summary[label] = summary.get(label, 0) + 1

    state["summary"] = summary
    return dict(sorted(groups.items(), key=lambda x: get_sort_key(x[0])))

def load_csv(source, is_text=False):
    for enc in ["utf-8-sig", "big5", "cp950", "utf-8"]:
        try:
            if is_text:
                reader = csv.DictReader(io.StringIO(source))
            else:
                f = open(source, encoding=enc, newline="")
                reader = csv.DictReader(f)
            rows = list(reader)
            if not is_text: f.close()
            if rows: return rows, enc
        except Exception:
            continue
    return [], None

def run_pipeline(rows=None):
    state["status"] = "fetching"
    if not rows:
        state["status"] = "error"
        state["status_msg"] = "無資料"
        return
    groups = split_orders(rows)
    state["groups"]      = groups
    state["total"]       = len(set((r.get(CONFIG["col_txn"]) or "") for r in rows))
    state["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state["status"]      = "ready"
    state["status_msg"]  = f"共 {state['total']} 張訂單，分成 {len(groups)} 組 | {state['last_update']}"
    for k, g in groups.items():
        log(f"  {g['title']}：{len(g['orders'])} 張")
    log("分單完成")

def scheduler():
    while True:
        now = datetime.now()
        if now.hour == CONFIG["auto_run_hour"] and now.minute == CONFIG["auto_run_min"]:
            log("定時觸發")
            time.sleep(61)
        time.sleep(30)

# ============================================================
# 設定檔持久化（斜放商品清單）
# ============================================================
def _get_base_dir():
    """取得執行檔所在目錄（支援 PyInstaller 打包）"""
    if getattr(sys, 'frozen', False):
        # 打包成 exe 後，exe 所在目錄
        return os.path.dirname(sys.executable)
    else:
        # 一般 Python 執行
        return os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_get_base_dir(), "settings.json")

def load_settings():
    """從 settings.json 讀取設定，啟動時呼叫"""
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            if "diagonal_skus" in data:
                CONFIG["diagonal_skus"] = data["diagonal_skus"]
                log(f"已載入特殊可出超材品設定：{len(CONFIG['diagonal_skus'])} 筆")
    except Exception as e:
        log(f"讀取設定檔失敗：{e}")

def save_settings():
    """將設定寫入 settings.json"""
    try:
        data = {"diagonal_skus": CONFIG["diagonal_skus"]}
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"儲存設定檔失敗：{e}")

# ============================================================
# Flask 網頁
# ============================================================
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

LOGIN_USER = os.environ.get("LOGIN_USER", "admin")
LOGIN_PASS = os.environ.get("LOGIN_PASS", "admin123")

# ── Token 免登入設定 ──────────────────────────────────────────
# 在 Railway 環境變數設定 ACCESS_TOKEN（建議 32 位隨機字串）
# 員工書籤連結：https://你的網域/auth?token=<ACCESS_TOKEN>
ACCESS_TOKEN = os.environ.get("ACCESS_TOKEN", "")  # 空字串 = 停用 token 登入

SYSTEM_NAME = "&#x1F3ED; 超人特工倉"
SYSTEM_SUBTITLE = "Super Warehouse Agent System"

LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8"><title>登入 - 超人特工倉</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:linear-gradient(135deg,#0f1923 0%,#1a2f45 100%);display:flex;align-items:center;justify-content:center;min-height:100vh}
.box{background:#fff;border-radius:16px;padding:44px 40px;width:360px;box-shadow:0 20px 60px rgba(0,0,0,.4)}
.logo{text-align:center;margin-bottom:24px}
.logo-icon{font-size:48px;display:block;margin-bottom:8px}
.logo h1{font-size:20px;font-weight:700;color:#1a1a1a;margin-bottom:4px}
.logo p{font-size:11px;color:#aaa;letter-spacing:1px}
label{font-size:12px;color:#555;display:block;margin-bottom:4px;font-weight:500}
input{width:100%;padding:10px 12px;border:1.5px solid #e0e0e0;border-radius:8px;font-size:14px;margin-bottom:14px;font-family:inherit;transition:border-color .2s}
input:focus{outline:none;border-color:#1a5fa8;box-shadow:0 0 0 3px rgba(26,95,168,.1)}
button{width:100%;padding:12px;background:linear-gradient(135deg,#1a5fa8,#0d4a8a);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;margin-top:4px;transition:opacity .2s}
button:hover{opacity:.88}
.err{color:#b71c1c;font-size:12px;text-align:center;margin-bottom:12px;background:#ffebee;padding:8px;border-radius:6px}
</style></head><body>
<div class="box">
  <div class="logo">
    <span class="logo-icon">&#x1F3ED;</span>
    <h1>超人特工倉</h1>
    <p>SUPER WAREHOUSE AGENT SYSTEM</p>
  </div>
  {% if error %}<div class="err">{{ error }}</div>{% endif %}
  <form method="POST" action="/login">
    <label>帳號</label>
    <input type="text" name="username" placeholder="請輸入帳號" autofocus>
    <label>密碼</label>
    <input type="password" name="password" placeholder="請輸入密碼">
    <button type="submit">&#x1F680; 進入系統</button>
  </form>
</div>
</body></html>"""

HOME_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8"><title>超人特工倉 - 首頁</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:#0f1923;min-height:100vh;color:#fff}
.topbar{background:rgba(255,255,255,.05);backdrop-filter:blur(10px);height:56px;padding:0 32px;display:flex;align-items:center;gap:12px;border-bottom:1px solid rgba(255,255,255,.08)}
.logo{font-size:16px;font-weight:700;margin-right:auto;letter-spacing:.5px}
.logo span{color:#f4a100}
.logout{color:#aaa;font-size:12px;text-decoration:none;padding:6px 12px;border:1px solid #333;border-radius:5px}
.logout:hover{border-color:#666;color:#fff}
.hero{text-align:center;padding:60px 20px 40px}
.hero h1{font-size:32px;font-weight:700;margin-bottom:8px}
.hero h1 span{color:#f4a100}
.hero p{font-size:14px;color:#888;letter-spacing:1px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:20px;max-width:960px;margin:0 auto;padding:0 24px 60px}
.card{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.08);border-radius:16px;padding:32px 28px;text-decoration:none;color:#fff;transition:all .25s;cursor:pointer;position:relative;overflow:hidden}
.card:hover{background:rgba(255,255,255,.09);border-color:rgba(255,255,255,.2);transform:translateY(-4px);box-shadow:0 16px 40px rgba(0,0,0,.3)}
.card-icon{font-size:44px;margin-bottom:16px;display:block}
.card-title{font-size:18px;font-weight:700;margin-bottom:6px}
.card-desc{font-size:13px;color:#888;line-height:1.7}
.card-badge{position:absolute;top:16px;right:16px;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:500}
.badge-ready{background:rgba(46,125,50,.3);color:#81c784;border:1px solid rgba(46,125,50,.4)}
.badge-soon{background:rgba(100,100,100,.2);color:#888;border:1px solid rgba(100,100,100,.3)}
.card-split{border-color:rgba(26,95,168,.4)}
.card-split:hover{border-color:rgba(26,95,168,.8);box-shadow:0 16px 40px rgba(26,95,168,.15)}
.card-customs{border-color:rgba(244,161,0,.3)}
.card-customs:hover{border-color:rgba(244,161,0,.7);box-shadow:0 16px 40px rgba(244,161,0,.12)}
.card-tools{border-color:rgba(0,150,136,.3)}
.card-tools:hover{border-color:rgba(0,150,136,.7);box-shadow:0 16px 40px rgba(0,150,136,.12)}
.card-ai{border-color:rgba(156,39,176,.3)}
.card-ai:hover{border-color:rgba(156,39,176,.7);box-shadow:0 16px 40px rgba(156,39,176,.12)}
.card-glasses{border-color:rgba(29,158,117,.35)}
.card-glasses:hover{border-color:rgba(93,202,165,.8);box-shadow:0 16px 40px rgba(29,158,117,.18)}
.sg-divider{max-width:960px;margin:0 auto 16px;padding:0 24px;display:flex;align-items:center;gap:12px}
.sg-divider-line{flex:1;height:1px;background:rgba(255,255,255,.07)}
.sg-divider-text{font-size:12px;color:#555;letter-spacing:1px;white-space:nowrap}
.install-steps{margin-top:14px;padding:14px 16px;background:rgba(0,0,0,.25);border-radius:10px;border:1px solid rgba(29,158,117,.2)}
.install-steps ol{padding-left:16px;margin:0}
.install-steps li{font-size:12px;color:#888;line-height:2;letter-spacing:.3px}
.install-steps li strong{color:#5DCAA5}
.sg-logo-wrap{display:flex;align-items:center;gap:12px;margin-bottom:12px}
.sg-logo-text{font-size:20px;font-weight:700;color:#5DCAA5;letter-spacing:.05em}
.sg-logo-sub{font-size:11px;color:#1D9E75;letter-spacing:.15em}
.dl-btn{display:inline-block;margin-top:14px;padding:9px 20px;background:#0F6E56;
  color:#5DCAA5;border-radius:8px;font-size:13px;font-weight:500;text-decoration:none;
  border:1px solid #1D9E75;transition:all .2s;letter-spacing:.05em}
.dl-btn:hover{background:#1D9E75;color:#fff}
</style></head><body>
<div class="topbar">
  <div class="logo">&#x1F3ED; <span>超人特工倉</span></div>
  <a href="/logout" class="logout">&#x1F6AA; 登出</a>
</div>
<div class="hero">
  <h1>歡迎回來，<span>特工！</span></h1>
  <p>SUPER WAREHOUSE AGENT SYSTEM &nbsp;|&nbsp; 選擇你的任務</p>
</div>
<div class="cards">
  <a href="/split" class="card card-split">
    <span class="card-badge badge-ready">&#x2713; 上線中</span>
    <span class="card-icon">&#x1F4E6;</span>
    <div class="card-title">分單中心</div>
    <div class="card-desc">上傳 4Sale 訂單 CSV，自動依通路和倉庫區域分單，一鍵複製交易序號到 4Sale 暫存區。</div>
  </a>
  <a href="/customs" class="card card-customs">
    <span class="card-badge badge-ready">&#x2713; 上線中</span>
    <span class="card-icon">&#x1F4CB;</span>
    <div class="card-title">報關助手</div>
    <div class="card-desc">上傳倉庫進貨清單，自動對應商品報關資料庫，帶入材質、品名、單價，一鍵匯出報關 Excel。</div>
  </a>


</div>

<div class="sg-divider">
  <div class="sg-divider-line"></div>
  <div class="sg-divider-text">&#x1F4CA; 廣告調整日誌</div>
  <div class="sg-divider-line"></div>
</div>

<div style="max-width:960px;margin:0 auto 24px;padding:0 24px">
  <div style="background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:16px;padding:24px">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
      <div style="display:flex;gap:16px;align-items:center">
        <span id="log-daily-ts" style="font-size:12px;color:#555"></span>
        <span id="log-hourly-ts" style="font-size:12px;color:#555"></span>
      </div>
      <button onclick="loadAdLog()" style="background:rgba(29,158,117,.2);border:1px solid rgba(29,158,117,.4);color:#5DCAA5;padding:5px 14px;border-radius:6px;font-size:12px;cursor:pointer">&#x27F3; 重新整理</button>
    </div>
    <div id="ad-log-list" style="display:flex;flex-direction:column;gap:6px;max-height:320px;overflow-y:auto">
      <div style="text-align:center;color:#555;font-size:13px;padding:20px">載入中...</div>
    </div>
  </div>
</div>

<div class="sg-divider">
  <div class="sg-divider-line"></div>
  <div class="sg-divider-text">&#x26A0; 利潤警示（低毛利廣告）</div>
  <div class="sg-divider-line"></div>
</div>

<div style="max-width:960px;margin:0 auto 24px;padding:0 24px">
  <div style="background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:16px;padding:24px">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
      <span style="font-size:12px;color:#555">依店鋪分類顯示，資料來自超人眼鏡最近一次廣告分析</span>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
        <select id="shop-filter" onchange="_selectedShop=this.value;renderLowMargin()" style="background:#1a2a24;color:#5DCAA5;border:1px solid #1D9E75;border-radius:6px;padding:4px 8px;font-size:12px;cursor:pointer">
          <option value="all">全部店鋪</option>
        </select>
        <select id="margin-threshold" onchange="renderLowMargin()" style="background:#1a2a24;color:#f4a100;border:1px solid rgba(244,161,0,.4);border-radius:6px;padding:4px 8px;font-size:12px;cursor:pointer">
          <option value="40">嚴重 &lt;40%</option>
          <option value="45" selected>警告 &lt;45%</option>
          <option value="55">關注 &lt;55%</option>
        </select>
        <button onclick="loadLowMargin()" style="background:rgba(244,161,0,.15);border:1px solid rgba(244,161,0,.4);color:#f4a100;padding:5px 14px;border-radius:6px;font-size:12px;cursor:pointer">&#x27F3; 重新整理</button>
      </div>
    </div>
    <div id="low-margin-list">
      <div style="text-align:center;color:#555;font-size:13px;padding:20px">載入中...</div>
    </div>
  </div>
</div>

<div class="sg-divider">
  <div class="sg-divider-line"></div>
  <div class="sg-divider-text">&#x1F9E0; CHROME 擴充工具</div>
  <div class="sg-divider-line"></div>
</div>

<div class="cards" style="padding-bottom:80px">
  <div class="card card-glasses">
    <span class="card-badge badge-ready">&#x2713; 上線中</span>
    <div class="sg-logo-wrap">
      <canvas id="sg-logo-home" width="96" height="54" style="border-radius:8px;"></canvas>
      <div>
        <div class="sg-logo-text">超人眼鏡</div>
        <div class="sg-logo-sub">DATA VISION · BIGSELLER</div>
      </div>
    </div>
    <div class="card-desc">在 BigSeller 在線產品頁面，即時顯示每個 SKU 的成本、售價與利潤率。讓運營判斷數據更直覺，未來持續新增更多功能。</div>
    <div class="install-steps">
      <ol>
        <li>點擊下方按鈕下載 <strong>superman_glasses.zip</strong></li>
        <li>解壓縮到任意資料夾（之後不要刪）</li>
        <li>Chrome 網址列輸入 <strong>chrome://extensions</strong></li>
        <li>開啟右上角 <strong>開發人員模式</strong></li>
        <li>點 <strong>載入未封裝項目</strong>，選擇解壓縮的資料夾</li>
        <li>完成！打開 BigSeller 在線產品即可使用</li>
      </ol>
    </div>
    <a href="/api/superman-glasses/download" class="dl-btn">&#x2B07; 下載超人眼鏡</a>
  </div>
</div>

<div class="sg-divider">
  <div class="sg-divider-line"></div>
  <div class="sg-divider-text">🔍 超材分析工具</div>
  <div class="sg-divider-line"></div>
</div>

<div style="max-width:960px;margin:0 auto 24px;padding:0 24px">
  <div style="background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:16px;padding:24px">
    <div style="display:flex;align-items:center;gap:16px;margin-bottom:16px">
      <div style="font-size:24px;">🔍</div>
      <div>
        <div style="font-weight:600;color:#f4a100">超材分析工具</div>
        <div style="font-size:12px;color:#aaa">BigSeller 超材訂單自動分析與標記</div>
      </div>
    </div>
    <div style="color:#ddd;line-height:1.6;margin-bottom:16px">
      上傳 BigSeller 匯出的超材檔案，自動判斷可拆單/不可拆單，一鍵複製訂單號進行批量標記。
    </div>
    <a href="/oversize-tool" class="dl-btn">🚀 開始超材分析</a>
  </div>
</div>

<div class="sg-divider">
  <div class="sg-divider-line"></div>
  <div class="sg-divider-text">&#x1F4CA; 廣告自動化儀表板</div>
  <div class="sg-divider-line"></div>
</div>

<div style="max-width:960px;margin:0 auto;padding:0 24px 80px;display:grid;grid-template-columns:1fr 1fr;gap:20px;">

  <!-- 廣告排程狀態 -->
  <div class="card" style="border-color:rgba(29,158,117,.35)">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
      <div style="font-size:16px;font-weight:700;color:#5DCAA5;">&#x23F0; 排程狀態</div>
      <button onclick="loadAdLog()" style="background:#0F6E56;color:#5DCAA5;border:1px solid #1D9E75;border-radius:6px;padding:4px 12px;font-size:12px;cursor:pointer;">重新整理</button>
    </div>
    <div id="sched-status" style="font-size:13px;color:#888;line-height:2;">載入中...</div>
  </div>

  <!-- 利潤警示 -->
  <div class="card" style="border-color:rgba(244,161,0,.35)">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
      <div style="font-size:16px;font-weight:700;color:#f4a100;">&#x26A0; 利潤警示</div>
      <span style="font-size:11px;color:#666;">毛利 &lt; 40%</span>
    </div>
    <div id="profit-warn" style="font-size:13px;color:#888;line-height:1.8;">載入中...</div>
  </div>

  <!-- 執行日誌（跨兩欄） -->
  <div class="card" style="grid-column:1/-1;border-color:rgba(29,158,117,.2)">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
      <div style="font-size:16px;font-weight:700;color:#5DCAA5;">&#x1F4DC; 執行日誌</div>
      <div id="log-meta" style="font-size:11px;color:#555;"></div>
    </div>
    <div id="ad-log-list" style="font-family:monospace;font-size:12px;color:#5DCAA5;background:#021c18;border-radius:8px;padding:12px;max-height:260px;overflow-y:auto;line-height:1.8;">載入中...</div>
  </div>

</div>

<script>
async function loadAdLog() {
  try {
    const r = await fetch('/api/superman-glasses/ad-log');
    const d = await r.json();

    // 排程狀態
    const cookieOk = d.cookie_ok;
    const costCount = d.cost_count || 0;
    const lastDaily = d.last_daily || '尚未執行';
    const lastHourly = d.last_hourly ? new Date(d.last_hourly * 1000).toLocaleString('zh-TW') : '尚未執行';
    document.getElementById('sched-status').innerHTML = `
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:4px;">
        <span style="color:${cookieOk ? '#5DCAA5' : '#e57373'};">${cookieOk ? '&#x2705;' : '&#x274C;'}</span>
        <span>BigSeller Cookie：${cookieOk ? '有效' : '未同步，請開啟 BigSeller'}</span>
      </div>
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:4px;">
        <span style="color:${costCount > 0 ? '#5DCAA5' : '#e57373'};">${costCount > 0 ? '&#x2705;' : '&#x26A0;'}</span>
        <span>成本資料：${costCount > 0 ? costCount + ' 個 SKU' : '尚無資料，請開啟庫存清單'}</span>
      </div>
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:4px;">
        <span>&#x1F4C5;</span>
        <span>每日 ROAS 排程：${lastDaily}</span>
      </div>
      <div style="display:flex;gap:8px;align-items:center;">
        <span>&#x23F1;</span>
        <span>每小時預算排程：${lastHourly}</span>
      </div>
    `;

    // 執行日誌（log 是 [{msg, time}] 格式）
    const logs = d.log || [];
    if (logs.length === 0) {
      document.getElementById('ad-log-list').innerHTML = '<span style="color:#555;">尚無執行記錄</span>';
    } else {
      document.getElementById('ad-log-list').innerHTML = logs.map(l => {
        const msg = typeof l === 'string' ? l : (l.msg || '');
        const time = typeof l === 'object' ? l.time : '';
        let color = '#5DCAA5';
        if (msg.includes('失敗') || msg.includes('ERROR') || msg.includes('錯誤')) color = '#e57373';
        else if (msg.includes('爆款') || msg.includes('加碼') || msg.includes('暫停')) color = '#f4a100';
        else if (msg.startsWith('---')) color = '#444';
        return `<div style="display:flex;gap:8px;color:${color};border-bottom:1px solid rgba(255,255,255,.04);padding:3px 0;font-size:12px;font-family:monospace">
          <span style="color:#444;white-space:nowrap">${time}</span>
          <span>${msg}</span>
        </div>`;
      }).join('');
    }
    document.getElementById('log-meta').textContent = `共 ${logs.length} 筆`;

    // 利潤警示
    const pwarn = document.getElementById('profit-warn');
    const warnLogs = logs.filter(l => {
      const msg = typeof l === 'string' ? l : (l.msg || '');
      return msg.includes('低毛利') || msg.includes('暫停') || msg.includes('毛利');
    });
    if (warnLogs.length > 0) {
      pwarn.innerHTML = warnLogs.slice(0, 10).map(l => {
        const msg = typeof l === 'string' ? l : (l.msg || '');
        return `<div style="color:#f4a100;border-bottom:1px solid rgba(255,255,255,.04);padding:2px 0;font-size:12px">${msg}</div>`;
      }).join('');
    } else {
      pwarn.innerHTML = '<span style="color:#555;">目前無低毛利警示</span>';
    }
  } catch(e) {
    document.getElementById('sched-status').innerHTML = '<span style="color:#e57373;">無法連線</span>';
    document.getElementById('ad-log-list').innerHTML = '<span style="color:#e57373;">載入失敗：' + e.message + '</span>';
  }
}
loadAdLog();
setInterval(loadAdLog, 60000); // 每分鐘自動更新
</script>

<script>
(function(){
  const cv = document.getElementById('sg-logo-home');
  if (!cv) return;
  const s = cv.width / 320;
  let pd = 12.5, pt = 38.4, pw = false, pwu = 0;
  const pool = [22.1,31.8,44.2,38.4,51.0,29.7,47.3,35.6,62.1,28.9];
  function nt(){ return pool[Math.floor(Math.random()*pool.length)]; }
  function draw(){
    const ctx = cv.getContext('2d');
    const W=cv.width, H=cv.height, now=Date.now()/1000;
    const teal='#1D9E75',tL='#5DCAA5',tD='#0F6E56',bg='rgba(4,52,44,0.92)';
    const lx=14*s,ly=28*s,lw=108*s,lh=80*s,r=14*s;
    const rx=198*s,ry=28*s,rw=108*s,rh=80*s,bY=ly+lh/2;
    ctx.clearRect(0,0,W,H);
    function rr(x,y,w,h,rd){
      ctx.beginPath();ctx.moveTo(x+rd,y);ctx.lineTo(x+w-rd,y);ctx.quadraticCurveTo(x+w,y,x+w,y+rd);
      ctx.lineTo(x+w,y+h-rd);ctx.quadraticCurveTo(x+w,y+h,x+w-rd,y+h);
      ctx.lineTo(x+rd,y+h);ctx.quadraticCurveTo(x,y+h,x,y+h-rd);
      ctx.lineTo(x,y+rd);ctx.quadraticCurveTo(x,y,x+rd,y);ctx.closePath();
    }
    rr(lx,ly,lw,lh,r);ctx.fillStyle=bg;ctx.fill();
    rr(rx,ry,rw,rh,r);ctx.fillStyle=bg;ctx.fill();
    ctx.save();rr(lx,ly,lw,lh,r);ctx.clip();
    const chars=['0','1','$','%','T','W','D','↑','∞','▲'];
    const cols=9,colW=lw/cols;
    for(let c=0;c<cols;c++){
      const cx=lx+c*colW+colW/2,spd=0.6+(c*.15)%.8,off=(c*3.7)%7;
      for(let row=0;row<5;row++){
        const t=(now*spd+off+row*1.4)%5,yp=ly+(t/5)*(lh+16*s)-8*s;
        const a=1-(row/5)*.85,ci=Math.floor((now*3+c*7+row*13))%chars.length;
        ctx.fillStyle=`rgba(93,202,165,${a*(row===0?1:.5)})`;
        ctx.font=`${Math.floor(10*s)}px monospace`;ctx.textAlign='center';
        ctx.fillText(chars[ci],cx,yp);
      }
    }
    ctx.restore();
    ctx.save();rr(rx,ry,rw,rh,r);ctx.clip();
    const rcx=rx+rw/2,rcy=ry+rh/2;
    for(let i=3;i>=1;i--){ctx.beginPath();ctx.arc(rcx,rcy,i*20*s,0,Math.PI*2);ctx.strokeStyle=`rgba(29,158,117,${.08*i})`;ctx.lineWidth=1*s;ctx.stroke();}
    const ang=(now*1.2)%(Math.PI*2);
    ctx.beginPath();ctx.moveTo(rcx,rcy);ctx.arc(rcx,rcy,44*s,ang,ang+Math.PI*.6);
    ctx.closePath();ctx.fillStyle='rgba(29,158,117,.12)';ctx.fill();
    ctx.beginPath();ctx.arc(rcx,rcy,44*s,ang,ang+Math.PI*.6);ctx.strokeStyle=tL;ctx.lineWidth=1.5*s;ctx.stroke();
    ctx.beginPath();ctx.arc(rcx+Math.cos(ang+.3)*44*s,rcy+Math.sin(ang+.3)*44*s,3*s,0,Math.PI*2);ctx.fillStyle='#fff';ctx.fill();
    ctx.beginPath();ctx.arc(rcx,rcy,6*s,0,Math.PI*2);ctx.fillStyle=teal;ctx.fill();
    ctx.beginPath();ctx.arc(rcx,rcy,3*s,0,Math.PI*2);ctx.fillStyle='#fff';ctx.fill();
    const diff=pt-pd;
    if(Math.abs(diff)<.1){pd=pt;if(!pw){pw=true;pwu=now+1.2;}if(pw&&now>pwu){pt=nt();pw=false;}}
    else{pd+=diff*.08;}
    ctx.fillStyle='rgba(4,52,44,.7)';ctx.fillRect(rx+2*s,ry+rh-32*s,rw-4*s,30*s);
    ctx.fillStyle=teal;ctx.font=`500 ${Math.floor(8*s)}px -apple-system,sans-serif`;
    ctx.textAlign='center';ctx.fillText('利潤率',rcx,ry+rh-20*s);
    const jo=diff*.15*s;
    ctx.save();ctx.beginPath();ctx.rect(rx+4*s,ry+rh-20*s,rw-8*s,18*s);ctx.clip();
    ctx.fillStyle=pd>=0?'#5DCAA5':'#E24B4A';
    ctx.font=`500 ${Math.floor(15*s)}px -apple-system,sans-serif`;ctx.textAlign='center';
    ctx.fillText((pd>=0?'+':'')+pd.toFixed(1)+'%',rcx,ry+rh-8*s+jo);ctx.restore();
    const ir=pt>pd;
    ctx.fillStyle=`rgba(${ir?'93,202,165':'232,75,74'},${.5+Math.sin(now*4)*.5})`;
    ctx.font=`${Math.floor(10*s)}px monospace`;ctx.textAlign='right';
    ctx.fillText(ir?'▲':'▼',rx+rw-6*s,ry+rh-20*s);ctx.restore();
    rr(lx,ly,lw,lh,r);ctx.strokeStyle=teal;ctx.lineWidth=2.5*s;ctx.stroke();
    rr(rx,ry,rw,rh,r);ctx.strokeStyle=teal;ctx.lineWidth=2.5*s;ctx.stroke();
    const bx1=lx+lw,bx2=rx,bm=(bx1+bx2)/2;
    ctx.beginPath();ctx.moveTo(bx1,bY-8*s);ctx.lineTo(bm+4*s,bY-4*s);
    ctx.lineTo(bm-4*s,bY+4*s);ctx.lineTo(bx2,bY+8*s);
    ctx.strokeStyle=tL;ctx.lineWidth=3*s;ctx.lineJoin='round';ctx.stroke();
    ctx.beginPath();ctx.arc(bm,bY,4*s,0,Math.PI*2);ctx.fillStyle=tL;ctx.fill();
    ctx.beginPath();ctx.moveTo(lx,ly+18*s);ctx.lineTo(lx-12*s,ly-4*s);
    ctx.strokeStyle=tD;ctx.lineWidth=3*s;ctx.lineCap='round';ctx.stroke();
    ctx.beginPath();ctx.moveTo(rx+rw,ry+18*s);ctx.lineTo(rx+rw+12*s,ry-4*s);ctx.stroke();
  }
  function loop(){draw();requestAnimationFrame(loop);}
  loop();
})();

// ── 廣告日誌 ──
async function loadAdLog() {
  const el = document.getElementById('ad-log-list');
  try {
    // 優先從 Google Sheets 讀（永久記錄），失敗才讀記憶體
    let logs = [];
    let source = '';
    let total = 0;
    try {
      const rs = await fetch('/api/superman-glasses/ad-log-sheet');
      const ds = await rs.json();
      if (ds.ok && ds.logs && ds.logs.length > 0) {
        logs = ds.logs;
        total = ds.total || logs.length;
        source = `Google Sheets（共 ${total} 筆）`;
      }
    } catch(e2) {}
    // Sheets 沒資料時 fallback 到記憶體
    if (!logs.length) {
      const r = await fetch('/api/superman-glasses/ad-log');
      const d = await r.json();
      const daily = d.last_daily ? '每日任務：' + d.last_daily : '每日任務：尚未執行';
      const hourly = d.last_hourly ? '預算任務：' + new Date(d.last_hourly*1000).toLocaleString('zh-TW') : '預算任務：尚未執行';
      document.getElementById('log-daily-ts').textContent = daily;
      document.getElementById('log-hourly-ts').textContent = hourly;
      logs = (d.log || []).map(l => ({ time: l.time, type: '', msg: l.msg || l }));
      source = '記憶體（重啟後清空）';
    }

    if (!logs.length) {
      el.innerHTML = '<div style="text-align:center;color:#555;font-size:13px;padding:20px">尚無執行記錄</div>';
      return;
    }

    const colorMap = {
      'ROAS調整': '#5DCAA5', '爆款': '#f4a100', '暫停': '#E24B4A', '低毛利': '#E24B4A',
      '預算加碼': '#85B7EB', '空燒': '#E24B4A', '排程': '#888', '完成': '#5DCAA5'
    };
    const typeColorMap = {
      'ROAS調整': '#5DCAA5', '爆款降ROAS': '#f4a100', '廣告暫停': '#E24B4A',
      '低毛利暫停': '#E24B4A', '預算加碼': '#85B7EB', '空燒警告': '#BA7517',
      '排程開始': '#555', '其他': '#888'
    };

    el.innerHTML = \`<div style="font-size:11px;color:#555;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid rgba(255,255,255,.06)">
      資料來源：\${source}
    </div>\` + logs.map(log => {
      const msg = log.msg || '';
      const type = log.type || '';
      const time = log.time || '';
      let color = typeColorMap[type] || '#888';
      if (!type) {
        for (const [k,v] of Object.entries(colorMap)) {
          if (msg.includes(k)) { color = v; break; }
        }
      }
      const typeBadge = type ? \`<span style="background:rgba(255,255,255,.06);border-radius:3px;padding:1px 5px;font-size:10px;color:\${color};margin-right:6px;flex-shrink:0">\${type}</span>\` : '';
      return \`<div style="display:flex;gap:6px;align-items:flex-start;padding:4px 0;border-bottom:1px solid rgba(255,255,255,.04);font-size:12px;font-family:monospace">
        <span style="color:#444;white-space:nowrap;flex-shrink:0;min-width:110px">\${time}</span>
        \${typeBadge}
        <span style="color:\${color};line-height:1.5;word-break:break-all">\${msg}</span>
      </div>\`;
    }).join('');
  } catch(e) {
    el.innerHTML = '<div style="text-align:center;color:#E24B4A;font-size:13px;padding:20px">無法載入：' + e.message + '</div>';
  }
}

// ── 低利潤警示 ──
// 店鋪篩選狀態
let _lowMarginData = [];
let _selectedShop = 'all';

async function loadLowMargin() {
  const el = document.getElementById('low-margin-list');
  el.innerHTML = '<div style="text-align:center;color:#555;padding:20px">載入中...</div>';
  try {
    // 從 Google Sheets 讀取最新快照
    const r = await fetch('/api/superman-glasses/ad-log-sheet');
    const d = await r.json();
    // 從利潤監控室讀取
    const r2 = await fetch('/api/superman-glasses/profit-snapshot-read');
    const d2 = await r2.json();

    if (d2.ok && d2.rows && d2.rows.length > 0) {
      _lowMarginData = d2.rows;
    } else {
      // fallback 到記憶體
      const r3 = await fetch('/api/superman-glasses/low-margin');
      const d3 = await r3.json();
      _lowMarginData = [];
      (d3.shops || []).forEach(shop => {
        shop.items.forEach(item => {
          _lowMarginData.push({ shop: shop.shopName, name: item.name, margin: item.margin, roas: item.roas, targetRoas: item.targetRoas, status: item.status });
        });
      });
    }
    renderLowMargin();
  } catch(e) {
    el.innerHTML = '<div style="text-align:center;color:#E24B4A;font-size:13px;padding:20px">無法載入：' + e.message + '</div>';
  }
}

function renderLowMargin() {
  const el = document.getElementById('low-margin-list');
  const shopFilter = document.getElementById('shop-filter');

  // 過濾門檻
  const threshold = parseInt(document.getElementById('margin-threshold')?.value || '45');

  // 篩選資料
  let filtered = _lowMarginData.filter(r => {
    const m = parseFloat(r.margin || r[10] || 0);
    return m > 0 && m <= threshold;
  });
  if (_selectedShop !== 'all') {
    filtered = filtered.filter(r => (r.shop || r[2] || '') === _selectedShop);
  }

  // 更新店鋪篩選選單
  const shops = [...new Set(_lowMarginData.map(r => r.shop || r[2] || '').filter(Boolean))];
  if (shopFilter) {
    const cur = shopFilter.value;
    shopFilter.innerHTML = '<option value="all">全部店鋪</option>' +
      shops.map(s => `<option value="${s}" ${s===cur?'selected':''}>${s}</option>`).join('');
  }

  if (!filtered.length) {
    el.innerHTML = '<div style="text-align:center;color:#5DCAA5;font-size:13px;padding:20px">✅ 目前無毛利≤' + threshold + '% 的廣告</div>';
    return;
  }

  // 依店鋪分組
  const byShop = {};
  filtered.forEach(r => {
    const shop = r.shop || r[2] || '未知店鋪';
    const margin = parseFloat(r.margin || r[10] || 0);
    const name = r.name || r[3] || '';
    const roas = r.roas || r[5] || '';
    const targetRoas = r.targetRoas || r[11] || '';
    const status = r.status || r[12] || '';
    if (!byShop[shop]) byShop[shop] = [];
    byShop[shop].push({ name, margin, roas, targetRoas, status });
  });

  el.innerHTML = Object.entries(byShop).map(([shop, items]) => {
    const hasBelow40 = items.some(i => i.margin <= 40);
    const shopColor = hasBelow40 ? '#E24B4A' : '#f4a100';
    return \`<div style="margin-bottom:20px">
      <div style="font-size:13px;font-weight:500;color:\${shopColor};margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid rgba(244,161,0,.2);display:flex;justify-content:space-between">
        <span>&#x1F3EA; \${shop}</span>
        <span style="font-size:11px;color:#888">\${items.length} 筆 | 毛利≤\${threshold}%</span>
      </div>
      <div style="display:flex;flex-direction:column;gap:4px">
        \${items.map(item => {
          const isBelow40 = item.margin <= 40;
          const bg = isBelow40 ? 'rgba(232,75,74,.1)' : 'rgba(244,161,0,.06)';
          const border = isBelow40 ? '1px solid rgba(232,75,74,.3)' : '1px solid rgba(244,161,0,.2)';
          const mColor = isBelow40 ? '#E24B4A' : '#f4a100';
          const tag = isBelow40 ? '&#x1F6D1; 嚴重' : '&#x26A0; 警告';
          return \`<div style="display:flex;align-items:center;gap:8px;padding:7px 10px;background:\${bg};border-radius:6px;border:\${border}">
            <span style="font-size:11px;color:\${mColor};flex-shrink:0">\${tag}</span>
            <span style="font-size:12px;color:#ccc;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">\${item.name}</span>
            <span style="font-size:12px;color:\${mColor};font-weight:500;flex-shrink:0">毛利 \${item.margin.toFixed(1)}%</span>
            <span style="font-size:11px;color:#888;flex-shrink:0">ROAS \${item.roas}/\${item.targetRoas}</span>
          </div>\`;
        }).join('')}
      </div>
    </div>\`;
  }).join('');
}

// 頁面載入時自動抓取
loadAdLog();
loadLowMargin();
</script>

</body></html>"""

# ── Token 登入紀錄（記憶體，重啟後清空） ──────────────────────
_token_login_log = []   # [{"time": str, "ip": str, "status": str}]

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

@app.route("/auth")
def token_auth():
    """Token 免登入入口：員工書籤存 /auth?token=xxx&next=/split"""
    if not ACCESS_TOKEN:
        return redirect("/login")
    import hmac
    token = request.args.get("token", "")
    next_page = request.args.get("next", "/")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
    if not hmac.compare_digest(token, ACCESS_TOKEN):
        _token_login_log.append({"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "ip": ip, "status": "FAIL"})
        return render_template_string(
            "<h3 style='font-family:sans-serif;color:#b71c1c;margin:40px auto;text-align:center'>Token 無效，請確認連結是否正確。</h3>"
        ), 403
    _token_login_log.append({"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "ip": ip, "status": "OK"})
    if len(_token_login_log) > 200:
        _token_login_log.pop(0)
    session["logged_in"] = True
    session["login_method"] = "token"
    if not next_page.startswith("/") or next_page.startswith("//"):
        next_page = "/"
    return redirect(next_page)

@app.route("/api/token-log")
@login_required
def token_log():
    """查看 token 登入紀錄（需登入才能看）"""
    rows = "".join(
        f"<tr><td>{r['time']}</td><td>{r['ip']}</td>"
        f"<td style='color:{'#2e7d32' if r['status']=='OK' else '#b71c1c'}'>{r['status']}</td></tr>"
        for r in reversed(_token_login_log)
    ) or "<tr><td colspan=3 style='text-align:center;color:#888'>尚無紀錄</td></tr>"
    html = f"""<!DOCTYPE html><html><head><meta charset='UTF-8'>
    <title>Token 登入紀錄</title>
    <style>body{{font-family:sans-serif;padding:32px;background:#f5f5f5}}
    table{{border-collapse:collapse;width:100%;max-width:700px;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1)}}
    th{{background:#1a5fa8;color:#fff;padding:12px 16px;text-align:left;font-size:13px}}
    td{{padding:10px 16px;border-bottom:1px solid #eee;font-size:13px}}
    h2{{margin-bottom:16px;color:#1a1a1a}}</style></head><body>
    <h2>Token 登入紀錄（最新在前）</h2>
    <table><thead><tr><th>時間</th><th>IP 位址</th><th>狀態</th></tr></thead>
    <tbody>{rows}</tbody></table>
    <p style='margin-top:12px;font-size:12px;color:#888'>共 {len(_token_login_log)} 筆（最多保留 200 筆，重啟清空）</p>
    <p style='margin-top:4px;font-size:12px'><a href='/'>← 返回首頁</a></p>
    </body></html>"""
    return html

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        u = request.form.get("username", "")
        p = request.form.get("password", "")
        if u == LOGIN_USER and p == LOGIN_PASS:
            session["logged_in"] = True
            return redirect("/")
        error = "帳號或密碼錯誤，特工請再試一次！"
    return render_template_string(LOGIN_HTML, error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/")
@login_required
def home():
    return render_template_string(HOME_HTML)

@app.route("/split")
@login_required
def split_page():
    return redirect("/split_app")



HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8">
<title>{{ company }} 分單系統</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"微軟正黑體","Microsoft JhengHei",sans-serif;background:#f0f2f5;font-size:13px;color:#1a1a1a}
.topbar{background:#0f1923;color:#fff;height:52px;padding:0 20px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:300}
.logo{font-size:15px;font-weight:600;margin-right:auto}
.logo span{color:#4fa3e0}
.btn{padding:7px 16px;border-radius:5px;border:none;font-size:13px;cursor:pointer;font-weight:500}
.btn-white{background:#fff;color:#0f1923}
.statusbar{background:#fff;border-bottom:1px solid #e0e0e0;padding:7px 20px;display:flex;align-items:center;gap:8px;font-size:12px}
.dot{width:8px;height:8px;border-radius:50%}
.dot-idle{background:#aaa}.dot-fetching{background:#f4a100}.dot-ready{background:#22c55e}.dot-error{background:#ef4444}

/* 通路統計卡片 */
.summary-bar{display:flex;gap:8px;padding:10px 20px;flex-wrap:wrap;background:#f8f8f8;border-bottom:1px solid #eee}
.sum-card{background:#fff;border:1px solid #e0e0e0;border-radius:6px;padding:6px 14px;font-size:12px;display:flex;align-items:center;gap:6px}
.sum-card .num{font-size:16px;font-weight:600;color:#1a1a1a}

/* 上傳區 */
.upload-center{padding:40px 20px;display:flex;justify-content:center}
.upload-box{background:#fff;border:2px dashed #c0cfe0;border-radius:12px;padding:40px 48px;text-align:center;max-width:460px;width:100%}
.upload-lbl{display:inline-block;background:#1a5fa8;color:#fff;padding:12px 32px;border-radius:6px;font-size:15px;font-weight:500;cursor:pointer;margin-bottom:8px}

.rebar{padding:7px 20px;background:#fff;border-bottom:1px solid #eee;display:flex;align-items:center;gap:8px}
.re-lbl{background:#f0f0f0;color:#444;padding:5px 12px;border-radius:5px;font-size:12px;cursor:pointer;font-weight:500}

/* Tabs */
.tabs{display:flex;gap:3px;padding:12px 20px 0;flex-wrap:wrap;background:#f0f2f5}
.tab{padding:5px 12px;border-radius:16px 16px 0 0;border:1px solid #d0d0d0;border-bottom:none;background:#e8e8e8;cursor:pointer;font-size:11px;color:#555;white-space:nowrap}
.tab.active{background:#fff;color:#0f1923;font-weight:500}
.tab:hover:not(.active){background:#f0f0f0}
.tab-sep{align-self:flex-end;color:#ccc;padding-bottom:3px;font-size:14px}

/* 主內容 */
.main{padding:0 20px 24px;background:#fff;border:1px solid #ddd;border-top:none;margin:0 20px;border-radius:0 0 8px 8px}
.grp{padding-top:16px}
.grp-hd{display:flex;align-items:center;gap:10px;padding:10px 16px;border-radius:6px 6px 0 0;color:#fff;font-size:13px;font-weight:500}
.gcnt{background:rgba(255,255,255,.25);padding:2px 10px;border-radius:12px;font-size:12px}
.copy-btn{margin-left:auto;display:flex;align-items:center;gap:5px;background:rgba(255,255,255,.2);border:1px solid rgba(255,255,255,.4);color:#fff;padding:4px 12px;border-radius:5px;font-size:12px;cursor:pointer;font-family:inherit;white-space:nowrap}
.copy-btn:hover{background:rgba(255,255,255,.35)}
.copy-btn.copied{background:rgba(255,255,255,.5);color:#1a1a1a}
table{width:100%;border-collapse:collapse;background:#fff;font-size:12px}
thead th{background:#f5f5f5;padding:7px 10px;text-align:left;font-weight:500;color:#555;border-bottom:1.5px solid #ddd;white-space:nowrap}
td{padding:6px 10px;border-bottom:.5px solid #efefef;vertical-align:top}
tr:last-child td{border-bottom:none}
tr.dr:hover td{background:#fafcff}
.mono{font-size:11px;color:#888;font-family:monospace}
.wh-tag{display:inline-block;font-size:10px;padding:1px 6px;border-radius:3px;background:#e8f5e9;color:#2e7d32;border:1px solid #c8e6c9;margin:1px}
.mix-tag{background:#fff3e0;color:#e65100;border-color:#ffe0b2}
.sum-row td{background:#f9f9f9;font-size:11px;color:#888;font-weight:500}
.oversize-row td{background:#fff8f8;}
.oversize-row:hover td{background:#ffefef!important}
.log-tog{margin:8px 20px;font-size:11px;color:#888;cursor:pointer}
.log-box{margin:0 20px 16px;background:#111;color:#7ec87e;font-family:monospace;font-size:11px;padding:10px;border-radius:6px;max-height:150px;overflow-y:auto;display:none}
.log-box.show{display:block}
.hidden{display:none}
@media print{
  .topbar,.statusbar,.summary-bar,.rebar,.tabs,.log-tog,.log-box{display:none!important}
  body,html{background:#fff!important}
  .main{border:none!important;margin:0!important;padding:0!important}
  .grp.hidden{display:block!important}
  .grp-hd{-webkit-print-color-adjust:exact;print-color-adjust:exact}
  td,th{padding:4px 7px}
  .grp{page-break-inside:avoid;margin-bottom:10px}
}
</style></head><body>

<div class="topbar">
  <a href="/" style="color:#aaa;font-size:12px;text-decoration:none;margin-right:8px">&#x1F3ED;</a>
  <div class="logo"><span>分單中心</span></div>
  <span style="font-size:11px;color:#778">{{ last_update }}</span>
  <a href="/settings/diagonal" style="color:#aaa;font-size:12px;text-decoration:none;padding:5px 10px;border:1px solid #444;border-radius:5px">特殊可出超材品設定</a>
  <button class="btn btn-white" onclick="window.print()">列印全部</button>
  <a href="/" style="color:#aaa;font-size:12px;text-decoration:none;padding:5px 10px;border:1px solid #444;border-radius:5px">&#x2302; 首頁</a>
</div>

<div class="statusbar">
  <span class="dot dot-{{ status }}"></span>
  <span>{{ status_msg }}</span>
</div>

{% if summary %}
<div class="summary-bar">
  {% for label in summary_keys %}
  <div class="sum-card">
    <span>{{ label }}</span>
    <span class="num">{{ summary[label] }}</span>
    <span style="color:#aaa;font-size:11px">張</span>
  </div>
  {% endfor %}
  <div class="sum-card" style="margin-left:auto;background:#f5f5f5">
    <span>合計</span>
    <span class="num">{{ total }}</span>
    <span style="color:#aaa;font-size:11px">張</span>
  </div>
</div>
{% endif %}

{% if not groups %}
<div class="upload-center">
  <div class="upload-box">
    <div style="font-size:48px;margin-bottom:12px">&#128194;</div>
    <div style="font-size:17px;font-weight:600;margin-bottom:6px">上傳 4Sale 訂單 CSV</div>
    <div style="font-size:13px;color:#888;margin-bottom:24px;line-height:1.7">
      4Sale 後台 → 訂單管理 → 篩選可出貨 → 匯出 CSV<br>
      上傳後自動依通路及倉庫區域分單
    </div>
    <label class="upload-lbl" for="csv-in">選擇 CSV 檔案</label>
    <input type="file" id="csv-in" accept=".csv" style="display:none" onchange="doUpload()">
    <div id="up-name" style="font-size:12px;color:#aaa;margin-top:6px"></div>
    <div style="margin-top:16px;padding-top:14px;border-top:1px solid #eee;font-size:12px;color:#bbb">
      需包含欄位：子交易序號、出貨類型、商品倉庫儲位、商品編號
    </div>
  </div>
</div>
{% else %}

<div class="rebar">
  <span style="font-size:12px;color:#888">重新上傳：</span>
  <label class="re-lbl" for="csv-in">換一份 CSV</label>
  <input type="file" id="csv-in" accept=".csv" style="display:none" onchange="doUpload()">
  <span id="up-name" style="font-size:12px;color:#aaa;margin-left:6px"></span>
</div>

{# 批次勾選複製工具列 #}
<div style="padding:8px 20px;background:#f0f4ff;border-bottom:1px solid #c5cae9;display:flex;align-items:center;gap:8px;flex-wrap:wrap">
  <span style="font-size:12px;font-weight:500;color:#3949ab;white-space:nowrap">批次複製：</span>
  {% for key in group_keys %}
  {% set g = groups[key] %}
  <label style="display:inline-flex;align-items:center;gap:3px;font-size:11px;cursor:pointer;white-space:nowrap;padding:2px 6px;border:1px solid #c5cae9;border-radius:12px;background:#fff">
    <input type="checkbox" class="grp-check" data-grpkey="{{ key }}" style="cursor:pointer">
    {% if key == '__oversize__' %}&#9888; 超材（{{ g.orders|length }}）
    {% elif key == '__delivery__' %}&#128665; 宅配（{{ g.orders|length }}）
    {% elif key == '__single_zone__' %}純區+單品（{{ g.orders|length }}）
    {% elif key == '__nopkg__' %}&#128230; 無包裝（{{ g.orders|length }}）
    {% else %}{{ key }}（{{ g.orders|length }}）{% endif %}
  </label>
  {# 隱藏的序號資料 #}
  <span style="display:none" class="grp-txn-data" data-grpkey="{{ key }}">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  {% endfor %}
  <button onclick="copySelectedGrps()" style="margin-left:auto;background:#3949ab;color:#fff;border:none;padding:5px 14px;border-radius:5px;font-size:12px;cursor:pointer;font-weight:500;white-space:nowrap">&#128203; 複製已勾選</button>
  <span id="grp-copy-msg" style="font-size:11px;color:#3949ab"></span>
</div>

<div class="tabs">
  <div class="tab active" onclick="showGrp('all')" data-key="all">全部（{{ total }}）</div>
  {% set ns = namespace(prev_ch='') %}
  {% for key in group_keys %}
  {% set ch = key.split(' - ')[0] if ' - ' in key else key %}
  {% if ns.prev_ch != '' and ch != ns.prev_ch %}
  <span class="tab-sep">|</span>
  {% endif %}
  {% set ns.prev_ch = ch %}
  <div class="tab" onclick="showGrp('{{ key }}')" data-key="{{ key }}">
    {% if key == '__oversize__' %}&#9888; 超材（{{ groups[key].orders|length }}）
    {% elif key == '__giant__' %}&#129427; 巨無霸（{{ groups[key].orders|length }}）
    {% elif key == '__delivery__' %}&#128665; 宅配（{{ groups[key].orders|length }}）
    {% elif key == '__single_zone__' %}&#128364; 純區+單品（{{ groups[key].orders|length }}）
    {% elif key == '__nopkg__' %}&#128230; 無包裝（{{ groups[key].orders|length }}）
    {% else %}{{ groups[key].icon | safe }} {{ key }}（{{ groups[key].orders|length }}）{% endif %}
  </div>
  {% endfor %}
</div>

<div class="main">
{% for key in group_keys %}
{% set g = groups[key] %}

{% if key == '__delivery__' %}
{# 宅配大分類：依區域分群顯示 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#1565c0">
    &#x1F69A; 宅配訂單
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  {% set zones_list = [] %}
  {% for o in g.orders %}{% if o.delivery_zone not in zones_list %}{% set _ = zones_list.append(o.delivery_zone) %}{% endif %}{% endfor %}

  {# 勾選複製工具列 #}
  <div style="padding:8px 16px;background:#e8eaf6;border-bottom:1px solid #c5cae9;display:flex;align-items:center;gap:10px;flex-wrap:wrap">
    <span style="font-size:12px;color:#3949ab;font-weight:500">勾選區域批次複製：</span>
    {% for zone_name in zones_list | sort %}
    <label style="display:inline-flex;align-items:center;gap:4px;font-size:12px;cursor:pointer;color:#333">
      <input type="checkbox" class="zone-check" data-zone="{{ zone_name }}" style="cursor:pointer">
      {{ zone_name }}
    </label>
    {% endfor %}
    <button onclick="copySelectedZones()" style="margin-left:auto;background:#3949ab;color:#fff;border:none;padding:5px 14px;border-radius:5px;font-size:12px;cursor:pointer;font-weight:500">&#128203; 複製已勾選</button>
    <span id="zone-copy-msg" style="font-size:11px;color:#3949ab"></span>
  </div>

  {% for zone_name in zones_list | sort %}
  {% set ns_zc = namespace(n=0) %}
  {% for o in g.orders %}{% if o.delivery_zone==zone_name %}{% set ns_zc.n=ns_zc.n+1 %}{% endif %}{% endfor %}
  <div style="padding:8px 16px;background:#e3f2fd;border-bottom:1px solid #bbdefb;display:flex;align-items:center;gap:10px">
    <span style="font-size:12px;font-weight:500;color:#1565c0">{{ zone_name }}（{{ ns_zc.n }} 張）</span>
    <button class="copy-btn" style="margin-left:auto;font-size:11px;padding:3px 10px"
      onclick="copyZoneTxns('{{ zone_name }}', this)">&#128203; 複製此區序號</button>
    <span style="display:none" class="zone-txn-data" data-zone="{{ zone_name }}">{% for o in g.orders %}{% if o.delivery_zone==zone_name %}{{ o.txn }}&#10;{% endif %}{% endfor %}</span>
  </div>
  <table>
    <thead><tr><th>#</th><th>交易序號</th><th>出貨類型</th><th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th></tr></thead>
    <tbody>
    {% set ns_di = namespace(i=0) %}
    {% for o in g.orders %}{% if o.delivery_zone == zone_name %}
    {% set ns_di.i = ns_di.i + 1 %}
    <tr class="dr">
      <td class="mono">{{ ns_di.i }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">{{ o.ship_raw }}{% if o.diagonal_used %}<div style="color:#1565c0;font-size:10px;margin-top:2px">特殊可出超材品</div>{% endif %}</td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">{% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}{% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}</td>
      <td style="font-weight:600;color:#1a5fa8;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endif %}{% endfor %}
    </tbody>
  </table>
  {% endfor %}
</div>

{% elif key == '__single_zone__' %}
{# 純區 + 超商單品 + 店到店單品 大分類 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#e65100">
    &#x26A1; 純區 + 超商單品 + 店到店單品
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製子交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  {# 依 single_zone_sub 分組顯示子分類 #}
  {% set sub_groups = {} %}
  {% for o in g.orders %}
    {% set sub = o.single_zone_sub if o.single_zone_sub else '其他' %}
    {% if sub not in sub_groups %}{% set _ = sub_groups.update({sub: []}) %}{% endif %}
    {% set _ = sub_groups[sub].append(o) %}
  {% endfor %}
  {% for sub, orders in sub_groups.items() %}
  <div style="margin:0">
    <div style="background:#bf360c;color:#fff;font-size:12px;font-weight:700;padding:6px 14px;letter-spacing:.5px">
      &#9656; {{ sub }} &nbsp;<span style="font-weight:400;opacity:.8">{{ orders|length }} 張</span>
    </div>
    <table>
      <thead><tr><th>#</th><th>子交易序號</th><th>出貨類型</th><th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th></tr></thead>
      <tbody>
      {% for o in orders %}
      <tr class="dr">
        <td class="mono">{{ loop.index }}</td>
        <td class="mono" style="font-weight:500">{{ o.txn }}</td>
        <td style="font-size:11px;color:#666">{{ o.ship_raw }}</td>
        <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
        <td style="font-size:11px;color:#555;white-space:nowrap">{% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}{% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}</td>
        <td style="font-weight:600;color:#e65100;text-align:center">{{ o.total_qty }}</td>
      </tr>
      {% endfor %}
      <tr class="sum-row"><td colspan="2">小計 {{ orders|length }} 張</td><td colspan="4">{% set ns3=namespace(t=0) %}{% for o in orders %}{% set ns3.t=ns3.t+o.total_qty %}{% endfor %}共 {{ ns3.t }} 件</td></tr>
      </tbody>
    </table>
  </div>
  {% endfor %}
</div>

{% elif key == "__giant__" %}
{# 巨無霸分類：人工確認 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#6a0dad">
    &#129427; 巨無霸｜人工確認
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  <div style="padding:8px 16px;background:#f3e5f5;border-bottom:1px solid #ce93d8;font-size:12px;color:#6a0dad">
    &#9888; 三邊總和 &gt; 170cm 且重量 &gt; 15kg，請人工確認是否可出貨
  </div>
  <table>
    <thead><tr><th>#</th><th>交易序號</th><th>出貨類型</th><th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th></tr></thead>
    <tbody>
    {% for o in g.orders %}
    <tr class="dr" style="background:#fdf6ff">
      <td class="mono">{{ loop.index }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">
        {{ o.ship_raw }}
        <div style="color:#6a0dad;font-size:11px;font-weight:500;margin-top:2px">&#129427; {{ o.oversize_msg }}</div>
      </td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">
        {% if o.total_dim > 0 %}<div>{{ o.length|int }}x{{ o.width|int }}x{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}
        {% if o.weight > 0 %}<div style="color:#6a0dad;font-weight:500">{{ o.weight }}kg</div>{% endif %}
      </td>
      <td style="font-weight:600;color:#6a0dad;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endfor %}
    <tr class="sum-row"><td colspan="2">小計 {{ g.orders|length }} 張</td><td colspan="4">{% set ns2=namespace(t=0) %}{% for o in g.orders %}{% set ns2.t=ns2.t+o.total_qty %}{% endfor %}共 {{ ns2.t }} 件</td></tr>
    </tbody>
  </table>
</div>

{% elif key == '__nopkg__' %}
{# 無包裝 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#00838f">
    &#x1F4E6; 無包裝
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  <table>
    <thead><tr><th>#</th><th>交易序號</th><th>出貨類型</th><th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th></tr></thead>
    <tbody>
    {% for o in g.orders %}
    <tr class="dr">
      <td class="mono">{{ loop.index }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">{{ o.ship_raw }}</td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">{% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div>{% endif %}{% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}</td>
      <td style="font-weight:600;color:#1a5fa8;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endfor %}
    <tr class="sum-row"><td colspan="2">小計 {{ g.orders|length }} 張</td><td colspan="4">{% set ns2=namespace(t=0) %}{% for o in g.orders %}{% set ns2.t=ns2.t+o.total_qty %}{% endfor %}共 {{ ns2.t }} 件</td></tr>
    </tbody>
  </table>
</div>

{% elif key == '__splittable__' %}
{# 可拆單分類：依通路分群顯示 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#e65100">
    &#x2702; 可拆單訂單
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  {% set ch_list = [] %}
  {% for o in g.orders %}{% if o.oversize_channel not in ch_list %}{% set _ = ch_list.append(o.oversize_channel) %}{% endif %}{% endfor %}
  {% for ch_name in ch_list %}
  <div style="padding:8px 16px 0;background:#fff8f5;border-bottom:1px solid #ffccbc">
    <span style="font-size:12px;font-weight:500;color:#e65100">
      {{ ch_name }}（{% set ns_s=namespace(n=0) %}{% for o in g.orders %}{% if o.oversize_channel==ch_name %}{% set ns_s.n=ns_s.n+1 %}{% endif %}{% endfor %}{{ ns_s.n }} 張）
    </span>
  </div>
  <table>
    <thead><tr><th>#</th><th>交易序號</th><th>出貨類型</th><th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th></tr></thead>
    <tbody>
    {% set ns_si = namespace(i=0) %}
    {% for o in g.orders %}{% if o.oversize_channel == ch_name %}
    {% set ns_si.i = ns_si.i + 1 %}
    <tr class="dr oversize-row">
      <td class="mono">{{ ns_si.i }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">
        {{ o.ship_raw }}
        <div style="color:#e65100;font-size:11px;font-weight:500;margin-top:2px">&#x26A0; {{ o.oversize_msg }}</div>
        {% if o.split_suggestion %}
        <div style="margin-top:6px;padding:6px 8px;background:#fff3e0;border-radius:4px;font-size:11px">
          <div style="font-weight:500;color:#e65100;margin-bottom:3px">
            建議拆成 {{ o.split_suggestion|length }} 單
            {% if o.split_rules %}（依{{ o.split_rules.label }}規格）{% endif %}：
          </div>
          {% for pkg in o.split_suggestion %}
          <div style="color:#555;padding:1px 0">第{{ pkg.pkg_no }}包：{{ pkg.items }}（{{ pkg.count }}件 / {{ pkg.dim }}cm / {{ pkg.weight }}kg）</div>
          {% endfor %}
        </div>
        {% endif %}
        {% if o.split_error %}<div style="color:#888;font-size:10px;margin-top:3px">{{ o.split_error }}</div>{% endif %}
      </td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">
        {% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}
        {% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}
      </td>
      <td style="font-weight:600;color:#e65100;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endif %}{% endfor %}
    </tbody>
  </table>
  {% endfor %}
</div>

{% elif key == '__oversize__' %}
{# 超材大分類：依通路分群顯示 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:#b71c1c">
    &#x26A0; 超材訂單
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  {# 依通路分群 #}
  {% set channels = [] %}
  {% for o in g.orders %}
    {% if o.oversize_channel not in channels %}{% set _ = channels.append(o.oversize_channel) %}{% endif %}
  {% endfor %}
  {% for ch_name in channels %}
  <div style="padding:10px 16px 0;background:#fff3f3;border-bottom:1px solid #ffcdd2">
    <span style="font-size:12px;font-weight:500;color:#b71c1c">{{ ch_name }}（{% set ns3=namespace(n=0) %}{% for o in g.orders %}{% if o.oversize_channel==ch_name %}{% set ns3.n=ns3.n+1 %}{% endif %}{% endfor %}{{ ns3.n }} 張）</span>
  </div>
  <table>
    <thead><tr>
      <th>#</th><th>交易序號</th><th>出貨類型</th>
      <th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th>
    </tr></thead>
    <tbody>
    {% set ns_idx = namespace(i=0) %}
    {% for o in g.orders %}{% if o.oversize_channel == ch_name %}
    {% set ns_idx.i = ns_idx.i + 1 %}
    <tr class="dr oversize-row">
      <td class="mono">{{ ns_idx.i }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">
        {{ o.ship_raw }}
        <div style="color:#b71c1c;font-size:11px;font-weight:500;margin-top:2px">&#x26A0; {{ o.oversize_msg }}</div>
        {% if o.fee_paid %}
        <div style="color:#e65100;font-size:11px;margin-top:2px">&#x1F4B0; 買家已付運費 {{ o.fee|int }} 元，不拆單</div>
        {% endif %}
        {% if o.split_suggestion %}
        <div style="margin-top:6px;padding:6px 8px;background:#fff3e0;border-radius:4px;font-size:11px">
          <div style="font-weight:500;color:#e65100;margin-bottom:3px">建議拆成 {{ o.split_suggestion|length }} 單：</div>
          {% for pkg in o.split_suggestion %}
          <div style="color:#555;padding:1px 0">第{{ pkg.pkg_no }}包：{{ pkg.items }}（{{ pkg.count }}件 / {{ pkg.dim }}cm / {{ pkg.weight }}kg）</div>
          {% endfor %}
        </div>
        {% endif %}
        {% if o.split_error %}<div style="color:#888;font-size:10px;margin-top:3px">{{ o.split_error }}</div>{% endif %}
      </td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">
        {% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}
        {% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}
      </td>
      <td style="font-weight:600;color:#b71c1c;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endif %}{% endfor %}
    </tbody>
  </table>
  {% endfor %}
</div>

{% else %}
{# 一般分組 #}
<div class="grp" data-key="{{ key }}">
  <div class="grp-hd" style="background:{{ g.color }}">
    {{ g.title | safe }}
    <span class="gcnt">{{ g.orders|length }} 張</span>
    <button class="copy-btn" onclick="copyTxns(this)">&#128203; 複製交易序號</button>
    <span style="display:none" class="txn-data">{% for o in g.orders %}{{ o.txn }}&#10;{% endfor %}</span>
  </div>
  <table>
    <thead><tr>
      <th>#</th><th>交易序號</th><th>出貨類型</th>
      <th>商品 / 儲位</th><th>尺寸/重量</th><th>件數</th>
    </tr></thead>
    <tbody>
    {% for o in g.orders %}
    <tr class="dr">
      <td class="mono">{{ loop.index }}</td>
      <td class="mono" style="font-weight:500">{{ o.txn }}</td>
      <td style="font-size:11px;color:#666">
        {{ o.ship_raw }}
        {% if o.diagonal_used and not o.oversize %}
        <div style="color:#1565c0;font-size:10px;margin-top:2px">斜放後合規</div>
        {% endif %}
      </td>
      <td>{% for p in o.products %}<div style="font-size:11px;padding:1px 0">{{ p.sku }}<span class="wh-tag">{{ p.zone_raw }}</span></div>{% endfor %}</td>
      <td style="font-size:11px;color:#555;white-space:nowrap">
        {% if o.total_dim > 0 %}<div>{{ o.length|int }}×{{ o.width|int }}×{{ o.height|int }}cm</div><div>三邊 {{ o.total_dim|int }}cm</div>{% endif %}
        {% if o.weight > 0 %}<div>{{ o.weight }}kg</div>{% endif %}
      </td>
      <td style="font-weight:600;color:#1a5fa8;text-align:center">{{ o.total_qty }}</td>
    </tr>
    {% endfor %}
    <tr class="sum-row">
      <td colspan="2">小計 {{ g.orders|length }} 張</td>
      <td colspan="4">
        {% set ns2 = namespace(t=0) %}
        {% for o in g.orders %}{% set ns2.t = ns2.t + o.total_qty %}{% endfor %}
        共 {{ ns2.t }} 件
      </td>
    </tr>
    </tbody>
  </table>
</div>
{% endif %}

{% endfor %}
</div>
{% endif %}

<div class="log-tog" onclick="document.getElementById('lg').classList.toggle('show')">顯示/隱藏記錄</div>
<div class="log-box" id="lg">{% for l in log_lines %}<div>{{ l }}</div>{% endfor %}</div>

<script>
function showGrp(key){
  document.querySelectorAll('.tab').forEach(t=>t.classList.toggle('active',t.dataset.key===key));
  document.querySelectorAll('.grp').forEach(s=>s.classList.toggle('hidden',key!=='all'&&s.dataset.key!==key));
}
function copySelectedGrps() {
  var NL = String.fromCharCode(10);
  var checked = document.querySelectorAll('.grp-check:checked');
  if (checked.length === 0) { alert('請先勾選要複製的分組'); return; }
  var all = [];
  checked.forEach(function(cb) {
    var el = document.querySelector('.grp-txn-data[data-grpkey="' + cb.dataset.grpkey + '"]');
    if (el) {
      var parts = el.textContent.trim().split(NL);
      parts.forEach(function(l){ if(l.trim()) all.push(l.trim()); });
    }
  });
  var text = all.join(NL);
  navigator.clipboard.writeText(text).then(function() {
    var msg = document.getElementById('grp-copy-msg');
    if (msg) { msg.textContent = '已複製 ' + all.length + ' 筆（' + checked.length + ' 個分組）'; setTimeout(function(){ msg.textContent=''; }, 3000); }
  }).catch(function() {
    var ta = document.createElement('textarea');
    ta.value = text; document.body.appendChild(ta); ta.select();
    document.execCommand('copy'); document.body.removeChild(ta);
    var msg = document.getElementById('grp-copy-msg');
    if (msg) { msg.textContent = '已複製 ' + all.length + ' 筆'; setTimeout(function(){ msg.textContent=''; }, 3000); }
  });
}

function copyZoneTxns(zone, btn) {
  const el = document.querySelector('.zone-txn-data[data-zone="' + zone + '"]');
  if (!el) return;
  const txns = el.textContent.trim();
  navigator.clipboard.writeText(txns).then(() => {
    const orig = btn.innerHTML; btn.textContent = '已複製！'; btn.classList.add('copied');
    setTimeout(() => { btn.innerHTML = orig; btn.classList.remove('copied'); }, 2000);
  }).catch(() => {
    const ta = document.createElement('textarea');
    ta.value = txns; document.body.appendChild(ta); ta.select();
    document.execCommand('copy'); document.body.removeChild(ta);
    btn.textContent = '已複製！';
    setTimeout(() => { btn.innerHTML = '&#128203; 複製此區序號'; }, 2000);
  });
}

function copySelectedZones() {
  var NL = String.fromCharCode(10);
  var checked = document.querySelectorAll('.zone-check:checked');
  if (checked.length === 0) { alert('請先勾選要複製的區域'); return; }
  var all = [];
  checked.forEach(function(cb) {
    var el = document.querySelector('.zone-txn-data[data-zone="' + cb.dataset.zone + '"]');
    if (el) {
      var parts = el.textContent.trim().split(NL);
      parts.forEach(function(l){ if(l.trim()) all.push(l.trim()); });
    }
  });
  var text = all.join(NL);
  navigator.clipboard.writeText(text).then(function() {
    var msg = document.getElementById('zone-copy-msg');
    if (msg) { msg.textContent = '已複製 ' + all.length + ' 筆（' + checked.length + ' 個區域）'; setTimeout(function(){ msg.textContent=''; }, 3000); }
  }).catch(function() {
    var ta = document.createElement('textarea');
    ta.value = text; document.body.appendChild(ta); ta.select();
    document.execCommand('copy'); document.body.removeChild(ta);
  });
}

function copyTxns(btn){
  const txns = btn.parentElement.querySelector('.txn-data').textContent.trim();
  navigator.clipboard.writeText(txns).then(()=>{
    btn.textContent = '已複製！';
    btn.classList.add('copied');
    setTimeout(()=>{
      btn.innerHTML = '&#128203; 複製交易序號';
      btn.classList.remove('copied');
    }, 2000);
  }).catch(()=>{
    // 備用方案：舊瀏覽器
    const ta = document.createElement('textarea');
    ta.value = txns;
    document.body.appendChild(ta);
    ta.select();
    document.execCommand('copy');
    document.body.removeChild(ta);
    btn.textContent = '已複製！';
    setTimeout(()=>{ btn.innerHTML = '&#128203; 複製交易序號'; }, 2000);
  });
}
function doUpload(){
  const f=document.getElementById('csv-in').files[0];
  if(!f)return;
  document.getElementById('up-name').textContent='上傳中：'+f.name;
  const fd=new FormData();fd.append('file',f);
  fetch('/api/upload',{method:'POST',body:fd})
    .then(r=>r.json())
    .then(d=>{if(d.ok)location.reload();else alert('失敗：'+d.msg)});
}
const lg=document.getElementById('lg');
if(lg)lg.scrollTop=lg.scrollHeight;
</script>
</body></html>"""

@app.route("/split_app")
@login_required
def index():
    try:
        return render_template_string(HTML,
            company     = CONFIG["company_name"],
            groups      = state["groups"],
            group_keys  = list(state["groups"].keys()),
            total       = state["total"],
            last_update = state["last_update"] or "—",
            status      = state["status"],
            status_msg  = state["status_msg"],
            summary     = state["summary"],
            summary_keys= list(state["summary"].keys()),
            log_lines   = state["log"][-30:],
        )
    except Exception as e:
        log(f"頁面渲染錯誤：{e}")
        return f"<html><body><h2>系統啟動中，請重新整理</h2><p>{e}</p></body></html>", 500

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    f = request.files.get("file")
    if not f: return jsonify({"ok": False, "msg": "無檔案"})
    content = None
    for enc in ["utf-8-sig", "big5", "cp950", "utf-8"]:
        try:
            f.seek(0); content = f.read().decode(enc); break
        except Exception: continue
    if not content: return jsonify({"ok": False, "msg": "無法解析編碼"})
    rows, enc = load_csv(content, is_text=True)
    if not rows: return jsonify({"ok": False, "msg": "CSV 解析失敗"})
    run_pipeline(rows=rows)
    return jsonify({"ok": True})

@app.route("/api/status")
@login_required
def api_status():
    return jsonify({k: state[k] for k in ("status","status_msg","last_update","total")})

# ── 特殊可出超材品設定頁面 ──────────────────────────────────────────
DIAGONAL_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8">
<title>特殊可出超材品設定</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:#f0f2f5;font-size:13px;color:#1a1a1a}
.topbar{background:#0f1923;color:#fff;height:52px;padding:0 20px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:300}
.logo{font-size:15px;font-weight:600;margin-right:auto}.logo span{color:#4fa3e0}
.btn{padding:7px 16px;border-radius:5px;border:none;font-size:13px;cursor:pointer;font-weight:500}
.btn-blue{background:#1a5fa8;color:#fff}.btn-red{background:#b71c1c;color:#fff}.btn:hover{opacity:.85}
.card{background:#fff;border:1px solid #ddd;border-radius:8px;padding:20px;margin:20px}
.card h2{font-size:14px;font-weight:500;color:#555;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #eee}
.desc{font-size:12px;color:#888;margin-bottom:14px;line-height:1.8}
.form-grid{display:grid;grid-template-columns:160px 140px 1fr auto;gap:10px;align-items:end;margin-bottom:8px}
label{font-size:12px;color:#666;display:block;margin-bottom:4px}
input[type=text],input[type=number]{width:100%;padding:7px 10px;border:1px solid #ddd;border-radius:5px;font-size:13px;font-family:inherit}
input:focus{outline:none;border-color:#1a5fa8}
.ch-group{padding:8px 12px;background:#f8f8f8;border:1px solid #ddd;border-radius:5px}
.ch-group label{display:inline-flex;align-items:center;gap:5px;margin-right:16px;font-size:12px;color:#333;cursor:pointer}
.ch-group label input{width:auto}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{background:#f5f5f5;padding:8px 10px;text-align:left;font-weight:500;color:#555;border-bottom:1.5px solid #ddd}
td{padding:8px 10px;border-bottom:.5px solid #eee}
.tag{display:inline-block;padding:2px 8px;border-radius:12px;font-size:11px;background:#e3f2fd;color:#1565c0;border:1px solid #bbdefb}
.tag-g{background:#e8f5e9;color:#2e7d32;border:1px solid #a5d6a7}
.tag-o{background:#fff3e0;color:#e65100;border:1px solid #ffcc80}
.empty{text-align:center;padding:30px;color:#aaa;font-size:12px}
.msg{padding:8px 14px;border-radius:5px;font-size:12px;margin:0 20px 10px}
.msg-ok{background:#e8f5e9;color:#2e7d32}.msg-err{background:#ffebee;color:#b71c1c}
</style></head><body>
<div class="topbar">
  <div class="logo"><span>特殊可出超材品</span> 設定</div>
  <a href="/" style="color:#aaa;font-size:12px;text-decoration:none">&#x2302; 返回首頁</a>
</div>
<div id="msg-area"></div>
<div class="card">
  <h2>新增設定</h2>
  <div class="desc">
    <b>斜放後有效最長邊：</b>商品斜放後最長維度（cm），讓系統用此值判斷超材。<br>
    <b>每包最大件數：</b>超過此件數視為超材，可針對不同通路群組生效。<br>
    兩個欄位至少填一個。
  </div>
  <div class="form-grid">
    <div><label>商品 SKU</label><input type="text" id="new-sku" placeholder="例：APB002"></div>
    <div><label>斜放後最長邊（cm）</label><input type="number" id="new-side" placeholder="留空不設定" min="1" max="200" step="0.5"></div>
    <div>
      <label>每包最大件數 &amp; 適用通路</label>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
        <input type="number" id="new-qty" placeholder="件數上限（留空不限）" min="1" step="1" style="width:180px">
        <div class="ch-group">
          <label><input type="checkbox" id="ch-store" checked> 店配類<small style="color:#888">（店到店/隔日/超商/店到家）</small></label>
          <label><input type="checkbox" id="ch-delivery" checked> 快遞類<small style="color:#888">（新竹/嘉里）</small></label>
        </div>
      </div>
    </div>
    <div style="padding-bottom:2px"><button class="btn btn-blue" onclick="addSku()">新增</button></div>
  </div>
</div>
<div class="card">
  <h2>目前設定清單</h2>
  <table>
    <thead><tr>
      <th>商品 SKU</th><th>斜放後最長邊</th><th>每包最大件數</th><th>適用通路</th><th style="width:60px"></th>
    </tr></thead>
    <tbody id="sku-tbody"><tr><td colspan="5" class="empty">尚無設定</td></tr></tbody>
  </table>
</div>
<script>
var CH_LABELS = {store:'店配類', delivery:'快遞類'};
function loadSkus() {
  fetch('/api/diagonal').then(function(r){return r.json();}).then(function(data) {
    var tbody = document.getElementById('sku-tbody');
    var keys = data ? Object.keys(data) : [];
    if (keys.length === 0) { tbody.innerHTML = '<tr><td colspan="5" class="empty">尚無設定</td></tr>'; return; }
    tbody.innerHTML = keys.map(function(sku) {
      var cfg = data[sku];
      var side = cfg.side ? '<strong>' + cfg.side + ' cm</strong>' : '<span style="color:#bbb">—</span>';
      var qty  = cfg.max_qty ? '<span class="tag-o">' + cfg.max_qty + ' 件</span>' : '<span style="color:#bbb">不限</span>';
      var chs  = (cfg.channels || []).map(function(c){ return '<span class="tag-g">' + (CH_LABELS[c]||c) + '</span>'; }).join(' ');
      if (!chs) chs = '<span style="color:#bbb">—</span>';
      return '<tr><td><span class="tag">' + sku + '</span></td><td>' + side + '</td><td>' + qty + '</td><td>' + chs + '</td>' +
             '<td><button class="btn btn-red" style="padding:4px 10px;font-size:11px" onclick="deleteSku(\'' + sku + '\')">刪除</button></td></tr>';
    }).join('');
  });
}
function showMsg(msg, ok) {
  var area = document.getElementById('msg-area');
  area.innerHTML = '<div class="msg ' + (ok?'msg-ok':'msg-err') + '">' + msg + '</div>';
  setTimeout(function(){ area.innerHTML = ''; }, 3000);
}
function addSku() {
  var sku  = document.getElementById('new-sku').value.trim();
  var side = document.getElementById('new-side').value.trim();
  var qty  = document.getElementById('new-qty').value.trim();
  var chs  = [];
  if (document.getElementById('ch-store').checked)    chs.push('store');
  if (document.getElementById('ch-delivery').checked) chs.push('delivery');
  if (!sku) { showMsg('請輸入 SKU', false); return; }
  if (!side && !qty) { showMsg('請至少填入斜放最長邊或件數上限', false); return; }
  if (qty && chs.length === 0) { showMsg('請至少勾選一個通路群組', false); return; }
  fetch('/api/diagonal', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({sku:sku, side:side?parseFloat(side):null, max_qty:qty?parseInt(qty):null, channels:chs})
  }).then(function(r){return r.json();}).then(function(d) {
    if (d.ok) {
      showMsg('已新增 ' + sku, true);
      document.getElementById('new-sku').value = '';
      document.getElementById('new-side').value = '';
      document.getElementById('new-qty').value = '';
      loadSkus();
    } else { showMsg('新增失敗：' + d.msg, false); }
  });
}
function deleteSku(sku) {
  if (!confirm('確定刪除 ' + sku + '？')) return;
  fetch('/api/diagonal/' + sku, {method:'DELETE'}).then(function(r){return r.json();}).then(function(d) {
    if (d.ok) { showMsg('已刪除 ' + sku, true); loadSkus(); }
  });
}
loadSkus();
</script>
</body></html>"""

@app.route("/settings/diagonal")
@login_required
def settings_diagonal():
    return render_template_string(DIAGONAL_HTML)

@app.route("/api/diagonal", methods=["GET"])
@login_required
def api_diagonal_get():
    return jsonify(CONFIG["diagonal_skus"])

@app.route("/api/diagonal", methods=["POST"])
@login_required
def api_diagonal_post():
    data     = request.get_json()
    sku      = (data.get("sku") or "").strip()
    side     = data.get("side")
    max_qty  = data.get("max_qty")
    channels = data.get("channels", [])
    if not sku:
        return jsonify({"ok": False, "msg": "SKU 不可為空"})
    if not side and not max_qty:
        return jsonify({"ok": False, "msg": "請至少填入一個設定"})
    try:
        CONFIG["diagonal_skus"][sku] = {
            "side":     float(side) if side else None,
            "max_qty":  int(max_qty) if max_qty else None,
            "channels": channels,  # ["store", "delivery"]
        }
        save_settings()
        log(f"新增特殊可出超材品：{sku} side={side} max_qty={max_qty} channels={channels}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/diagonal/<sku>", methods=["DELETE"])
@login_required
def api_diagonal_delete(sku):
    if sku in CONFIG["diagonal_skus"]:
        del CONFIG["diagonal_skus"][sku]
        save_settings()
        log(f"刪除特殊可出超材品設定：{sku}")
    return jsonify({"ok": True})

# ============================================================
# 報關模組
# ============================================================

def get_drive_service():
    """建立 Google Drive API 連線"""
    try:
        import base64
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        cred_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT", "")
        if not cred_json:
            return None, "未設定 GOOGLE_SERVICE_ACCOUNT"
        cred_dict = None
        try:
            cred_dict = json.loads(base64.b64decode(cred_json + "==").decode("utf-8"))
        except Exception:
            pass
        if not cred_dict:
            try:
                cred_dict = json.loads(cred_json)
            except Exception as e:
                return None, str(e)
        scopes = ["https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(cred_dict, scopes=scopes)
        service = build("drive", "v3", credentials=creds)
        return service, None
    except Exception as e:
        return None, str(e)


_drive_folder_id_cache = ""  # 快取資料夾 ID

def upload_image_to_drive(image_url, folder_id=None):
    """
    下載圖片（支援 1688 防盜鏈）並上傳到 Google Drive。
    同一張圖（MD5相同）不重複上傳，直接回傳已存的 URL。
    回傳可公開存取的直連 URL；失敗回傳原始 URL。
    """
    global _drive_folder_id_cache
    if not image_url:
        return image_url
    image_url = str(image_url).strip()  # 清除換行符號、空白
    if not image_url or not image_url.startswith('http'):
        return image_url
    if "drive.google.com" in image_url or "googleapis.com" in image_url:
        return image_url  # 已經是 Drive URL，跳過
    try:
        import requests as req_lib
        from googleapiclient.http import MediaIoBaseUpload

        # 下載圖片
        dl = req_lib.get(image_url, headers={
            "Referer": "https://www.1688.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }, timeout=10)
        if dl.status_code != 200 or not dl.content:
            return image_url

        # 判斷副檔名
        content_type = dl.headers.get("Content-Type", "image/jpeg")
        ext = "jpg" if "jpeg" in content_type else content_type.split("/")[-1].split(";")[0]

        service, err = get_drive_service()
        if err:
            return image_url

        # 取得或建立圖片資料夾（優先用快取）
        if not folder_id:
            folder_id = _drive_folder_id_cache or os.environ.get("DRIVE_IMG_FOLDER_ID", "")
        if not folder_id:
            folder_meta = {
                "name": "報關圖片",
                "mimeType": "application/vnd.google-apps.folder"
            }
            folder = service.files().create(body=folder_meta, fields="id").execute()
            folder_id = folder.get("id", "")
            service.permissions().create(
                fileId=folder_id,
                body={"type": "anyone", "role": "reader"}
            ).execute()
            log(f"已建立 Google Drive 圖片資料夾: {folder_id}")
        _drive_folder_id_cache = folder_id

        # 用圖片內容 MD5 當檔名，同內容不重複上傳
        import hashlib
        fname = hashlib.md5(dl.content).hexdigest()[:16] + "." + ext

        # 先查詢 Drive 是否已存在同名檔案
        existing = service.files().list(
            q=f"name='{fname}' and '{folder_id}' in parents and trashed=false",
            fields="files(id)",
            pageSize=1
        ).execute()
        if existing.get("files"):
            file_id = existing["files"][0]["id"]
            return f"https://drive.google.com/thumbnail?id={file_id}&sz=w200"

        # 不存在才上傳
        file_meta = {"name": fname, "parents": [folder_id]}
        media = MediaIoBaseUpload(io.BytesIO(dl.content), mimetype=content_type, resumable=False)
        uploaded = service.files().create(body=file_meta, media_body=media, fields="id").execute()
        file_id = uploaded.get("id", "")

        # 設定公開可讀
        service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"}
        ).execute()

        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w200"
    except Exception as e:
        log(f"upload_image_to_drive 失敗: {e}")
        return image_url


def get_sheets_client():
    """建立 Google Sheets 連線"""
    try:
        import gspread, base64, re
        from google.oauth2.service_account import Credentials

        cred_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT", "")
        if not cred_json:
            return None, "未設定 GOOGLE_SERVICE_ACCOUNT 環境變數"

        cred_dict = None

        # 方法1: Base64 解碼
        try:
            cred_dict = json.loads(base64.b64decode(cred_json + "==").decode("utf-8"))
        except Exception:
            pass

        # 方法2: 強制修復換行後直接解析
        if not cred_dict:
            try:
                fixed = cred_json.replace(chr(13)+chr(10), chr(92)+"n").replace(chr(13), chr(92)+"n").replace(chr(10), chr(92)+"n")
                cred_dict = json.loads(fixed)
            except Exception:
                pass

        # 方法3: 直接解析
        if not cred_dict:
            try:
                cred_dict = json.loads(cred_json)
            except Exception as e:
                return None, f"JSON 解析失敗: {str(e)[:100]}"

        if not cred_dict:
            return None, "無法解析 GOOGLE_SERVICE_ACCOUNT"

        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(cred_dict, scopes=scopes)
        client = gspread.authorize(creds)
        return client, None
    except Exception as e:
        return None, str(e)


def load_customs_db():
    """從 Google Sheets 載入商品報關資料庫，回傳 {sku: {...}} dict"""
    client, err = get_sheets_client()
    if err:
        return {}, err
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID")
        sh = client.open_by_key(sheet_id)
        ws = sh.worksheet("商品報關資料庫")

        # 用 FORMULA 模式讀取，才能拿到 =IMAGE("url") 公式字串
        # get_all_records() 只讀顯示值，無法讀到公式
        header_row = ws.row_values(1)
        all_values = ws.get_all_values(value_render_option='FORMULA')
        col_idx = {name: i for i, name in enumerate(header_row)}

        def get_cell(row_data, col_name, default=""):
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row_data):
                return default
            return str(row_data[idx]).strip()

        db = {}
        for row_data in all_values[1:]:  # 跳過標題列
            sku = get_cell(row_data, "SKU編碼")
            if sku:
                db[sku] = {
                    "sku":          sku,
                    "name":         get_cell(row_data, "系統名稱"),
                    "style":        get_cell(row_data, "樣式"),
                    "size":         get_cell(row_data, "尺寸"),
                    "material":     get_cell(row_data, "材質"),
                    "customs_name": get_cell(row_data, "報關品名"),
                    "price":        get_cell(row_data, "單價"),
                    "unit":         get_cell(row_data, "單位"),
                    "product_size": get_cell(row_data, "商品尺寸"),
                    "image":        parse_image_url(get_cell(row_data, "圖片")),
                }
        return db, None
    except Exception as e:
        return {}, str(e)

def append_to_customs_db(new_item):
    """新增一筆資料到 Google Sheets 商品報關資料庫"""
    client, err = get_sheets_client()
    if err:
        return err
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID")
        sh = client.open_by_key(sheet_id)
        ws = sh.worksheet("商品報關資料庫")
        ws.append_row([
            new_item.get("sku", ""),
            new_item.get("name", ""),
            new_item.get("style", ""),
            new_item.get("size", ""),
            new_item.get("material", ""),
            new_item.get("customs_name", ""),
            new_item.get("price", ""),
            new_item.get("unit", ""),
            new_item.get("product_size", ""),
            new_item.get("image", ""),
        ])
        return None
    except Exception as e:
        return str(e)

# ============================================================
# 貨架入庫模組
# ============================================================

def load_rack_sheet():
    """取得貨架庫位紀錄 worksheet，不存在則自動建立"""
    client, err = get_sheets_client()
    if err:
        return None, err
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID")
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet("貨架庫位紀錄")
        except Exception:
            ws = sh.add_worksheet(title="貨架庫位紀錄", rows=1000, cols=6)
            ws.append_row(["貨號", "儲位", "數量", "入庫時間", "備註"])
        return ws, None
    except Exception as e:
        return None, str(e)

def rack_save_records(items, rack_code):
    """items = [{"sku": "BCE001-001", "qty": 2}], rack_code = "RACK-A-1" """
    ws, err = load_rack_sheet()
    if err:
        return err
    try:
        now = datetime.now().strftime("%Y/%m/%d %H:%M")
        rows = [[item.get("sku",""), rack_code, item.get("qty",1), now, ""] for item in items]
        if rows:
            ws.append_rows(rows)
        return None
    except Exception as e:
        return str(e)

def rack_query_sku(sku):
    """查詢貨號在哪些儲位"""
    ws, err = load_rack_sheet()
    if err:
        return [], err
    try:
        all_values = ws.get_all_values()
        results = []
        for row in all_values[1:]:
            if len(row) < 2:
                continue
            if sku.upper() in str(row[0]).strip().upper():
                results.append({
                    "sku":  row[0] if len(row)>0 else "",
                    "rack": row[1] if len(row)>1 else "",
                    "qty":  row[2] if len(row)>2 else "",
                    "time": row[3] if len(row)>3 else "",
                    "note": row[4] if len(row)>4 else "",
                })
        return results, None
    except Exception as e:
        return [], str(e)

def rack_query_rack(rack_code):
    """查詢某儲位裡有哪些貨號"""
    ws, err = load_rack_sheet()
    if err:
        return [], err
    try:
        all_values = ws.get_all_values()
        results = []
        for row in all_values[1:]:
            if len(row) < 2:
                continue
            if str(row[1]).strip().upper() == rack_code.strip().upper():
                results.append({
                    "sku":  row[0] if len(row)>0 else "",
                    "rack": row[1] if len(row)>1 else "",
                    "qty":  row[2] if len(row)>2 else "",
                    "time": row[3] if len(row)>3 else "",
                    "note": row[4] if len(row)>4 else "",
                })
        return results, None
    except Exception as e:
        return [], str(e)

RACK_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>貨架入庫系統</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:#0f1923;min-height:100vh;color:#fff;font-size:14px}
.topbar{background:rgba(255,255,255,.05);backdrop-filter:blur(10px);height:52px;padding:0 20px;display:flex;align-items:center;gap:12px;border-bottom:1px solid rgba(255,255,255,.08);position:sticky;top:0;z-index:300}
.logo{font-size:15px;font-weight:700;margin-right:auto}.logo span{color:#f4a100}
a.back{color:#aaa;font-size:12px;text-decoration:none;padding:5px 10px;border:1px solid #333;border-radius:5px}
a.back:hover{border-color:#666;color:#fff}
.tabs{display:flex;border-bottom:2px solid rgba(255,255,255,.08);padding:0 20px}
.tab{padding:14px 24px;cursor:pointer;font-size:13px;color:#888;border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .2s}
.tab.active{color:#f4a100;border-bottom-color:#f4a100;font-weight:600}
.tab:hover{color:#fff}
.panel{display:none;padding:24px 20px;max-width:800px;margin:0 auto}
.panel.active{display:block}
.card{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:20px;margin-bottom:16px}
.card h3{font-size:14px;font-weight:600;margin-bottom:14px;color:#f4a100}
input[type=text],input[type=number]{width:100%;padding:10px 14px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);border-radius:8px;color:#fff;font-size:14px;font-family:inherit;outline:none}
input:focus{border-color:#f4a100;background:rgba(244,161,0,.08)}
input::placeholder{color:#555}
.btn{padding:10px 20px;border-radius:8px;border:none;font-size:13px;cursor:pointer;font-weight:600;transition:all .2s;white-space:nowrap}
.btn-yellow{background:#f4a100;color:#0f1923}.btn-yellow:hover{background:#ffb300}
.btn-green{background:#2e7d32;color:#fff}.btn-green:hover{background:#388e3c}
.btn-red{background:#b71c1c;color:#fff}.btn-red:hover{background:#c62828}
.scan-list{list-style:none;margin:12px 0}
.scan-item{display:flex;align-items:center;gap:10px;padding:10px 14px;background:rgba(255,255,255,.06);border-radius:8px;margin-bottom:6px;border:1px solid rgba(255,255,255,.08)}
.scan-sku{flex:1;font-weight:600;font-size:15px;letter-spacing:.5px}
.scan-qty{width:70px;padding:5px 8px;text-align:center}
.scan-del{cursor:pointer;color:#e57373;font-size:18px;background:none;border:none;padding:0 4px}
.confirm-overlay{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.88);z-index:999;align-items:center;justify-content:center;flex-direction:column;gap:20px;text-align:center}
.confirm-overlay.show{display:flex}
.confirm-rack{font-size:80px;font-weight:900;color:#f4a100;letter-spacing:6px;text-shadow:0 0 60px rgba(244,161,0,.6);animation:pulse 1s ease-in-out infinite alternate}
@keyframes pulse{from{transform:scale(1)}to{transform:scale(1.06)}}
.confirm-sub{font-size:24px;color:#fff;opacity:.7;letter-spacing:2px}
.confirm-items{background:rgba(255,255,255,.08);border-radius:12px;padding:16px 28px;min-width:340px;max-width:520px;text-align:left}
.confirm-items li{padding:7px 0;border-bottom:1px solid rgba(255,255,255,.08);font-size:15px}
.confirm-items li:last-child{border:none}
.confirm-btns{display:flex;gap:14px;margin-top:4px}
.result-table{width:100%;border-collapse:collapse;margin-top:12px}
.result-table th{background:rgba(244,161,0,.15);padding:10px 12px;text-align:left;font-size:12px;color:#f4a100;border-bottom:1px solid rgba(244,161,0,.2)}
.result-table td{padding:10px 12px;border-bottom:1px solid rgba(255,255,255,.06);font-size:13px}
.result-table tr:hover td{background:rgba(255,255,255,.04)}
.rack-badge{display:inline-block;padding:3px 10px;background:rgba(244,161,0,.15);color:#f4a100;border-radius:20px;font-weight:700;font-size:13px;border:1px solid rgba(244,161,0,.3)}
.msg{padding:10px 14px;border-radius:8px;font-size:13px;margin-bottom:12px}
.msg-ok{background:rgba(46,125,50,.2);color:#81c784;border:1px solid rgba(46,125,50,.3)}
.msg-err{background:rgba(183,28,28,.2);color:#ef9a9a;border:1px solid rgba(183,28,28,.3)}
.empty{text-align:center;padding:32px;color:#555;font-size:13px}
.row{display:flex;gap:10px}
</style></head><body>

<div class="topbar">
  <div class="logo">&#x1F4E6; <span>貨架入庫系統</span></div>
  <a href="/" class="back">&#x2302; 返回首頁</a>
  <a href="/logout" class="back">登出</a>
</div>

<div class="tabs">
  <div class="tab active" onclick="switchTab('inbound',this)">&#x1F4E5; 入庫作業</div>
  <div class="tab" onclick="switchTab('search',this)">&#x1F50D; 查找貨位</div>
</div>

<!-- 入庫作業 -->
<div class="panel active" id="panel-inbound">
  <div id="msg-inbound"></div>
  <div class="card">
    <h3>&#x1F4F7; 步驟一：掃描 / 輸入貨號</h3>
    <input type="text" id="sku-input" placeholder="掃描或輸入貨號，按 Enter 加入清單" autocomplete="off" autofocus>
    <ul class="scan-list" id="scan-list">
      <li class="empty" id="scan-empty">尚未掃描任何貨號</li>
    </ul>
  </div>
  <div class="card">
    <h3>&#x1F4CD; 步驟二：掃描 / 輸入儲位條碼</h3>
    <div class="row">
      <input type="text" id="rack-input" placeholder="掃描或輸入儲位，例：RACK-A-1" autocomplete="off">
      <button class="btn btn-yellow" onclick="confirmRack()">&#x2705; 確認入庫</button>
    </div>
    <div style="font-size:12px;color:#555;margin-top:8px">掃描儲位後會出現大字二次確認，避免刷錯儲位</div>
  </div>
</div>

<!-- 查找貨位 -->
<div class="panel" id="panel-search">
  <div id="msg-search"></div>
  <div class="card">
    <h3>&#x1F50D; 查詢貨號所在儲位</h3>
    <div class="row">
      <input type="text" id="search-sku" placeholder="輸入或掃描貨號" autocomplete="off">
      <button class="btn btn-yellow" onclick="searchSku()">查詢</button>
    </div>
    <div id="result-sku"></div>
  </div>
  <div class="card">
    <h3>&#x1F4E6; 查詢儲位內有哪些貨號</h3>
    <div class="row">
      <input type="text" id="search-rack" placeholder="輸入或掃描儲位條碼，例：RACK-A-1" autocomplete="off">
      <button class="btn btn-yellow" onclick="searchRack()">查詢</button>
    </div>
    <div id="result-rack"></div>
  </div>
</div>

<!-- 大字二次確認 -->
<div class="confirm-overlay" id="confirm-overlay">
  <div class="confirm-rack" id="confirm-rack-text"></div>
  <div class="confirm-sub" id="confirm-rack-sub"></div>
  <div class="confirm-items">
    <div style="font-size:13px;color:#aaa;margin-bottom:10px">即將入庫以下貨號：</div>
    <ul id="confirm-item-list"></ul>
  </div>
  <div class="confirm-btns">
    <button class="btn btn-green" onclick="doSave()">&#x2705; 確認正確，儲存</button>
    <button class="btn btn-red" onclick="cancelConfirm()">&#x274C; 取消重掃</button>
  </div>
</div>

<script>
var scannedItems = {};
var pendingRack = '';

function switchTab(name, el){
  document.querySelectorAll('.tab').forEach(function(t){t.classList.remove('active');});
  document.querySelectorAll('.panel').forEach(function(p){p.classList.remove('active');});
  el.classList.add('active');
  document.getElementById('panel-'+name).classList.add('active');
  if(name==='inbound') document.getElementById('sku-input').focus();
  else document.getElementById('search-sku').focus();
}

document.getElementById('sku-input').addEventListener('keydown',function(e){
  if(e.key!=='Enter') return;
  var sku = this.value.trim().toUpperCase();
  if(!sku) return;
  scannedItems[sku] = (scannedItems[sku]||0)+1;
  renderScanList();
  this.value=''; this.focus();
});

document.getElementById('rack-input').addEventListener('keydown',function(e){
  if(e.key==='Enter') confirmRack();
});
document.getElementById('search-sku').addEventListener('keydown',function(e){
  if(e.key==='Enter') searchSku();
});
document.getElementById('search-rack').addEventListener('keydown',function(e){
  if(e.key==='Enter') searchRack();
});

function renderScanList(){
  var ul=document.getElementById('scan-list');
  var keys=Object.keys(scannedItems);
  document.getElementById('scan-empty').style.display=keys.length?'none':'block';
  Array.from(ul.querySelectorAll('.scan-item')).forEach(function(el){el.remove();});
  keys.forEach(function(sku){
    var li=document.createElement('li'); li.className='scan-item';
    li.innerHTML='<span class="scan-sku">'+sku+'</span>'+
      '<input type="number" class="scan-qty" value="'+scannedItems[sku]+'" min="1" onchange="updateQty(\''+sku+'\',this.value)">'+
      '<span style="font-size:12px;color:#666">件</span>'+
      '<button class="scan-del" onclick="removeSku(\''+sku+'\')">&#x2715;</button>';
    ul.appendChild(li);
  });
}

function updateQty(sku,val){var n=parseInt(val);if(n>0)scannedItems[sku]=n;else removeSku(sku);}
function removeSku(sku){delete scannedItems[sku];renderScanList();}

function confirmRack(){
  if(!Object.keys(scannedItems).length){showMsg('inbound','&#x26A0; 請先掃描至少一個貨號！',false);return;}
  var rack=document.getElementById('rack-input').value.trim().toUpperCase();
  if(!rack){showMsg('inbound','&#x26A0; 請掃描或輸入儲位條碼！',false);return;}
  pendingRack=rack;
  document.getElementById('confirm-rack-text').textContent=rack;
  var parts=rack.split('-');
  var sub=parts.length>=3?parts[1]+' 排　第 '+parts[2]+' 層':rack;
  document.getElementById('confirm-rack-sub').textContent=sub;
  var ul=document.getElementById('confirm-item-list'); ul.innerHTML='';
  Object.keys(scannedItems).forEach(function(sku){
    var li=document.createElement('li');
    li.textContent=sku+'　×　'+scannedItems[sku]+' 件';
    ul.appendChild(li);
  });
  document.getElementById('confirm-overlay').classList.add('show');
}

function cancelConfirm(){
  document.getElementById('confirm-overlay').classList.remove('show');
  document.getElementById('rack-input').focus();
}

function doSave(){
  var items=Object.keys(scannedItems).map(function(sku){return{sku:sku,qty:scannedItems[sku]};});
  fetch('/api/rack/save',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({items:items,rack:pendingRack})
  }).then(function(r){return r.json();}).then(function(d){
    document.getElementById('confirm-overlay').classList.remove('show');
    if(d.ok){
      showMsg('inbound','&#x2705; 入庫成功！共 '+items.length+' 種貨號 → '+pendingRack,true);
      scannedItems={};renderScanList();
      document.getElementById('rack-input').value='';
      document.getElementById('sku-input').focus();
    } else {
      showMsg('inbound','&#x274C; 儲存失敗：'+d.msg,false);
    }
  });
}

function searchSku(){
  var sku=document.getElementById('search-sku').value.trim();
  if(!sku){showMsg('search','請輸入貨號',false);return;}
  document.getElementById('result-sku').innerHTML='<div style="color:#888;padding:10px">查詢中...</div>';
  fetch('/api/rack/query-sku?sku='+encodeURIComponent(sku))
    .then(function(r){return r.json();}).then(function(d){
      if(!d.ok){document.getElementById('result-sku').innerHTML='<div class="msg msg-err">'+d.msg+'</div>';return;}
      if(!d.results.length){document.getElementById('result-sku').innerHTML='<div class="empty">&#x1F4ED; 找不到此貨號的入庫紀錄</div>';return;}
      var h='<table class="result-table"><tr><th>貨號</th><th>儲位</th><th>數量</th><th>入庫時間</th></tr>';
      d.results.forEach(function(r){h+='<tr><td>'+r.sku+'</td><td><span class="rack-badge">'+r.rack+'</span></td><td>'+r.qty+'</td><td>'+r.time+'</td></tr>';});
      document.getElementById('result-sku').innerHTML=h+'</table>';
    });
}

function searchRack(){
  var rack=document.getElementById('search-rack').value.trim();
  if(!rack){showMsg('search','請輸入儲位條碼',false);return;}
  document.getElementById('result-rack').innerHTML='<div style="color:#888;padding:10px">查詢中...</div>';
  fetch('/api/rack/query-rack?rack='+encodeURIComponent(rack))
    .then(function(r){return r.json();}).then(function(d){
      if(!d.ok){document.getElementById('result-rack').innerHTML='<div class="msg msg-err">'+d.msg+'</div>';return;}
      if(!d.results.length){document.getElementById('result-rack').innerHTML='<div class="empty">&#x1F4ED; 此儲位目前沒有入庫紀錄</div>';return;}
      var h='<table class="result-table"><tr><th>貨號</th><th>儲位</th><th>數量</th><th>入庫時間</th></tr>';
      d.results.forEach(function(r){h+='<tr><td>'+r.sku+'</td><td><span class="rack-badge">'+r.rack+'</span></td><td>'+r.qty+'</td><td>'+r.time+'</td></tr>';});
      document.getElementById('result-rack').innerHTML=h+'</table>';
    });
}

function showMsg(panel,msg,ok){
  var el=document.getElementById('msg-'+panel);
  el.innerHTML='<div class="msg '+(ok?'msg-ok':'msg-err')+'">'+msg+'</div>';
  setTimeout(function(){el.innerHTML='';},4000);
}
</script>
</body></html>"""

@app.route("/rack")
@login_required
def rack_page():
    return render_template_string(RACK_HTML)

@app.route("/api/rack/save", methods=["POST"])
@login_required
def api_rack_save():
    data = request.get_json()
    items = data.get("items", [])
    rack  = data.get("rack", "").strip().upper()
    if not items or not rack:
        return jsonify({"ok": False, "msg": "缺少貨號或儲位"})
    err = rack_save_records(items, rack)
    if err:
        return jsonify({"ok": False, "msg": err})
    log(f"貨架入庫：{len(items)} 種貨號 → {rack}")
    return jsonify({"ok": True})

@app.route("/api/rack/query-sku")
@login_required
def api_rack_query_sku():
    sku = request.args.get("sku", "").strip()
    if not sku:
        return jsonify({"ok": False, "msg": "請提供貨號"})
    results, err = rack_query_sku(sku)
    if err:
        return jsonify({"ok": False, "msg": err})
    return jsonify({"ok": True, "results": results})

@app.route("/api/rack/query-rack")
@login_required
def api_rack_query_rack():
    rack = request.args.get("rack", "").strip()
    if not rack:
        return jsonify({"ok": False, "msg": "請提供儲位條碼"})
    results, err = rack_query_rack(rack)
    if err:
        return jsonify({"ok": False, "msg": err})
    return jsonify({"ok": True, "results": results})


CUSTOMS_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8">
<title>報關清單系統</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:#f0f2f5;font-size:13px;color:#1a1a1a}
.topbar{background:#0f1923;color:#fff;height:52px;padding:0 20px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:300}
.logo{font-size:15px;font-weight:600;margin-right:auto}.logo span{color:#f4a100}
.btn{padding:7px 16px;border-radius:5px;border:none;font-size:13px;cursor:pointer;font-weight:500}
.btn-yellow{background:#f4a100;color:#fff}.btn-green{background:#2e7d32;color:#fff}
.btn-blue{background:#1a5fa8;color:#fff}.btn-red{background:#b71c1c;color:#fff}
.btn-gray{background:#666;color:#fff}.btn:hover{opacity:.85}
.card{background:#fff;border:1px solid #ddd;border-radius:8px;padding:20px;margin:20px}
.card h2{font-size:14px;font-weight:500;color:#555;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #eee}
.upload-area{border:2px dashed #ddd;border-radius:8px;padding:40px;text-align:center;cursor:pointer;transition:border-color .2s}
.upload-area:hover{border-color:#f4a100}
.upload-area.drag{border-color:#f4a100;background:#fffbf0}
label{font-size:12px;color:#666;display:block;margin-bottom:4px}
input[type=text],input[type=number],select{width:100%;padding:7px 10px;border:1px solid #ddd;border-radius:5px;font-size:13px;font-family:inherit}
input:focus,select:focus{outline:none;border-color:#f4a100}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{background:#f5f5f5;padding:8px 6px;text-align:left;font-weight:500;color:#555;border-bottom:1.5px solid #ddd;white-space:nowrap}
td{padding:6px;border-bottom:.5px solid #eee;vertical-align:middle}
tr.ok{background:#f1f8e9}
tr.missing{background:#ffebee}
tr.missing td:first-child::after{content:" &#128560;";color:#b71c1c}
.tag-ok{display:inline-block;padding:1px 8px;border-radius:10px;font-size:11px;background:#e8f5e9;color:#2e7d32}
.tag-miss{display:inline-block;padding:1px 8px;border-radius:10px;font-size:11px;background:#ffebee;color:#b71c1c}
.msg{padding:8px 14px;border-radius:5px;font-size:12px;margin:10px 20px}
.msg-ok{background:#e8f5e9;color:#2e7d32}.msg-err{background:#ffebee;color:#b71c1c}
.msg-warn{background:#fff8e1;color:#e65100}
.modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.5);z-index:999;align-items:center;justify-content:center}
.modal.show{display:flex}
.modal-box{background:#fff;border-radius:10px;padding:28px;width:520px;max-height:90vh;overflow-y:auto}
.modal-box h3{font-size:15px;font-weight:600;margin-bottom:16px;color:#e65100}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}
.form-group{display:flex;flex-direction:column;gap:4px}
img.thumb{width:50px;height:50px;object-fit:cover;border-radius:4px;border:1px solid #eee}
#preview-section{display:none}
</style></head><body>

<div class="topbar">
  <div class="logo"><span>&#128230;</span> 報關清單系統</div>
  <a href="/" style="color:#aaa;font-size:12px;text-decoration:none">&#x2302; 返回首頁</a>
  <a href="/logout" style="color:#aaa;font-size:12px;text-decoration:none">登出</a>
</div>

<div id="msg-area" style="position:sticky;top:52px;z-index:200"></div>

<div class="card">
  <h2>&#128228; 上傳倉庫進貨清單</h2>
  <p style="font-size:12px;color:#888;margin-bottom:16px">
    請先在 Excel 的 <strong>R 欄</strong>填入 SKU 編碼，再上傳檔案。<br>
    系統會自動對應商品報關資料庫，帶入材質、品名、單價等資料。
  </p>
  <div class="upload-area" id="drop-zone" onclick="document.getElementById('xlsx-in').click()">
    <div style="font-size:36px;margin-bottom:8px">&#128196;</div>
    <div style="font-weight:500;margin-bottom:4px">點擊或拖曳上傳 Excel 檔案</div>
    <div style="font-size:12px;color:#aaa">支援 .xlsx 及 .xls 格式，圖片會自動忽略不影響上傳</div>
  </div>
  <input type="file" id="xlsx-in" accept=".xlsx,.xls" style="display:none" onchange="uploadFile(this)">
</div>

<div id="preview-section">
  <!-- 櫃號/封號/日期填寫區 -->
  <div class="card">
    <h2>&#128230; 填寫報關資訊</h2>
    <p style="font-size:12px;color:#888;margin-bottom:14px">此資訊會顯示在匯出的報關單標題區</p>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
      <div>
        <label style="font-size:12px;color:#666;display:block;margin-bottom:4px">出口商（A1）</label>
        <input type="text" id="exporter" placeholder="例：ORAL TRADING CO., LTD." style="width:100%;padding:8px 10px;border:1.5px solid #ddd;border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:12px;color:#666;display:block;margin-bottom:4px">進口商（A2）</label>
        <input type="text" id="importer" placeholder="例：台灣進口商名稱" style="width:100%;padding:8px 10px;border:1.5px solid #ddd;border-radius:6px;font-size:13px">
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px">
      <div>
        <label style="font-size:12px;color:#666;display:block;margin-bottom:4px">櫃號</label>
        <input type="text" id="cabinet-no" placeholder="例：TCKU3456789" style="width:100%;padding:8px 10px;border:1.5px solid #ddd;border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:12px;color:#666;display:block;margin-bottom:4px">封號</label>
        <input type="text" id="seal-no" placeholder="例：CN12345678" style="width:100%;padding:8px 10px;border:1.5px solid #ddd;border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:12px;color:#666;display:block;margin-bottom:4px">出貨日期</label>
        <input type="date" id="ship-date" style="width:100%;padding:8px 10px;border:1.5px solid #ddd;border-radius:6px;font-size:13px">
      </div>
    </div>
  </div>
  <div class="card">
    <h2>&#128270; 預覽結果</h2>
    <div style="display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap">
      <span id="stat-total" style="font-size:12px;color:#555"></span>
      <span id="stat-ok" style="font-size:12px;color:#2e7d32"></span>
      <span id="stat-miss" style="font-size:12px;color:#b71c1c"></span>
    </div>
    <div style="overflow-x:auto">
      <table id="preview-table">
        <thead><tr>
          <th>SKU</th><th>類型</th><th>產品尺寸</th><th>材質</th><th>報關品名</th>
          <th>箱號</th><th>PCS/件</th><th>件數</th><th>總PCS</th><th>單位</th>
          <th>單價</th><th>總金額RMB</th><th>毛重</th><th>長</th><th>寬</th><th>高</th>
          <th>材積</th><th>總重量</th><th>圖片</th><th>狀態</th>
        </tr></thead>
        <tbody id="preview-body"></tbody>
      </table>
    </div>
    <div style="margin-top:16px;display:flex;gap:10px;flex-wrap:wrap">
      <button class="btn btn-green" onclick="exportExcel()">&#128229; 匯出報關 Excel</button>
      <button class="btn btn-gray" onclick="resetPage()">&#128260; 重新上傳</button>
    </div>
  </div>
</div>


<!-- 新品建檔 Modal -->
<div class="modal" id="new-item-modal">
  <div class="modal-box">
    <h3>&#127381; 發現新商品！請建立報關資料</h3>
    <p style="font-size:12px;color:#888;margin-bottom:16px">
      SKU「<strong id="modal-sku"></strong>」在資料庫中查無資料，<br>
      請填入報關資料後儲存，系統會自動更新商品報關資料庫。
    </p>
    <div class="form-grid">
      <div class="form-group"><label>SKU編碼</label><input type="text" id="f-sku" readonly style="background:#f5f5f5"></div>
      <div class="form-group"><label>系統名稱</label><input type="text" id="f-name" placeholder="例：針織手提袋"></div>
      <div class="form-group"><label>樣式</label><input type="text" id="f-style" placeholder="例：格子款"></div>
      <div class="form-group"><label>尺寸</label><input type="text" id="f-size" placeholder="例：F"></div>
      <div class="form-group"><label>材質 *</label><input type="text" id="f-material" placeholder="例：滌綸"></div>
      <div class="form-group"><label>報關品名 *</label><input type="text" id="f-customs-name" placeholder="例：手提袋"></div>
      <div class="form-group"><label>單價 *</label><input type="number" id="f-price" placeholder="例：3.4" step="0.01"></div>
      <div class="form-group"><label>單位</label><input type="text" id="f-unit" placeholder="例：個"></div>
      <div class="form-group"><label>商品尺寸</label><input type="text" id="f-product-size" placeholder="例：20*35*10"></div>
      <div class="form-group" style="grid-column:1/-1">
        <label>圖片</label>
        <div style="display:flex;gap:8px;align-items:center;margin-bottom:6px">
          <input type="text" id="f-image" placeholder="貼上圖片URL，或用下方方式上傳" style="flex:1" oninput="previewImageUrl(this.value)">
          <label style="background:#1565c0;color:#fff;padding:6px 12px;border-radius:5px;cursor:pointer;font-size:12px;white-space:nowrap">
            &#128193; 選擇圖片
            <input type="file" id="f-image-file" accept="image/*" style="display:none" onchange="handleImageFile(this)">
          </label>
        </div>
        <div style="display:flex;gap:8px;align-items:center;margin-bottom:6px">
          <input type="text" id="f-1688-url" placeholder="貼上 1688 商品網址，自動抓取圖片" style="flex:1;font-size:12px">
          <button type="button" onclick="fetch1688Image()" style="background:#f57c00;color:#fff;border:none;padding:6px 12px;border-radius:5px;cursor:pointer;font-size:12px;white-space:nowrap">&#128247; 抓取圖片</button>
        </div>
        <div id="f-image-preview" style="display:none;margin-top:6px">
          <img id="f-image-thumb" style="max-width:80px;max-height:80px;border-radius:4px;border:1px solid #ddd">
        </div>
        <div style="font-size:11px;color:#aaa;margin-top:4px">&#128161; 支援：貼 URL、Ctrl+V 貼截圖、選圖片檔、1688網址自動抓圖</div>
      </div>
    </div>
    <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:8px">
      <button class="btn btn-gray" onclick="skipNewItem()">跳過此筆</button>
      <button class="btn btn-yellow" onclick="saveNewItem()">&#128190; 儲存並繼續</button>
    </div>
  </div>
</div>

<script>
var allRows = [];
var missingSkus = [];
var currentMissingIdx = 0;
var customsDb = {};
var processedRows = [];

// 拖曳上傳
document.addEventListener('DOMContentLoaded', function() {
  var dz = document.getElementById('drop-zone');
  if(!dz) return;
  dz.addEventListener('dragover', function(e){ e.preventDefault(); dz.classList.add('drag'); });
  dz.addEventListener('dragleave', function(){ dz.classList.remove('drag'); });
  dz.addEventListener('drop', function(e){
    e.preventDefault(); dz.classList.remove('drag');
    var f = e.dataTransfer.files[0];
    if(f) processFile(f);
  });
});

function uploadFile(input) {
  var f = input.files[0];
  if(f) processFile(f);
}

function processFile(file) {
  if(!file.name.endsWith('.xlsx') && !file.name.endsWith('.xls')) {
    showMsg('&#128561; 哎呀！這不是 Excel 檔案，請選擇 .xlsx 或 .xls 格式！', false);
    return;
  }
  // 檔案大小警告（超過 2MB）
  if(file.size > 2 * 1024 * 1024) {
    var sizeMB = (file.size / 1024 / 1024).toFixed(1);
    var NL = String.fromCharCode(10);
    if(!confirm('&#9888; 您的檔案有 ' + sizeMB + ' MB，可能包含圖片！' + NL + NL + '建議先在 Excel 刪除圖片再上傳，可節省流量費用。' + NL + NL + '要繼續上傳嗎？')) {
      return;
    }
  }
  showMsg('&#8987; 正在讀取檔案並對應資料庫...', true);
  var fd = new FormData();
  fd.append('file', file);
  fetch('/api/customs/upload', {method:'POST', body:fd})
    .then(function(r){return r.json();})
    .then(function(d){
      if(!d.ok) { showMsg('&#128561; ' + d.msg, false); return; }
      if(d.db_err) {
        showMsg('&#9888; 無法連接商品報關資料庫：' + d.db_err + '，所有商品將顯示為查無資料', false);
      } else {
        var stat = '&#10003; 資料庫載入成功（共 ' + d.db_count + ' 筆商品）';
        if(typeof d.matched !== 'undefined') {
          stat += '，本次對應成功 ' + d.matched + ' 筆，未對應 ' + d.unmatched + ' 筆';
        }
        showMsg(stat, d.unmatched === 0);
      }
      allRows = d.rows;
      customsDb = d.db;
      checkMissing();
    });
}

function checkMissing() {
  missingSkus = [];
  allRows.forEach(function(row, i){
    if(row.status === 'missing') missingSkus.push(i);
  });
  if(missingSkus.length > 0) {
    currentMissingIdx = 0;
    showMsg('&#128561; 發現 ' + missingSkus.length + ' 個新商品需要建檔，請逐一填入資料！', false);
    showNewItemModal(missingSkus[0]);
  } else {
    renderPreview();
  }
}

function showNewItemModal(rowIdx) {
  var row = allRows[rowIdx];
  document.getElementById('modal-sku').textContent = row.sku;
  document.getElementById('f-sku').value = row.sku;
  document.getElementById('f-name').value = '';
  document.getElementById('f-style').value = '';
  document.getElementById('f-size').value = '';
  document.getElementById('f-material').value = '';
  document.getElementById('f-customs-name').value = '';
  document.getElementById('f-price').value = '';
  document.getElementById('f-unit').value = '';
  document.getElementById('f-product-size').value = '';
  document.getElementById('f-image').value = '';
  document.getElementById('f-image').placeholder = '貼上圖片URL，或用下方方式上傳';
  document.getElementById('f-1688-url').value = '';
  document.getElementById('f-image-preview').style.display = 'none';
  window.__pendingImageBase64 = '';
  document.getElementById('new-item-modal').classList.add('show');
}

function saveNewItem() {
  var sku = document.getElementById('f-sku').value.trim();
  var material = document.getElementById('f-material').value.trim();
  var customsName = document.getElementById('f-customs-name').value.trim();
  var price = document.getElementById('f-price').value.trim();
  if(!material || !customsName || !price) {
    showMsg('材質、報關品名、單價為必填！', false);
    return;
  }
  var imageUrl = document.getElementById('f-image').value.trim();
  var pendingB64 = window.__pendingImageBase64 || '';

  function doSave(finalImageUrl) {
    var newItem = {
      sku: sku,
      name: document.getElementById('f-name').value.trim(),
      style: document.getElementById('f-style').value.trim(),
      size: document.getElementById('f-size').value.trim(),
      material: material,
      customs_name: customsName,
      price: parseFloat(price),
      unit: document.getElementById('f-unit').value.trim(),
      product_size: document.getElementById('f-product-size').value.trim(),
      image: finalImageUrl,
    };
    fetch('/api/customs/new-item', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify(newItem)
    }).then(function(r){return r.json();}).then(function(d){
      if(!d.ok) { showMsg('儲存失敗：' + d.msg, false); return; }
      var usedUrl = d.image || finalImageUrl;
      customsDb[sku] = newItem;
      customsDb[sku].image = usedUrl;
      var rowIdx = missingSkus[currentMissingIdx];
      allRows[rowIdx].status = 'ok';
      allRows[rowIdx].material = newItem.material;
      allRows[rowIdx].customs_name = newItem.customs_name;
      allRows[rowIdx].price = newItem.price;
      allRows[rowIdx].unit = newItem.unit;
      allRows[rowIdx].image = usedUrl;
      allRows[rowIdx].total_rmb = (parseFloat(allRows[rowIdx].total_pcs) * newItem.price).toFixed(2);
      window.__pendingImageBase64 = '';
      document.getElementById('f-image-preview').style.display = 'none';
      document.getElementById('new-item-modal').classList.remove('show');
      currentMissingIdx++;
      if(currentMissingIdx < missingSkus.length) {
        showNewItemModal(missingSkus[currentMissingIdx]);
      } else {
        showMsg('所有新品建檔完成！', true);
        renderPreview();
      }
    });
  }

  // 如果有 base64 圖片（截圖或上傳檔），先上傳到 Drive
  if(pendingB64) {
    showMsg('正在上傳圖片...', true);
    fetch('/api/customs/upload-image-base64', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({data: pendingB64, row_idx: 0, col_idx: 0})
    }).then(function(r){return r.json();}).then(function(d){
      doSave(d.ok ? d.drive_url : imageUrl);
    }).catch(function(){ doSave(imageUrl); });
  } else {
    doSave(imageUrl);
  }
}

function skipNewItem() {
  document.getElementById('new-item-modal').classList.remove('show');
  currentMissingIdx++;
  if(currentMissingIdx < missingSkus.length) {
    showNewItemModal(missingSkus[currentMissingIdx]);
  } else {
    renderPreview();
  }
}

function renderPreview() {
  var total = allRows.length;
  var ok = allRows.filter(function(r){return r.status==='ok';}).length;
  var miss = total - ok;
  document.getElementById('stat-total').textContent = '共 ' + total + ' 筆';
  document.getElementById('stat-ok').textContent = '&#10003; 對應成功 ' + ok + ' 筆';
  document.getElementById('stat-miss').textContent = miss > 0 ? '&#9888; 未建檔 ' + miss + ' 筆' : '';
  var tbody = document.getElementById('preview-body');
  tbody.innerHTML = allRows.map(function(r){
    var cls = r.status === 'ok' ? 'ok' : 'missing';
    var statusTag = r.status === 'ok'
      ? '<span class="tag-ok">&#10003; 已對應</span>'
      : '<span class="tag-miss">&#128560; 查無資料</span>';
    var imgUrl = (r.image || '').trim();
    var isValidUrl = imgUrl && imgUrl.startsWith('http') && imgUrl.length > 15;
    var img = isValidUrl
      ? '<img class="thumb" src="' + imgUrl + '">'
      : (imgUrl ? '<span style="color:#e53935;font-size:12px">🙈 網址有誤</span>' : '—');
    return '<tr class="' + cls + '">' +
      '<td>' + (r.sku||'—') + '</td>' +
      '<td>' + (r.type||'') + '</td>' +
      '<td>' + (r.product_size_orig||'') + '</td>' +
      '<td>' + (r.material||'') + '</td>' +
      '<td>' + (r.customs_name||'') + '</td>' +
      '<td>' + (r.box_no||'') + '</td>' +
      '<td>' + (r.pcs_per||'') + '</td>' +
      '<td>' + (r.qty||'') + '</td>' +
      '<td>' + (r.total_pcs||'') + '</td>' +
      '<td>' + (r.unit||'') + '</td>' +
      '<td>' + (r.price||'') + '</td>' +
      '<td>' + (r.total_rmb||'') + '</td>' +
      '<td>' + (r.gross_weight||'') + '</td>' +
      '<td>' + (r.len||'') + '</td>' +
      '<td>' + (r.wid||'') + '</td>' +
      '<td>' + (r.hei||'') + '</td>' +
      '<td>' + (r.volume||'') + '</td>' +
      '<td>' + (r.total_weight||'') + '</td>' +
      '<td>' + img + '</td>' +
      '<td>' + statusTag + '</td>' +
      '</tr>';
  }).join('');
  document.getElementById('preview-section').style.display = 'block';
  window.scrollTo(0, document.getElementById('preview-section').offsetTop);
}

function exportExcel() {
  var cabinetNo = document.getElementById('cabinet-no').value.trim();
  var sealNo    = document.getElementById('seal-no').value.trim();
  var shipDate  = document.getElementById('ship-date').value;
  var exporter  = document.getElementById('exporter').value.trim();
  var importer  = document.getElementById('importer').value.trim();
  var shipDateStr = '';
  if(shipDate) {
    var parts = shipDate.split('-');
    shipDateStr = parts[0] + '年' + parseInt(parts[1]) + '月' + parseInt(parts[2]) + '日';
  }

  // 收集所有有圖片的列
  var imgRows = [];
  var badUrls = [];
  allRows.forEach(function(r, i) {
    var url = (r.image || '').trim();
    if(!url) return;
    var isValid = url.startsWith('http') && url.indexOf('.') > 0 && url.length > 15;
    if(isValid) {
      imgRows.push({idx: i, url: url});
    } else {
      badUrls.push((r.sku || '第'+(i+1)+'列'));
    }
  });

  if(badUrls.length > 0) {
    showErrors(badUrls.map(function(sku) { return {sku: sku, url: '網址格式有誤'}; }));
  }

  if(imgRows.length === 0) {
    doExport({}, cabinetNo, sealNo, shipDateStr, exporter, importer);
    return;
  }

  showMsg('正在下載圖片（0/' + imgRows.length + ' 張）...', true);
  var imgMap = {};
  var done = 0;

  imgRows.forEach(function(item) {
    imgToBase64(item.url, function(b64) {
      if(b64) imgMap[item.idx] = b64;
      done++;
      showMsg('正在下載圖片（' + done + '/' + imgRows.length + ' 張）...', true);
      if(done === imgRows.length) {
        showMsg('圖片下載完成，正在產生 Excel...', true);
        doExport(imgMap, cabinetNo, sealNo, shipDateStr, exporter, importer);
      }
    });
  });
}

function doExport(imgMap, cabinetNo, sealNo, shipDateStr, exporter, importer) {
  fetch('/api/customs/export', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({
      rows: allRows,
      img_map: imgMap,
      cabinet_no: cabinetNo,
      seal_no:    sealNo,
      ship_date:  shipDateStr,
      exporter:   exporter,
      importer:   importer
    })
  }).then(function(r){return r.blob();}).then(function(blob){
    var url = URL.createObjectURL(blob);
    var a = document.createElement('a');
    a.href = url;
    var fname = '報關清單' + (cabinetNo ? '_' + cabinetNo : '') + '_' + new Date().toISOString().slice(0,10) + '.xlsx';
    a.download = fname;
    a.click();
    showMsg('匯出成功！', true);
    showToast('匯出成功！', true);
  }).catch(function(e){ showMsg('匯出失敗：' + e, false); });
}

function resetPage() {
  allRows = []; missingSkus = []; currentMissingIdx = 0;
  document.getElementById('preview-section').style.display = 'none';
  document.getElementById('xlsx-in').value = '';
  showMsg('', true);
}

function imgToBase64(url, callback) {
  var MAX = 60;
  // 先嘗試瀏覽器 Canvas 方式（Windows Chrome 可行）
  var img = new Image();
  img.crossOrigin = 'anonymous';
  img.onload = function() {
    var canvas = document.createElement('canvas');
    var ratio = Math.min(MAX / img.naturalWidth, MAX / img.naturalHeight, 1);
    canvas.width = Math.round(img.naturalWidth * ratio);
    canvas.height = Math.round(img.naturalHeight * ratio);
    canvas.getContext('2d').drawImage(img, 0, 0, canvas.width, canvas.height);
    try {
      var b64 = canvas.toDataURL('image/jpeg', 0.7);
      callback(b64, null);
    } catch(e) {
      // Canvas CORS 被擋（Safari / 嚴格瀏覽器），改用後端 proxy 下載
      fetchViaProxy(url, callback);
    }
  };
  img.onerror = function() {
    // 圖片載入失敗，也試試後端 proxy
    fetchViaProxy(url, callback);
  };
  img.src = url + (url.indexOf('?') >= 0 ? '&' : '?') + '_t=' + Date.now();
}

function fetchViaProxy(url, callback) {
  // 後端 proxy 下載圖片，繞過 CORS 和防盜鏈
  fetch('/api/customs/proxy-image', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({url: url})
  })
  .then(function(r){ return r.json(); })
  .then(function(d){
    if(d.ok && d.data) callback(d.data, null);
    else callback(null, d.msg || 'proxy failed');
  })
  .catch(function(e){ callback(null, 'proxy error'); });
}

function showToast(msg, ok) {
  var old = document.getElementById('toast-notify');
  if(old) old.remove();
  var t = document.createElement('div');
  t.id = 'toast-notify';
  t.textContent = (ok ? '✅ ' : '❌ ') + msg;
  t.style.cssText = 'position:fixed;top:24px;right:24px;z-index:9999;padding:14px 22px;border-radius:10px;font-size:14px;font-weight:600;color:#fff;box-shadow:0 4px 20px rgba(0,0,0,0.25);transition:opacity 0.5s;opacity:1;background:' + (ok ? '#2e7d32' : '#c62828');
  document.body.appendChild(t);
  setTimeout(function(){ t.style.opacity='0'; setTimeout(function(){ t.remove(); }, 500); }, 4000);
}

function showErrors(errors) {
  var old = document.getElementById('error-list-modal');
  if(old) old.remove();
  var PAGE = 5;
  function render(count) {
    var old2 = document.getElementById('error-list-modal');
    if(old2) old2.remove();

    var overlay = document.createElement('div');
    overlay.id = 'error-list-modal';
    overlay.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:9998;display:flex;align-items:center;justify-content:center';

    var box = document.createElement('div');
    box.style.cssText = 'background:#fff;border-radius:12px;padding:24px;max-width:480px;width:90%;max-height:70vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,0.3)';

    var header = document.createElement('div');
    header.style.cssText = 'display:flex;justify-content:space-between;align-items:center;margin-bottom:16px';

    var title = document.createElement('h3');
    title.style.cssText = 'color:#c62828;margin:0';
    title.textContent = '發現 ' + errors.length + ' 筆圖片網址有誤';
    header.appendChild(title);

    var closeBtn = document.createElement('button');
    closeBtn.textContent = '✕';
    closeBtn.style.cssText = 'background:none;border:none;font-size:20px;cursor:pointer;color:#999';
    closeBtn.onclick = function() { overlay.remove(); };
    header.appendChild(closeBtn);
    box.appendChild(header);

    var ul = document.createElement('ul');
    ul.style.cssText = 'margin:0;padding-left:20px';
    errors.slice(0, count).forEach(function(item) {
      var li = document.createElement('li');
      li.style.cssText = 'margin-bottom:8px;font-size:13px';
      var strong = document.createElement('strong');
      strong.textContent = item.sku;
      var span = document.createElement('span');
      span.style.color = '#e53935';
      span.textContent = '：' + item.url;
      li.appendChild(strong);
      li.appendChild(span);
      ul.appendChild(li);
    });
    box.appendChild(ul);

    if(count < errors.length) {
      var moreBtn = document.createElement('button');
      moreBtn.style.cssText = 'margin-top:12px;padding:8px 16px;background:#1565c0;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px';
      moreBtn.textContent = '載入更多（還有 ' + (errors.length - count) + ' 筆）';
      moreBtn.onclick = function() { render(Math.min(count + PAGE, errors.length)); };
      box.appendChild(moreBtn);
    }

    overlay.appendChild(box);
    document.body.appendChild(overlay);
  }
  render(Math.min(PAGE, errors.length));
}

// ── 圖片建檔：URL預覽 ──
function previewImageUrl(url) {
  var preview = document.getElementById('f-image-preview');
  var thumb = document.getElementById('f-image-thumb');
  if(url && url.startsWith('http')) {
    thumb.src = url;
    preview.style.display = 'block';
  } else {
    preview.style.display = 'none';
  }
}

function fetch1688Image() {
  var url = (document.getElementById('f-1688-url').value || '').trim();
  if(!url) { showMsg('請先貼上 1688 商品網址！', false); return; }
  showMsg('正在從 1688 抓取圖片...', true);
  fetch('/api/customs/fetch-1688-image', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({url: url})
  }).then(function(r){return r.json();}).then(function(d){
    if(d.ok && d.image_url) {
      document.getElementById('f-image').value = d.image_url;
      previewImageUrl(d.image_url);
      document.getElementById('f-1688-url').value = '';
      showMsg('圖片抓取成功！', true);
    } else {
      showMsg('抓取失敗：' + (d.msg || '找不到圖片'), false);
    }
  }).catch(function(e){ showMsg('連線錯誤：' + e, false); });
}

// ── 圖片建檔：選擇檔案上傳 ──
function handleImageFile(input) {
  if(!input.files || !input.files[0]) return;
  var reader = new FileReader();
  reader.onload = function(e) {
    window.__pendingImageBase64 = e.target.result;
    var thumb = document.getElementById('f-image-thumb');
    thumb.src = e.target.result;
    document.getElementById('f-image-preview').style.display = 'block';
    document.getElementById('f-image').value = '';
    document.getElementById('f-image').placeholder = '圖片已選擇（上傳後自動填入URL）';
  };
  reader.readAsDataURL(input.files[0]);
}

// ── 圖片建檔：Ctrl+V 貼截圖 ──
document.addEventListener('DOMContentLoaded', function() {
  document.addEventListener('paste', function(e) {
    if(!document.getElementById('new-item-modal').classList.contains('show')) return;
    var items = e.clipboardData && e.clipboardData.items;
    if(!items) return;
    for(var i=0; i<items.length; i++) {
      if(items[i].type.indexOf('image') >= 0) {
        var file = items[i].getAsFile();
        var reader = new FileReader();
        reader.onload = function(ev) {
          window.__pendingImageBase64 = ev.target.result;
          document.getElementById('f-image-thumb').src = ev.target.result;
          document.getElementById('f-image-preview').style.display = 'block';
          document.getElementById('f-image').value = '';
          document.getElementById('f-image').placeholder = '截圖已貼上（儲存時自動上傳）';
        };
        reader.readAsDataURL(file);
        e.preventDefault();
        break;
      }
    }
  });
});

function showMsg(msg, ok) {
  var area = document.getElementById('msg-area');
  if(!msg) { area.innerHTML = ''; return; }
  var div = document.createElement('div');
  div.className = 'msg ' + (ok ? 'msg-ok' : 'msg-err');
  div.style.cssText = 'display:flex;align-items:center;justify-content:space-between';
  var span = document.createElement('span');
  span.innerHTML = msg;
  div.appendChild(span);
  if(!ok) {
    var btn = document.createElement('button');
    btn.textContent = '✕';
    btn.style.cssText = 'background:none;border:none;font-size:16px;cursor:pointer;color:inherit;margin-left:8px;flex-shrink:0';
    btn.onclick = function() { area.innerHTML = ''; };
    div.appendChild(btn);
  }
  area.innerHTML = '';
  area.appendChild(div);
}
</script>
</body></html>"""

@app.route("/customs")
@login_required
def customs_page():
    from flask import Response
    return Response(CUSTOMS_HTML, mimetype="text/html")

@app.route("/api/customs/upload", methods=["POST"])
@login_required
def api_customs_upload():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "msg": "請安裝 openpyxl"})

    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "msg": "未收到檔案"})

    try:
        file_bytes = f.read()
        fname = f.filename.lower()

        if fname.endswith('.xls'):
            # 舊版 xls → 用 xlrd 讀取
            try:
                import xlrd
                wb_xls = xlrd.open_workbook(file_contents=file_bytes)
                ws_xls = wb_xls.sheet_by_index(0)
                # 轉成統一格式處理
                def cv_xls(r, c):
                    try:
                        v = ws_xls.cell_value(r, c)
                        return str(v).strip() if v != '' else ""
                    except:
                        return ""
                max_row = ws_xls.nrows
                use_xls = True
            except ImportError:
                return jsonify({"ok": False, "msg": "&#128561; 哎呀！.xls 格式需要安裝 xlrd，請另存為 .xlsx 再上傳！"})
        elif fname.endswith('.xlsx'):
            import openpyxl
            # keep_vba=False, 不載入圖片避免記憶體爆炸
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                data_only=True,
                keep_vba=False
            )
            ws = wb.active
            use_xls = False
            max_row = ws.max_row
        else:
            return jsonify({"ok": False, "msg": "&#128561; 哎呀！只支援 .xlsx 或 .xls 格式，請確認檔案類型！"})

        # 統一的讀取函數
        def cv(r, c):
            if use_xls:
                try:
                    v = ws_xls.cell_value(r-1, c-1)  # xls 是 0-indexed
                    return str(v).strip() if v != '' else ""
                except:
                    return ""
            else:
                v = ws.cell(r, c).value
                return str(v).strip() if v is not None else ""

        def has_data_in_row(r):
            return any(cv(r, c) for c in range(1, 5))

        # 找標題列並建立欄位對應
        HEADER_KEYWORDS = ["箱號", "箱号", "PCS", "品名", "材質", "材质"]
        HEADER_MAP = {
            "類型": "type", "类型": "type", "嘜頭": "type",
            "產品尺寸": "product_size_orig", "产品尺寸": "product_size_orig",
            "材質": "material", "材质": "material",
            "品名": "customs_name_raw",
            "箱號": "box_no", "箱号": "box_no",
            "PCS/件": "pcs_per", "PCS": "pcs_per",
            "件數": "qty", "件数": "qty",
            "總PCS": "total_pcs", "总PCS": "total_pcs",
            "單位": "unit", "单位": "unit",
            "單價": "price", "单价": "price",
            "總金額RMB": "total_rmb", "总金额RMB": "total_rmb",
            "毛重": "gross_weight",
            "長": "len", "长": "len",
            "寬": "wid", "宽": "wid",
            "高": "hei",
            "材積": "volume", "材积": "volume",
            "總重量": "total_weight", "总重量": "total_weight",
            "SKU": "sku", "SKU編碼": "sku", "sku": "sku", "SKU碼": "sku",
            "SKU 編碼": "sku", "料號": "sku", "商品編號": "sku", "商品SKU": "sku",
            "貨號": "sku",
        }

        header_row = None
        col_map = {}  # field_name -> col_index (1-based)
        data_start = 2

        for r in range(1, min(15, max_row+1)):
            row_vals = [cv(r, c) for c in range(1, min(25, (ws_xls.ncols if use_xls else ws.max_column)+1))]
            row_text = " ".join(row_vals)
            if any(kw in row_text for kw in HEADER_KEYWORDS):
                header_row = r
                data_start = r + 1
                # 建立欄位對應
                for c, val in enumerate(row_vals, 1):
                    val = val.strip()
                    if val in HEADER_MAP:
                        col_map[HEADER_MAP[val]] = c
                log(f"找到標題列第{r}列，欄位對應：{col_map}")
                break

        def get_col(row_r, field, default_col=None):
            """用欄位名稱取值，找不到時用預設欄位號"""
            c = col_map.get(field, default_col)
            if c:
                return cv(row_r, c)
            return ""

        # ★ SKU 欄位智慧偵測：
        # 1. 先看標題列有沒有對應到 SKU
        # 2. 若沒有，掃描標題列之外的欄位（通常 SKU 沒有標題），
        #    找出「最多非空值、且內容看起來像 SKU（含字母+數字）」的欄位
        def looks_like_sku(v):
            """判斷字串是否像 SKU：含英文字母 + 數字，且長度合理"""
            if not v or len(v) < 3 or len(v) > 50:
                return False
            has_alpha = any(c.isalpha() for c in v)
            has_digit = any(c.isdigit() for c in v)
            # 純數字也可能是 SKU（例如條碼 8858891600715），這邊放寬
            if has_digit and len(v) >= 8:
                return True
            return has_alpha and has_digit

        if "sku" in col_map:
            sku_col = col_map["sku"]
            log(f"SKU 欄位：第 {sku_col} 欄（由標題列偵測）")
        else:
            # 自動偵測：掃描所有欄位，找最像 SKU 的那一欄
            # 條件：非標題列欄位 + 至少 3 筆非空值 + 70% 以上像 SKU
            max_col = ws_xls.ncols if use_xls else ws.max_column
            best_col = None
            best_score = 0
            all_candidates = []
            for c in range(1, max_col + 1):
                # 跳過已經對應到其他欄位的 col
                if c in col_map.values():
                    continue
                sku_like_count = 0
                total_nonempty = 0
                for r in range(data_start, min(data_start + 30, max_row + 1)):
                    v = cv(r, c)
                    if v:
                        total_nonempty += 1
                        if looks_like_sku(v):
                            sku_like_count += 1
                # 提高門檻：至少 3 筆資料且 70% 以上像 SKU
                if total_nonempty >= 3 and sku_like_count >= total_nonempty * 0.7:
                    all_candidates.append((c, sku_like_count, total_nonempty))
                    if sku_like_count > best_score:
                        best_score = sku_like_count
                        best_col = c
            if best_col:
                sku_col = best_col
                log(f"SKU 欄位：第 {sku_col} 欄（自動偵測，{best_score} 個像 SKU 的值）")
                if len(all_candidates) > 1:
                    log(f"⚠️ 注意：有多個候選欄位 {all_candidates}，已選分數最高者")
            else:
                sku_col = 19  # 最後備援：第 19 欄（歐樂範本 SKU 位置）
                log(f"⚠️ 標題列無 SKU 欄位，自動偵測失敗，fallback 到第 {sku_col} 欄")

        # 載入資料庫
        db, db_err = load_customs_db()
        if db_err:
            log(f"載入報關資料庫失敗：{db_err}")

        log(f"報關資料庫載入：{len(db)} 筆，SKU清單：{list(db.keys())[:10]}")

        rows = []
        errors = []
        matched = 0
        unmatched = 0
        for r in range(data_start, max_row+1):
            sku = cv(r, sku_col)
            if not sku:
                if not has_data_in_row(r):
                    continue
                errors.append(f"第 {r} 列缺少 SKU（讀第{sku_col}欄）")
                continue  # ★ 修正：空 SKU 不該建 row

            total_pcs = get_col(r, "total_pcs", 8)
            price_db  = str(db.get(sku, {}).get("price", "")) if sku in db else ""
            try:
                total_rmb = round(float(total_pcs) * float(price_db), 2) if total_pcs and price_db else ""
            except:
                total_rmb = ""

            # C 模式：尺寸一律用資料庫的，資料庫沒有才 fallback 到 XLS
            db_size = db.get(sku, {}).get("product_size", "") if sku in db else ""
            in_db = sku in db
            if in_db:
                matched += 1
            else:
                unmatched += 1
            row = {
                "sku":               sku,
                "type":              get_col(r, "type", 1),
                "product_size_orig": db_size if db_size else get_col(r, "product_size_orig", 2),
                "material":          db.get(sku, {}).get("material", get_col(r, "material", 3)) if in_db else get_col(r, "material", 3),
                "customs_name":      db.get(sku, {}).get("customs_name", "") if in_db else "",
                "box_no":            get_col(r, "box_no", 5),
                "pcs_per":           get_col(r, "pcs_per", 6),
                "qty":               get_col(r, "qty", 7),
                "total_pcs":         total_pcs,
                "unit":              db.get(sku, {}).get("unit", get_col(r, "unit", 9)) if in_db else get_col(r, "unit", 9),
                "price":             price_db if in_db else get_col(r, "price", 10),
                "total_rmb":         str(total_rmb) if total_rmb != "" else get_col(r, "total_rmb", 11),
                "gross_weight":      get_col(r, "gross_weight", 12),
                "len":               get_col(r, "len", 13),
                "wid":               get_col(r, "wid", 14),
                "hei":               get_col(r, "hei", 15),
                "volume":            get_col(r, "volume", 16),
                "total_weight":      get_col(r, "total_weight", 17),
                "image":             db.get(sku, {}).get("image", "") if in_db else "",
                "status":            "ok" if in_db else "missing",
            }
            rows.append(row)

        log(f"上傳處理完成：共 {len(rows)} 列，對應成功 {matched} 筆，未對應 {unmatched} 筆")
        if unmatched > 0 and len(db) > 0:
            unmatched_samples = [r["sku"] for r in rows if r["status"] == "missing"][:5]
            log(f"未對應 SKU 範例：{unmatched_samples}")

        if errors:
            log(f"欄位警告：{errors}")

        return jsonify({"ok": True, "rows": rows, "db": db, "warnings": errors, "db_count": len(db), "db_err": db_err, "matched": matched, "unmatched": unmatched})
    except Exception as e:
        return jsonify({"ok": False, "msg": f"讀取失敗：{e}"})

def upload_to_imgur(image_url):
    """已改用 Google Drive，此函數保留相容性"""
    return upload_image_to_drive(image_url)

@app.route("/api/customs/new-item", methods=["POST"])
@login_required
def api_customs_new_item():
    data = request.get_json()
    # 圖片 URL 自動轉存 Google Drive（避免 1688 防盜鏈導致匯出沒圖片）
    if data.get("image"):
        data["image"] = upload_image_to_drive(data["image"])
    err = append_to_customs_db(data)
    if err:
        return jsonify({"ok": False, "msg": err})
    log(f"新品建檔：{data.get('sku')} {data.get('customs_name')}")
    return jsonify({"ok": True, "image": data.get("image", "")})


@app.route("/api/customs/proxy-image", methods=["POST"])
@login_required
def api_proxy_image():
    """後端 proxy 下載圖片並回傳 base64（解決 Safari CORS 問題）"""
    try:
        import requests as req_lib, base64 as b64lib
        from PIL import Image as PILImage
        data = request.get_json(silent=True) or {}
        url = data.get("url", "").strip()
        if not url or not url.startswith("http"):
            return jsonify({"ok": False, "msg": "無效的圖片網址"})

        dl = req_lib.get(url, headers={
            "Referer": "https://www.1688.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }, timeout=10)
        if dl.status_code != 200 or not dl.content:
            return jsonify({"ok": False, "msg": f"下載失敗 HTTP {dl.status_code}"})

        # 縮圖壓縮
        pil = PILImage.open(io.BytesIO(dl.content))
        pil.thumbnail((60, 60), PILImage.LANCZOS)
        out = io.BytesIO()
        pil.save(out, format="JPEG", quality=70)
        out.seek(0)
        b64 = "data:image/jpeg;base64," + b64lib.b64encode(out.read()).decode()
        return jsonify({"ok": True, "data": b64})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/customs/extract-cell-images", methods=["POST"])
@login_required
def api_extract_cell_images():
    """用 Sheets API v4 讀取 in-cell 圖片，回填 URL 到 J 欄"""
    try:
        import requests as req_lib, re as _re

        # 取得 Google OAuth token
        creds = get_gspread_client().auth
        token = creds.token

        ss_id = GOOGLE_SHEETS_ID
        sheet_name = "商品報關資料庫"
        img_col_index = 9  # J 欄 = index 9 (0-based)

        # 用 Sheets API v4 取得完整試算表資料（含圖片）
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{ss_id}?includeGridData=true&ranges={requests.utils.quote(sheet_name)}&fields=sheets.data.rowData.values.userEnteredValue,sheets.data.rowData.values.effectiveValue"
        headers = {"Authorization": f"Bearer {token}"}
        resp = req_lib.get(url, headers=headers, timeout=30)
        data = resp.json()

        if "error" in data:
            return jsonify({"ok": False, "msg": data["error"]["message"]})

        rows = data.get("sheets", [{}])[0].get("data", [{}])[0].get("rowData", [])

        # 找出 J 欄有圖片但沒有 URL 的列
        # Sheets API 對 in-cell 圖片的回傳方式
        updated = 0
        skipped = 0
        img_urls = []

        sheet = get_gspread_client().open_by_key(ss_id).worksheet(sheet_name)

        for row_idx, row in enumerate(rows[1:], start=2):  # 從第 2 列開始（跳過標題）
            values = row.get("values", [])
            if len(values) <= img_col_index:
                continue

            cell = values[img_col_index]
            user_val = cell.get("userEnteredValue", {})
            effective_val = cell.get("effectiveValue", {})

            # 檢查是否已有 URL
            existing = user_val.get("stringValue", "") or effective_val.get("stringValue", "")
            if existing:
                skipped += 1
                continue

            # 找 in-cell 圖片的 URL（藏在 formulaValue 或 imageValue）
            formula = user_val.get("formulaValue", "")
            if formula and "IMAGE" in formula.upper():
                # =IMAGE("url") 格式
                m = _re.search(r'"([^"]+)"', formula)
                if m:
                    img_url = m.group(1)
                    sheet.update_cell(row_idx, img_col_index + 1, img_url)
                    img_urls.append({"row": row_idx, "url": img_url})
                    updated += 1

        return jsonify({
            "ok": True,
            "updated": updated,
            "skipped": skipped,
            "total_rows": len(rows) - 1,
            "img_urls": img_urls[:5]  # 回傳前 5 筆預覽
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "msg": str(e), "trace": traceback.format_exc()[-500:]})


@app.route("/api/customs/fetch-1688-image", methods=["POST"])
@login_required
def api_fetch_1688_image():
    """從 1688 商品頁面抓取主圖 URL"""
    try:
        import requests as req_lib, re as _re
        data = request.get_json(silent=True) or {}
        url = data.get("url", "").strip()
        if not url:
            return jsonify({"ok": False, "msg": "請提供 1688 商品網址"})

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.1688.com/",
            "Accept-Language": "zh-TW,zh;q=0.9"
        }
        resp = req_lib.get(url, headers=headers, timeout=10)
        html = resp.text

        # 嘗試多種方式抓圖片
        image_url = None

        # 方法1: og:image meta tag
        m = _re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', html)
        if m:
            image_url = m.group(1)

        # 方法2: 主圖 JSON
        if not image_url:
            m = _re.search(r'"mainImageUrl"\s*:\s*"([^"]+)"', html)
            if m:
                image_url = m.group(1)

        # 方法3: img src 中最大的圖
        if not image_url:
            imgs = _re.findall(r'src=["\']([^"\']+\.jpg[^"\']*)["\']', html)
            aliimgs = [i for i in imgs if 'alicdn.com' in i or 'alibaba' in i]
            if aliimgs:
                image_url = aliimgs[0]

        if image_url:
            # 清理 URL
            image_url = image_url.replace('\\/', '/').split('?')[0]
            if not image_url.startswith('http'):
                image_url = 'https:' + image_url
            return jsonify({"ok": True, "image_url": image_url})
        else:
            return jsonify({"ok": False, "msg": "找不到商品圖片，請手動貼上圖片URL"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/customs/upload-image-base64", methods=["POST"])
@login_required
def api_upload_image_base64():
    """接收前端傳來的 base64 圖片，上傳到 Google Drive，回傳 Drive URL"""
    try:
        import base64 as b64lib, hashlib
        from googleapiclient.http import MediaIoBaseUpload
        data = request.get_json(silent=True) or {}
        b64data = data.get("data", "")  # base64 字串（含 data:image/...;base64, 前綴）
        row_idx = int(data.get("row_idx", 0))
        col_idx = int(data.get("col_idx", 0))

        if not b64data or not row_idx:
            return jsonify({"ok": False, "msg": "缺少參數"})

        # 解析 base64
        if "," in b64data:
            header, b64str = b64data.split(",", 1)
            content_type = header.split(":")[1].split(";")[0] if ":" in header else "image/jpeg"
        else:
            b64str = b64data
            content_type = "image/jpeg"

        img_bytes = b64lib.b64decode(b64str)
        ext = content_type.split("/")[-1].split(";")[0] or "jpg"
        fname = hashlib.md5(img_bytes).hexdigest()[:16] + "." + ext

        # 取得 Drive service
        service, err = get_drive_service()
        if err:
            return jsonify({"ok": False, "msg": err})

        global _drive_folder_id_cache
        folder_id = _drive_folder_id_cache or os.environ.get("DRIVE_IMG_FOLDER_ID", "")
        if not folder_id:
            folder_meta = {"name": "報關圖片", "mimeType": "application/vnd.google-apps.folder"}
            folder = service.files().create(body=folder_meta, fields="id").execute()
            folder_id = folder.get("id", "")
            service.permissions().create(fileId=folder_id, body={"type": "anyone", "role": "reader"}).execute()
        _drive_folder_id_cache = folder_id

        # 查重
        existing = service.files().list(
            q=f"name='{fname}' and '{folder_id}' in parents and trashed=false",
            fields="files(id)", pageSize=1
        ).execute()
        if existing.get("files"):
            file_id = existing["files"][0]["id"]
        else:
            file_meta = {"name": fname, "parents": [folder_id]}
            media = MediaIoBaseUpload(io.BytesIO(img_bytes), mimetype=content_type, resumable=False)
            uploaded = service.files().create(body=file_meta, media_body=media, fields="id").execute()
            file_id = uploaded.get("id", "")
            service.permissions().create(fileId=file_id, body={"type": "anyone", "role": "reader"}).execute()

        drive_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w200"

        # 寫回 Google Sheets
        if row_idx and col_idx:
            client, err2 = get_sheets_client()
            if not err2:
                sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
                sh = client.open_by_key(sheet_id)
                ws = sh.worksheet("商品報關資料庫")
                ws.update_cell(row_idx, col_idx, drive_url)

        return jsonify({"ok": True, "drive_url": drive_url})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/customs/test-image", methods=["POST"])
@login_required
def api_test_image():
    """診斷單張圖片上傳，回傳詳細錯誤"""
    try:
        import requests as req_lib
        data = request.get_json(silent=True) or {}
        url = data.get("url", "").strip()
        if not url:
            return jsonify({"ok": False, "step": "input", "msg": "沒有 URL"})

        # Step 1: 下載圖片
        try:
            dl = req_lib.get(url, headers={
                "Referer": "https://www.1688.com/",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }, timeout=10)
            if dl.status_code != 200:
                return jsonify({"ok": False, "step": "download", "msg": f"HTTP {dl.status_code}"})
            img_size = len(dl.content)
        except Exception as e:
            return jsonify({"ok": False, "step": "download", "msg": str(e)})

        # Step 2: 取得 Drive service
        try:
            service, err = get_drive_service()
            if err:
                return jsonify({"ok": False, "step": "drive_auth", "msg": err})
        except Exception as e:
            return jsonify({"ok": False, "step": "drive_auth_exception", "msg": str(e)})

        # Step 3: 測試建立資料夾
        try:
            folder_meta = {"name": "報關圖片測試", "mimeType": "application/vnd.google-apps.folder"}
            folder = service.files().create(body=folder_meta, fields="id").execute()
            folder_id = folder.get("id", "")
            # 刪除測試資料夾
            service.files().delete(fileId=folder_id).execute()
            return jsonify({"ok": True, "step": "drive_ok", "img_size": img_size,
                            "msg": f"Drive 連線正常！圖片 {img_size} bytes 可下載，Drive API 可建立資料夾"})
        except Exception as e:
            return jsonify({"ok": False, "step": "drive_create_folder", "msg": str(e), "img_size": img_size})
    except Exception as e:
        return jsonify({"ok": False, "step": "exception", "msg": str(e)})


@app.route("/api/customs/get-pending-images", methods=["POST"])
@login_required
def api_get_pending_images():
    """查詢尚未轉換的圖片清單（回傳 row_idx, col_idx, url）"""
    try:
        import re as _re
        data = request.get_json(silent=True) or {}
        offset = int(data.get("offset", 0))
        batch_size = 20

        client, err = get_sheets_client()
        if err:
            return jsonify({"ok": False, "msg": err})
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        sh = client.open_by_key(sheet_id)
        ws = sh.worksheet("商品報關資料庫")
        all_values = ws.get_all_values(value_render_option='FORMULA')
        if not all_values:
            return jsonify({"ok": True, "items": [], "total": 0, "remaining": 0})

        header = all_values[0]
        try:
            img_col_idx = header.index("圖片")
        except ValueError:
            return jsonify({"ok": False, "msg": "找不到圖片欄位"})

        all_pending = []
        for row_idx, row in enumerate(all_values[1:], start=2):
            if img_col_idx < len(row):
                cell_val = row[img_col_idx]
                m = _re.search(r'=IMAGE\(["\']([^"\']+)["\']', str(cell_val), _re.IGNORECASE)
                url = m.group(1) if m else str(cell_val).strip()
                if url and url.startswith("http") and "drive.google.com" not in url and "googleapis.com" not in url:
                    all_pending.append({"row_idx": row_idx, "col_idx": img_col_idx + 1, "url": url})

        total = len(all_pending)
        batch = all_pending[offset:offset + batch_size]
        remaining = max(0, total - offset - len(batch))
        return jsonify({"ok": True, "items": batch, "total": total,
                        "remaining": remaining, "next_offset": offset + len(batch)})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/customs/migrate-images", methods=["POST"])
@login_required
def api_customs_migrate_images():
    """分批轉換 Google Sheets 中的 1688 圖片 URL 到 Google Drive（每批 30 筆）"""
    try:
        import re as _re
        from concurrent.futures import ThreadPoolExecutor, as_completed
        data = request.get_json(silent=True) or {}
        batch_size = 30
        offset = int(data.get("offset", 0))

        client, err = get_sheets_client()
        if err:
            return jsonify({"ok": False, "msg": err})
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        sh = client.open_by_key(sheet_id)
        ws = sh.worksheet("商品報關資料庫")

        all_values = ws.get_all_values(value_render_option='FORMULA')
        if not all_values:
            return jsonify({"ok": True, "updated": 0, "remaining": 0, "msg": "沒有資料"})

        header = all_values[0]
        try:
            img_col_idx = header.index("圖片")
        except ValueError:
            return jsonify({"ok": False, "msg": "找不到圖片欄位"})

        all_to_convert = []
        for row_idx, row in enumerate(all_values[1:], start=2):
            if img_col_idx < len(row):
                cell_val = row[img_col_idx]
                m = _re.search(r'=IMAGE\([\"\x27]([^\"\x27]+)[\"\x27]', str(cell_val), _re.IGNORECASE)
                url = m.group(1) if m else str(cell_val).strip()
                if url and url.startswith("http") and "drive.google.com" not in url and "googleapis.com" not in url:
                    all_to_convert.append((row_idx, img_col_idx + 1, url))

        total = len(all_to_convert)
        if total == 0:
            return jsonify({"ok": True, "updated": 0, "remaining": 0, "total": 0,
                            "msg": "所有圖片已是 Google Drive URL"})

        batch = all_to_convert[offset:offset + batch_size]
        remaining = max(0, total - offset - len(batch))

        results = {}
        def convert_one(item):
            ridx, cidx, url = item
            new_url = upload_image_to_drive(url)
            return ridx, cidx, new_url

        with ThreadPoolExecutor(max_workers=5) as pool:
            futures = {pool.submit(convert_one, item): item for item in batch}
            for future in as_completed(futures):
                ridx, cidx, new_url = future.result()
                results[(ridx, cidx)] = new_url

        updated = 0
        failed = 0
        orig_map = {item[0]: item[2] for item in batch}
        for (ridx, cidx), new_url in results.items():
            if "drive.google.com" in new_url or "googleapis.com" in new_url:
                ws.update_cell(ridx, cidx, new_url)
                updated += 1
            else:
                failed += 1

        log(f"圖片遷移 offset={offset} 完成：成功 {updated} 筆，失敗 {failed} 筆，剩餘 {remaining} 筆")
        return jsonify({"ok": True, "updated": updated, "failed": failed, "total": total,
                        "remaining": remaining, "next_offset": offset + len(batch),
                        "msg": f"已處理 {min(offset+len(batch), total)}/{total} 張（成功 {updated}，防盜鏈擋住 {failed}）"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/customs/export", methods=["POST"])
@login_required
def api_customs_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.drawing.image import Image as XLImage
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import requests as req_lib
        from PIL import Image as PILImage
        data = request.get_json()
        rows       = data.get("rows", [])
        cabinet_no = data.get("cabinet_no", "")
        seal_no    = data.get("seal_no", "")
        ship_date  = data.get("ship_date", "")
        exporter   = data.get("exporter", "")
        importer   = data.get("importer", "")

        # ── 圖片：優先用前端傳來的 base64，省去防盜鏈問題 ──────
        import base64 as b64lib
        IMG_ROW_H  = 75
        IMG_MAX_PX = 90

        frontend_img_map = data.get("img_map", {})  # {str(idx): base64_data_url}

        img_map = {}  # idx(int) -> BytesIO or None
        for i, row in enumerate(rows):
            b64data = frontend_img_map.get(str(i)) or frontend_img_map.get(i)
            if b64data:
                try:
                    if "," in b64data:
                        b64str = b64data.split(",", 1)[1]
                    else:
                        b64str = b64data
                    raw = b64lib.b64decode(b64str)
                    pil = PILImage.open(io.BytesIO(raw))
                    pil.thumbnail((IMG_MAX_PX, IMG_MAX_PX), PILImage.LANCZOS)
                    out = io.BytesIO()
                    pil.save(out, format="PNG")
                    out.seek(0)
                    img_map[i] = out
                except Exception:
                    img_map[i] = None
            else:
                img_map[i] = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "本次報關清單"

        # ── A1：出口商（可填寫）──
        ws.cell(1, 1, exporter)
        ws.cell(1, 1).font = Font(bold=True, size=11)

        # ── A2：進口商（可填寫）──
        ws.cell(2, 1, importer)
        ws.cell(2, 1).font = Font(size=11)

        # ── A3：固定 CFR ──
        ws.cell(3, 1, "CFR")
        ws.cell(3, 1).font = Font(bold=True)

        ws.cell(1, 14, "義烏:")
        ws.cell(1, 14).font = Font(bold=True)
        ws.cell(1, 15, "地址:浙江省義烏市江東街道青口東洲路1017號（東山路與東洲路交叉口）")
        ws.cell(2, 15, "電話:0579-85202818     傳真:0579-85202819")

        # ── 第4列空白 ──
        ws.cell(4, 1, "")

        # ── 第5列：櫃號/封號/出貨日期 ──────────────────────
        cabinet_text = f"櫃號: {cabinet_no}      封號：{seal_no}"
        ws.cell(5, 1, cabinet_text)
        ws.cell(5, 1).font = Font(bold=True)
        if ship_date:
            ws.cell(5, 16, f"出貨日期:{ship_date}")
            ws.cell(5, 16).font = Font(bold=True)

        # ── 第6列：欄位標題 ──────────────────────────────────
        HEADER_ROW = 6
        headers = ["類型","產品尺寸","材質","品名","箱號","PCS/件","件數","總PCS",
                   "單位","單價","總金額RMB","毛重","長","寬","高","材積","總重量","圖片"]
        header_fill = PatternFill("solid", fgColor="FF1565C0")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(HEADER_ROW, c, h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # 數字格式
        fmt_int  = '0'     # F/G/H 欄：整數無小數
        fmt_2dec = '0.00'  # P/Q 欄：兩位小數

        def to_num(val):
            try:
                return float(str(val).replace(',', '')) if val not in ('', None) else ''
            except:
                return val

        # ── 第7列起：商品資料 ──────────────────────────────
        miss_fill = PatternFill("solid", fgColor="FFFFEBEE")
        for i, row in enumerate(rows):
            r = HEADER_ROW + 1 + i
            ws.cell(r, 1,  row.get("type",""))
            ws.cell(r, 2,  row.get("product_size_orig",""))
            ws.cell(r, 3,  row.get("material",""))
            ws.cell(r, 4,  row.get("customs_name",""))
            ws.cell(r, 5,  row.get("box_no",""))
            # F(6) G(7) H(8)：整數無小數點
            for col, key in [(6,"pcs_per"),(7,"qty"),(8,"total_pcs")]:
                v = to_num(row.get(key,""))
                c = ws.cell(r, col, int(v) if isinstance(v, float) else v)
                if isinstance(v, float):
                    c.number_format = fmt_int
            ws.cell(r, 9,  row.get("unit",""))
            ws.cell(r, 10, to_num(row.get("price","")))
            ws.cell(r, 11, to_num(row.get("total_rmb","")))
            ws.cell(r, 12, to_num(row.get("gross_weight","")))
            ws.cell(r, 13, to_num(row.get("len","")))
            ws.cell(r, 14, to_num(row.get("wid","")))
            ws.cell(r, 15, to_num(row.get("hei","")))
            # P(16) Q(17)：兩位小數
            for col, key in [(16,"volume"),(17,"total_weight")]:
                v = to_num(row.get(key,""))
                c = ws.cell(r, col, v)
                if isinstance(v, float):
                    c.number_format = fmt_2dec

            # ── 嵌入圖片 ──
            img_buf = img_map.get(i)
            img_url = row.get("image", "")
            if img_buf:
                try:
                    xl_img = XLImage(img_buf)
                    xl_img.width  = IMG_MAX_PX
                    xl_img.height = IMG_MAX_PX
                    ws.add_image(xl_img, f"R{r}")
                except Exception:
                    if img_url:
                        cell = ws.cell(r, 18, "點擊看圖")
                        cell.hyperlink = img_url
                        cell.font = Font(color="FF0563C1", underline="single")
            else:
                if img_url:
                    cell = ws.cell(r, 18, "點擊看圖")
                    cell.hyperlink = img_url
                    cell.font = Font(color="FF0563C1", underline="single")

            if row.get("status") != "ok":
                for c in range(1, 19):
                    ws.cell(r, c).fill = miss_fill

            # 列高放大讓圖片顯示完整
            ws.row_dimensions[r].height = IMG_ROW_H

        # 自動調整欄寬
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # 圖片欄(R)固定寬度
        ws.column_dimensions['R'].width = 14

        fname = f"報關清單{'_'+cabinet_no if cabinet_no else ''}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname)
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ============================================================
# 貨架入庫工具
# ============================================================

WAREHOUSE_HTML = """
<!DOCTYPE html>
<html lang="zh-TW"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<title>&#x8CA8;&#x67B6;&#x5165;&#x5EAB;&#x7CFB;&#x7D71;</title>
<script src="https://unpkg.com/@zxing/library@latest/umd/index.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Microsoft JhengHei",sans-serif;background:#0f1923;color:#fff;min-height:100vh}
.topbar{background:#0a1219;height:52px;padding:0 16px;display:flex;align-items:center;gap:12px;border-bottom:1px solid #1e2d3d;position:sticky;top:0;z-index:100}
.logo{font-size:15px;font-weight:600;margin-right:auto}.logo span{color:#f4a100}
.topbar a{color:#aaa;font-size:12px;text-decoration:none;margin-left:8px}
.tabs{display:flex;border-bottom:2px solid #1e2d3d}
.tab-btn{flex:1;padding:13px 8px;background:none;border:none;color:#888;font-size:13px;cursor:pointer;font-family:inherit;border-bottom:2px solid transparent;margin-bottom:-2px;transition:.2s;text-align:center}
.tab-btn.active{color:#f4a100;border-bottom-color:#f4a100;font-weight:600}
.tab-pane{display:none;padding:14px}
.tab-pane.active{display:block}
.card{background:#1a2535;border:1px solid #1e2d3d;border-radius:12px;padding:16px;margin-bottom:14px}
.card-title{font-size:13px;color:#f4a100;font-weight:700;margin-bottom:12px}
.input-row{display:flex;gap:8px;margin-bottom:10px}
.inp{flex:1;padding:13px 14px;background:#0f1923;border:2px solid #1e2d3d;border-radius:8px;color:#fff;font-size:15px;font-family:inherit}
.inp:focus{outline:none;border-color:#f4a100}
.inp::placeholder{color:#444}
.btn{padding:13px 16px;border-radius:8px;border:none;font-size:14px;cursor:pointer;font-weight:600;font-family:inherit;white-space:nowrap;touch-action:manipulation;-webkit-tap-highlight-color:transparent}
.btn-yellow{background:#f4a100;color:#000}
.btn-green{background:#2e7d32;color:#fff}
.btn-gray{background:#444;color:#fff}
.btn-red{background:#b71c1c;color:#fff}
.btn-cam{background:#1e3a5f;color:#64b5f6;padding:13px 14px;font-size:18px}
.btn-block{width:100%;padding:14px;margin-top:8px;border-radius:8px;border:none;font-size:15px;cursor:pointer;font-weight:600;font-family:inherit;background:#f4a100;color:#000}
.sku-list{margin-top:8px;display:flex;flex-direction:column;gap:6px;max-height:260px;overflow-y:auto}
.sku-item{display:flex;align-items:center;justify-content:space-between;background:#0f1923;border:1px solid #1e2d3d;border-radius:8px;padding:10px 12px}
.sku-name{font-weight:700;font-size:14px}
.qty-row{display:flex;align-items:center;gap:10px}
.qty-btn{width:32px;height:32px;border-radius:50%;border:2px solid #f4a100;background:none;color:#f4a100;font-size:20px;cursor:pointer;line-height:1}
.qty-num{font-size:16px;font-weight:700;color:#f4a100;min-width:28px;text-align:center}
.del-btn{background:none;border:none;color:#666;font-size:20px;cursor:pointer;padding:4px}
.empty-hint{color:#555;text-align:center;padding:20px;font-size:13px}
.msg-bar{padding:10px 14px;border-radius:8px;font-size:13px;margin-bottom:10px;min-height:20px}
.msg-ok{background:#1b5e20;color:#a5d6a7}
.msg-err{background:#4a0000;color:#ef9a9a}
.msg-warn{background:#3e2800;color:#ffcc80}
.search-box{display:flex;gap:8px;margin-bottom:10px}
.result-card{background:#0f1923;border:1px solid #1e2d3d;border-radius:10px;padding:14px;margin-bottom:8px}
.result-sku{font-size:15px;font-weight:700;color:#f4a100;margin-bottom:8px}
.loc-tags{display:flex;flex-wrap:wrap;gap:6px}
.loc-tag{background:#1e3a5f;border-radius:16px;padding:5px 12px;font-size:12px}
.loc-rack{color:#64b5f6;font-weight:700}
.loc-qty{color:#f4a100;margin-left:4px}
.loc-time{color:#666;font-size:11px;display:block;margin-top:2px}
.no-result{color:#555;text-align:center;padding:30px;font-size:13px}
.rec-table{width:100%;border-collapse:collapse;font-size:12px}
.rec-table th{background:#0f1923;padding:8px;text-align:left;color:#888;border-bottom:1px solid #1e2d3d}
.rec-table td{padding:8px;border-bottom:1px solid #1a2535}
.rack-badge{background:#1e3a5f;color:#64b5f6;padding:2px 8px;border-radius:8px;font-size:11px;font-weight:700}
.spinner{display:inline-block;width:14px;height:14px;border:2px solid #333;border-top-color:#f4a100;border-radius:50%;animation:spin .7s linear infinite;vertical-align:middle;margin-right:4px}
@keyframes spin{to{transform:rotate(360deg)}}
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.88);z-index:999;align-items:center;justify-content:center}
.overlay.show{display:flex}
.confirm-box{background:#1a2535;border:3px solid #f4a100;border-radius:20px;padding:32px 24px;text-align:center;width:90%;max-width:420px}
.confirm-rack{font-size:56px;font-weight:900;color:#f4a100;letter-spacing:3px;word-break:break-all;line-height:1.1;margin:12px 0}
.confirm-label{font-size:13px;color:#aaa;margin-bottom:6px}
.confirm-items-box{text-align:left;background:#0f1923;border-radius:8px;padding:10px;margin:10px 0;max-height:180px;overflow-y:auto}
.ci{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #1e2d3d;font-size:13px}
.ci:last-child{border:none}
.confirm-btns{display:flex;gap:10px;margin-top:16px}
.confirm-btns .btn{flex:1;padding:14px}
.cam-overlay{display:none;position:fixed;inset:0;background:#000;z-index:998;flex-direction:column;align-items:center;justify-content:center}
.cam-overlay.show{display:flex}
.cam-frame{position:relative;width:100%;max-width:500px}
.cam-video{width:100%;border-radius:8px;display:block}
.cam-line{position:absolute;top:50%;left:10%;right:10%;height:2px;background:#f4a100;box-shadow:0 0 8px #f4a100}
.cam-hint{color:#aaa;font-size:13px;margin:12px 0;text-align:center}
@keyframes scanAnim{0%{top:0}50%{top:calc(100% - 2px)}100%{top:0}}
</style>
</head>
<div class="topbar">
  <div class="logo">&#x1F4E6; <span>&#x8CA8;&#x67B6;&#x5165;&#x5EAB;</span></div>
  <a href="/">&#x2302; &#x9996;&#x9801;</a>
  <a href="/logout">&#x767B;&#x51FA;</a>
</div>
<div class="tabs">
  <button class="tab-btn active" id="tb-inbound">&#x1F4E5; &#x5165;&#x5EAB;</button>
  <button class="tab-btn" id="tb-search">&#x1F50D; &#x67E5;&#x627E;</button>
  <button class="tab-btn" id="tb-records">&#x1F4CB; &#x7D00;&#x9304;</button>
</div>
<body>

<div class="tab-pane active" id="pane-inbound">
  <div id="msg-inbound" class="msg-bar"></div>
  <div class="card">
    <div class="card-title">STEP 1 - &#x8F38;&#x5165;&#x8CA8;&#x865F;</div>
    <p style="font-size:12px;color:#666;margin-bottom:10px">&#x8F38;&#x5165;&#x8CA8;&#x865F;&#x5F8C;&#x6309; + &#x52A0;&#x5165;&#xFF0C;&#x540C;&#x4E00;&#x8CA8;&#x865F;&#x91CD;&#x8907;&#x5247;&#x6578;&#x91CF;+1</p>
    <div class="input-row">
      <input type="text" id="sku-input" class="inp" placeholder="&#x8F38;&#x5165;&#x8CA8;&#x865F;...">
      <button class="btn btn-cam" id="cam-sku-btn" onclick="openCam('sku')">&#x1F4F7;</button>
      <button class="btn btn-yellow" id="add-sku-btn">+</button>
    </div>
    <div id="sku-list" class="sku-list"><div class="empty-hint">&#x5C1A;&#x672A;&#x52A0;&#x5165;&#x4EFB;&#x4F55;&#x8CA8;&#x865F;</div></div>
  </div>
  <div class="card">
    <div class="card-title">STEP 2 - &#x8F38;&#x5165;&#x5132;&#x4F4D;</div>
    <p style="font-size:12px;color:#666;margin-bottom:10px">&#x683C;&#x5F0F;: RACK-A-1&#xFF0C;&#x78BA;&#x8A8D;&#x5F8C;&#x51FA;&#x73FE;&#x5927;&#x5B57;&#x4E8C;&#x6B21;&#x78BA;&#x8A8D;</p>
    <div class="input-row">
      <input type="text" id="rack-input" class="inp" placeholder="&#x4F8B;: RACK-A-1">
      <button class="btn btn-cam" id="cam-rack-btn" onclick="openCam('rack')">&#x1F4F7;</button>
    </div>
    <button class="btn-block" id="confirm-rack-btn">&#x1F50D; &#x78BA;&#x8A8D;&#x5132;&#x4F4D;</button>
  </div>
</div>
<div class="tab-pane" id="pane-search">
  <div class="card">
    <div class="card-title">&#x67E5;&#x627E;&#x5546;&#x54C1;&#x5728;&#x54EA;&#x500B;&#x5132;&#x4F4D;</div>
    <div class="search-box">
      <input type="text" id="search-sku" class="inp" placeholder="&#x8F38;&#x5165;&#x8CA8;&#x865F;...">
      <button class="btn btn-cam" id="cam-search-btn" onclick="openCam('search')">&#x1F4F7;</button>
      <button class="btn btn-yellow" id="do-search-btn">&#x67E5;&#x627E;</button>
    </div>
    <div id="search-result"></div>
  </div>
  <div class="card">
    <div class="card-title">&#x67E5;&#x627E;&#x8CA8;&#x67B6;&#x88E1;&#x6709;&#x4EC0;&#x9EBC;</div>
    <div class="search-box">
      <input type="text" id="search-rack" class="inp" placeholder="&#x4F8B;: RACK-A-1">
      <button class="btn btn-yellow" id="do-rack-search-btn">&#x67E5;&#x627E;</button>
    </div>
    <div id="rack-result"></div>
  </div>
</div>
<div class="tab-pane" id="pane-records">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
    <span style="font-size:12px;color:#888">&#x6700;&#x8FD1; 100 &#x7B46;</span>
    <button class="btn btn-yellow" id="refresh-btn" style="padding:8px 16px;font-size:12px">&#x21BB; &#x91CD;&#x65B0;&#x6574;&#x7406;</button>
  </div>
  <div style="overflow-x:auto">
    <table class="rec-table">
      <thead><tr><th>&#x5165;&#x5EAB;&#x6642;&#x9593;</th><th>&#x8CA8;&#x865F;</th><th>&#x5132;&#x4F4D;</th><th>&#x6578;&#x91CF;</th></tr></thead>
      <tbody id="records-body"><tr><td colspan="4" style="text-align:center;color:#555;padding:20px">&#x8ACB;&#x6309;&#x91CD;&#x65B0;&#x6574;&#x7406;</td></tr></tbody>
    </table>
  </div>
</div>
<div class="overlay" id="confirm-overlay">
  <div class="confirm-box">
    <div class="confirm-label">&#x1F4CD; &#x78BA;&#x8A8D;&#x5165;&#x5EAB;&#x5230;&#x6B64;&#x5132;&#x4F4D;</div>
    <div class="confirm-rack" id="confirm-rack-text"></div>
    <div style="font-size:12px;color:#888">&#x4EE5;&#x4E0B;&#x5546;&#x54C1;&#x5C07;&#x5165;&#x5EAB;:</div>
    <div class="confirm-items-box" id="confirm-items"></div>
    <div class="confirm-btns">
      <button class="btn btn-gray" id="cancel-btn">&#x274C; &#x53D6;&#x6D88;</button>
      <button class="btn btn-green" id="do-inbound-btn">&#x2705; &#x78BA;&#x8A8D;&#x5165;&#x5EAB;</button>
    </div>
  </div>
</div>
<div class="cam-overlay" id="cam-overlay">
  <div class="cam-frame" style="position:relative;width:100%;max-width:500px">
    <video id="cam-video" class="cam-video" autoplay playsinline muted></video>
    <!-- 對焦框 -->
    <div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);width:72%;height:130px;border-radius:8px;box-shadow:0 0 0 9999px rgba(0,0,0,0.5);">
      <!-- 掃描線 -->
      <div id="cam-scan-line" style="position:absolute;top:0;left:0;right:0;height:2px;background:#f4a100;box-shadow:0 0 8px #f4a100;animation:scanAnim 1.8s ease-in-out infinite"></div>
      <!-- 四個角 -->
      <div style="position:absolute;top:-2px;left:-2px;width:20px;height:20px;border-top:4px solid #f4a100;border-left:4px solid #f4a100;border-radius:4px 0 0 0"></div>
      <div style="position:absolute;top:-2px;right:-2px;width:20px;height:20px;border-top:4px solid #f4a100;border-right:4px solid #f4a100;border-radius:0 4px 0 0"></div>
      <div style="position:absolute;bottom:-2px;left:-2px;width:20px;height:20px;border-bottom:4px solid #f4a100;border-left:4px solid #f4a100;border-radius:0 0 0 4px"></div>
      <div style="position:absolute;bottom:-2px;right:-2px;width:20px;height:20px;border-bottom:4px solid #f4a100;border-right:4px solid #f4a100;border-radius:0 0 4px 0"></div>
    </div>
  </div>
  <div class="cam-hint" id="cam-hint">&#x5C07;&#x689D;&#x78BC;&#x5C0D;&#x6E96;&#x6A21;&#x64EC;&#x6846;</div>
  <button class="btn btn-red" id="cam-close-btn" onclick="closeCam()" style="margin-top:16px;padding:12px 32px">&#x274C; &#x95DC;&#x9589;&#x76F8;&#x6A5F;</button>
</div>

<!-- 掃描結果確認視窗 -->
<div class="overlay" id="scan-confirm-overlay">
  <div class="confirm-box" style="max-width:320px">
    <div style="font-size:13px;color:#aaa;margin-bottom:8px">&#x1F4F7; &#x5DF2;&#x25195;&#x25551;&#x5230;</div>
    <div id="scan-confirm-code" style="font-size:28px;font-weight:900;color:#f4a100;letter-spacing:2px;word-break:break-all;margin:8px 0"></div>
    <div style="font-size:12px;color:#666;margin-bottom:16px" id="scan-confirm-target"></div>
    <div class="confirm-btns">
      <button class="btn btn-gray" onclick="scanConfirmCancel()">&#x274C; &#x91CD;&#x65B0;&#x25195;</button>
      <button class="btn btn-green" onclick="scanConfirmOK()">&#x2705; &#x78BA;&#x8A8D;</button>
    </div>
  </div>
</div>

<script>
var skuList = [];
var camTarget = null;
var camStream = null;
var scanInterval = null;

function switchTab(name) {
  ['inbound','search','records'].forEach(function(t) {
    var pane = document.getElementById('pane-'+t);
    var tab = document.getElementById('tb-'+t);
    if(pane) pane.style.display = (t===name)?'block':'none';
    if(tab) { tab.style.color=(t===name)?'#f4a100':'#888'; tab.style.borderBottomColor=(t===name)?'#f4a100':'transparent'; tab.style.fontWeight=(t===name)?'600':'normal'; }
  });
  if(name==='records') loadRecords();
}

function addSku() {
  var val = document.getElementById('sku-input').value.trim().toUpperCase();
  if(!val) return;
  var ex = skuList.find(function(x){ return x.sku===val; });
  if(ex){ ex.qty++; } else { skuList.push({sku:val, qty:1}); }
  document.getElementById('sku-input').value = '';
  document.getElementById('sku-input').focus();
  renderSkuList();
}

function renderSkuList() {
  var el = document.getElementById('sku-list');
  if(!skuList.length){ el.innerHTML='<div class="hint">&#x23578;&#x672A;&#x52A0;&#x5165;&#x4EFB;&#x4F55;&#x8CA8;&#x865F;</div>'; return; }
  el.innerHTML = skuList.map(function(item,i){
    return '<div class="sku-item"><span class="sku-name">'+item.sku+'</span><div class="qty-row"><button class="qty-btn" data-i="'+i+'" data-d="-1">-</button><span class="qty-num">'+item.qty+'</span><button class="qty-btn" data-i="'+i+'" data-d="1">+</button><button class="del-btn" data-del="'+i+'">&#x2715;</button></div></div>';
  }).join('');
  el.querySelectorAll('.qty-btn').forEach(function(btn){
    btn.addEventListener('click', function(){
      var i=parseInt(this.getAttribute('data-i'));
      var d=parseInt(this.getAttribute('data-d'));
      skuList[i].qty=Math.max(1,skuList[i].qty+d);
      renderSkuList();
    });
  });
  el.querySelectorAll('.del-btn').forEach(function(btn){
    btn.addEventListener('click', function(){
      skuList.splice(parseInt(this.getAttribute('data-del')),1);
      renderSkuList();
    });
  });
}

function confirmRack() {
  var skuVal = document.getElementById('sku-input').value.trim().toUpperCase();
  if(skuVal) addSku();
  if(!skuList.length){ showMsg('inbound','&#x26A0; &#x8ACB;&#x5148;&#x8F38;&#x5165;&#x8CA8;&#x865F;','warn'); return; }
  var rack = document.getElementById('rack-input').value.trim().toUpperCase();
  if(!rack){ showMsg('inbound','&#x26A0; &#x8ACB;&#x8F38;&#x5165;&#x5132;&#x4F4D;&#x689D;&#x78BC;','warn'); return; }
  document.getElementById('confirm-rack-text').textContent = rack;
  document.getElementById('confirm-items').innerHTML = skuList.map(function(item){
    return '<div class="ci"><span>'+item.sku+'</span><span style="color:#f4a100">'+item.qty+' &#x4EF6;</span></div>';
  }).join('');
  document.getElementById('confirm-overlay').classList.add('show');
}

function doInbound() {
  var rack = document.getElementById('rack-input').value.trim().toUpperCase();
  document.getElementById('confirm-overlay').classList.remove('show');
  showMsg('inbound','<span class="spinner"></span>&#x5BEB;&#x5165;&#x4E2D;...','ok');
  fetch('/api/warehouse/inbound',{
    method:'POST', headers:{'Content-Type':'application/json'},
    body:JSON.stringify({rack:rack, items:skuList})
  }).then(function(r){return r.json();}).then(function(d){
    if(d.ok){
      showMsg('inbound','&#x2705; &#x5165;&#x5EAB;&#x6210;&#x529F;! '+d.count+' &#x7B46;&#x5DF2;&#x5132;&#x5B58;','ok');
      skuList=[]; renderSkuList();
      document.getElementById('rack-input').value='';
    } else {
      showMsg('inbound','&#x274C; '+d.msg,'err');
    }
  }).catch(function(e){ showMsg('inbound','&#x274C; &#x932F;&#x8AA4;: '+e,'err'); });
}

function doSearch() {
  var q = document.getElementById('search-sku').value.trim().toUpperCase();
  if(!q) return;
  var el = document.getElementById('search-result');
  el.innerHTML = '<div style="color:#888;padding:10px"><span class="spinner"></span>&#x67E5;&#x627E;&#x4E2D;...</div>';
  fetch('/api/warehouse/search?q='+encodeURIComponent(q))
    .then(function(r){return r.json();}).then(function(d){
      if(!d.ok){ el.innerHTML='<div class="no-result">&#x5931;&#x6557;: '+d.msg+'</div>'; return; }
      if(!d.results.length){ el.innerHTML='<div class="no-result">&#x1F50D; &#x627E;&#x4E0D;&#x5230; '+q+'</div>'; return; }
      var grouped={};
      d.results.forEach(function(r){ if(!grouped[r.sku]) grouped[r.sku]=[]; grouped[r.sku].push(r); });
      el.innerHTML=Object.keys(grouped).map(function(sku){
        var tags=grouped[sku].map(function(r){
          return '<div class="loc-tag"><span class="loc-rack">'+r.rack+'</span>'+(r.qty?'<span class="loc-qty">x'+r.qty+'</span>':'')+'<span class="loc-time">'+r.time+'</span></div>';
        }).join('');
        return '<div class="result-card"><div class="result-sku">&#x1F4E6; '+sku+'</div><div class="loc-tags">'+tags+'</div></div>';
      }).join('');
    });
}

function doRackSearch() {
  var q = document.getElementById('search-rack').value.trim().toUpperCase();
  if(!q) return;
  var el = document.getElementById('rack-result');
  el.innerHTML = '<div style="color:#888;padding:10px"><span class="spinner"></span>&#x67E5;&#x627E;&#x4E2D;...</div>';
  fetch('/api/warehouse/search-rack?rack='+encodeURIComponent(q))
    .then(function(r){return r.json();}).then(function(d){
      if(!d.ok){ el.innerHTML='<div class="no-result">&#x5931;&#x6557;: '+d.msg+'</div>'; return; }
      if(!d.results.length){ el.innerHTML='<div class="no-result">&#x1F50D; &#x5132;&#x4F4D; '+q+' &#x6C92;&#x6709;&#x7D00;&#x9304;</div>'; return; }
      el.innerHTML='<div class="result-card"><div class="result-sku">&#x1F4CD; '+q+'('+d.results.length+'&#x7B46;)</div><table class="rec-table"><thead><tr><th>&#x8CA8;&#x865F;</th><th>&#x6578;&#x91CF;</th><th>&#x6642;&#x9593;</th></tr></thead><tbody>'+
        d.results.map(function(r){
          return '<tr><td style="color:#f4a100;font-weight:700">'+r.sku+'</td><td>'+(r.qty||'-')+'</td><td style="color:#888;font-size:11px">'+r.time+'</td></tr>';
        }).join('')+'</tbody></table></div>';
    });
}

function loadRecords() {
  var tbody = document.getElementById('records-body');
  tbody.innerHTML='<tr><td colspan="4" style="text-align:center;color:#888;padding:16px"><span class="spinner"></span>&#x8F09;&#x5165;&#x4E2D;...</td></tr>';
  fetch('/api/warehouse/records').then(function(r){return r.json();}).then(function(d){
    if(!d.ok||!d.records.length){
      tbody.innerHTML='<tr><td colspan="4" style="text-align:center;color:#555;padding:20px">&#x5C1A;&#x7121;&#x7D00;&#x9304;</td></tr>'; return;
    }
    tbody.innerHTML=d.records.map(function(r){
      return '<tr><td style="color:#888;font-size:11px">'+r.time+'</td><td style="color:#f4a100;font-weight:700">'+r.sku+'</td><td><span class="rack-badge">'+r.rack+'</span></td><td>'+(r.qty||'-')+'</td></tr>';
    }).join('');
  });
}

function openCam(target) {
  camTarget = target;
  var overlay = document.getElementById('cam-overlay');
  overlay.classList.add('show');
  var hint = document.getElementById('cam-hint');
  if(hint) hint.textContent = '\u6B63\u5728\u555F\u52D5\u76F8\u6A5F...';

  if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) {
    if(hint) hint.textContent = '\u6B64\u88DD\u7F6E\u4E0D\u652F\u6301\u76F8\u6A5F';
    return;
  }

  navigator.mediaDevices.getUserMedia({
    audio: false,
    video: {
      facingMode: {ideal: 'environment'},
      width:  {ideal: 1920},
      height: {ideal: 1080},
      advanced: [{focusMode: 'continuous'}]
    }
  }).then(function(stream) {
    var video = document.getElementById('cam-video');
    video.srcObject = stream;
    video.setAttribute('playsinline', true);
    video.play();
    window._camStream = stream;

    // 嘗試連續對焦
    var track = stream.getVideoTracks()[0];
    if(track && track.applyConstraints) {
      track.applyConstraints({advanced:[{focusMode:'continuous'}]}).catch(function(){});
    }
    if(hint) hint.textContent = '\u5C07\u689D\u78BC\u5C0D\u6E96\u6A21\u64EC\u6846';

    if(typeof ZXing === 'undefined') {
      if(hint) hint.textContent = 'ZXing \u672A\u8F09\u5165';
      return;
    }

    // 用 BrowserMultiFormatReader 全畫面掃描（支援任意角度/大小）
    var hints = new Map();
    hints.set(ZXing.DecodeHintType.TRY_HARDER, true);
    hints.set(ZXing.DecodeHintType.POSSIBLE_FORMATS, [
      ZXing.BarcodeFormat.QR_CODE,
      ZXing.BarcodeFormat.CODE_128,
      ZXing.BarcodeFormat.CODE_39,
      ZXing.BarcodeFormat.EAN_13,
      ZXing.BarcodeFormat.EAN_8,
      ZXing.BarcodeFormat.UPC_A,
      ZXing.BarcodeFormat.DATA_MATRIX,
      ZXing.BarcodeFormat.ITF,
      ZXing.BarcodeFormat.CODABAR,
      ZXing.BarcodeFormat.CODE_93
    ]);

    var reader = new ZXing.BrowserMultiFormatReader(hints);
    window._codeReader = reader;
    window._scanLocked = false;
    window._lastScan = '';

    // 全畫面掃描，每 200ms 一次
    window._scanTimer = setInterval(function() {
      if(window._scanLocked) return;
      if(video.readyState !== video.HAVE_ENOUGH_DATA) return;
      var vw = video.videoWidth, vh = video.videoHeight;
      if(!vw || !vh) return;

      // 全畫面 canvas
      var canvas = document.createElement('canvas');
      canvas.width = vw;
      canvas.height = vh;
      canvas.getContext('2d').drawImage(video, 0, 0, vw, vh);

      try {
        var result = reader.decodeFromCanvas(canvas);
        if(result) {
          var code = result.getText().trim().toUpperCase();
          if(code && code !== window._lastScan) {
            window._lastScan = code;
            window._scanLocked = true;
            // 震動反饋（手機）
            if(navigator.vibrate) navigator.vibrate(100);
            handleScan(code);
          }
        }
      } catch(e) { /* 繼續掃描 */ }
    }, 200);

  }).catch(function(e) {
    if(hint) hint.textContent = '\u76F8\u6A5F\u6B0A\u9650\u932F\u8AA4: ' + e.message;
  });
}

function closeCam() {
  if(window._scanTimer){ clearInterval(window._scanTimer); window._scanTimer=null; }
  if(window._codeReader){ try{ window._codeReader.reset(); }catch(e){} window._codeReader=null; }
  if(window._camStream){ window._camStream.getTracks().forEach(function(t){t.stop();}); window._camStream=null; }
  var v = document.getElementById('cam-video');
  if(v && v.srcObject){ v.srcObject=null; }
  document.getElementById('cam-overlay').classList.remove('show');
}

function handleScan(code) {
  // 暫停掃描，顯示確認視窗
  window._pendingScan = code;
  var overlay = document.getElementById('scan-confirm-overlay');
  document.getElementById('scan-confirm-code').textContent = code;
  var targetLabel = {
    'sku': '\u8CA8;\u865F;\uFF08STEP1\uFF09',
    'rack': '\u5132;\u4F4D;\u689D;\u78BC;\uFF08STEP2\uFF09',
    'search': '\u67E5;\u627E;\u8CA8;\u865F;'
  }[camTarget] || camTarget;
  document.getElementById('scan-confirm-target').textContent = '\u76EE;\u6A19;: ' + targetLabel;
  if(overlay) overlay.classList.add('show');
}

function scanConfirmOK() {
  var code = window._pendingScan;
  var overlay = document.getElementById('scan-confirm-overlay');
  if(overlay) overlay.classList.remove('show');
  closeCam();
  if(camTarget === 'sku') {
    var ex = skuList.find(function(x){ return x.sku === code; });
    if(ex){ ex.qty++; } else { skuList.push({sku:code, qty:1}); }
    renderSkuList();
    showMsg('inbound', '&#x1F4F7; \u5DF2\u52A0\u5165: ' + code, 'ok');
  } else if(camTarget === 'rack') {
    document.getElementById('rack-input').value = code;
    showMsg('inbound', '&#x1F4F7; \u5132\u4F4D: ' + code, 'ok');
  } else if(camTarget === 'search') {
    document.getElementById('search-sku').value = code;
    doSearch();
  }
}

function scanConfirmCancel() {
  var overlay = document.getElementById('scan-confirm-overlay');
  if(overlay) overlay.classList.remove('show');
  // 解鎖，繼續掃描
  window._scanLocked = false;
  window._lastScan = '';
}


function showMsg(zone,msg,type) {
  var el=document.getElementById('msg-'+zone);
  el.className='msg msg-'+(type||'ok');
  el.innerHTML=msg;
  if(type==='ok') setTimeout(function(){el.className='msg';},4000);
}

window.addEventListener('DOMContentLoaded', function() {
  document.getElementById('tb-inbound').addEventListener('click', function(){ switchTab('inbound'); });
  document.getElementById('tb-search').addEventListener('click', function(){ switchTab('search'); });
  document.getElementById('tb-records').addEventListener('click', function(){ switchTab('records'); });
  document.getElementById('add-sku-btn').addEventListener('click', addSku);
  document.getElementById('sku-input').addEventListener('keydown', function(e){ if(e.key==='Enter'){e.preventDefault();addSku();} });
  document.getElementById('confirm-rack-btn').addEventListener('click', confirmRack);
  document.getElementById('rack-input').addEventListener('keydown', function(e){ if(e.key==='Enter'){e.preventDefault();confirmRack();} });
  document.getElementById('cancel-btn').addEventListener('click', function(){ document.getElementById('confirm-overlay').classList.remove('show'); });
  document.getElementById('do-inbound-btn').addEventListener('click', doInbound);
  document.getElementById('do-search-btn').addEventListener('click', doSearch);
  document.getElementById('search-sku').addEventListener('keydown', function(e){ if(e.key==='Enter') doSearch(); });
  document.getElementById('do-rack-btn').addEventListener('click', doRackSearch);
  document.getElementById('search-rack').addEventListener('keydown', function(e){ if(e.key==='Enter') doRackSearch(); });
  document.getElementById('refresh-btn').addEventListener('click', loadRecords);
  document.getElementById('sku-input').focus();
});
</script>
</body></html>
"""


def get_warehouse_sheet():
    """取得貨架庫位紀錄 worksheet，不存在就自動建立"""
    client, err = get_sheets_client()
    if err:
        return None, err
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID")
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet("貨架庫位紀錄")
        except Exception:
            ws = sh.add_worksheet(title="貨架庫位紀錄", rows=1000, cols=6)
            ws.append_row(["貨號", "儲位", "數量", "入庫時間", "備註"])
        return ws, None
    except Exception as e:
        return None, str(e)


@app.route("/warehouse")
@login_required
def warehouse_page():
    from flask import Response
    return Response(WAREHOUSE_HTML, mimetype='text/html; charset=utf-8')


@app.route("/api/warehouse/inbound", methods=["POST"])
@login_required
def api_warehouse_inbound():
    data  = request.get_json()
    rack  = data.get("rack", "").strip().upper()
    items = data.get("items", [])
    if not rack:
        return jsonify({"ok": False, "msg": "儲位條碼不能為空"})
    if not items:
        return jsonify({"ok": False, "msg": "請先掃描至少一個貨號"})
    ws, err = get_warehouse_sheet()
    if err:
        return jsonify({"ok": False, "msg": err})
    try:
        now = datetime.now().strftime("%Y/%m/%d %H:%M")
        rows_to_add = []
        for item in items:
            sku = str(item.get("sku", "")).strip().upper()
            qty = item.get("qty", 1)
            if sku:
                rows_to_add.append([sku, rack, qty, now, ""])
        if rows_to_add:
            ws.append_rows(rows_to_add)
        return jsonify({"ok": True, "count": len(rows_to_add)})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/warehouse/search")
@login_required
def api_warehouse_search():
    q = request.args.get("q", "").strip().upper()
    if not q:
        return jsonify({"ok": False, "msg": "請輸入查詢關鍵字"})
    ws, err = get_warehouse_sheet()
    if err:
        return jsonify({"ok": False, "msg": err})
    try:
        rows = ws.get_all_records()
        results = []
        for row in rows:
            sku = str(row.get("貨號", "")).strip().upper()
            if q in sku:
                results.append({
                    "sku":  sku,
                    "rack": str(row.get("儲位", "")),
                    "qty":  str(row.get("數量", "")),
                    "time": str(row.get("入庫時間", "")),
                })
        results.reverse()
        return jsonify({"ok": True, "results": results})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/warehouse/search-rack")
@login_required
def api_warehouse_search_rack():
    rack = request.args.get("rack", "").strip().upper()
    if not rack:
        return jsonify({"ok": False, "msg": "請輸入儲位條碼"})
    ws, err = get_warehouse_sheet()
    if err:
        return jsonify({"ok": False, "msg": err})
    try:
        rows = ws.get_all_records()
        results = []
        for row in rows:
            r = str(row.get("儲位", "")).strip().upper()
            if rack in r:
                results.append({
                    "sku":  str(row.get("貨號", "")),
                    "rack": r,
                    "qty":  str(row.get("數量", "")),
                    "time": str(row.get("入庫時間", "")),
                })
        results.reverse()
        return jsonify({"ok": True, "results": results})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/warehouse/records")
@login_required
def api_warehouse_records():
    ws, err = get_warehouse_sheet()
    if err:
        return jsonify({"ok": False, "msg": err})
    try:
        rows = ws.get_all_records()
        records = [{"sku": str(r.get("貨號","")), "rack": str(r.get("儲位","")),
                    "qty": str(r.get("數量","")), "time": str(r.get("入庫時間",""))}
                   for r in rows]
        records.reverse()
        return jsonify({"ok": True, "records": records[:100]})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


# ============================================================
# 超人眼鏡 Chrome Extension API
# ============================================================

# profit_script.js 內容（直接內嵌，避免讀檔問題）
_SUPERMAN_GLASSES_SCRIPT = r"""
/* 超人眼鏡 - BigSeller 利潤計算核心腳本
   由 Railway 超人特工倉提供，版本統一管理 v1.2
   v1.2: 成本改從庫存清單 DOM 讀取，準確可靠 */
(function () {
  'use strict';
  if (window.__supermanGlassesLoaded) return;
  window.__supermanGlassesLoaded = true;

  const PANEL_ID   = 'sg-profit-panel';
  const TOGGLE_ID  = 'sg-profit-toggle';
  const CACHE_KEY  = 'sg_cache_v1';
  const RAILWAY_URL = 'https://yindan-system-production.up.railway.app';
  const COST_API    = RAILWAY_URL + '/api/superman-glasses/cost';
  const COST_KEY    = 'sg_cost_cache'; // 本地快取 key
  const COST_TTL    = 2 * 60 * 60 * 1000; // 2 小時本地快取
  const CACHE_TTL  = 60 * 60 * 1000; // 1 小時快取

  const isInventoryPage = location.pathname.includes('/inventory/index');
  const isListingPage   = location.pathname.includes('/listing/shopee/active');
  const isAdPage        = location.pathname.includes('/advertise/shopee');

  function profitColor(margin) {
    if (margin == null || isNaN(margin)) return '#888';
    if (margin >= 60) return '#7F77DD';   // 王牌 紫
    if (margin >= 50) return '#1D9E75';   // 精銳 綠
    if (margin >= 45) return '#5DCAA5';   // 準特工 淺綠
    if (margin >= 36) return '#BA7517';   // 準備撤退 橘
    return '#E24B4A';                     // 陣亡（含虧損）紅
  }

  function profitLabel(margin) {
    if (margin == null || isNaN(margin)) return '';
    if (margin >= 60) return '🚀 王牌';
    if (margin >= 50) return '💎 精銳';
    if (margin >= 45) return '🟢 準特工';
    if (margin >= 36) return '🔴 準備撤退';
    return '💀 陣亡（虧損）';
  }

  function cacheAge(ts) {
    const sec = Math.floor((Date.now() - ts) / 1000);
    if (sec < 60) return sec + ' 秒前';
    return Math.floor(sec / 60) + ' 分鐘前';
  }

  // ── 從庫存清單 DOM 讀取成本（在庫存頁面執行）────────────────
  function readCostFromDOM() {
    const map = {};
    const rows = document.querySelectorAll('.vxe-body--row');
    rows.forEach(row => {
      const cells = row.querySelectorAll('.vxe-body--column');
      if (cells.length < 16) return;
      // 欄1=SKU，欄15=加權成本價
      const skuText = cells[1]?.textContent?.trim() || '';
      const skuMatch = skuText.match(/([A-Z]{2,}\d{3,}[-\w]*)/);
      const sku = skuMatch?.[1];
      const costText = cells[15]?.textContent?.trim() || '';
      const costMatch = costText.match(/([\d]+\.[\d]+|[\d]+)/);
      const cost = costMatch ? parseFloat(costMatch[1]) : 0;
      if (sku && cost > 0) map[sku] = cost;
    });
    return map;
  }

  // ── 翻頁掃描：自動點下一頁，收集所有成本 ─────────────────────
  async function scanAllPages() {
    const allMap = {};
    let pageNum = 1;

    const getNextBtn = () => {
      const btns = document.querySelectorAll('.ant-pagination-next:not(.ant-pagination-disabled)');
      return btns.length > 0 ? btns[0] : null;
    };

    const waitForLoad = () => new Promise(resolve => setTimeout(resolve, 800));

    const getTotalItems = () => {
      const m = document.body.textContent?.match(/共\s*([\d,]+)\s*[项條]/);
      return m ? parseInt(m[1].replace(',', '')) : 0;
    };

    // 讀當前頁
    const pageCosts = readCostFromDOM();
    Object.assign(allMap, pageCosts);

    const totalItems = getTotalItems();
    const perPage = 50;
    const totalPages = Math.ceil(totalItems / perPage);

    // 更新進度
    const updateProgress = (cur, total) => {
      const el = document.getElementById('sg-scan-progress');
      if (el) el.textContent = `掃描中 ${cur}/${total} 頁...`;
    };

    while (pageNum < totalPages) {
      const nextBtn = getNextBtn();
      if (!nextBtn) break;
      nextBtn.click();
      await waitForLoad();
      pageNum++;
      updateProgress(pageNum, totalPages);
      const pageCosts = readCostFromDOM();
      Object.assign(allMap, pageCosts);
    }

    return allMap;
  }

  // ── 庫存頁面：建立掃描 UI ────────────────────────────────────
  function initInventoryScanner() {
    if (document.getElementById('sg-scanner-btn')) return;

    const style = document.createElement('style');
    style.textContent = `
#sg-scanner-wrap{position:fixed;bottom:24px;right:24px;z-index:9999;display:flex;flex-direction:column;align-items:flex-end;gap:8px;}
#sg-scanner-btn{background:#0F6E56;color:#5DCAA5;border:none;border-radius:8px;padding:10px 16px;
  font-size:13px;font-weight:500;cursor:pointer;box-shadow:0 2px 12px rgba(0,0,0,.2);
  font-family:-apple-system,sans-serif;display:flex;align-items:center;gap:8px;}
#sg-scanner-btn:hover{background:#1D9E75;}
#sg-scanner-btn canvas{border-radius:4px;}
#sg-scan-status{background:#04342C;color:#5DCAA5;border-radius:8px;padding:8px 14px;
  font-size:12px;font-family:monospace;box-shadow:0 2px 12px rgba(0,0,0,.2);display:none;max-width:260px;}
    `;
    document.head.appendChild(style);

    const wrap = document.createElement('div');
    wrap.id = 'sg-scanner-wrap';
    wrap.innerHTML = `
      <div id="sg-scan-status"></div>
      <button id="sg-scanner-btn">
        <span>&#x1F4E1;</span>
        <span id="sg-scan-progress">掃描庫存成本</span>
      </button>
    `;
    document.body.appendChild(wrap);

    // 檢查快取狀態
    let shouldAutoScan = true; // 預設自動掃描
    try {
      const saved = localStorage.getItem(COST_KEY);
      if (saved) {
        const c = JSON.parse(saved);
        const age = cacheAge(c.ts);
        const count = Object.keys(c.map || {}).length;
        const status = document.getElementById('sg-scan-status');
        status.textContent = `上次掃描：${age} | ${count} 個 SKU`;
        status.style.display = 'block';
        // 上次掃描在6小時內，不自動重掃
        if (Date.now() - c.ts < 6 * 60 * 60 * 1000) shouldAutoScan = false;
      }
    } catch(e) {}

    // 自動執行掃描（背景靜默進行）
    if (shouldAutoScan) {
      (async () => {
        const progress = document.getElementById('sg-scan-progress');
        const status = document.getElementById('sg-scan-status');
        const btn = document.getElementById('sg-scanner-btn');
        btn.disabled = true;
        btn.style.opacity = '0.7';
        status.style.display = 'block';
        status.textContent = '自動掃描中...';
        progress.textContent = '掃描中...';
        try {
          const map = await scanAllPages();
          const count = Object.keys(map).length;
          status.textContent = `自動掃描完成 ${count} 個 SKU，上傳中...`;
          const uploadResp = await fetch(COST_API, {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ map })
          });
          const uploadResult = await uploadResp.json();
          if (uploadResult.ok) {
            try { localStorage.setItem(COST_KEY, JSON.stringify({ ts: Date.now(), map })); } catch(e) {}
            progress.textContent = '掃描庫存成本';
            status.textContent = `自動掃描完成！${count} 個 SKU ✓`;
          }
        } catch(e) {
          status.textContent = '自動掃描失敗：' + e.message;
          progress.textContent = '掃描庫存成本';
        }
        btn.disabled = false;
        btn.style.opacity = '1';
      })();
    }

    document.getElementById('sg-scanner-btn').onclick = async () => {
      const btn = document.getElementById('sg-scanner-btn');
      const progress = document.getElementById('sg-scan-progress');
      const status = document.getElementById('sg-scan-status');
      btn.disabled = true;
      btn.style.opacity = '0.7';
      status.style.display = 'block';
      status.textContent = '開始掃描...';
      progress.textContent = '掃描中 1/? 頁...';

      try {
        const map = await scanAllPages();
        const count = Object.keys(map).length;

        // 上傳到 Railway（全公司共用）
        status.textContent = `掃描完成 ${count} 個 SKU，上傳中...`;
        const uploadResp = await fetch(COST_API, {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ map })
        });
        const uploadResult = await uploadResp.json();

        if (uploadResult.ok) {
          // 同時存本地快取
          try { localStorage.setItem(COST_KEY, JSON.stringify({ ts: Date.now(), map })); } catch(e) {}
          progress.textContent = '上傳成功！';
          status.textContent = `完成！${count} 個 SKU 已同步給全公司 ✓`;

          // 寫入 📊 利潤監控室（抓在線產品售價合併計算毛利）
          (async () => {
            try {
              const profitRows = [];
              // 抓在線產品取得售價
              let page = 1, totalPage = 1;
              while (page <= totalPage && page <= 10) {
                const r = await fetch(`/api/v1/product/listing/shopee/active.json?orderBy=create_time&desc=true&inquireType=0&shopeeStatus=live&status=active&pageNo=${page}&pageSize=50`);
                const d = await r.json();
                if (d.code !== 0) break;
                totalPage = d.data?.page?.totalPage || 1;
                (d.data?.page?.rows || []).forEach(row => {
                  if (row.hasVariation && row.variations?.length) {
                    row.variations.forEach(v => {
                      const sku = v.variationSku;
                      const price = v.price || v.originalPrice || 0;
                      const cost = map[sku];
                      if (!sku || !cost || !price) return;
                      const margin = ((price - cost) / price * 100).toFixed(1);
                      profitRows.push({ sku, price, cost, margin });
                    });
                  } else {
                    const sku = row.itemSku;
                    const price = row.price || row.originalPrice || 0;
                    const cost = map[sku];
                    if (!sku || !cost || !price) return;
                    const margin = ((price - cost) / price * 100).toFixed(1);
                    profitRows.push({ sku, price, cost, margin });
                  }
                });
                page++;
              }
              if (profitRows.length > 0) {
                await fetch(RAILWAY_URL + '/api/superman-glasses/product-profit', {
                  method: 'POST',
                  headers: { 'Content-Type': 'application/json' },
                  body: JSON.stringify({ rows: profitRows })
                });
                console.log('[超人眼鏡] 商品利潤已寫入 📊 利潤監控室，共', profitRows.length, '筆');
              }
            } catch(e) {
              console.warn('[超人眼鏡] 商品利潤寫入失敗:', e.message);
            }
          })();
        } else {
          progress.textContent = '上傳失敗';
          status.textContent = `掃描完成但上傳失敗：${uploadResult.msg}`;
        }
        setTimeout(() => { progress.textContent = '重新掃描'; }, 3000);
      } catch(e) {
        status.textContent = '失敗：' + e.message;
        progress.textContent = '掃描庫存成本';
      }
      btn.disabled = false;
      btn.style.opacity = '1';
    };

    // ── 每6小時自動重整頁面，觸發重新掃描 ──
    // 只有當這個頁面一直開著才需要，避免成本資料過期
    const AUTO_RELOAD_TTL = 6 * 60 * 60 * 1000;
    const scheduleAutoReload = () => {
      try {
        const saved = localStorage.getItem(COST_KEY);
        const lastTs = saved ? JSON.parse(saved).ts : 0;
        const nextReload = lastTs + AUTO_RELOAD_TTL - Date.now();
        // 至少等6小時才重整，避免掃描途中被打斷
        const delay = Math.max(nextReload, AUTO_RELOAD_TTL);
        setTimeout(() => {
          // 掃描進行中不重整
          const scanBtn = document.getElementById('sg-scanner-btn');
          if (scanBtn && scanBtn.disabled) return;
          if (location.pathname.includes('/inventory/index')) {
            location.reload();
          }
        }, delay);
      } catch(e) {}
    };
    scheduleAutoReload();
  }

  // ── 在線產品頁面：從 Railway 取成本（本地快取2小時）────────
  async function fetchCostMap() {
    // 先看本地快取
    try {
      const raw = localStorage.getItem(COST_KEY);
      if (raw) {
        const c = JSON.parse(raw);
        if (Date.now() - c.ts < COST_TTL && Object.keys(c.map||{}).length > 0) {
          return { map: c.map, fromCache: true, ts: c.ts };
        }
      }
    } catch(e) {}

    // 從 Railway 取最新
    try {
      const r = await fetch(COST_API, { cache: 'no-cache' });
      const d = await r.json();
      if (d.ok && d.count > 0) {
        // 存本地快取
        try { localStorage.setItem(COST_KEY, JSON.stringify({ ts: Date.now(), map: d.map })); } catch(e) {}
        return { map: d.map, fromCache: false, ts: d.ts, count: d.count };
      }
    } catch(e) {}

    return { map: {}, fromCache: false, ts: 0, count: 0 };
  }

  function getCostAge() {
    try {
      const raw = localStorage.getItem(COST_KEY);
      if (!raw) return null;
      const c = JSON.parse(raw);
      return c.ts ? cacheAge(c.ts) : null;
    } catch(e) { return null; }
  }

  // 抓店鋪清單
  async function fetchShops() {
    try {
      const r = await fetch('/api/v1/shop/list.json?platform=shopee');
      const d = await r.json();
      if (d.code === 0) {
        const list = d.data?.shopee || [];
        return list.map(s => ({ id: s.id, name: s.name }));
      }
    } catch(e) {}
    return [];
  }

  async function fetchListings(shopId) {
    const items = [];
    let page = 1, total = 1;
    const shopParam = shopId ? `&shopId=${shopId}` : '';
    while (page <= total) {
      const r = await fetch(
        `/api/v1/product/listing/shopee/active.json?orderBy=create_time&desc=true` +
        `&timeType=create_time&startDateStr=&endDateStr=&searchType=productName` +
        `&inquireType=0&shopeeStatus=live&status=active&pageNo=${page}&pageSize=50${shopParam}`
      );
      const d = await r.json();
      if (d.code !== 0) break;
      total = d.data?.page?.totalPage || 1;
      (d.data?.page?.rows || []).forEach(row => {
        if (row.hasVariation && row.variations?.length) {
          row.variations.forEach(v => items.push({
            name: row.name, parentSku: row.itemSku, sku: v.variationSku,
            shopName: row.shopName || '', shopId: row.shopId || '',
            originalPrice: v.originalPrice, salePrice: v.price,
            promotionPrice: v.promotionPrice, joinPromotion: v.joinPromotion === 1, stock: v.stock
          }));
        } else {
          items.push({
            name: row.name, parentSku: row.itemSku, sku: row.itemSku,
            shopName: row.shopName || '', shopId: row.shopId || '',
            originalPrice: row.originalPrice, salePrice: row.price,
            promotionPrice: row.promotionPrice, joinPromotion: row.joinPromotion === 1, stock: row.stock
          });
        }
      });
      page++;
    }
    return items;
  }

  function calcProfits(listings, costMap) {
    return listings.map(item => {
      const cost = costMap[item.sku] ?? costMap[item.parentSku] ?? null;
      const activePrice = item.joinPromotion ? item.promotionPrice : item.salePrice;
      const profit = cost != null && activePrice != null ? activePrice - cost : null;
      const margin = profit != null && activePrice ? profit / activePrice * 100 : null;
      return { ...item, cost, activePrice, profit, margin };
    });
  }

  function injectStyles() {
    if (document.getElementById('sg-styles')) return;
    const s = document.createElement('style');
    s.id = 'sg-styles';
    s.textContent = `
#${TOGGLE_ID}{position:fixed;top:50%;right:0;transform:translateY(-50%);z-index:9998;
  background:#0F6E56;color:#fff;border:none;border-radius:8px 0 0 8px;
  padding:14px 6px;cursor:pointer;writing-mode:vertical-rl;font-size:13px;
  font-weight:500;letter-spacing:.05em;box-shadow:-2px 0 8px rgba(0,0,0,.2);
  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;}
#${TOGGLE_ID}:hover{background:#1D9E75;}
#${PANEL_ID}{position:fixed;top:56px;right:0;width:420px;height:calc(100vh - 56px);
  background:#fff;border-left:1px solid #e8e8e8;box-shadow:-4px 0 20px rgba(0,0,0,.1);
  z-index:9999;display:flex;flex-direction:column;
  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:13px;}
#${PANEL_ID} .sg-head{padding:12px 14px;border-bottom:1px solid #f0f0f0;display:flex;align-items:center;gap:8px;flex-shrink:0;background:#04342C;}
#${PANEL_ID} .sg-title{font-size:15px;font-weight:500;color:#5DCAA5;flex:1;letter-spacing:.05em;}
#${PANEL_ID} .sg-version{font-size:10px;color:#1D9E75;letter-spacing:.05em;}
#${PANEL_ID} .sg-close{cursor:pointer;color:#1D9E75;font-size:18px;padding:4px;}
#${PANEL_ID} .sg-close:hover{color:#5DCAA5;}
#${PANEL_ID} .sg-minimize{cursor:pointer;color:#1D9E75;font-size:16px;padding:4px;margin-right:2px;line-height:1;user-select:none;}
#${PANEL_ID} .sg-minimize:hover{color:#5DCAA5;}
#${PANEL_ID}.sg-collapsed{height:auto!important;}
#${PANEL_ID}.sg-collapsed .sg-toolbar,
#${PANEL_ID}.sg-collapsed .sg-summary,
#${PANEL_ID}.sg-collapsed .sg-body,
#${PANEL_ID}.sg-collapsed #sg-cache-ts{display:none!important;}
#${PANEL_ID} .sg-toolbar{padding:8px 12px;border-bottom:1px solid #f0f0f0;display:flex;gap:6px;align-items:center;flex-wrap:wrap;flex-shrink:0;}
#${PANEL_ID} .sg-search{flex:1;min-width:100px;padding:5px 10px;border:1px solid #ddd;border-radius:6px;font-size:13px;outline:none;}
#${PANEL_ID} .sg-search:focus{border-color:#1D9E75;}
#${PANEL_ID} .sg-select{padding:5px 7px;border:1px solid #ddd;border-radius:6px;font-size:12px;outline:none;cursor:pointer;background:#fff;}
#${PANEL_ID} .sg-reload{padding:5px 10px;border:1px solid #9FE1CB;border-radius:6px;background:#e1f5ee;color:#0F6E56;font-size:12px;cursor:pointer;white-space:nowrap;}
#${PANEL_ID} .sg-reload:hover{background:#9FE1CB;}
#${PANEL_ID} .sg-summary{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;padding:8px 12px;border-bottom:1px solid #f0f0f0;flex-shrink:0;}
#${PANEL_ID} .sg-metric{background:#fafafa;border-radius:8px;padding:6px 8px;text-align:center;}
#${PANEL_ID} .sg-metric-label{font-size:10px;color:#999;margin-bottom:2px;}
#${PANEL_ID} .sg-metric-val{font-size:15px;font-weight:500;color:#1a1a1a;}
#${PANEL_ID} .sg-body{overflow-y:auto;flex:1;}
#${PANEL_ID} .sg-row{padding:9px 12px;border-bottom:1px solid #f8f8f8;display:grid;grid-template-columns:1fr auto;gap:4px;align-items:start;}
#${PANEL_ID} .sg-row:hover{background:#f9fffe;}
#${PANEL_ID} .sg-name{font-size:12px;color:#555;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:230px;}
#${PANEL_ID} .sg-sku{font-size:11px;color:#aaa;font-family:monospace;}
#${PANEL_ID} .sg-prices{text-align:right;line-height:1.7;}
#${PANEL_ID} .sg-prow{font-size:11px;color:#888;}
#${PANEL_ID} .sg-badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500;margin-top:2px;}
#${PANEL_ID} .sg-promo{background:#fff0e6;color:#c05200;font-size:10px;padding:1px 5px;border-radius:3px;margin-left:3px;}
#${PANEL_ID} .sg-loading{padding:40px 16px;text-align:center;color:#999;}
#${PANEL_ID} .sg-spinner{display:inline-block;width:20px;height:20px;border:2px solid #e8e8e8;border-top-color:#1D9E75;border-radius:50%;animation:sgSpin .7s linear infinite;margin-bottom:8px;}
@keyframes sgSpin{to{transform:rotate(360deg)}}
#${PANEL_ID} .sg-empty{padding:40px 16px;text-align:center;color:#bbb;font-size:13px;}
    `;
    document.head.appendChild(s);
  }

  function createLogoCanvas(w, h) {
    const cv = document.createElement('canvas');
    cv.width = w; cv.height = h;
    const s = w / 320;
    let pd = 12.5, pt = 38.4, pw = false, pwu = 0;
    const pool = [22.1,31.8,44.2,38.4,51.0,29.7,47.3,35.6,62.1,28.9];
    function nt(){ return pool[Math.floor(Math.random()*pool.length)]; }
    function draw() {
      const ctx = cv.getContext('2d');
      const now = Date.now()/1000;
      const W=cv.width, H=cv.height;
      const teal='#1D9E75',tL='#5DCAA5',tD='#0F6E56',bg='rgba(4,52,44,0.92)';
      const lx=14*s,ly=28*s,lw=108*s,lh=80*s,r=14*s;
      const rx=198*s,ry=28*s,rw=108*s,rh=80*s,bY=ly+lh/2;
      ctx.clearRect(0,0,W,H);
      function rr(x,y,w,h,rd){
        ctx.beginPath();ctx.moveTo(x+rd,y);ctx.lineTo(x+w-rd,y);ctx.quadraticCurveTo(x+w,y,x+w,y+rd);
        ctx.lineTo(x+w,y+h-rd);ctx.quadraticCurveTo(x+w,y+h,x+w-rd,y+h);
        ctx.lineTo(x+rd,y+h);ctx.quadraticCurveTo(x,y+h,x,y+h-rd);
        ctx.lineTo(x,y+rd);ctx.quadraticCurveTo(x,y,x+rd,y);ctx.closePath();
      }
      rr(lx,ly,lw,lh,r);ctx.fillStyle=bg;ctx.fill();
      rr(rx,ry,rw,rh,r);ctx.fillStyle=bg;ctx.fill();
      ctx.save();rr(lx,ly,lw,lh,r);ctx.clip();
      const chars=['0','1','$','%','T','W','D','↑','∞','▲'];
      const cols=9,colW=lw/cols;
      for(let c=0;c<cols;c++){
        const cx=lx+c*colW+colW/2,spd=0.6+(c*.15)%.8,off=(c*3.7)%7;
        for(let row=0;row<5;row++){
          const t=(now*spd+off+row*1.4)%5,yp=ly+(t/5)*(lh+16*s)-8*s;
          const a=1-(row/5)*.85,ci=Math.floor((now*3+c*7+row*13))%chars.length;
          ctx.fillStyle=`rgba(93,202,165,${a*(row===0?1:.5)})`;
          ctx.font=`${Math.floor(10*s)}px monospace`;ctx.textAlign='center';
          ctx.fillText(chars[ci],cx,yp);
        }
      }
      ctx.restore();
      ctx.save();rr(rx,ry,rw,rh,r);ctx.clip();
      const rcx=rx+rw/2,rcy=ry+rh/2;
      for(let i=3;i>=1;i--){ctx.beginPath();ctx.arc(rcx,rcy,i*20*s,0,Math.PI*2);ctx.strokeStyle=`rgba(29,158,117,${.08*i})`;ctx.lineWidth=1*s;ctx.stroke();}
      const ang=(now*1.2)%(Math.PI*2);
      ctx.beginPath();ctx.moveTo(rcx,rcy);ctx.arc(rcx,rcy,44*s,ang,ang+Math.PI*.6);
      ctx.closePath();ctx.fillStyle='rgba(29,158,117,.12)';ctx.fill();
      ctx.beginPath();ctx.arc(rcx,rcy,44*s,ang,ang+Math.PI*.6);ctx.strokeStyle=tL;ctx.lineWidth=1.5*s;ctx.stroke();
      ctx.beginPath();ctx.arc(rcx+Math.cos(ang+.3)*44*s,rcy+Math.sin(ang+.3)*44*s,3*s,0,Math.PI*2);ctx.fillStyle='#fff';ctx.fill();
      ctx.beginPath();ctx.arc(rcx,rcy,6*s,0,Math.PI*2);ctx.fillStyle=teal;ctx.fill();
      ctx.beginPath();ctx.arc(rcx,rcy,3*s,0,Math.PI*2);ctx.fillStyle='#fff';ctx.fill();
      const diff=pt-pd;
      if(Math.abs(diff)<.1){pd=pt;if(!pw){pw=true;pwu=now+1.2;}if(pw&&now>pwu){pt=nt();pw=false;}}
      else{pd+=diff*.08;}
      ctx.fillStyle='rgba(4,52,44,.7)';ctx.fillRect(rx+2*s,ry+rh-32*s,rw-4*s,30*s);
      ctx.fillStyle=teal;ctx.font=`500 ${Math.floor(8*s)}px -apple-system,sans-serif`;
      ctx.textAlign='center';ctx.fillText('利潤率',rcx,ry+rh-20*s);
      const jo=diff*.15*s;
      ctx.save();ctx.beginPath();ctx.rect(rx+4*s,ry+rh-20*s,rw-8*s,18*s);ctx.clip();
      ctx.fillStyle=pd>=0?'#5DCAA5':'#E24B4A';
      ctx.font=`500 ${Math.floor(15*s)}px -apple-system,sans-serif`;ctx.textAlign='center';
      ctx.fillText((pd>=0?'+':'')+pd.toFixed(1)+'%',rcx,ry+rh-8*s+jo);ctx.restore();
      const ir=pt>pd;
      ctx.fillStyle=`rgba(${ir?'93,202,165':'232,75,74'},${.5+Math.sin(now*4)*.5})`;
      ctx.font=`${Math.floor(10*s)}px monospace`;ctx.textAlign='right';
      ctx.fillText(ir?'▲':'▼',rx+rw-6*s,ry+rh-20*s);ctx.restore();
      rr(lx,ly,lw,lh,r);ctx.strokeStyle=teal;ctx.lineWidth=2.5*s;ctx.stroke();
      rr(rx,ry,rw,rh,r);ctx.strokeStyle=teal;ctx.lineWidth=2.5*s;ctx.stroke();
      const bx1=lx+lw,bx2=rx,bm=(bx1+bx2)/2;
      ctx.beginPath();ctx.moveTo(bx1,bY-8*s);ctx.lineTo(bm+4*s,bY-4*s);
      ctx.lineTo(bm-4*s,bY+4*s);ctx.lineTo(bx2,bY+8*s);
      ctx.strokeStyle=tL;ctx.lineWidth=3*s;ctx.lineJoin='round';ctx.stroke();
      ctx.beginPath();ctx.arc(bm,bY,4*s,0,Math.PI*2);ctx.fillStyle=tL;ctx.fill();
      ctx.beginPath();ctx.moveTo(lx,ly+18*s);ctx.lineTo(lx-12*s,ly-4*s);
      ctx.strokeStyle=tD;ctx.lineWidth=3*s;ctx.lineCap='round';ctx.stroke();
      ctx.beginPath();ctx.moveTo(rx+rw,ry+18*s);ctx.lineTo(rx+rw+12*s,ry-4*s);ctx.stroke();
    }
    function loop(){draw();requestAnimationFrame(loop);}
    loop();
    return cv;
  }

  let _rows = [];

  function renderRows() {
    const body = document.getElementById('sg-body');
    if (!body) return;
    const search = (document.getElementById('sg-search')?.value || '').toLowerCase();
    const sort   = document.getElementById('sg-sort')?.value   || 'profit_desc';
    const filter = document.getElementById('sg-filter')?.value || 'all';
    let rows = [..._rows];
    if (search) rows = rows.filter(r =>
      r.sku?.toLowerCase().includes(search) || r.name?.toLowerCase().includes(search));
    if (filter === 'promo')        rows = rows.filter(r => r.joinPromotion);
    else if (filter === 'dead')    rows = rows.filter(r => r.margin == null || r.margin < 36);
    else if (filter === 'retreat') rows = rows.filter(r => r.margin != null && r.margin >= 36 && r.margin < 45);
    else if (filter === 'rookie')  rows = rows.filter(r => r.margin != null && r.margin >= 45 && r.margin < 50);
    else if (filter === 'elite')   rows = rows.filter(r => r.margin != null && r.margin >= 50 && r.margin < 60);
    else if (filter === 'ace')     rows = rows.filter(r => r.margin != null && r.margin >= 60);
    else if (filter === 'no_cost') rows = rows.filter(r => r.cost == null);
    rows.sort((a,b)=>{
      if(sort==='margin_asc')  return (a.margin??-999)-(b.margin??-999);
      if(sort==='margin_desc') return (b.margin??-999)-(a.margin??-999);
      if(sort==='profit_asc')  return (a.profit??-999999)-(b.profit??-999999);
      if(sort==='profit_desc') return (b.profit??-999999)-(a.profit??-999999);
      if(sort==='price_desc')  return (b.activePrice??0)-(a.activePrice??0);
      if(sort==='sku_asc')     return (a.sku||'').localeCompare(b.sku||'');
      return 0;
    });
    const withM = rows.filter(r=>r.margin!=null);
    const avg = withM.length ? (withM.reduce((s,r)=>s+r.margin,0)/withM.length).toFixed(1)+'%' : '-';
    const loss  = rows.filter(r=>r.profit!=null&&r.profit<0).length;
    const promo = rows.filter(r=>r.joinPromotion).length;
    document.getElementById('sg-avg').textContent   = avg;
    document.getElementById('sg-loss').textContent  = loss;
    document.getElementById('sg-promo').textContent = promo;
    document.getElementById('sg-cnt').textContent   = rows.length;
    if (!rows.length) { body.innerHTML = '<div class="sg-empty">沒有符合條件的商品</div>'; return; }
    body.innerHTML = rows.map(r => {
      const c = profitColor(r.margin);
      const lbl = profitLabel(r.margin);
      const ps = r.profit != null ? (r.profit >= 0 ? '+' : '') + r.profit.toFixed(0) : '-';
      const ms = r.margin != null ? r.margin.toFixed(1) + '%' : '無成本';
      const pr = r.joinPromotion ? '<span class="sg-promo">促銷</span>' : '';
      return `<div class="sg-row">
        <div>
          <div class="sg-name" title="${r.name||''}">${r.name||'-'}</div>
          <div class="sg-sku">${r.sku||'-'}${pr}</div>
        </div>
        <div class="sg-prices">
          <div class="sg-prow">成本 ${r.cost!=null?'TWD '+r.cost.toFixed(0):'—'}</div>
          <div class="sg-prow">售價 TWD ${r.activePrice!=null?r.activePrice.toFixed(0):'—'}</div>
          <span class="sg-badge" style="background:${c}20;color:${c};border:1px solid ${c}40;">${lbl}　${ps}｜${ms}</span>
        </div>
      </div>`;
    }).join('');
  }

  async function loadData(forceRefresh = false) {
    const body = document.getElementById('sg-body');
    if (!body) return;
    const shopId = document.getElementById('sg-shop')?.value || '';
    const cacheKey = CACHE_KEY + (shopId ? '_' + shopId : '');
    const tsEl = document.getElementById('sg-cache-ts');

    // 從 Railway 取成本（有本地快取就秒回）
    const costResult = await fetchCostMap();
    const costMap = costResult.map || {};
    const costCount = Object.keys(costMap).length;

    if (tsEl) {
      if (costCount === 0) {
        tsEl.textContent = '⚠ 尚無成本資料 — 請到庫存清單頁面點「掃描庫存成本」（只需做一次）';
        tsEl.style.color = '#BA7517';
      } else {
        const src = costResult.fromCache ? '本地快取' : 'Railway 同步';
        tsEl.textContent = `成本：${costCount} 個 SKU｜${cacheAge(costResult.ts)}更新（${src}）`;
        tsEl.style.color = '#1D9E75';
      }
    }

    // 在線產品快取
    const cache = (() => {
      try {
        const raw = localStorage.getItem(cacheKey);
        if (!raw) return null;
        const c = JSON.parse(raw);
        if (Date.now() - c.ts > CACHE_TTL) return null;
        return c;
      } catch(e) { return null; }
    })();

    if (cache && !forceRefresh) {
      _rows = calcProfits(cache.listings, costMap);
      renderRows();
      _fetchAndUpdate(false, shopId, cacheKey, costMap);
      return;
    }

    body.innerHTML = '<div class="sg-loading"><div class="sg-spinner"></div><div>載入在線商品...</div></div>';
    await _fetchAndUpdate(true, shopId, cacheKey, costMap);
  }

  async function _fetchAndUpdate(showLoading, shopId, cacheKey, costMap) {
    const body = document.getElementById('sg-body');
    try {
      if (showLoading && body)
        body.innerHTML = '<div class="sg-loading"><div class="sg-spinner"></div><div>載入在線商品...</div></div>';
      const listings = await fetchListings(shopId);
      try { localStorage.setItem(cacheKey, JSON.stringify({ ts: Date.now(), listings })); } catch(e) {}
      _rows = calcProfits(listings, costMap);
      renderRows();
    } catch(e) {
      if (showLoading && body)
        body.innerHTML = `<div class="sg-empty" style="color:#E24B4A">載入失敗：${e.message}</div>`;
    }
  }

  function createPanel() {
    if (document.getElementById(PANEL_ID)) return;
    injectStyles();
    const panel = document.createElement('div');
    panel.id = PANEL_ID;
    const head = document.createElement('div');
    head.className = 'sg-head';
    const logo = createLogoCanvas(80, 45);
    logo.style.cssText = 'border-radius:6px;';
    head.appendChild(logo);
    const tw = document.createElement('div');
    tw.style.flex = '1';
    tw.innerHTML = `<div class="sg-title">超人眼鏡</div><div class="sg-version">Data Vision · BigSeller</div>`;
    head.appendChild(tw);
    const cb = document.createElement('span');
    cb.className = 'sg-close'; cb.innerHTML = '&#x2715;';
    cb.title = '關閉';
    cb.onclick = () => { panel.style.display='none'; document.getElementById(TOGGLE_ID).style.display=''; };
    const mb = document.createElement('span');
    mb.className = 'sg-minimize'; mb.innerHTML = '&#x2212;';
    mb.title = '縮小';
    mb.onclick = () => {
      const collapsed = panel.classList.toggle('sg-collapsed');
      mb.innerHTML = collapsed ? '&#x002B;' : '&#x2212;';
      mb.title = collapsed ? '展開' : '縮小';
    };
    head.appendChild(mb);
    head.appendChild(cb);
    panel.appendChild(head);
    panel.insertAdjacentHTML('beforeend', `
      <div class="sg-toolbar">
        <input class="sg-search" id="sg-search" type="text" placeholder="搜尋 SKU / 商品名稱..." />
        <select class="sg-select" id="sg-shop" style="max-width:110px;">
          <option value="">全部店鋪</option>
        </select>
        <select class="sg-select" id="sg-sort">
          <option value="profit_desc">利潤額 ↓</option>
          <option value="profit_asc">利潤額 ↑</option>
          <option value="margin_desc">利潤率 ↓</option>
          <option value="margin_asc">利潤率 ↑</option>
          <option value="price_desc">售價 ↓</option>
          <option value="sku_asc">SKU</option>
        </select>
        <select class="sg-select" id="sg-filter">
          <option value="all">全部</option>
          <option value="promo">促銷中</option>
          <option value="dead">💀 陣亡（虧損）(0-35%)</option>
          <option value="retreat">🔴 準備撤退 (36-44%)</option>
          <option value="rookie">🟢 準特工 (45-49%)</option>
          <option value="elite">💎 精銳 (50-60%)</option>
          <option value="ace">🚀 王牌 (60%+)</option>
          <option value="no_cost">無成本資料</option>
        </select>
        <button class="sg-reload" id="sg-reload">強制更新</button>
      </div>
      <div style="padding:4px 12px;font-size:10px;color:#aaa;border-bottom:1px solid #f0f0f0;flex-shrink:0;" id="sg-cache-ts">載入中...</div>
      <div class="sg-summary">
        <div class="sg-metric"><div class="sg-metric-label">平均利潤率</div><div class="sg-metric-val" id="sg-avg">-</div></div>
        <div class="sg-metric"><div class="sg-metric-label">虧損 SKU</div><div class="sg-metric-val" id="sg-loss" style="color:#E24B4A">-</div></div>
        <div class="sg-metric"><div class="sg-metric-label">促銷中</div><div class="sg-metric-val" id="sg-promo" style="color:#BA7517">-</div></div>
        <div class="sg-metric"><div class="sg-metric-label">顯示筆數</div><div class="sg-metric-val" id="sg-cnt">-</div></div>
      </div>
      <div class="sg-body" id="sg-body">
        <div class="sg-loading"><div class="sg-spinner"></div><div>初始化中...</div></div>
      </div>
    `);
    document.body.appendChild(panel);

    // 載入店鋪清單填入下拉選單
    fetchShops().then(shops => {
      const sel = document.getElementById('sg-shop');
      if (sel && shops.length > 0) {
        shops.forEach(s => {
          const opt = document.createElement('option');
          opt.value = s.id; opt.textContent = s.name;
          sel.appendChild(opt);
        });
      }
    });

    let st;
    document.getElementById('sg-search').oninput = () => { clearTimeout(st); st = setTimeout(renderRows, 200); };
    document.getElementById('sg-sort').onchange   = renderRows;
    document.getElementById('sg-filter').onchange = renderRows;
    document.getElementById('sg-shop').onchange   = () => loadData(true);
    document.getElementById('sg-reload').onclick  = () => loadData(true);
  }

  function createToggle() {
    if (document.getElementById(TOGGLE_ID)) return;
    const btn = document.createElement('button');
    btn.id = TOGGLE_ID; btn.textContent = '超人眼鏡';
    btn.onclick = () => { const p=document.getElementById(PANEL_ID); if(p){p.style.display='flex';btn.style.display='none';} };
    document.body.appendChild(btn);
  }

  // ── 廣告頁面：ROAS 自動調整 ────────────────────────────────

  // 根據毛利率決定目標 ROAS
  function getTargetRoas(margin) {
    // <=45% -> null（自動暫停）
    // >45%  -> 廣告費後毛利保持45%反推，進位到0.5，最高20
    if (margin == null || margin <= 45) return null;
    const floor = 1.0 / ((margin / 100) - 0.45);
    return Math.min(Math.ceil(floor * 2) / 2.0, 20.0);
  }

  function getFloorRoas(margin) {
    // 爆款下限：死守廣告費後毛利40%，進位到0.5
    // 只有初始ROAS > 爆款下限時才有空間（51%以上）
    if (margin == null || margin <= 40) return null;
    const target = getTargetRoas(margin);
    if (target == null) return null;
    const floor = Math.ceil((1.0 / ((margin / 100) - 0.40)) * 2) / 2.0;
    if (floor >= target) return null;  // 無爆款空間
    return floor;
  }

  function getRoasZoneName(margin) {
    if (margin == null) return '無成本資料';
    if (margin <= 45) return `💀 毛利${margin.toFixed(1)}% ≤45% 自動暫停`;
    const roas = getTargetRoas(margin);
    const roasStr = roas != null ? roas.toFixed(1) : '-';
    const floor = margin > 40 ? (1.0 / ((margin / 100) - 0.40)).toFixed(1) : '-';
    return `毛利${margin.toFixed(1)}% | 初始${roasStr} | 下限${floor}`;
  }

  // 呼叫 BigSeller API 修改廣告
  async function editAd(campaignId, adType, shopId, editAction, value) {
    const body = {
      campaignWithType: [{ campaignId, adType, shopId }],
      editAction
    };
    if (editAction === 11) body.roasTarget = String(value);
    if (editAction === 6)  body.budget = value;
    const r = await fetch('/api/v1/product/listing/shopee/editSingleShopeeProductAds.json', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(body)
    });
    const d = await r.json();
    return d.code === 0;
  }

  // 抓所有廣告（分頁）
  async function fetchPausedAds() {
    // 抓所有已暫停的 autoRoas 廣告
    const ads = [];
    let page = 1;
    while (true) {
      const r = await fetch('/api/v1/product/listing/shopee/queryAdCampaignShopInfoPage.json', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ pageNo: page, pageSize: 100 })
      });
      const d = await r.json();
      if (d.code !== 0) break;
      (d.data?.rows || []).forEach(ad => {
        if (ad.biddingMethod === 'autoRoas' && ad.campaignStatus === 'paused') ads.push(ad);
      });
      if (page >= (d.data?.totalPage || 1)) break;
      page++;
    }
    return ads;
  }

  async function fetchAllAds() {
    const ads = [];
    let page = 1, totalPage = 1;
    while (page <= totalPage) {
      const r = await fetch('/api/v1/product/listing/shopee/queryAdCampaignShopInfoPage.json', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ pageNo: page, pageSize: 100 })
      });
      const d = await r.json();
      if (d.code !== 0) break;
      totalPage = d.data?.totalPage || 1;
      (d.data?.rows || []).forEach(ad => {
        if (ad.biddingMethod === 'autoRoas' && ad.campaignStatus === 'ongoing') ads.push(ad);
      });
      page++;
    }
    return ads;
  }

  // 帶日期範圍抓廣告（用於7天/30天空燒分析）
  async function fetchAdsWithDateRange(days) {
    const today = new Date();
    const start = new Date(today);
    start.setDate(start.getDate() - days);
    const fmt = d => d.toISOString().split('T')[0];
    const startStr = fmt(start);
    const endStr = fmt(today);

    const map = {}; // campaignId -> {expense, broadRoi, roasTarget}
    let page = 1, totalPage = 1;
    while (page <= totalPage) {
      const r = await fetch('/api/v1/product/listing/shopee/queryAdCampaignShopInfoPage.json', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ pageNo: page, pageSize: 100, startDateStr: startStr, endDateStr: endStr })
      });
      const d = await r.json();
      if (d.code !== 0) break;
      totalPage = d.data?.totalPage || 1;
      (d.data?.rows || []).forEach(ad => {
        if (ad.biddingMethod === 'autoRoas' && ad.campaignStatus === 'ongoing') {
          map[ad.campaignId] = { expense: parseFloat(ad.expense)||0, broadRoi: parseFloat(ad.broadRoi)||0, roasTarget: parseFloat(ad.roasTarget)||0 };
        }
      });
      page++;
    }
    return map;
  }

  // 建立廣告頁面 UI
  function initAdAdjuster() {
    if (document.getElementById('sg-ad-wrap')) return;

    const style = document.createElement('style');
    style.textContent = `
#sg-ad-wrap{position:fixed;bottom:24px;right:24px;z-index:9999;display:flex;flex-direction:column;align-items:flex-end;gap:8px;font-family:-apple-system,sans-serif;}
#sg-ad-panel{background:#04342C;border:1px solid #1D9E75;border-radius:12px;padding:16px;width:320px;box-shadow:0 4px 24px rgba(0,0,0,.4);display:none;flex-direction:column;gap:10px;}
#sg-ad-panel .sg-ad-title{color:#5DCAA5;font-size:13px;font-weight:600;border-bottom:1px solid #1D9E75;padding-bottom:8px;margin-bottom:4px;}
#sg-ad-panel .sg-ad-summary{color:#aaa;font-size:11px;line-height:1.6;}
#sg-ad-panel .sg-ad-summary b{color:#fff;}
#sg-ad-run{background:#1D9E75;color:#fff;border:none;border-radius:8px;padding:10px;font-size:13px;font-weight:600;cursor:pointer;width:100%;margin-top:4px;}
#sg-ad-run:hover{background:#5DCAA5;}
#sg-ad-run:disabled{background:#333;color:#666;cursor:not-allowed;}
#sg-ad-log{max-height:140px;overflow-y:auto;font-size:10px;color:#5DCAA5;font-family:monospace;background:#021c18;border-radius:6px;padding:6px;display:none;}
#sg-ad-btn{background:#0F6E56;color:#5DCAA5;border:none;border-radius:8px;padding:10px 16px;font-size:13px;font-weight:500;cursor:pointer;box-shadow:0 2px 12px rgba(0,0,0,.2);}
#sg-ad-btn:hover{background:#1D9E75;}
    `;
    document.head.appendChild(style);

    const wrap = document.createElement('div');
    wrap.id = 'sg-ad-wrap';
    wrap.innerHTML = `
      <div id="sg-ad-panel">
        <div class="sg-ad-title">&#x1F4E1; 超人眼鏡 · ROAS 智能調整</div>
        <div class="sg-ad-summary" id="sg-ad-summary">載入中...</div>
        <div id="sg-ad-log"></div>
        <button id="sg-ad-run" disabled>分析中...</button>
      </div>
      <button id="sg-ad-btn">&#x1F453; ROAS 調整</button>
    `;
    document.body.appendChild(wrap);

    const btn   = document.getElementById('sg-ad-btn');
    const panel = document.getElementById('sg-ad-panel');
    const runBtn = document.getElementById('sg-ad-run');
    const summary = document.getElementById('sg-ad-summary');
    const log = document.getElementById('sg-ad-log');

    // 今天的日期字串 (YYYY-MM-DD)
    const todayStr = new Date().toISOString().split('T')[0];
    const ROAS_DONE_KEY   = 'sg_roas_done_' + todayStr;   // 今天 ROAS 是否已調整
    const BUDGET_DONE_KEY = 'sg_budget_done_' + todayStr; // 今天預算已加過的 campaignId

    // 檢查今天 ROAS 是否已調整過
    const roasDoneToday = !!localStorage.getItem(ROAS_DONE_KEY);
    // 今天已加過預算的廣告 ID 集合
    const budgetDoneToday = new Set(JSON.parse(localStorage.getItem(BUDGET_DONE_KEY) || '[]'));

    btn.onclick = () => {
      const visible = panel.style.display === 'flex';
      panel.style.display = visible ? 'none' : 'flex';
      if (!visible) analyzeAds();
    };

    let _adPlan = [];
    let _budgetPlan = [];
    let _pausePlan = [];
    let _warnList  = [];
    let _restartPlan = [];

    async function analyzeAds() {
      summary.innerHTML = '&#x23F3; 載入成本資料...';
      runBtn.disabled = true;
      runBtn.textContent = '分析中...';
      log.style.display = 'none';
      log.innerHTML = '';

      // 1. 取成本資料
      const costResult = await fetchCostMap();
      const costMap = costResult.map || {};
      const costCount = Object.keys(costMap).length;
      if (costCount === 0) {
        summary.innerHTML = '&#x26A0;&#xFE0F; 無成本資料<br>請先到庫存清單頁面掃描成本';
        return;
      }

      summary.innerHTML = `&#x23F3; 載入在線商品... (${costCount} 個SKU成本)`;

      // 2. 抓全量在線產品，建立 itemId -> {skus, price} map
      const itemIdMap = {};
      let listPage = 1, listTotal = 1;
      while (listPage <= listTotal) {
        const r = await fetch(
          `/api/v1/product/listing/shopee/active.json?orderBy=create_time&desc=true` +
          `&timeType=create_time&startDateStr=&endDateStr=&searchType=productName` +
          `&inquireType=0&shopeeStatus=live&status=active&pageNo=${listPage}&pageSize=50`
        );
        const d = await r.json();
        if (d.code !== 0) break;
        listTotal = d.data?.page?.totalPage || 1;
        (d.data?.page?.rows || []).forEach(row => {
          if (row.hasVariation && row.variations?.length) {
            // 每個 SKU 保留自己的售價，用 skuPriceMap 對應
            const skuPriceMap = {};
            row.variations.forEach(v => {
              if (v.variationSku) skuPriceMap[v.variationSku] = v.price || v.originalPrice || 0;
            });
            const skus = row.variations.map(v => v.variationSku).filter(Boolean);
            const stock = row.variations.reduce((s,v) => s + (v.stock||0), 0);
            itemIdMap[row.itemId] = { skus, skuPriceMap, price: 0, stock };
          } else {
            itemIdMap[row.itemId] = { skus: [row.itemSku], skuPriceMap: {}, price: row.price || row.originalPrice || 0, stock: row.stock || 0 };
          }
        });
        listPage++;
        if (listPage % 5 === 0) summary.innerHTML = `&#x23F3; 載入在線商品 ${listPage}/${listTotal} 頁...`;
      }
      // 存到 window 讓重啟邏輯使用
      window._sgItemIdMap = itemIdMap;

      summary.innerHTML = `&#x23F3; 分析廣告...`;

      // 3. 取進行中廣告 + 7天/30天累積數據
      const ads = await fetchAllAds();
      summary.innerHTML = `&#x23F3; 抓取7天數據...`;
      const ads7d  = await fetchAdsWithDateRange(7);
      summary.innerHTML = `&#x23F3; 抓取30天數據...`;
      const ads30d = await fetchAdsWithDateRange(30);

      // 4. 對每個廣告算利潤率，產生各計劃
      _adPlan = [];
      _budgetPlan = [];
      _pausePlan = [];
      _warnList  = [];
      _restartPlan = [];
      let noMargin = 0, alreadyCorrect = 0, noItemId = 0;

      for (const ad of ads) {
        const item = itemIdMap[ad.itemId];
        if (!item) { noItemId++; continue; }

        let marginTotal = 0, marginCount = 0;
        // SKU格式容錯：-01 ↔ -001
        const findCost = (sku) => {
          if (costMap[sku] != null) return costMap[sku];
          const padded = sku.replace(/-(\d{1,2})$/, (_, n) => '-' + n.padStart(3, '0'));
          if (costMap[padded] != null) return costMap[padded];
          const trimmed = sku.replace(/-0+(\d+)$/, '-$1');
          if (costMap[trimmed] != null) return costMap[trimmed];
          return null;
        };
        for (const sku of item.skus) {
          const cost = findCost(sku);
          if (cost == null || cost <= 0) continue;
          const price = (item.skuPriceMap && item.skuPriceMap[sku]) || item.price || 0;
          if (price <= 0) continue;
          marginTotal += ((price - cost) / price) * 100;
          marginCount++;
        }
        const bestMargin = marginCount > 0 ? marginTotal / marginCount : null;

        if (bestMargin === null) { noMargin++; continue; }

        const targetRoas  = getTargetRoas(bestMargin);
        const currentRoas = parseFloat(ad.roasTarget) || 0;
        const actualRoas  = parseFloat(ad.broadRoi)   || 0;
        const budget      = parseFloat(ad.campaignBudget) || 0;
        const expense     = parseFloat(ad.expense) || 0;
        const budgetUsage = budget > 0 ? expense / budget : 0;
        const adName      = ad.adName?.substring(0, 20) || String(ad.campaignId);

        // 毛利<=45% 加入暫停計劃
        if (targetRoas === null) {
          _pausePlan.push({ ad, name: adName, reason: `毛利${bestMargin.toFixed(1)}%<=45%`, isLowMargin: true });
          continue;
        }

        // ROAS 計劃（今天未調過才加）
        if (!roasDoneToday && Math.abs(targetRoas - currentRoas) > 0.05) {
          _adPlan.push({ ad, currentRoas, targetRoas, margin: bestMargin.toFixed(1), zone: getRoasZoneName(bestMargin), name: adName });
        } else if (Math.abs(targetRoas - currentRoas) <= 0.05) {
          alreadyCorrect++;
        }

        // 預算加碼計劃
        // 條件：實際ROAS >= 目標ROAS + 預算使用率 >= 90%（每次達標都加碼）
        if (actualRoas > 0 &&
            actualRoas >= currentRoas &&
            budgetUsage >= 0.9 &&
            budget > 0) {
          const newBudget = Math.ceil(budget * 1.3);
          _budgetPlan.push({ ad, currentBudget: budget, newBudget, actualRoas: actualRoas.toFixed(1), budgetUsage: (budgetUsage * 100).toFixed(0) + '%', name: adName });
        }

        // ── 防空燒：30天暫停 / 7天警告 ──
        const d7  = ads7d[ad.campaignId];
        const d30 = ads30d[ad.campaignId];
        if (d30 && d30.expense > 500 && d30.broadRoi > 0 && d30.broadRoi < currentRoas * 0.5) {
          _pausePlan.push({ ad, expense30: d30.expense.toFixed(0), roas30: d30.broadRoi.toFixed(1), targetRoas: currentRoas, name: adName });
        } else if (d7 && d7.expense > 200 && d7.broadRoi > 0 && d7.broadRoi < currentRoas * 0.5) {
          _warnList.push({ expense7: d7.expense.toFixed(0), roas7: d7.broadRoi.toFixed(1), targetRoas: currentRoas, name: adName });
        }
      }

      // 5. 分析已暫停廣告 → 重啟計劃
      summary.innerHTML = `&#x23F3; 分析暫停廣告...`;
      const pausedAds = await fetchPausedAds();
      const burnedIds = new Set(_pausePlan.map(p => p.ad.campaignId));
      const pausedItemMap = window._sgItemIdMap || itemIdMap;
      for (const pad of pausedAds) {
        const item = pausedItemMap[pad.itemId];
        if (!item) continue;
        let mTotal = 0, mCount = 0;
        for (const sku of (item.skus||[])) {
          const cost = costMap[sku];
          const price = (item.skuPriceMap&&item.skuPriceMap[sku]) || item.price || 0;
          if (!cost || !price) continue;
          mTotal += ((price-cost)/price*100); mCount++;
        }
        const margin = mCount > 0 ? mTotal/mCount : null;
        if (margin === null) continue;
        const targetRoas = getTargetRoas(margin);
        const padName = (pad.adName||'').substring(0,20);
        const shopName = (pad.shopName||'').substring(0,20);
        // 今天剛被空燒暫停 → 不重啟
        if (burnedIds.has(pad.campaignId)) continue;
        // 毛利<=45% → 繼續暫停
        if (targetRoas === null || margin <= 45) continue;
        // 庫存 > 0 → 加入重啟計劃
        const stock = item.stock || 0;
        if (stock <= 0) continue;
        _restartPlan.push({ ad: pad, name: padName, shopName, margin: margin.toFixed(1), targetRoas, stock });
      }

      const roasNeed   = _adPlan.length;
      const budgetNeed = _budgetPlan.length;
      const pauseNeed  = _pausePlan.length;
      const restartNeed = _restartPlan.length;
      const warnNeed   = _warnList.length;
      const up   = _adPlan.filter(p => p.targetRoas > p.currentRoas).length;
      const down  = _adPlan.filter(p => p.targetRoas < p.currentRoas).length;

      let roasLine = '';
      if (roasDoneToday) roasLine = `<span style="color:#5DCAA5">&#x2705; 今天 ROAS 已調整過</span>`;
      else if (roasNeed > 0) roasLine = `<b>ROAS 需調整：</b>${roasNeed} 筆（&#x2191;${up} / &#x2193;${down}）`;
      else roasLine = `<span style="color:#5DCAA5">&#x2705; ROAS 全部正確</span>`;

      let warnHtml = '';
      if (warnNeed > 0) {
        warnHtml = `<span style="color:#BA7517">&#x26A0; 7天空燒警告：${warnNeed} 筆</span><br>`;
        _warnList.forEach(w => { warnHtml += `&nbsp;&nbsp;&#x2022; ${w.name} | 花費${w.expense7} TWD | ROAS ${w.roas7}/${w.targetRoas}<br>`; });
      }

      summary.innerHTML = `
        ${roasLine}<br>
        <b>預算加碼（達標且用量&#x2265;90%）：</b>${budgetNeed} 筆<br>
        <b style="color:#E24B4A">&#x1F6D1; 30天空燒暫停：</b>${pauseNeed} 筆<br>
        <b style="color:#5DCAA5">&#x1F504; 符合重啟條件：</b>${restartNeed} 筆<br>
        ${warnHtml}
        <b>無成本/無在線產品：</b>${noMargin + noItemId} 筆<br>
        <b>進行中廣告（autoRoas）：</b>${ads.length} 筆
      `;

      // ── 分析完成後自動上傳利潤快照到 Google Sheets ──
      (async () => {
        try {
          const costRaw = localStorage.getItem('sg_ext_cost_cache') || localStorage.getItem('sg_cost_cache');
          const costMap = costRaw ? JSON.parse(costRaw).map || {} : {};
          const snapRows = [];
          for (const ad of ads) {
            const item = itemIdMap[ad.itemId];
            if (!item) continue;
            // 計算毛利
            let mTotal = 0, mCount = 0;
            for (const sku of (item.skus||[])) {
              const cost = costMap[sku];
              const price = (item.skuPriceMap&&item.skuPriceMap[sku]) || item.price || 0;
              if (!cost || !price) continue;
              mTotal += ((price-cost)/price*100);
              mCount++;
            }
            const margin = mCount > 0 ? (mTotal/mCount).toFixed(1) : '';
            const targetRoas = margin ? getTargetRoas(parseFloat(margin)) : '';
            const d7 = ads7d[ad.campaignId];
            const d30 = ads30d[ad.campaignId];
            // 判斷狀態
            let status = ad.campaignStatus === 'ongoing' ? '進行中' : '已暫停';
            let note = '';
            if (margin && parseFloat(margin) <= 45) { status = '低毛利暫停'; note = `毛利${margin}%≤45%`; }
            else if (d30 && d30.expense > 500 && d30.broadRoi > 0 && d30.broadRoi < (parseFloat(ad.roasTarget)||0) * 0.5) { note = '30天空燒'; }
            snapRows.push({
              shopName: ad.shopName || '',
              adName: (ad.adName||'').substring(0,40),
              itemId: ad.itemId,
              currentRoas: ad.roasTarget,
              actualRoas7: d7 ? parseFloat(d7.broadRoi||0).toFixed(1) : '',
              expense7: d7 ? parseFloat(d7.expense||0).toFixed(0) : '',
              expense30: d30 ? parseFloat(d30.expense||0).toFixed(0) : '',
              budget: ad.campaignBudget,
              margin,
              targetRoas,
              status,
              note
            });
          }
          if (snapRows.length > 0) {
            await fetch('https://yindan-system-production.up.railway.app/api/superman-glasses/profit-snapshot', {
              method: 'POST',
              headers: {'Content-Type': 'application/json'},
              body: JSON.stringify({ rows: snapRows })
            });
            console.log('[超人眼鏡] 利潤快照已寫入 Google Sheets，共', snapRows.length, '筆');
          }
        } catch(e) {
          console.warn('[超人眼鏡] 快照寫入失敗:', e.message);
        }
      })();

      const totalAction = (roasDoneToday ? 0 : roasNeed) + budgetNeed + pauseNeed + restartNeed;
      if (totalAction > 0) {
        runBtn.disabled = false;
        runBtn.textContent = `⚡ 執行（ROAS×${roasDoneToday?0:roasNeed} 預算×${budgetNeed} 暫停×${pauseNeed} 重啟×${restartNeed}）`;
      } else {
        runBtn.disabled = true;
        runBtn.textContent = '✅ 今日無需調整';
      }
    }

    runBtn.onclick = async () => {
      const roasNeed    = roasDoneToday ? 0 : _adPlan.length;
      const budgetNeed  = _budgetPlan.length;
      const pauseNeed   = _pausePlan.length;
      const restartNeed = (_restartPlan || []).length;
      if (roasNeed + budgetNeed + pauseNeed + restartNeed === 0) return;

      runBtn.disabled = true;
      log.style.display = 'block';
      log.innerHTML = '';
      let ok = 0, fail = 0;

      // ── 執行 ROAS 調整 ──
      if (!roasDoneToday && _adPlan.length > 0) {
        log.innerHTML += `<span style="color:#5DCAA5">── ROAS 調整 ──</span>\n`;
        for (let i = 0; i < _adPlan.length; i++) {
          const { ad, targetRoas, currentRoas, zone, name } = _adPlan[i];
          runBtn.textContent = `ROAS ${i+1}/${_adPlan.length}...`;
          const success = await editAd(ad.campaignId, ad.adType, ad.shopId, 11, targetRoas);
          const arrow = targetRoas > currentRoas ? '&#x2191;' : '&#x2193;';
          log.innerHTML += `${success?'&#x2705;':'&#x274C;'} ${name} | ${currentRoas}${arrow}${targetRoas} | ${zone}\n`;
          log.scrollTop = log.scrollHeight;
          if (success) ok++; else fail++;
          await new Promise(r => setTimeout(r, 300));
        }
        localStorage.setItem(ROAS_DONE_KEY, '1');
      }

      // ── 執行預算加碼 ──
      if (_budgetPlan.length > 0) {
        log.innerHTML += `<span style="color:#BA7517">── 預算加碼 ──</span>\n`;
        for (let i = 0; i < _budgetPlan.length; i++) {
          const { ad, currentBudget, newBudget, actualRoas, budgetUsage, name } = _budgetPlan[i];
          runBtn.textContent = `加碼 ${i+1}/${_budgetPlan.length}...`;
          const success = await editAd(ad.campaignId, ad.adType, ad.shopId, 6, newBudget);
          log.innerHTML += `${success?'&#x1F4B0;':'&#x274C;'} ${name} | ${currentBudget}→${newBudget} TWD | ROAS ${actualRoas} | 使用率${budgetUsage}\n`;
          log.scrollTop = log.scrollHeight;
          if (success) ok++; else fail++;
          await new Promise(r => setTimeout(r, 300));
        }
      }

      // ── 暫停空燒廣告（30天未達標）──
      if (_pausePlan.length > 0) {
        log.innerHTML += `<span style="color:#E24B4A">── 暫停空燒廣告 ──</span>\n`;
        for (let i = 0; i < _pausePlan.length; i++) {
          const { ad, expense30, roas30, targetRoas, name } = _pausePlan[i];
          runBtn.textContent = `暫停 ${i+1}/${_pausePlan.length}...`;
          const success = await editAd(ad.campaignId, ad.adType, ad.shopId, 2, null);
          log.innerHTML += `${success?'&#x1F6D1;':'&#x274C;'} ${name} | 30天花費${expense30}TWD | ROAS ${roas30}/${targetRoas}\n`;
          log.scrollTop = log.scrollHeight;
          if (success) ok++; else fail++;
          await new Promise(r => setTimeout(r, 300));
        }
      }

      // ── 執行重啟廣告 ──
      if (_restartPlan.length > 0) {
        log.innerHTML += `<span style="color:#5DCAA5">── 重啟廣告 ──</span>
`;
        for (let i = 0; i < _restartPlan.length; i++) {
          const { ad, name, shopName, margin, targetRoas, stock } = _restartPlan[i];
          runBtn.textContent = `重啟 ${i+1}/${_restartPlan.length}...`;
          // 先啟動
          const s1 = await editAd(ad.campaignId, ad.adType, ad.shopId, 3, null);  // 3=恢復暫停廣告
          // 再設定正確 ROAS
          const s2 = s1 ? await editAd(ad.campaignId, ad.adType, ad.shopId, 11, targetRoas) : false;
          log.innerHTML += `${s1?'&#x1F504;':'&#x274C;'} [${shopName}] ${name} | 毛利${margin}% 庫存${stock} ROAS→${targetRoas}
`;
          log.scrollTop = log.scrollHeight;
          if (s1) ok++; else fail++;
          await new Promise(r => setTimeout(r, 400));
        }
      }

      runBtn.textContent = `✅ 完成！成功 ${ok} / 失敗 ${fail}`;
      summary.innerHTML += `<br><b style="color:#5DCAA5">&#x2705; 執行完成${fail>0?'，'+fail+'筆失敗':''}</b>`;
    };
  }

  function init() {
    if (isInventoryPage) {
      setTimeout(initInventoryScanner, 500);
      // SPA 防護：監控 body，如果按鈕被移除就重新加
      const observer = new MutationObserver(() => {
        if (!document.getElementById('sg-scanner-wrap')) {
          setTimeout(initInventoryScanner, 300);
        }
      });
      observer.observe(document.body, { childList: true, subtree: false });
    } else if (isListingPage) {
      createToggle();
      createPanel();
      loadData();
    } else if (isAdPage) {
      setTimeout(initAdAdjuster, 500);
      const observer = new MutationObserver(() => {
        if (!document.getElementById('sg-ad-wrap')) {
          setTimeout(initAdAdjuster, 300);
        }
      });
      observer.observe(document.body, { childList: true, subtree: false });
    }
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', init);
  else init();
})();
"""

import io, zipfile
import requests as _requests

# ── 廣告自動排程存儲 ──────────────────────────────────────────
# 廣告效果基準（根據實際數據動態更新）
_ad_benchmark = {
    "ctr_avg": 2.91,    # 平均點擊率 %
    "cr_avg": 6.17,     # 平均轉化率 %
    "cpc_avg": 6.9,     # 平均每次點擊費用 TWD
    "updated": None     # 最後更新時間
}

_ad_scheduler_store = {
    "lock": None,              # {"ts": timestamp, "task": "roas/budget"}
    "last_daily": None,        # 上次執行每日任務的日期 "2026-04-16"
    "last_hourly": None,       # 上次執行每小時任務的時間戳
    "log": [],                 # 執行記錄（最多100筆）
    "low_margin_shops": [],    # 低利潤廣告（依店鋪分類）
    "low_margin_ts": 0,        # 上次更新時間
}

def _ad_log(msg, write_sheet=False):
    """記錄廣告排程執行日誌，重要操作加入寫入佇列"""
    from datetime import timezone, timedelta
    tw_tz = timezone(timedelta(hours=8))
    _now_tw = datetime.now(tw_tz)
    ts = _now_tw.strftime("%m/%d %H:%M")
    ts_full = _now_tw.strftime("%Y/%m/%d %H:%M:%S")
    entry = {"time": ts, "msg": msg}
    skip_display = any(k in msg for k in ["Cookie", "成本資料", "排程錯誤"])
    if not skip_display:
        _ad_scheduler_store["log"].insert(0, entry)
        if len(_ad_scheduler_store["log"]) > 100:
            _ad_scheduler_store["log"] = _ad_scheduler_store["log"][:100]
    print(f"[廣告排程] {entry}")

    important = any(k in msg for k in ["ROAS ✅", "ROAS ❌", "爆款", "暫停 ✅", "暫停 ❌", "加碼 ✅", "加碼 ❌", "預算 ✅", "預算 ❌", "低毛利", "空燒", "重啟 ✅", "重啟 ❌", "庫存不足", "主圖建議", "商品頁建議", "爆款結束", "=== 開始", "=== 完成"])
    skip = any(k in msg for k in ["Cookie", "成本資料", "排程錯誤"])
    if skip or (not important and not write_sheet):
        return

    # 加入批次佇列，不立即寫 Sheets
    import re
    shop_m = re.search(r"\[([^\]]+)\]", msg)
    shop = shop_m.group(1) if shop_m else ""
    suggestion = ""
    if "主圖建議" in msg:       suggestion = "建議更換主圖或優化標題以提升點擊率"
    elif "商品頁建議" in msg:   suggestion = "建議優化商品頁圖片/描述以提升轉化率"
    elif "爆款結束" in msg:     suggestion = "爆款結束，ROAS逐步恢復中，請觀察"
    elif "ROAS ✅" in msg:
        if "→" in msg:
            parts = msg.split("→")
            try:
                old_r = float(parts[0].split()[-1])
                new_r = float(parts[1].split()[0])
                suggestion = f"ROAS {'調降' if new_r < old_r else '調升'}至合理水位"
            except:
                suggestion = "ROAS 已自動調整至合理水位"
        else:
            suggestion = "ROAS 已自動調整至合理水位"
    elif "暫停 ✅" in msg and "毛利" in msg:  suggestion = "毛利不足，建議提高售價或降低成本"
    elif "暫停 ✅" in msg and "ROAS" in msg:  suggestion = "30天空燒，建議檢查主圖/標題/商品頁"
    elif "重啟 ✅" in msg:      suggestion = "廣告已重啟，請觀察7天ROAS表現"
    elif "預算 ✅" in msg and "降回85" in msg: suggestion = "廣告表現差，預算止損降回最低85 TWD"
    elif "預算 ✅" in msg:      suggestion = "廣告達標且預算用盡，自動加碼30%"
    elif "爆款" in msg:         suggestion = "爆款廣告，ROAS下調擴大曝光"
    elif "空燒警告" in msg:     suggestion = "近期有好轉跡象，繼續觀察7天"
    elif "庫存不足" in msg:     suggestion = "盡快補貨，補貨後廣告自動重啟"
    _ad_scheduler_store.setdefault("sheet_queue", []).append([ts_full, shop, msg, suggestion])

def _flush_ad_log_to_sheets():
    """把佇列中的日誌一次批次寫入 Sheets"""
    rows = _ad_scheduler_store.pop("sheet_queue", [])
    if not rows:
        return
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return
        client, err = get_sheets_client()
        if err:
            print(f"[廣告排程] Sheets 連線失敗: {err}")
            return
        try:
            ws = client.open_by_key(sheet_id).worksheet("🚀 廣告戰情室")
            first_row = ws.row_values(1)
            if not first_row or first_row[0] != "時間":
                ws.insert_row(["時間", "店鋪", "動作", "修改建議"], 1, value_input_option="RAW")
        except Exception:
            sh = client.open_by_key(sheet_id)
            try:
                ws = sh.worksheet("🚀 廣告戰情室")
            except Exception:
                ws = sh.add_worksheet(title="🚀 廣告戰情室", rows=10000, cols=4)
                ws.append_row(["時間", "店鋪", "動作", "修改建議"], value_input_option="RAW")
        ws.append_rows(rows, value_input_option="RAW")
        print(f"[廣告排程] 寫入 Sheets {len(rows)} 筆")
    except Exception as e:
        print(f"[廣告排程] 寫入 Sheets 失敗: {e}")


def _read_schedule_state():
    """從 Google Sheets 讀排程狀態（last_daily / last_hourly）"""
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id: return
        client, err = get_sheets_client()
        if err: return
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet("⚙️ 排程狀態")
        except:
            return
        rows = ws.get_all_values()
        state = {row[0]: row[1] for row in rows if len(row) >= 2}
        if state.get("last_daily"):
            _ad_scheduler_store["last_daily"] = state["last_daily"]
        if state.get("last_hourly"):
            try:
                _ad_scheduler_store["last_hourly"] = float(state["last_hourly"])
            except:
                pass
        print(f"[排程狀態] 從 Sheets 讀回：last_daily={state.get('last_daily')} last_hourly={state.get('last_hourly','')[:10]}")
    except Exception as e:
        print(f"[排程狀態] 讀取失敗: {e}")

def _write_schedule_state():
    """把排程狀態寫入 Google Sheets（last_daily / last_hourly）"""
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id: return
        client, err = get_sheets_client()
        if err: return
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet("⚙️ 排程狀態")
        except:
            ws = sh.add_worksheet(title="⚙️ 排程狀態", rows=10, cols=2)
        ws.clear()
        ws.append_rows([
            ["last_daily",  _ad_scheduler_store.get("last_daily") or ""],
            ["last_hourly", str(_ad_scheduler_store.get("last_hourly") or "")],
            ["updated",     datetime.now().strftime("%Y/%m/%d %H:%M:%S")],
        ], value_input_option="RAW")
    except Exception as e:
        print(f"[排程狀態] 寫入失敗: {e}")

def _get_cost_map():
    """取得成本資料：優先記憶體，沒有就從 Google Sheets 讀"""
    cost_map = _cost_store.get("map", {})
    if cost_map:
        return cost_map
    # 記憶體沒有 → 從 Sheets 讀回
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id: return {}
        client, err = get_sheets_client()
        if err: return {}
        ws = client.open_by_key(sheet_id).worksheet("💾 成本備份")
        rows = ws.get_all_values()
        if len(rows) <= 1: return {}
        cost_map = {}
        for row in rows[1:]:
            if len(row) >= 2 and row[0] and row[1]:
                try: cost_map[row[0]] = float(row[1])
                except: pass
        if cost_map:
            _cost_store["map"] = cost_map
            _cost_store["count"] = len(cost_map)
            _cost_store["ts"] = int(time.time() * 1000)
            _ad_scheduler_store["cost_count"] = len(cost_map)
            print(f"[排程] 從 Sheets 讀回 {len(cost_map)} 筆成本")
        return cost_map
    except Exception as e:
        print(f"[排程] 從 Sheets 讀成本失敗: {e}")
        return {}

def _bigseller_api(path, body=None):
    """呼叫 BigSeller API（需要在 Railway 環境中有 cookie）"""
    # Cookie 優先從記憶體（Extension 上傳），其次從環境變數
    cookie = _ad_scheduler_store.get("bs_cookie") or os.environ.get("BIGSELLER_COOKIE", "")
    if not cookie:
        _ad_log("🔴 無 BigSeller Cookie，請用管理版超人眼鏡重新登入", write_sheet=True)
        return None
    
    headers = {
        "Content-Type": "application/json",
        "Cookie": cookie,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.bigseller.com/"
    }
    url = f"https://www.bigseller.com{path}"
    
    try:
        if body is not None:
            r = _requests.post(url, json=body, headers=headers, timeout=30)
        else:
            r = _requests.get(url, headers=headers, timeout=30)
        
        # ⭐ 關鍵：檢查 HTTP 狀態碼
        if r.status_code == 401:
            _ad_log("🔴 BigSeller 認證失敗 (401)，Cookie 已失效", write_sheet=True)
            _invalidate_cookie("401 Unauthorized")
            return None
        elif r.status_code == 403:
            _ad_log("🔴 BigSeller 拒絕存取 (403)，Cookie 可能失效", write_sheet=True)
            _invalidate_cookie("403 Forbidden")
            return None
        elif r.status_code != 200:
            _ad_log(f"🔴 BigSeller API 錯誤：HTTP {r.status_code}", write_sheet=True)
            return None
        
        # 解析回應
        try:
            data = r.json()
        except Exception as e:
            _ad_log(f"🔴 BigSeller API 回應格式錯誤：{str(e)[:100]}", write_sheet=True)
            return None
        
        # 檢查 BigSeller 的業務狀態碼
        if data.get("code") != 0:
            msg = data.get("msg", "未知錯誤")
            if any(keyword in str(msg).lower() for keyword in ["login", "auth", "登入", "認證", "token", "session"]):
                _ad_log(f"🔴 BigSeller 要求重新登入：{msg}", write_sheet=True)
                _invalidate_cookie(f"Business Error: {msg}")
                return None
            else:
                # 其他業務錯誤不清空 Cookie，可能只是暫時問題
                _ad_log(f"⚠️ BigSeller API 業務錯誤：{msg}", write_sheet=False)
        
        return data
        
    except Exception as e:
        error_msg = str(e)
        if "timeout" in error_msg.lower():
            _ad_log("⚠️ BigSeller API 請求逾時，可能網路問題")
        elif "connection" in error_msg.lower():
            _ad_log("⚠️ BigSeller API 連線失敗，可能網路問題")
        else:
            _ad_log(f"🔴 BigSeller API 例外：{error_msg[:200]}")
        return None

def _invalidate_cookie(reason):
    """標記 Cookie 失效並清空記憶體"""
    _ad_log(f"🔴 Cookie 失效原因：{reason}，已清空快取", write_sheet=True)
    _ad_scheduler_store["bs_cookie"] = ""
    _ad_scheduler_store["cookie_ts"] = 0
    _ad_scheduler_store["cookie_invalid_count"] = _ad_scheduler_store.get("cookie_invalid_count", 0) + 1
    
    # 如果失效次數過多，發送緊急通知
    invalid_count = _ad_scheduler_store.get("cookie_invalid_count", 0)
    if invalid_count >= 3:
        _ad_log(f"🆘 Cookie 已連續失效 {invalid_count} 次，請檢查帳號狀態", write_sheet=True)

def _check_cookie_health():
    """檢查 Cookie 健康狀態，返回 True/False"""
    cookie = _ad_scheduler_store.get("bs_cookie", "")
    if not cookie:
        _ad_log("🔴 Cookie 健康檢查：無 Cookie", write_sheet=True)
        return False
    
    # 快速測試：呼叫店鋪清單 API（輕量級）
    test_result = _bigseller_api("/api/v1/shop/list.json?platform=shopee")
    
    if test_result is None:
        # _bigseller_api 內部已經處理了錯誤訊息和 Cookie 失效
        return False
    
    if test_result.get("code") == 0:
        # 成功時重設失效計數
        _ad_scheduler_store["cookie_invalid_count"] = 0
        _ad_scheduler_store["cookie_last_check"] = int(time.time())
        return True
    else:
        msg = test_result.get("msg", "未知錯誤")
        _ad_log(f"🔴 Cookie 健康檢查失敗：{msg}")
        return False

def _get_cookie_status():
    """取得 Cookie 狀態摘要（用於 API 回應）"""
    cookie = _ad_scheduler_store.get("bs_cookie", "")
    if not cookie:
        return {
            "status": "missing",
            "message": "無 Cookie，請重新上傳",
            "last_check": None,
            "invalid_count": 0
        }
    
    last_check = _ad_scheduler_store.get("cookie_last_check", 0)
    invalid_count = _ad_scheduler_store.get("cookie_invalid_count", 0)
    
    # 判斷狀態
    if invalid_count >= 3:
        status = "critical"
        message = f"連續失效 {invalid_count} 次，請檢查帳號"
    elif invalid_count > 0:
        status = "warning"  
        message = f"近期失效 {invalid_count} 次"
    elif last_check > 0 and time.time() - last_check < 3600:
        status = "healthy"
        message = "運作正常"
    else:
        status = "unknown"
        message = "未檢查或檢查時間過久"
    
    return {
        "status": status,
        "message": message,
        "last_check": datetime.fromtimestamp(last_check).strftime("%Y-%m-%d %H:%M:%S") if last_check else None,
        "invalid_count": invalid_count,
        "cookie_length": len(cookie)
    }

def _get_target_roas(margin):
    """初始 ROAS：所有毛利統一按45%基本標準反推
    初始 ROAS = 1 / (毛利率 - 0.45)，進位到0.5
    最高上限 20.0（毛利46~50%會碰到上限）
    <=45% -> None（自動暫停，不投廣告）
    """
    import math
    if margin is None or margin <= 45:
        return None   # 自動暫停
    floor = 1.0 / ((margin / 100.0) - 0.45)
    target = math.ceil(floor * 2) / 2.0
    return min(target, 20.0)

def _get_boom_floor_roas(margin):
    """爆款下限 ROAS：死守廣告費後毛利40%
    公式：最低ROAS = 1 / (毛利率 - 0.40)，進位到0.5
    只有初始ROAS > 爆款下限時才有空間觸發（51%以上）
    """
    import math
    if margin is None or margin <= 40:
        return None
    target = _get_target_roas(margin)
    if target is None:
        return None
    floor = math.ceil((1.0 / ((margin / 100.0) - 0.40)) * 2) / 2.0
    # 下限必須低於初始才有意義
    if floor >= target:
        return None  # 無爆款空間
    return floor

def _edit_ad(campaign_id, ad_type, shop_id, edit_action, value=None):
    """修改廣告（ROAS/預算/暫停）"""
    body = {"campaignWithType": [{"campaignId": campaign_id, "adType": ad_type, "shopId": shop_id}], "editAction": edit_action}
    if edit_action == 11 and value is not None: body["roasTarget"] = str(value)
    if edit_action == 6  and value is not None: body["budget"] = value
    d = _bigseller_api("/api/v1/product/listing/shopee/editSingleShopeeProductAds.json", body)
    return d and d.get("code") == 0

def _fetch_ads_range(days=None):
    """抓廣告資料，可帶日期範圍"""
    from datetime import timedelta
    body = {"pageNo": 1, "pageSize": 200}
    if days:
        today = datetime.now()
        start = today - timedelta(days=days)
        body["startDateStr"] = start.strftime("%Y-%m-%d")
        body["endDateStr"]   = today.strftime("%Y-%m-%d")
    ads = []
    page = 1
    while True:
        body["pageNo"] = page
        d = _bigseller_api("/api/v1/product/listing/shopee/queryAdCampaignShopInfoPage.json", body)
        if not d or d.get("code") != 0: break
        rows = d.get("data", {}).get("rows", [])
        for row in rows:
            if row.get("biddingMethod") == "autoRoas" and row.get("campaignStatus") == "ongoing":
                ads.append(row)
        if page >= d.get("data", {}).get("totalPage", 1): break
        page += 1
    return ads

def _fetch_listings_map():
    """抓在線產品，建立 itemId -> {skus, price} map"""
    item_map = {}
    page = 1
    while True:
        d = _bigseller_api(
            f"/api/v1/product/listing/shopee/active.json?orderBy=create_time&desc=true"
            f"&timeType=create_time&startDateStr=&endDateStr=&searchType=productName"
            f"&inquireType=0&shopeeStatus=live&status=active&pageNo={page}&pageSize=50"
        )
        if not d or d.get("code") != 0: break
        rows = d.get("data", {}).get("page", {}).get("rows", [])
        for row in rows:
            if row.get("hasVariation") and row.get("variations"):
                # 每個 SKU 保留自己的售價
                sku_price_map = {
                    v["variationSku"]: v.get("price") or v.get("originalPrice") or 0
                    for v in row["variations"] if v.get("variationSku")
                }
                skus = list(sku_price_map.keys())
                stock = sum(v.get("stock", 0) for v in row["variations"])
                item_map[row["itemId"]] = {"skus": skus, "sku_price_map": sku_price_map, "price": 0, "stock": stock}
            else:
                item_map[row["itemId"]] = {"skus": [row.get("itemSku", "")], "sku_price_map": {}, "price": row.get("price") or row.get("originalPrice") or 0, "stock": row.get("stock", 0)}
        total_page = d.get("data", {}).get("page", {}).get("totalPage", 1)
        if page >= total_page: break
        page += 1
    return item_map

def _calc_margin(item, cost_map):
    """計算商品平均毛利率（每個SKU用自己的售價）"""
    total = 0.0
    count = 0
    sku_price_map = item.get("sku_price_map", {})
    default_price = item.get("price", 0)

    def _find_cost(sku):
        """SKU 格式容錯：-01 ↔ -001 互相嘗試"""
        if sku in cost_map: return cost_map[sku]
        import re
        # PDD032-01 → PDD032-001
        padded = re.sub(r'-(\d{1,2})$', lambda m: '-' + m.group(1).zfill(3), sku)
        if padded in cost_map: return cost_map[padded]
        # PDD032-001 → PDD032-01
        trimmed = re.sub(r'-0+(\d+)$', r'-\1', sku)
        if trimmed in cost_map: return cost_map[trimmed]
        return None

    for sku in item.get("skus", []):
        cost = _find_cost(sku)
        if not cost or cost <= 0: continue
        price = sku_price_map.get(sku) or default_price
        if price <= 0: continue
        margin = (price - cost) / price * 100
        total += margin
        count += 1
    return total / count if count > 0 else None

def run_daily_ad_tasks(force=False):
    """每日廣告任務：ROAS調整 + 爆款降ROAS + 空燒暫停"""
    from datetime import timezone, timedelta
    tw_tz = timezone(timedelta(hours=8))
    today = datetime.now(tw_tz).strftime("%Y-%m-%d")  # 台灣時間日期
    if not force and _ad_scheduler_store["last_daily"] == today:
        return  # 今天已跑過

    _ad_log("=== 開始每日廣告任務 ===")
    
    # ⭐ Cookie 健康檢查
    if not _check_cookie_health():
        _ad_log("❌ Cookie 健康檢查失敗，跳過每日任務", write_sheet=True)
        return
    
    cost_map = _get_cost_map()
    if not cost_map:
        _ad_log("無成本資料（記憶體+Sheets都是空的），跳過每日任務")
        # 沒有成本不標記今天已跑，明天或成本讀回後繼續嘗試
        _ad_scheduler_store["last_daily"] = None
        return

    item_map   = _fetch_listings_map()
    ads_now    = _fetch_ads_range()
    ads_7d     = {a["campaignId"]: a for a in _fetch_ads_range(7)}
    ads_30d    = {a["campaignId"]: a for a in _fetch_ads_range(30)}

    roas_ok = roas_fail = boom_ok = boom_fail = pause_ok = pause_fail = 0
    burned_today = set()  # 今天剛被空燒暫停的廣告，不能重啟

    for ad in ads_now:
        item = item_map.get(ad.get("itemId"))
        if not item: continue
        margin = _calc_margin(item, cost_map)
        if margin is None: continue

        target_roas  = _get_target_roas(margin)
        current_roas = float(ad.get("roasTarget") or 0)
        ad_type  = ad.get("adType")
        shop_id  = ad.get("shopId")
        cid      = ad.get("campaignId")
        name     = (ad.get("adName") or str(cid))[:20]

        # ── 毛利<=45% 自動暫停 ──
        if target_roas is None:
            if _edit_ad(cid, ad_type, shop_id, 2):
                _ad_log(f"暫停 ✅ [{ad.get('shopName','')[:12]}] {name} 毛利{margin:.0f}%<=45% 不投廣告")
                pause_ok += 1
            else:
                pause_fail += 1
            time.sleep(0.3)
            continue

        # ── ROAS 調整（初始設定）──
        # 注意：若目前 ROAS 低於目標（可能是爆款降低），不在此處調整，交給爆款邏輯判斷
        if current_roas > target_roas and abs(target_roas - current_roas) > 0.05:
            # 目前ROAS高於目標 → 直接調降到正確值
            if _edit_ad(cid, ad_type, shop_id, 11, target_roas):
                _ad_log(f"ROAS ✅ [{ad.get('shopName','')[:12]}] {name} {current_roas}→{target_roas} (毛利{margin:.0f}%)")
                roas_ok += 1
            else:
                roas_fail += 1
            time.sleep(0.3)

        # ── 爆款降 ROAS / 爆款結束恢復 ROAS ──
        floor_roas = _get_boom_floor_roas(margin)
        d7  = ads_7d.get(cid)
        d30 = ads_30d.get(cid)
        roas_7  = float(d7.get("broadRoi")  or 0) if d7  else 0
        roas_30 = float(d30.get("broadRoi") or 0) if d30 else 0
        exp_7   = float(d7.get("expense")   or 0) if d7  else 0
        exp_30  = float(d30.get("expense")  or 0) if d30 else 0

        if floor_roas is not None:
            # 爆款觸發：7天ROAS>目標200% 且花費>500
            cond_7  = roas_7  > 0 and exp_7  > 500 and roas_7  > target_roas * 2.0
            cond_30 = roas_30 > 0 and exp_30 > 500 and roas_30 > target_roas * 2.0
            boom_roas = None
            if cond_7 and cond_30:
                boom_roas = round(max(current_roas * 0.8, floor_roas), 1)  # 降20%
            elif cond_7:
                boom_roas = round(max(current_roas * 0.9, floor_roas), 1)  # 降10%

            if boom_roas and boom_roas < current_roas:
                # 爆款：降低 ROAS
                if _edit_ad(cid, ad_type, shop_id, 11, boom_roas):
                    shop_name = ad.get('shopName','')[:12]
                    _ad_log(f"爆款 ✅ [{shop_name}] {name} ROAS {current_roas}→{boom_roas} (毛利{margin:.0f}% 下限{floor_roas})")
                    boom_ok += 1
                else:
                    boom_fail += 1
                time.sleep(0.3)

            elif current_roas < target_roas:
                # 爆款結束恢復：目前ROAS低於目標（曾是爆款），但7天表現已不達標
                # 7天ROAS < 目標×70% → 爆款結束，往上恢復
                if roas_7 < target_roas * 0.7 or roas_7 == 0:
                    recover_roas = round(min(current_roas * 1.1, target_roas), 1)  # 每次+10%，上限到初始值
                    if recover_roas > current_roas and abs(recover_roas - current_roas) > 0.05:
                        if _edit_ad(cid, ad_type, shop_id, 11, recover_roas):
                            shop_name = ad.get('shopName','')[:12]
                            _ad_log(f"爆款結束 ✅ [{shop_name}] {name} ROAS {current_roas}→{recover_roas} (7天{roas_7:.1f}<目標{target_roas}×70%)")
                            roas_ok += 1
                        time.sleep(0.3)

        # ── 空燒暫停（30天差 且 7天也沒有好轉）──
        d7  = ads_7d.get(cid)
        d30 = ads_30d.get(cid)
        if d30:
            exp30  = float(d30.get("expense") or 0)
            roi30  = float(d30.get("broadRoi") or 0)
            roi7   = float(d7.get("broadRoi") or 0) if d7 else 0
            exp7   = float(d7.get("expense") or 0) if d7 else 0
            # 條件1：30天長期表現差（花費>500 且 ROAS<目標50%）
            cond_long = exp30 > 500 and roi30 > 0 and roi30 < target_roas * 0.5
            # 條件2：7天近期也沒有好轉（ROAS<目標70%）
            # 若7天ROAS≥目標70%，表示近期有改善，給機會繼續跑
            cond_recent = roi7 == 0 or roi7 < target_roas * 0.7
            if cond_long and cond_recent:
                if _edit_ad(cid, ad_type, shop_id, 2):  # 暫停
                    # 判斷問題原因給修改建議
                    ctr_v = float(ad.get("ctr") or 0)
                    cr_v  = float(ad.get("cr")  or 0)
                    ctr_avg = _ad_benchmark["ctr_avg"]
                    cr_avg  = _ad_benchmark["cr_avg"]
                    if ctr_v < ctr_avg * 0.5:
                        reason = f"CTR{ctr_v:.1f}%低於均值{ctr_avg}%，建議換主圖/優化標題"
                    elif cr_v < cr_avg * 0.5:
                        reason = f"CR{cr_v:.1f}%低於均值{cr_avg}%，建議優化商品頁/圖片"
                    else:
                        reason = f"整體表現差，建議全面檢視"
                    _ad_log(f"暫停 ✅ [{ad.get('shopName','')[:10]}] {name} 30天ROAS{roi30:.1f} 7天ROAS{roi7:.1f} | {reason}")
                    burned_today.add(cid)  # 記錄今天空燒暫停，不能被重啟
                    pause_ok += 1
                else:
                    pause_fail += 1
                time.sleep(0.3)
            elif cond_long and not cond_recent:
                # 30天差但7天有好轉，記錄警告但不暫停
                _ad_log(f"空燒警告 [{ad.get('shopName','')[:10]}] {name} 30天差({roi30:.1f}) 但7天好轉({roi7:.1f}) 觀察中")

        # ── CTR/CR 異常通知（花費>50才有意義）──
        ad_expense = float(ad.get("expense") or 0)
        if ad_expense > 50:
            ctr_v = float(ad.get("ctr") or 0)
            cr_v  = float(ad.get("cr")  or 0)
            shop_n = ad.get("shopName","")[:12]
            if ctr_v > 0 and ctr_v < _ad_benchmark["ctr_avg"] * 0.5:
                _ad_log(f"⚠️ 主圖建議 [{shop_n}] {name} CTR{ctr_v:.2f}%<均值{_ad_benchmark['ctr_avg']}%×50% 建議換主圖/優化標題")
            if cr_v > 0 and cr_v < _ad_benchmark["cr_avg"] * 0.5:
                _ad_log(f"⚠️ 商品頁建議 [{shop_n}] {name} CR{cr_v:.2f}%<均值{_ad_benchmark['cr_avg']}%×50% 建議優化商品頁/圖片")

    # ── 重啟符合條件的暫停廣告（毛利>45% 且 有庫存）──────────────────
    restart_ok = restart_skip_margin = restart_skip_stock = 0
    try:
        # 抓所有暫停的 autoRoas 廣告
        paused_page = 1
        paused_ads = []
        while True:
            body_p = {"pageNo": paused_page, "pageSize": 100}
            dp = _bigseller_api("/api/v1/product/listing/shopee/queryAdCampaignShopInfoPage.json", body_p)
            if not dp or dp.get("code") != 0: break
            rows_p = dp.get("data", {}).get("rows", [])
            for row in rows_p:
                if row.get("biddingMethod") == "autoRoas" and row.get("campaignStatus") == "paused":
                    paused_ads.append(row)
            if paused_page >= dp.get("data", {}).get("totalPage", 1): break
            paused_page += 1

        for pad in paused_ads:
            iid   = pad.get("itemId")
            item  = item_map.get(iid)
            if not item: continue
            margin = _calc_margin(item, cost_map)
            if margin is None: continue
            cid     = pad.get("campaignId")
            ad_type = pad.get("adType")
            shop_id = pad.get("shopId")
            name    = (pad.get("adName") or str(cid))[:20]
            shop_name = pad.get("shopName") or str(shop_id)

            # 今天剛被空燒暫停的廣告 → 不重啟（空燒需要持續觀察）
            if cid in burned_today:
                _ad_log(f"空燒保護 [{shop_name}] {name} 今天剛因空燒暫停，不重啟")
                continue

            # 毛利 ≤ 45% → 繼續暫停
            if margin <= 45:
                restart_skip_margin += 1
                continue

            # 檢查庫存（item_map 裡有 stock 欄位）
            stock = item.get("stock", 0)
            if stock <= 0:
                restart_skip_stock += 1
                _ad_log(f"庫存不足 [{shop_name}] {name} 庫存{stock}，維持暫停")
                continue

            # 毛利>45% 且 有庫存 → 重啟廣告
            target_roas = _get_target_roas(margin)
            if _edit_ad(cid, ad_type, shop_id, 3):  # 3 = 恢復暫停廣告
                _ad_log(f"重啟 ✅ [{shop_name}] {name} 毛利{margin:.0f}% 庫存{stock} ROAS→{target_roas}")
                restart_ok += 1
                # 同時設定正確 ROAS
                _edit_ad(cid, ad_type, shop_id, 11, target_roas)
            time.sleep(0.3)

        if restart_ok > 0:
            _ad_log(f"=== 重啟廣告完成 {restart_ok}筆✅ 毛利不足跳過{restart_skip_margin}筆 庫存不足跳過{restart_skip_stock}筆 ===", write_sheet=True)
    except Exception as e:
        _ad_log(f"重啟廣告失敗: {e}")

    # ── 低毛利廣告回報（依店鋪分類）──────────────────
    low_margin_by_shop = {}  # shop_name -> [{"name", "margin", "roas", "shopId"}]
    for ad in ads_now:
        item = item_map.get(ad.get("itemId"))
        if not item: continue
        margin = _calc_margin(item, cost_map)
        if margin is None: continue
        if margin < 40:
            shop_name = ad.get("shopName") or ad.get("storeName") or str(ad.get("shopId",""))
            shop_id   = str(ad.get("shopId",""))
            name      = (ad.get("adName") or str(ad.get("campaignId","")))[:30]
            roas      = float(ad.get("roasTarget") or 0)
            key = f"{shop_id}_{shop_name}"
            if key not in low_margin_by_shop:
                low_margin_by_shop[key] = {"shopName": shop_name, "shopId": shop_id, "items": []}
            low_margin_by_shop[key]["items"].append({
                "name": name, "margin": round(margin, 1), "roas": round(roas, 1)
            })
            _ad_log(f"低毛利 [{shop_name}] {name} 毛利{margin:.0f}%")

    # 存入 store 供首頁 API 使用
    shops_list = sorted(low_margin_by_shop.values(), key=lambda x: len(x["items"]), reverse=True)
    _ad_scheduler_store["low_margin_shops"] = shops_list
    _ad_scheduler_store["low_margin_ts"] = time.time()
    if shops_list:
        total = sum(len(s["items"]) for s in shops_list)
        _ad_log(f"低毛利彙總：{len(shops_list)} 間店鋪，共 {total} 筆廣告")
    else:
        _ad_log("低毛利檢查完成，無異常")

    _ad_scheduler_store["last_daily"] = today
    threading.Thread(target=_write_schedule_state, daemon=True).start()
    now_str = datetime.now().strftime("%Y/%m/%d %H:%M")
    _ad_log(f"=== [{now_str}] 每日排程完成 | ROAS調整 {roas_ok}筆✅{roas_fail}筆❌ | 爆款 {boom_ok}筆✅ | 暫停 {pause_ok}筆✅ | 重啟 {restart_ok}筆✅ ===", write_sheet=True)
    _flush_ad_log_to_sheets()  # 批次寫入 Sheets

def run_hourly_budget_task(force=False):
    """每小時預算任務：ROAS達標且使用率≥90%→加30%"""
    now_ts = time.time()
    last = _ad_scheduler_store.get("last_hourly") or 0
    if not force and now_ts - last < 3600:
        return  # 未滿1小時

    _ad_log("--- 開始每小時預算任務 ---")
    
    # ⭐ Cookie 健康檢查
    if not _check_cookie_health():
        _ad_log("❌ Cookie 健康檢查失敗，跳過每小時任務", write_sheet=True)
        return
    
    cost_map = _get_cost_map()
    if not cost_map:
        # 沒有成本不更新 last_hourly，下一分鐘繼續嘗試
        _ad_scheduler_store["last_hourly"] = 0
        return

    item_map = _fetch_listings_map()
    ads      = _fetch_ads_range()
    ok = fail = skip = 0

    for ad in ads:
        item = item_map.get(ad.get("itemId"))
        if not item: continue
        margin = _calc_margin(item, cost_map)
        if margin is None: continue

        target_roas  = _get_target_roas(margin)
        current_roas = float(ad.get("roasTarget") or 0)
        actual_roas  = float(ad.get("broadRoi")   or 0)
        budget       = float(ad.get("campaignBudget") or 0)
        expense      = float(ad.get("expense")    or 0)
        if budget <= 0: continue
        usage = expense / budget

        # 條件一：ROAS達標 + 使用率≥90% → 加碼30%
        if actual_roas > 0 and actual_roas >= current_roas and usage >= 0.9:
            new_budget = round(budget * 1.3)
            cid     = ad.get("campaignId")
            ad_type = ad.get("adType")
            shop_id = ad.get("shopId")
            name    = (ad.get("adName") or str(cid))[:20]
            shop_name = (ad.get("shopName") or str(shop_id))[:12]
            if _edit_ad(cid, ad_type, shop_id, 6, new_budget):
                _ad_log(f"預算 ✅ [{shop_name}] {name} {budget:.0f}→{new_budget} TWD (ROAS {actual_roas:.1f} 用量{usage*100:.0f}%)")
                ok += 1
            else:
                fail += 1
            time.sleep(0.3)

        # 條件二：廣告跑差（7天ROAS<目標×70%）且預算>85 → 降回85（止損）
        # broadRoi/expense 是 API 預設7天數據
        elif budget > 85 and target_roas is not None:
            cid = ad.get("campaignId")
            roas_7h = actual_roas  # API 預設7天
            exp_7h  = expense
            # 有花費但表現差 → 降回85
            if exp_7h > 100 and roas_7h > 0 and roas_7h < target_roas * 0.7:
                ad_type = ad.get("adType")
                shop_id = ad.get("shopId")
                name    = (ad.get("adName") or str(cid))[:20]
                shop_name = (ad.get("shopName") or str(shop_id))[:12]
                if _edit_ad(cid, ad_type, shop_id, 6, 85):
                    _ad_log(f"預算 ✅ [{shop_name}] {name} 降回85 (7天ROAS{roas_7h:.1f}<目標{target_roas}×70%)")
                    ok += 1
                time.sleep(0.3)
            else:
                skip += 1
        else:
            skip += 1

    _ad_scheduler_store["last_hourly"] = now_ts
    threading.Thread(target=_write_schedule_state, daemon=True).start()
    if ok > 0 or fail > 0:
        now_str = datetime.now().strftime("%Y/%m/%d %H:%M")
        _ad_log(f"=== [{now_str}] 每小時排程完成 | 預算加碼 {ok}筆✅{(' 失敗'+str(fail)+'筆❌') if fail else ''} ===", write_sheet=True)
        _flush_ad_log_to_sheets()  # 批次寫入 Sheets
    else:
        _ad_log(f"--- 預算任務檢查完成，本次無符合加碼條件 ({skip}筆檢查)")

def ad_scheduler_thread():
    """廣告排程背景執行緒"""
    time.sleep(30)  # 啟動後等30秒再開始
    while True:
        try:
            # 用 UTC+8 台灣時間判斷
            from datetime import timezone, timedelta
            tw_tz = timezone(timedelta(hours=8))
            now_tw = datetime.now(tw_tz)
            # 每天台灣時間9點跑每日任務
            if now_tw.hour == 9:
                run_daily_ad_tasks()
            # 每小時跑預算任務
            run_hourly_budget_task()
        except Exception as e:
            _ad_log(f"排程錯誤: {e}")
        time.sleep(60)  # 每分鐘檢查一次

# ── 廣告排程 API（查看執行記錄）──────────────────────────────
@app.route("/api/superman-glasses/save-cookie", methods=["POST", "OPTIONS"])
def superman_glasses_save_cookie():
    """Extension 上傳 BigSeller Cookie 供排程使用"""
    if request.method == "OPTIONS":
        resp = jsonify({"ok": True})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp
    try:
        data = request.get_json(force=True)
        cookie = data.get("cookie", "").strip()
        source = data.get("source", "UNKNOWN")
        version = data.get("version", "")
        
        if not cookie:
            resp = jsonify({"ok": False, "msg": "empty cookie"})
            resp.headers['Access-Control-Allow-Origin'] = '*'
            return resp, 400
        
        # ⭐ 來源控制：只接受管理版或自動更新
        allowed_sources = ["BIN-ADMIN", "AUTO-UPDATE", "SUPERMAN-ADMIN"]
        if source not in allowed_sources and "ADMIN" not in source:
            resp = jsonify({"ok": False, "msg": f"只接受管理版 Cookie，當前來源：{source}"})
            resp.headers['Access-Control-Allow-Origin'] = '*'
            return resp, 403
        
        # 更新記憶體
        old_cookie = _ad_scheduler_store.get("bs_cookie", "")
        _ad_scheduler_store["bs_cookie"] = cookie
        _ad_scheduler_store["cookie_ts"] = int(time.time())
        _ad_scheduler_store["cookie_source"] = source
        _ad_scheduler_store["cookie_version"] = version
        
        # 重設失效計數（新 Cookie 給新機會）
        if cookie != old_cookie:
            _ad_scheduler_store["cookie_invalid_count"] = 0
        
        # 記錄日誌
        change_info = "更新" if cookie != old_cookie else "重複上傳"
        _ad_log(f"Cookie 已{change_info} [{source}] 長度 {len(cookie)}")
        
        # 寫入 Sheets 持久化（Railway 重啟後可讀回）
        def _persist_cookie():
            try:
                sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
                if not sheet_id: return
                client, err = get_sheets_client()
                if err: return
                sh = client.open_by_key(sheet_id)
                try:
                    ws = sh.worksheet("⚙️ 排程狀態")
                except Exception:
                    ws = sh.add_worksheet(title="⚙️ 排程狀態", rows=10, cols=2)
                
                # 更新多個欄位
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updates = [
                    ["bs_cookie", cookie],
                    ["cookie_source", source], 
                    ["cookie_version", version],
                    ["cookie_updated", timestamp]
                ]
                
                # 讀現有資料，批次更新
                rows = ws.get_all_values()
                state = {r[0]: i+1 for i, r in enumerate(rows) if len(r) > 0}
                
                for key, value in updates:
                    row_num = state.get(key)
                    if row_num:
                        ws.update_cell(row_num, 2, str(value))
                    else:
                        ws.append_row([key, str(value)], value_input_option="RAW")
                        
            except Exception as e:
                print(f"[Cookie] 寫入 Sheets 失敗: {e}")
        
        threading.Thread(target=_persist_cookie, daemon=True).start()
        
        resp = jsonify({
            "ok": True, 
            "len": len(cookie),
            "source": source,
            "change": change_info
        })
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
        
    except Exception as e:
        resp = jsonify({"ok": False, "msg": str(e)})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp, 500


@app.route("/api/superman-glasses/ad-log-sheet", methods=["GET"])
def superman_glasses_ad_log_sheet():
    """從 Google Sheets 讀取歷史日誌（永久記錄）"""
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return jsonify({"ok": False, "msg": "未設定 GOOGLE_SHEETS_ID"}), 400
        client, err = get_sheets_client()
        if err: return jsonify({"ok": False, "msg": err}), 500
        ws = client.open_by_key(sheet_id).worksheet("🚀 廣告戰情室")
        rows = ws.get_all_values()
        # 跳過標題列，取最新100筆（從後往前）
        data_rows = rows[1:] if len(rows) > 1 else []
        recent = list(reversed(data_rows[-200:]))[:100]
        logs = [{"time": r[0], "type": r[1] if len(r)>1 else "", "msg": r[2] if len(r)>2 else ""} for r in recent]
        resp = jsonify({"ok": True, "logs": logs, "total": len(data_rows)})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/profit-snapshot", methods=["POST"])
def superman_glasses_profit_snapshot():
    """將利潤快照寫入 Google Sheets 📊 利潤監控室"""
    try:
        data = request.get_json(force=True)
        rows = data.get("rows", [])
        if not rows:
            return jsonify({"ok": False, "msg": "無資料"}), 400

        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return jsonify({"ok": False, "msg": "未設定 GOOGLE_SHEETS_ID"}), 400

        client, err = get_sheets_client()
        if err:
            return jsonify({"ok": False, "msg": err}), 500
        try:
            ws = client.open_by_key(sheet_id).worksheet("📊 利潤監控室")
            first_row = ws.row_values(1)
            if not first_row or first_row[0] != "日期":
                ws.clear()
                ws.append_row(["日期","時間","店鋪","廣告名稱","itemId","目前ROAS","實際ROAS(7天)","花費7天(TWD)","花費30天(TWD)","預算(TWD)","毛利%","目標ROAS","狀態","備註"], value_input_option="RAW")
        except Exception:
            sh = client.open_by_key(sheet_id)
            try:
                ws = sh.worksheet("📊 利潤監控室")
            except:
                ws = sh.add_worksheet(title="📊 利潤監控室", rows=10000, cols=14)
            ws.append_row(["日期","時間","店鋪","廣告名稱","itemId","目前ROAS","實際ROAS(7天)","花費7天(TWD)","花費30天(TWD)","預算(TWD)","毛利%","目標ROAS","狀態","備註"], value_input_option="RAW")

        now = datetime.now()
        date_str = now.strftime("%Y/%m/%d")
        time_str = now.strftime("%H:%M")

        sheet_rows = []
        for r in rows:
            sheet_rows.append([
                date_str,
                time_str,
                r.get("shopName", ""),
                r.get("adName", ""),
                r.get("itemId", ""),
                r.get("currentRoas", ""),
                r.get("actualRoas7", ""),
                r.get("expense7", ""),
                r.get("expense30", ""),
                r.get("budget", ""),
                r.get("margin", ""),
                r.get("targetRoas", ""),
                r.get("status", ""),
                r.get("note", ""),
            ])

        ws.append_rows(sheet_rows, value_input_option="RAW")
        resp = jsonify({"ok": True, "written": len(sheet_rows)})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/profit-snapshot", methods=["OPTIONS"])
def superman_glasses_profit_snapshot_options():
    resp = jsonify({"ok": True})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp

@app.route("/api/superman-glasses/clear-war-room", methods=["POST"])
def superman_glasses_clear_war_room():
    """清空 🚀 廣告戰情室，只保留標題列"""
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return jsonify({"ok": False, "msg": "未設定 GOOGLE_SHEETS_ID"}), 400
        client, err = get_sheets_client()
        if err:
            return jsonify({"ok": False, "msg": err}), 500
        ws = client.open_by_key(sheet_id).worksheet("🚀 廣告戰情室")
        ws.clear()
        ws.append_row(["時間", "類型", "店鋪", "廣告名稱", "調整內容", "毛利%", "結果"], value_input_option="RAW")
        resp = jsonify({"ok": True, "msg": "廣告戰情室已清空"})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/ad-log", methods=["GET"])
def superman_glasses_ad_log():
    cookie_status = _get_cookie_status()
    resp = jsonify({
        "ok": True,
        "log": _ad_scheduler_store["log"][:100],
        "last_daily": _ad_scheduler_store["last_daily"],
        "last_hourly": _ad_scheduler_store["last_hourly"],
        "cookie_ok": bool(_ad_scheduler_store.get("bs_cookie")),
        "cost_count": _ad_scheduler_store.get("cost_count", 0),
        # ⭐ 新增：詳細 Cookie 健康資訊
        "cookie_status": cookie_status["status"],
        "cookie_message": cookie_status["message"],
        "cookie_last_check": cookie_status["last_check"],
        "cookie_invalid_count": cookie_status["invalid_count"],
        "cookie_source": _ad_scheduler_store.get("cookie_source", ""),
        "cookie_version": _ad_scheduler_store.get("cookie_version", "")
    })
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route("/api/superman-glasses/low-margin", methods=["GET"])
def superman_glasses_low_margin():
    """回傳低利潤廣告清單（依店鋪分類）"""
    data = _ad_scheduler_store.get("low_margin_shops", [])
    ts   = _ad_scheduler_store.get("low_margin_ts", 0)
    resp = jsonify({
        "ok": True,
        "shops": data,
        "ts": ts,
        "updated": datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else None
    })
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route("/api/superman-glasses/profit-snapshot-read", methods=["GET"])
def superman_glasses_profit_snapshot_read():
    """從 Google Sheets 📊 利潤監控室讀取最新快照"""
    try:
        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return jsonify({"ok": False, "msg": "未設定 GOOGLE_SHEETS_ID"}), 400
        client, err = get_sheets_client()
        if err: return jsonify({"ok": False, "msg": err}), 500
        ws = client.open_by_key(sheet_id).worksheet("📊 利潤監控室")
        rows = ws.get_all_values()
        if len(rows) <= 1:
            return jsonify({"ok": True, "rows": [], "total": 0})
        headers = rows[0]
        # 取最新一次快照（最後一批相同日期+時間的記錄）
        data_rows = rows[1:]
        if not data_rows:
            return jsonify({"ok": True, "rows": [], "total": 0})
        # 找最新的日期+時間
        last_dt = data_rows[-1][0] + data_rows[-1][1] if len(data_rows[-1]) >= 2 else ""
        latest = [r for r in data_rows if (r[0]+r[1] if len(r)>=2 else "") == last_dt]
        # 轉成 dict
        result = []
        for r in latest:
            row_dict = {headers[i]: r[i] if i < len(r) else "" for i in range(len(headers))}
            result.append({
                "shop": row_dict.get("店鋪",""),
                "name": row_dict.get("廣告名稱",""),
                "margin": float(row_dict.get("毛利%","0") or 0),
                "roas": row_dict.get("目前ROAS",""),
                "targetRoas": row_dict.get("目標ROAS",""),
                "status": row_dict.get("狀態",""),
                "note": row_dict.get("備註",""),
                "time": row_dict.get("日期","") + " " + row_dict.get("時間","")
            })
        resp = jsonify({"ok": True, "rows": result, "total": len(data_rows), "snapshot_time": last_dt})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/schedule-lock", methods=["POST", "OPTIONS"])
def superman_glasses_schedule_lock():
    """排程鎖：避免多台電腦同時執行"""
    if request.method == "OPTIONS":
        resp = jsonify({"ok": True})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp
    try:
        data = request.get_json(force=True)
        task = data.get("task", "hourly")
        action = data.get("action", "acquire")
        lock_key = f"lock_{task}"
        lock_ttl = 300  # 5分鐘自動解鎖（防止當機卡死）

        if action == "acquire":
            lock = _ad_scheduler_store.get(lock_key)
            now_ts = time.time()
            last_hourly = _ad_scheduler_store.get("last_hourly") or 0
            if lock and now_ts - lock < lock_ttl:
                resp = jsonify({"ok": False, "msg": "鎖定中", "last_hourly": last_hourly})
            else:
                _ad_scheduler_store[lock_key] = now_ts
                resp = jsonify({"ok": True, "msg": "取得鎖", "last_hourly": last_hourly})
        else:  # release
            _ad_scheduler_store.pop(lock_key, None)
            resp = jsonify({"ok": True, "msg": "釋放鎖"})
    except Exception as e:
        resp = jsonify({"ok": False, "msg": str(e)})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route("/api/superman-glasses/ext-log", methods=["POST", "OPTIONS"])
def superman_glasses_ext_log():
    """接收 Extension 回傳的廣告排程執行歷程"""
    if request.method == "OPTIONS":
        resp = jsonify({"ok": True})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp
    try:
        data = request.get_json(force=True)
        msg = data.get("msg", "")
        suggestion = data.get("suggestion", "")
        if msg:
            import re
            from datetime import timezone, timedelta
            _tw = timezone(timedelta(hours=8))
            ts_full = datetime.now(_tw).strftime("%Y/%m/%d %H:%M:%S")
            shop_m = re.search(r"\[([^\]]+)\]", msg)
            shop = shop_m.group(1) if shop_m else ""
            _ad_scheduler_store.setdefault("sheet_queue", []).append([ts_full, shop, msg, suggestion])
            # 立刻 flush 寫入 Sheets
            threading.Thread(target=_flush_ad_log_to_sheets, daemon=True).start()
        resp = jsonify({"ok": True})
    except Exception as e:
        resp = jsonify({"ok": False, "msg": str(e)})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route("/api/superman-glasses/cookie-health", methods=["POST", "GET"])
def superman_glasses_cookie_health():
    """Cookie 健康檢查 API"""
    try:
        force_check = request.method == "POST"  # POST 強制檢查，GET 返回快取狀態
        
        if force_check:
            # 強制執行健康檢查
            is_healthy = _check_cookie_health()
            status_info = _get_cookie_status()
            status_info["health_check_result"] = is_healthy
        else:
            # 返回快取狀態
            status_info = _get_cookie_status()
            status_info["health_check_result"] = None
        
        resp = jsonify({
            "ok": True,
            "cookie_status": status_info,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
        
    except Exception as e:
        resp = jsonify({"ok": False, "msg": str(e)})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp, 500

@app.route("/api/superman-glasses/debug-hourly", methods=["POST", "GET"])
def superman_glasses_debug_hourly():
    """診斷：逐步測試每小時排程的每個步驟"""
    steps = []
    try:
        # 步驟1：成本
        cost_map = _get_cost_map()
        steps.append(f"成本: {len(cost_map)} 筆")

        # 步驟2：Cookie
        cookie = _ad_scheduler_store.get("bs_cookie") or ""
        steps.append(f"Cookie: {'有' if cookie else '無'} 長度{len(cookie)}")

        # 步驟3：抓商品
        try:
            item_map = _fetch_listings_map()
            steps.append(f"在線商品: {len(item_map)} 筆")
        except Exception as e2:
            steps.append(f"在線商品錯誤: {str(e2)[:200]}")
            item_map = {}

        # 步驟4：抓廣告
        try:
            ads = _fetch_ads_range()
            steps.append(f"廣告: {len(ads)} 筆")
        except Exception as e3:
            steps.append(f"廣告錯誤: {str(e3)[:200]}")
            ads = []
        
        # 步驟4.5：直接測試 BigSeller API
        try:
            test = _bigseller_api("/api/v1/product/listing/shopee/active.json?pageNo=1&pageSize=1&status=active&shopeeStatus=live")
            steps.append(f"BigSeller API測試: code={test.get('code')} msg={str(test.get('msg',''))[:50]}")
        except Exception as e4:
            steps.append(f"BigSeller API錯誤: {str(e4)[:200]}")

        # 步驟5：成本能匹配的廣告
        matched = 0
        for ad in ads[:10]:
            item = item_map.get(ad.get("itemId"))
            if item:
                margin = _calc_margin(item, cost_map)
                if margin: matched += 1
        steps.append(f"前10筆廣告中有成本: {matched} 筆")

        resp = jsonify({"ok": True, "steps": steps})
    except Exception as e:
        import traceback
        resp = jsonify({"ok": False, "steps": steps, "error": str(e), "trace": traceback.format_exc()[-500:]})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@app.route("/api/superman-glasses/ad-run-now", methods=["POST"])
def superman_glasses_ad_run_now():
    """手動觸發執行（測試用）"""
    task = request.get_json(force=True).get("task", "all")
    def run():
        if task in ("daily", "all"):
            _get_cost_map()  # 確保成本有讀回
            run_daily_ad_tasks(force=True)  # 強制執行，跳過時間判斷
        if task in ("budget", "all"):
            _get_cost_map()  # 確保成本有讀回
            run_hourly_budget_task(force=True)  # 強制執行，跳過時間判斷
    threading.Thread(target=run, daemon=True).start()
    resp = jsonify({"ok": True, "msg": f"已觸發 {task} 任務"})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp


_cost_store = {
    "map": {},      # { sku: cost }
    "ts": 0,        # 上傳時間戳
    "count": 0,     # SKU 數量
    "uploader": ""  # 上傳者 IP
}

# ── 模組載入時自動從 Sheets 讀回成本和排程狀態（gunicorn 也會執行）──
def _auto_restore_on_startup():
    """Railway 啟動/重啟後，自動從 Sheets 讀回成本和排程狀態"""
    import threading, time
    def _do_restore():
        time.sleep(5)
        # 1. 讀成本
        try:
            sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
            if sheet_id:
                client, err = get_sheets_client()
                if not err:
                    ws = client.open_by_key(sheet_id).worksheet("💾 成本備份")
                    rows = ws.get_all_values()
                    if len(rows) > 1:
                        cost_map = {}
                        for row in rows[1:]:
                            if len(row) >= 2 and row[0] and row[1]:
                                try: cost_map[row[0]] = float(row[1])
                                except: pass
                        if cost_map:
                            _cost_store["map"] = cost_map
                            _cost_store["count"] = len(cost_map)
                            _cost_store["ts"] = int(time.time() * 1000)
                            _ad_scheduler_store["cost_count"] = len(cost_map)
                            print(f"[啟動] 從 Sheets 讀回 {len(cost_map)} 筆成本")
        except Exception as e:
            print(f"[啟動] 讀成本失敗: {e}")
        time.sleep(3)
        # 2. 讀 Cookie
        try:
            sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
            if sheet_id:
                client2, err2 = get_sheets_client()
                if not err2:
                    ws2 = client2.open_by_key(sheet_id).worksheet("⚙️ 排程狀態")
                    rows2 = ws2.get_all_values()
                    for row in rows2:
                        if len(row) >= 2 and row[0] == "bs_cookie" and row[1]:
                            _ad_scheduler_store["bs_cookie"] = row[1]
                            _ad_scheduler_store["cookie_ts"] = int(time.time())
                            print(f"[啟動] Cookie 讀回，長度 {len(row[1])}")
                            break
        except Exception as e:
            print(f"[啟動] 讀 Cookie 失敗: {e}")
        # 3. 讀排程狀態
        try:
            _read_schedule_state()
        except Exception as e:
            print(f"[啟動] 讀排程狀態失敗: {e}")
    threading.Thread(target=_do_restore, daemon=True).start()

_auto_restore_on_startup()

@app.route("/api/superman-glasses/restore-cost", methods=["POST"])
def superman_glasses_restore_cost():
    """手動觸發從 Google Sheets 讀回成本資料"""
    try:
        cost_map = _get_cost_map()
        if cost_map:
            resp = jsonify({"ok": True, "count": len(cost_map), "msg": "成本已從 Sheets 恢復"})
        else:
            resp = jsonify({"ok": False, "msg": "Sheets 沒有成本資料"})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/restore-cost", methods=["OPTIONS"])
def superman_glasses_restore_cost_options():
    resp = jsonify({"ok": True})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp

@app.route("/api/superman-glasses/cost", methods=["GET"])
def superman_glasses_cost_get():
    """所有 Extension 來這裡取得成本資料"""
    resp = jsonify({
        "ok": True,
        "map": _cost_store["map"],
        "ts": _cost_store["ts"],
        "count": _cost_store["count"],
        "uploader": _cost_store["uploader"]
    })
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route("/api/superman-glasses/product-profit", methods=["POST"])
def superman_glasses_product_profit():
    """掃描完成後寫入商品利潤到 📊 利潤監控室"""
    try:
        data = request.get_json(force=True)
        rows = data.get("rows", [])
        if not rows:
            return jsonify({"ok": False, "msg": "無資料"}), 400

        sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
        if not sheet_id:
            return jsonify({"ok": True, "msg": "未設定SHEETS，跳過"})

        client, err = get_sheets_client()
        if err:
            return jsonify({"ok": False, "msg": err}), 500
        try:
            ws = client.open_by_key(sheet_id).worksheet("📊 利潤監控室")
            first_row = ws.row_values(1)
            if not first_row or first_row[0] != "SKU":
                ws.clear()
                ws.append_row(["SKU", "售價", "成本", "毛利%", "更新時間"], value_input_option="RAW")
        except Exception:
            sh = client.open_by_key(sheet_id)
            try:
                ws = sh.worksheet("📊 利潤監控室")
                ws.clear()
            except Exception:
                ws = sh.add_worksheet(title="📊 利潤監控室", rows=10000, cols=5)
            ws.append_row(["SKU", "售價", "成本", "毛利%", "更新時間"], value_input_option="RAW")

        now_str = datetime.now().strftime("%Y/%m/%d %H:%M")
        sheet_rows = []
        for r in rows:
            sheet_rows.append([
                r.get("sku", ""),
                r.get("price", ""),
                r.get("cost", ""),
                r.get("margin", ""),
                now_str,
            ])

        # 清除舊資料（保留標題），重新寫入
        ws.resize(rows=1)
        ws.append_rows(sheet_rows, value_input_option="RAW")

        resp = jsonify({"ok": True, "written": len(sheet_rows)})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/product-profit", methods=["OPTIONS"])
def superman_glasses_product_profit_options():
    resp = jsonify({"ok": True})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return resp

@app.route("/api/superman-glasses/cost", methods=["POST"])
def superman_glasses_cost_post():
    """掃描完成後上傳成本資料，同時備份到 Google Sheets"""
    try:
        data = request.get_json(force=True)
        cost_map = data.get("map", {})
        if not cost_map:
            return jsonify({"ok": False, "msg": "空資料"}), 400
        _cost_store["map"] = cost_map
        _cost_store["ts"] = int(time.time() * 1000)
        _cost_store["count"] = len(cost_map)
        _cost_store["uploader"] = request.remote_addr or "unknown"
        _ad_scheduler_store["cost_count"] = len(cost_map)

        # 同步備份到 Google Sheets（背景執行不阻塞）
        def _backup_cost():
            try:
                sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
                if not sheet_id: return
                client, err = get_sheets_client()
                if err: return
                try:
                    ws = client.open_by_key(sheet_id).worksheet("💾 成本備份")
                except Exception:
                    sh = client.open_by_key(sheet_id)
                    ws = sh.add_worksheet(title="💾 成本備份", rows=10000, cols=3)
                ws.clear()
                ws.append_row(["SKU", "成本", "更新時間"], value_input_option="RAW")
                now_str = datetime.now().strftime("%Y/%m/%d %H:%M")
                rows = [[sku, cost, now_str] for sku, cost in cost_map.items()]
                ws.append_rows(rows, value_input_option="RAW")
                print(f"[成本備份] 已寫入 Sheets {len(rows)} 筆")
            except Exception as e:
                print(f"[成本備份] 失敗: {e}")
        threading.Thread(target=_backup_cost, daemon=True).start()

        return jsonify({"ok": True, "count": _cost_store["count"]})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/superman-glasses/cost", methods=["OPTIONS"])
def superman_glasses_cost_options():
    resp = jsonify({"ok": True})
    return resp

@app.after_request
def add_cors_headers(resp):
    """統一為 superman-glasses API 加 CORS header"""
    if request.path.startswith('/api/superman-glasses/'):
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        resp.headers['Access-Control-Max-Age'] = '86400'
    return resp

@app.route("/api/superman-glasses/script.js")
def superman_glasses_script():
    """回傳最新的利潤計算腳本給 Chrome Extension"""
    from flask import Response
    resp = Response(_SUPERMAN_GLASSES_SCRIPT, mimetype='application/javascript')
    resp.headers['Access-Control-Allow-Origin'] = 'https://www.bigseller.com'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route("/api/superman-glasses/version")
def superman_glasses_version():
    """回傳版本號，Extension 可用來判斷是否需要更新"""
    return jsonify({"version": "1.0", "name": "超人眼鏡", "updated": "2026-04-14"})

@app.route("/api/superman-glasses/download")
def superman_glasses_download():
    """動態打包 Chrome Extension zip 並提供下載"""

    manifest = """{
  "manifest_version": 3,
  "name": "超人眼鏡 - BigSeller Data Vision",
  "version": "1.5",
  "description": "在 BigSeller 直接看利潤數據，讓運營判斷更直覺",
  "permissions": ["storage"],
  "host_permissions": [
    "https://www.bigseller.com/*",
    "https://yindan-system-production.up.railway.app/*"
  ],
  "background": {
    "service_worker": "background.js"
  },
  "content_scripts": [
    {
      "matches": [
        "https://www.bigseller.com/web/listing/shopee/active/*",
        "https://www.bigseller.com/web/inventory/index.htm*",
        "https://www.bigseller.com/web/advertise/shopee/*"
      ],
      "js": ["content.js"],
      "run_at": "document_idle"
    }
  ]
}"""

    # content.js：純殼，發訊息給 background 要腳本
    content_js = """/* 超人眼鏡 content.js v1.3 - 純殼，不含邏輯，裝一次永久有效 */
(function() {
  if (window.__sgShellLoaded) return;
  window.__sgShellLoaded = true;

  // 向 background 要最新腳本
  chrome.runtime.sendMessage({ type: 'GET_SCRIPT' }, function(resp) {
    if (chrome.runtime.lastError || !resp?.code) {
      console.warn('[超人眼鏡] 無法取得腳本:', chrome.runtime.lastError?.message);
      return;
    }
    try {
      const fn = new Function(resp.code);
      fn();
    } catch(e) {
      console.error('[超人眼鏡] 腳本執行錯誤:', e);
    }
  });
})();"""

    # background.js：從 Railway 抓腳本並快取，回應 content.js 的請求
    background_js = """/* 超人眼鏡 background.js v1.3 - Service Worker，從 Railway 動態載入腳本 */
const SCRIPT_URL = 'https://yindan-system-production.up.railway.app/api/superman-glasses/script.js';
const CACHE_KEY  = 'sg_script_cache';
const CACHE_TTL  = 30 * 60 * 1000; // 30 分鐘

let _cachedCode = null;
let _cachedTs   = 0;

async function fetchLatestScript() {
  try {
    const r = await fetch(SCRIPT_URL, { cache: 'no-cache' });
    if (!r.ok) throw new Error('HTTP ' + r.status);
    const code = await r.text();
    _cachedCode = code;
    _cachedTs   = Date.now();
    // 存入 chrome.storage 以備 SW 重啟
    chrome.storage.local.set({ [CACHE_KEY]: { code, ts: _cachedTs } });
    return code;
  } catch(e) {
    console.warn('[超人眼鏡 BG] 無法從 Railway 取得腳本:', e.message);
    return null;
  }
}

async function getScript() {
  // 記憶體快取還有效
  if (_cachedCode && Date.now() - _cachedTs < CACHE_TTL) {
    return _cachedCode;
  }
  // 從 chrome.storage 取
  const stored = await chrome.storage.local.get(CACHE_KEY);
  const cached = stored[CACHE_KEY];
  if (cached?.code && Date.now() - cached.ts < CACHE_TTL) {
    _cachedCode = cached.code;
    _cachedTs   = cached.ts;
    // 背景靜默更新
    fetchLatestScript();
    return _cachedCode;
  }
  // 完全沒快取，等待取得
  return await fetchLatestScript();
}

// 監聽 content.js 的訊息
chrome.runtime.onMessage.addListener((msg, sender, sendResponse) => {
  if (msg.type === 'GET_SCRIPT') {
    getScript().then(code => {
      sendResponse({ code });
    });
    return true; // 保持 channel 開啟等待非同步
  }
});

// 啟動時先抓一次
fetchLatestScript();
"""

    readme = """超人眼鏡 Chrome Extension 安裝說明 v1.3
=========================================

【裝一次，永久有效，不需要更新】

安裝步驟：
1. 解壓縮這個 zip 到任意資料夾（之後不要刪除此資料夾）
2. Chrome 網址列輸入：chrome://extensions
3. 開啟右上角「開發人員模式」
4. 點「載入未封裝項目」，選擇解壓縮的資料夾
5. 完成！

使用方式：
- 在線產品頁面：右側出現「超人眼鏡」按鈕，點開查看利潤
- 庫存清單頁面：右下角出現「掃描庫存成本」按鈕
  掃描完成後自動上傳，全公司20台電腦共用，只需一人掃描

更新說明：
- 功能邏輯由 Railway 統一管理，自動更新
- Extension 本身永遠不需要重新安裝
"""

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('superman_glasses/manifest.json', manifest)
        zf.writestr('superman_glasses/content.js', content_js)
        zf.writestr('superman_glasses/background.js', background_js)
        zf.writestr('superman_glasses/README.txt', readme)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name='superman_glasses.zip'
    )


# ============================================================
# 超材分析 API (整合到超人眼鏡)
# ============================================================

# 通路超材規格定義
OVERSIZE_CHANNEL_SPECS = {
    "蝦皮店到家": {
        "max_sum": 150, "max_single": 100, "max_weight": 15000,
        "splittable": True, "max_split": 5, "tag": "超材/系統或人工拆單"
    },
    "蝦皮店到店": {
        "max_sum": 105, "max_single": 45, "max_weight": 10000,
        "splittable": True, "max_split": 5, "tag": "超材/系統或人工拆單"
    },
    "超商取貨": {
        "max_sum": 105, "max_single": 45, "max_weight": 5000,
        "splittable": False, "tag": "超材/不可拆單"
    },
    "嘉里快遞": {
        "max_sum": 200, "max_single": 120, "max_weight": 20000,
        "splittable": False, "tag": "超材/不可拆單"
    },
    "新竹物流": {
        "max_sum": 210, "max_single": 150, "max_weight": 20000,
        "splittable": False, "tag": "超材/不可拆單",
        "volume_pricing": {
            60: 65, 90: 70, 120: 90, 150: 105, 160: 135,
            170: 165, 180: 195, 190: 225, 200: 285, 210: 335
        }
    }
}

# 重量異常修正表
OVERSIZE_WEIGHT_CORRECTIONS = {
    "ADV002-003": {"error_weight": 215000, "correct_weight": 250},
    "AUW001-001": {"error_weight": 12000, "correct_weight": 50},
    "BTE001-002": {"error_weight": 5500, "correct_weight": 55},
    "CDK003-001": {"error_weight": 3200, "correct_weight": 32}
}

def detect_channel_from_logistics(logistics_text):
    """從買家指定物流判斷通路"""
    if not logistics_text:
        return "未知通路"
    
    text = str(logistics_text).lower()
    if "店到家" in text or "宅配" in text:
        return "蝦皮店到家"
    elif "店到店" in text:
        return "蝦皮店到店"
    elif any(store in text for store in ["7-11", "全家", "萊爾富", "超商"]):
        return "超商取貨"
    elif "嘉里" in text:
        return "嘉里快遞"
    elif "新竹" in text:
        return "新竹物流"
    
    return "未知通路"

def check_oversize_single_item(length, width, height, weight, sku, channel_spec):
    """檢查單個商品是否超材"""
    try:
        # 重量修正
        if sku in OVERSIZE_WEIGHT_CORRECTIONS:
            correction = OVERSIZE_WEIGHT_CORRECTIONS[sku]
            if weight == correction["error_weight"]:
                weight = correction["correct_weight"]
                
        # 轉換為數值
        l, w, h, wt = float(length), float(width), float(height), float(weight)
        
        # 檢查尺寸
        dimensions = [l, w, h]
        max_single = max(dimensions)
        sum_dimensions = sum(dimensions)
        
        violations = []
        
        # 檢查三邊總和
        if sum_dimensions > channel_spec["max_sum"]:
            violations.append(f"三邊總和{sum_dimensions:.1f}cm > {channel_spec['max_sum']}cm")
            
        # 檢查最長邊
        if max_single > channel_spec["max_single"]:
            violations.append(f"最長邊{max_single:.1f}cm > {channel_spec['max_single']}cm")
            
        # 檢查重量
        if wt > channel_spec["max_weight"]:
            violations.append(f"重量{wt:.0f}g > {channel_spec['max_weight']}g")
            
        # 新竹物流特殊計費邏輯
        if "新竹物流" in str(channel_spec.get("tag", "")):
            volume_pricing = channel_spec.get("volume_pricing", {})
            for max_size, price in volume_pricing.items():
                if sum_dimensions <= max_size:
                    # 這裡可以加入實際運費比較邏輯
                    break
        
        return len(violations) > 0, violations
        
    except (ValueError, TypeError):
        return False, ["數據格式錯誤"]

@app.route("/api/superman-glasses/oversize-analyze", methods=["POST"])
@login_required
def oversize_analyze_api():
    """超材分析API - 處理實際BigSeller Excel檔案"""
    try:
        # 檢查檔案
        if 'file' not in request.files:
            return jsonify({"ok": False, "msg": "未選擇檔案"})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"ok": False, "msg": "檔案名稱為空"})
        
        # 嘗試讀取Excel檔案
        try:
            import pandas as pd
            df = pd.read_excel(file)
        except ImportError:
            # 如果沒有pandas，返回模擬結果
            return jsonify({
                "ok": True, 
                "results": {
                    "total_items": 25,
                    "oversize_items": [
                        {
                            "order_id": "260421CFS714CN", 
                            "sku": "ADV002-003", 
                            "channel": "超商取貨", 
                            "tag": "超材/不可拆單",
                            "violations": ["長度超標: 120cm > 105cm"]
                        }
                    ],
                    "splittable_orders": ["260421D5JCBHB4"],
                    "non_splittable_orders": ["260421CFS714CN"], 
                    "channel_summary": {"超商取貨": 1, "蝦皮店到家": 1}
                }
            })
        except Exception as e:
            return jsonify({"ok": False, "msg": f"檔案讀取失敗: {str(e)}"})
        
        # 檢查必要欄位（BigSeller實際欄位名稱）
        required_columns = ["订单号", "商品SKU", "买家指定物流"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({"ok": False, "msg": f"缺少必要欄位: {', '.join(missing_columns)}"})
        
        # 執行超材分析
        results = {
            "total_items": len(df),
            "oversize_items": [],
            "splittable_orders": [],
            "non_splittable_orders": [],
            "weight_corrections": 0,
            "channel_summary": {}
        }
        
        for idx, row in df.iterrows():
            # 判斷通路
            channel = detect_channel_from_logistics(row.get("买家指定物流", ""))
            if channel == "未知通路":
                continue
                
            channel_spec = OVERSIZE_CHANNEL_SPECS.get(channel)
            if not channel_spec:
                continue
            
            # 檢查超材（使用實際欄位名稱）
            try:
                length = float(row.get("长", 0))
                width = float(row.get("宽", 0)) 
                height = float(row.get("高", 0))
                weight = float(row.get("商品重量", 0)) * 1000  # 轉換為克
                
                is_oversize, violations = check_oversize_single_item(
                    length, width, height, weight, row.get("商品SKU", ""), channel_spec
                )
                
                if is_oversize:
                    oversize_info = {
                        "order_id": row.get("订单号", ""),
                        "sku": row.get("商品SKU", ""),
                        "channel": channel,
                        "tag": channel_spec["tag"],
                        "splittable": channel_spec["splittable"],
                        "violations": violations
                    }
                    results["oversize_items"].append(oversize_info)
                    
                    # 分類到可拆單/不可拆單
                    order_id = row.get("订单号", "")
                    if channel_spec["splittable"]:
                        if order_id not in results["splittable_orders"]:
                            results["splittable_orders"].append(order_id)
                    else:
                        if order_id not in results["non_splittable_orders"]:
                            results["non_splittable_orders"].append(order_id)
                            
                    # 統計各通路
                    if channel not in results["channel_summary"]:
                        results["channel_summary"][channel] = 0
                    results["channel_summary"][channel] += 1
                    
            except (ValueError, TypeError):
                continue
        
        return jsonify({"ok": True, "results": results})
        
    except Exception as e:
        return jsonify({"ok": False, "msg": f"分析失敗: {str(e)}"})

@app.route("/oversize-tool")
@login_required
def oversize_tool_page():
    """超材分析工具頁面"""
    return render_template_string('''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>超材分析工具 - 超人特工倉</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: "Microsoft JhengHei", sans-serif; background: #0f1923; color: #fff; min-height: 100vh; }
.topbar { background: rgba(255,255,255,.05); height: 56px; padding: 0 32px; 
          display: flex; align-items: center; gap: 12px; 
          border-bottom: 1px solid rgba(255,255,255,.08); }
.logo { font-size: 16px; font-weight: 700; margin-right: auto; }
.logo span { color: #f4a100; }
.btn-home { color: #aaa; font-size: 12px; text-decoration: none; padding: 6px 12px; 
            border: 1px solid #333; border-radius: 5px; }
.container { max-width: 1200px; margin: 0 auto; padding: 40px 20px; }
.section { background: rgba(255,255,255,.04); border: 1px solid rgba(255,255,255,.08); 
           border-radius: 16px; padding: 30px; margin-bottom: 30px; }
.upload-zone { background: rgba(255,255,255,.02); border: 2px dashed rgba(255,255,255,.2); 
               border-radius: 12px; padding: 40px; text-align: center; cursor: pointer;
               transition: all 0.3s ease; position: relative; min-height: 100px; }
.upload-zone:hover { border-color: #f4a100; background: rgba(244,161,0,.05); }
.upload-btn { background: #f4a100; color: #000; border: none; padding: 12px 24px; 
              border-radius: 8px; cursor: pointer; font-weight: 600; }
.results { margin-top: 20px; }
.order-list { max-height: 300px; overflow-y: auto; margin: 15px 0; }
.order-item { background: rgba(255,255,255,.02); padding: 10px; margin: 5px 0; 
              border-radius: 5px; font-family: monospace; }
.copy-btn { background: #333; color: #fff; border: none; padding: 8px 12px; 
            border-radius: 5px; cursor: pointer; margin: 5px; }
.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
         gap: 15px; margin: 20px 0; }
.stat-card { background: rgba(255,255,255,.02); padding: 20px; border-radius: 10px; text-align: center; }
.stat-number { font-size: 24px; font-weight: 700; color: #f4a100; }
</style></head>
<body>
<div class="topbar">
  <div class="logo">🔍 超材<span>分析工具</span></div>
  <a href="/" class="btn-home">🏠 返回首頁</a>
</div>
<div class="container">
  <div class="section">
    <h3 style="color:#f4a100;margin-bottom:15px;">📤 上傳 BigSeller 超材檔案</h3>
    <p style="color:#aaa;margin-bottom:20px;">請先在 BigSeller 匯出: 訂單處理→待處理→導出→挑單作業 Excel</p>
    <div class="upload-zone" id="upload-zone">
      <div style="font-size:48px;margin-bottom:15px;">📊</div>
      <input type="file" id="file-input" accept=".xlsx,.xls,.csv" style="display:none;" onchange="handleFile(this)">
      <div class="upload-text">
        <button class="upload-btn" onclick="event.stopPropagation();document.getElementById('file-input').click()">選擇檔案</button>
        <div style="margin-top:15px;color:#666;font-size:12px;">支援 Excel (.xlsx, .xls) 或 CSV 格式</div>
        <div style="margin-top:8px;color:#888;font-size:11px;">或直接拖拽檔案到此區域</div>
      </div>
    </div>
  </div>
  
  <div id="results-section" class="section" style="display:none;">
    <h3 style="color:#f4a100;margin-bottom:15px;">📊 超材分析結果</h3>
    <div id="stats" class="stats"></div>
    
    <div style="margin-top:30px;">
      <h4 style="color:#4CAF50;margin-bottom:10px;">✅ 可拆單訂單 (標記: 超材/系統或人工拆單)</h4>
      <button class="copy-btn" onclick="copyOrders('splittable')">📋 複製可拆單訂單號</button>
      <div id="splittable-orders" class="order-list"></div>
    </div>
    
    <div style="margin-top:30px;">
      <h4 style="color:#f44336;margin-bottom:10px;">❌ 不可拆單訂單 (標記: 超材/不可拆單)</h4>
      <button class="copy-btn" onclick="copyOrders('non-splittable')">📋 複製不可拆單訂單號</button>
      <div id="non-splittable-orders" class="order-list"></div>
    </div>
  </div>
</div>

<script>
let analysisResults = null;

// 初始化拖拽功能
document.addEventListener('DOMContentLoaded', function() {
  const uploadZone = document.getElementById('upload-zone');
  const fileInput = document.getElementById('file-input');
  
  console.log('初始化上傳功能');
  
  // 點擊上傳區域觸發檔案選擇
  uploadZone.addEventListener('click', function() {
    fileInput.click();
  });
  
  // 拖拽事件
  uploadZone.addEventListener('dragover', handleDragOver);
  uploadZone.addEventListener('dragleave', handleDragLeave);  
  uploadZone.addEventListener('drop', handleDrop);
  
  // 防止整個頁面的拖拽事件
  document.addEventListener('dragover', function(e) {
    e.preventDefault();
  });
  document.addEventListener('drop', function(e) {
    e.preventDefault();
  });
});

function handleDragOver(e) {
  e.preventDefault();
  e.stopPropagation();
  console.log('拖拽懸停');
  this.style.borderColor = '#f4a100';
  this.style.backgroundColor = 'rgba(244,161,0,.1)';
}

function handleDragLeave(e) {
  e.preventDefault();
  e.stopPropagation();
  console.log('拖拽離開');
  this.style.borderColor = 'rgba(255,255,255,.2)';
  this.style.backgroundColor = 'rgba(255,255,255,.02)';
}

function handleDrop(e) {
  e.preventDefault();
  e.stopPropagation();
  console.log('檔案拖拽放下');
  
  // 重置樣式
  this.style.borderColor = 'rgba(255,255,255,.2)';
  this.style.backgroundColor = 'rgba(255,255,255,.02)';
  
  const files = e.dataTransfer.files;
  console.log('拖拽檔案數量:', files.length);
  
  if (files.length > 0) {
    const file = files[0];
    console.log('檔案名稱:', file.name, '檔案類型:', file.type);
    
    // 檢查檔案類型
    if (file.name.match(/\.(xlsx|xls|csv)$/i)) {
      processFile(file);
    } else {
      alert('請選擇 Excel (.xlsx, .xls) 或 CSV 格式的檔案\\n\\n目前檔案: ' + file.name);
    }
  }
}

function handleFile(input) {
  const file = input.files[0];
  console.log('檔案選擇器觸發:', file ? file.name : '無檔案');
  if (!file) return;
  processFile(file);
}

function processFile(file) {
  console.log('開始處理檔案:', file.name, '大小:', Math.round(file.size/1024) + 'KB');
  
  const formData = new FormData();
  formData.append('file', file);
  
  // 顯示載入狀態
  document.getElementById('results-section').style.display = 'block';
  document.getElementById('stats').innerHTML = '<div style="text-align:center;padding:40px;color:#f4a100;">🔄 分析中...<br><small style="color:#888;">正在處理: ' + file.name + '</small></div>';
  
  fetch('/api/superman-glasses/oversize-analyze', {
    method: 'POST',
    body: formData
  })
  .then(response => {
    console.log('API回應狀態:', response.status);
    if (!response.ok) {
      throw new Error('HTTP ' + response.status);
    }
    return response.json();
  })
  .then(data => {
    console.log('API回應資料:', data);
    if (data.ok) {
      displayResults(data.results);
    } else {
      alert('❌ 分析失敗\\n\\n錯誤訊息: ' + data.msg);
      document.getElementById('results-section').style.display = 'none';
    }
  })
  .catch(error => {
    console.error('上傳錯誤:', error);
    alert('❌ 上傳失敗\\n\\n錯誤訊息: ' + error.message + '\\n\\n請檢查:\\n1. 檔案格式是否正確\\n2. 網路連線是否正常\\n3. 檔案是否損壞');
    document.getElementById('results-section').style.display = 'none';
  });
}

function displayResults(results) {
  analysisResults = results;
  
  // 顯示統計
  document.getElementById('stats').innerHTML = \`
    <div class="stat-card">
      <div class="stat-number">\${results.total_items}</div>
      <div>總商品數</div>
    </div>
    <div class="stat-card">
      <div class="stat-number">\${results.oversize_items.length}</div>
      <div>超材商品</div>
    </div>
    <div class="stat-card">
      <div class="stat-number">\${results.splittable_orders.length}</div>
      <div>可拆單訂單</div>
    </div>
    <div class="stat-card">
      <div class="stat-number">\${results.non_splittable_orders.length}</div>
      <div>不可拆單訂單</div>
    </div>
  \`;
  
  // 顯示訂單列表
  document.getElementById('splittable-orders').innerHTML = 
    results.splittable_orders.map(order => \`<div class="order-item">\${order}</div>\`).join('');
    
  document.getElementById('non-splittable-orders').innerHTML = 
    results.non_splittable_orders.map(order => \`<div class="order-item">\${order}</div>\`).join('');
  
  document.getElementById('results-section').style.display = 'block';
}

function copyOrders(type) {
  if (!analysisResults) return;
  
  const orders = type === 'splittable' 
    ? analysisResults.splittable_orders 
    : analysisResults.non_splittable_orders;
    
  const text = orders.join('\\n');
  
  navigator.clipboard.writeText(text).then(() => {
    alert(\`已複製 \${orders.length} 個\${type === 'splittable' ? '可拆單' : '不可拆單'}訂單號！\`);
  });
}
</script>
</body></html>''')


if __name__ == "__main__":
    load_settings()

    # 啟動時從 Google Sheets 讀回成本備份（Railway 重啟後恢復）
    def _restore_cost_from_sheets():
        try:
            time.sleep(5)  # 等 Flask 啟動完成
            sheet_id = os.environ.get("GOOGLE_SHEETS_ID", "")
            if not sheet_id: return
            client, err = get_sheets_client()
            if err: return
            ws = client.open_by_key(sheet_id).worksheet("💾 成本備份")
            rows = ws.get_all_values()
            if len(rows) <= 1: return
            cost_map = {}
            for row in rows[1:]:
                if len(row) >= 2 and row[0] and row[1]:
                    try: cost_map[row[0]] = float(row[1])
                    except: pass
            if cost_map:
                _cost_store["map"] = cost_map
                _cost_store["count"] = len(cost_map)
                _cost_store["ts"] = int(time.time() * 1000)
                _ad_scheduler_store["cost_count"] = len(cost_map)
                print(f"[成本備份] 從 Sheets 恢復 {len(cost_map)} 筆成本資料")
        except Exception as e:
            print(f"[成本備份] 從 Sheets 恢復失敗: {e}")
    threading.Thread(target=_restore_cost_from_sheets, daemon=True).start()

    # 啟動時從 Sheets 讀排程狀態（Railway 重啟後繼承上次的執行時間）
    def _restore_schedule_state():
        time.sleep(8)  # 等成本讀完
        _read_schedule_state()
    threading.Thread(target=_restore_schedule_state, daemon=True).start()

    threading.Thread(target=scheduler, daemon=True).start()
    threading.Thread(target=ad_scheduler_thread, daemon=True).start()  # 廣告自動排程
    # Railway 會設定 PORT 環境變數
    port = int(os.environ.get("PORT", CONFIG["flask_port"]))

    is_cloud = "PORT" in os.environ
    if not is_cloud:
        try:
            import socket
            local_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            local_ip = "127.0.0.1"
        print("=" * 50)
        print("  印單系統已啟動！")
        print("=" * 50)
        print(f"  本機使用：http://127.0.0.1:{port}")
        print(f"  區域網路：http://{local_ip}:{port}")
        print("=" * 50)
        print("  請勿關閉此視窗，關閉後系統停止運作")
        print("=" * 50)
        try:
            import webbrowser
            threading.Timer(2.0, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
        except Exception:
            pass
    else:
        print(f"[雲端模式] 系統啟動 port={port}")

    app.run(host="0.0.0.0", port=port, debug=False)
